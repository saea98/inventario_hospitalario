"""
Reporte de inventario por ubicación (LoteUbicacion).

Lista existencias desglosadas por ubicación física dentro del almacén, con filtros y Excel.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Sum, F, Value, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import LoteUbicacion, Almacen, UbicacionAlmacen, Institucion, Lote


def _parse_decimal_param(raw):
    """Devuelve Decimal o None si vacío / inválido (acepta coma como separador)."""
    s = (raw or "").strip().replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _filtrar_lote_ubicacion_queryset(request):
    """
    Queryset base: filas LoteUbicacion con relaciones cargadas.
    Respeta GET: almacen_id, ubicacion_id, institucion_id, q, estado_lote, incluir_cero,
    precio_min, precio_max (rango inclusivo sobre precio unitario del lote),
    precio_mayor_que (precio unitario estrictamente mayor a este valor; AND con el rango).
    Los lotes sin precio se tratan como 0 para estos filtros.
    """
    qs = (
        LoteUbicacion.objects.select_related(
            "lote",
            "lote__producto",
            "lote__institucion",
            "ubicacion",
            "ubicacion__almacen",
        )
        .filter(ubicacion__activo=True)
    )

    if request.GET.get("incluir_cero") != "1":
        qs = qs.filter(cantidad__gt=0)

    almacen_id = (request.GET.get("almacen_id") or "").strip()
    if almacen_id.isdigit():
        qs = qs.filter(ubicacion__almacen_id=int(almacen_id))

    ubicacion_id = (request.GET.get("ubicacion_id") or "").strip()
    if ubicacion_id.isdigit():
        qs = qs.filter(ubicacion_id=int(ubicacion_id))

    institucion_id = (request.GET.get("institucion_id") or "").strip()
    if institucion_id.isdigit():
        qs = qs.filter(lote__institucion_id=int(institucion_id))

    estado_lote = (request.GET.get("estado_lote") or "").strip()
    if estado_lote.isdigit():
        qs = qs.filter(lote__estado=int(estado_lote))

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(lote__numero_lote__icontains=q)
            | Q(lote__producto__clave_cnis__icontains=q)
            | Q(lote__producto__descripcion__icontains=q)
            | Q(ubicacion__codigo__icontains=q)
            | Q(ubicacion__descripcion__icontains=q)
        )

    precio_min = _parse_decimal_param(request.GET.get("precio_min"))
    precio_max = _parse_decimal_param(request.GET.get("precio_max"))
    precio_mayor_que = _parse_decimal_param(request.GET.get("precio_mayor_que"))

    if precio_min is not None or precio_max is not None or precio_mayor_que is not None:
        qs = qs.annotate(
            precio_eff=Coalesce(
                F("lote__precio_unitario"),
                Value(Decimal("0")),
                output_field=DecimalField(max_digits=24, decimal_places=4),
            )
        )
        if precio_min is not None:
            qs = qs.filter(precio_eff__gte=precio_min)
        if precio_max is not None:
            qs = qs.filter(precio_eff__lte=precio_max)
        if precio_mayor_que is not None:
            qs = qs.filter(precio_eff__gt=precio_mayor_que)

    return qs.order_by(
        "ubicacion__almacen__nombre",
        "ubicacion__codigo",
        "lote__producto__clave_cnis",
        "lote__numero_lote",
    )


def _fila_dict(lu):
    """Una fila para template / Excel a partir de un LoteUbicacion."""
    lote = lu.lote
    prod = lote.producto
    ubi = lu.ubicacion
    alm = ubi.almacen
    inst = lote.institucion
    precio = lote.precio_unitario or Decimal("0")
    cant = lu.cantidad
    try:
        valor = (Decimal(str(cant)) * precio).quantize(Decimal("0.01"))
    except Exception:
        valor = Decimal("0")

    return {
        "lu_id": lu.id,
        "lote_id": lote.id,
        "almacen": alm.nombre if alm else "—",
        "ubicacion_codigo": ubi.codigo,
        "ubicacion_desc": (ubi.descripcion or "")[:200],
        "ubicacion_estado": ubi.get_estado_display() if hasattr(ubi, "get_estado_display") else ubi.estado,
        "institucion_clue": inst.clue if inst else "—",
        "institucion": inst.denominacion if inst else "—",
        "clave_cnis": prod.clave_cnis if prod else "—",
        "producto": prod.descripcion if prod else "—",
        "numero_lote": lote.numero_lote,
        "cantidad": cant,
        "cantidad_reservada": lu.cantidad_reservada,
        "estado_lote": lote.get_estado_display() if hasattr(lote, "get_estado_display") else str(lote.estado),
        "fecha_caducidad": lote.fecha_caducidad.strftime("%d/%m/%Y") if lote.fecha_caducidad else "—",
        "remision": (lote.remision or "").strip() or "—",
        "precio_unitario": precio,
        "valor_linea": valor,
    }


@login_required
def reporte_por_ubicacion(request):
    qs = _filtrar_lote_ubicacion_queryset(request)
    total_registros = qs.count()

    precio_nz = Coalesce(
        F("lote__precio_unitario"),
        Value(Decimal("0")),
        output_field=DecimalField(max_digits=24, decimal_places=4),
    )
    agg = qs.annotate(
        valor_linea=ExpressionWrapper(
            F("cantidad") * precio_nz,
            output_field=DecimalField(max_digits=24, decimal_places=2),
        )
    ).aggregate(
        sum_cant=Sum("cantidad"),
        sum_valor=Sum("valor_linea"),
    )
    suma_cantidad = agg["sum_cant"] or 0
    valor_total = agg["sum_valor"] or Decimal("0")

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    filas = [_fila_dict(lu) for lu in page_obj.object_list]

    qcopy = request.GET.copy()
    qcopy.pop("page", None)
    querystring = qcopy.urlencode()

    context = {
        "page_obj": page_obj,
        "filas": filas,
        "total_registros": total_registros,
        "suma_cantidad": suma_cantidad,
        "valor_total": valor_total,
        "almacenes": Almacen.objects.all().order_by("nombre"),
        "ubicaciones": UbicacionAlmacen.objects.select_related("almacen")
        .filter(activo=True)
        .order_by("almacen__nombre", "codigo"),
        "instituciones": Institucion.objects.filter(activo=True).order_by("denominacion"),
        "estados_lote": Lote.ESTADOS_CHOICES,
        # Preservar filtros en formulario
        "filtro_almacen_id": request.GET.get("almacen_id", ""),
        "filtro_ubicacion_id": request.GET.get("ubicacion_id", ""),
        "filtro_institucion_id": request.GET.get("institucion_id", ""),
        "filtro_estado_lote": request.GET.get("estado_lote", ""),
        "filtro_q": request.GET.get("q", ""),
        "filtro_incluir_cero": request.GET.get("incluir_cero") == "1",
        "filtro_precio_min": request.GET.get("precio_min", ""),
        "filtro_precio_max": request.GET.get("precio_max", ""),
        "filtro_precio_mayor_que": request.GET.get("precio_mayor_que", ""),
        "querystring": querystring,
    }
    return render(request, "inventario/reporte_por_ubicacion.html", context)


@login_required
def exportar_por_ubicacion_excel(request):
    qs = _filtrar_lote_ubicacion_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Por ubicación"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "ALMACÉN",
        "CÓDIGO UBICACIÓN",
        "DESCRIPCIÓN UBICACIÓN",
        "ESTADO UBICACIÓN",
        "CLUES",
        "INSTITUCIÓN",
        "CLAVE CNIS",
        "DESCRIPCIÓN PRODUCTO",
        "LOTE",
        "CANTIDAD",
        "RESERVADA",
        "ESTADO LOTE",
        "CADUCIDAD",
        "REMISIÓN",
        "PRECIO UNIT.",
        "VALOR LÍNEA",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row_num = 2
    for lu in qs.iterator(chunk_size=500):
        d = _fila_dict(lu)
        vals = [
            d["almacen"],
            d["ubicacion_codigo"],
            d["ubicacion_desc"],
            d["ubicacion_estado"],
            d["institucion_clue"],
            d["institucion"],
            d["clave_cnis"],
            d["producto"],
            d["numero_lote"],
            d["cantidad"],
            d["cantidad_reservada"],
            d["estado_lote"],
            d["fecha_caducidad"],
            d["remision"],
            float(d["precio_unitario"]),
            float(d["valor_linea"]),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = border
        row_num += 1

    # Anchos aproximados
    widths = [18, 14, 24, 12, 12, 28, 14, 36, 16, 10, 10, 14, 12, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = "inventario_por_ubicacion.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response
