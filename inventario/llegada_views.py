"""
Vistas para la Fase 2.2.2: Llegada de Proveedores
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import transaction, models
from django.db.models import Q
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os
from django.conf import settings
from .access_control import requiere_rol

from .llegada_models import LlegadaProveedor, ItemLlegada, DocumentoLlegada
from .models import Almacen
from .llegada_forms import (
    LlegadaProveedorForm,
    ItemLlegadaFormSet,
    ControlCalidadForm,
    FacturacionForm,
    ItemFacturacionFormSet,
    SupervisionForm,
    UbicacionFormSet,
    DocumentoLlegadaForm,
)


class ListaLlegadasView(LoginRequiredMixin, View):
    """Muestra la lista de llegadas de proveedores con filtros y paginación"""
    
    def get(self, request):
        llegadas = LlegadaProveedor.objects.select_related('proveedor', 'cita').all()
        
        # Filtros de búsqueda
        folio = request.GET.get('folio', '').strip()
        proveedor = request.GET.get('proveedor', '').strip()
        fecha_inicio = request.GET.get('fecha_inicio', '').strip()
        fecha_fin = request.GET.get('fecha_fin', '').strip()
        estado = request.GET.get('estado', '').strip()
        orden_suministro = request.GET.get('orden_suministro', '').strip()
        remision = request.GET.get('remision', '').strip()
        
        # Aplicar filtros
        if folio:
            llegadas = llegadas.filter(folio__icontains=folio)
        
        if proveedor:
            llegadas = llegadas.filter(
                Q(proveedor__razon_social__icontains=proveedor) |
                Q(proveedor__rfc__icontains=proveedor)
            )
        
        # Filtro de rango de fechas
        if fecha_inicio:
            try:
                from datetime import datetime
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                llegadas = llegadas.filter(cita__fecha_cita__date__gte=fecha_inicio_obj)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                from datetime import datetime, time as dt_time
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                # Incluir todo el día
                fecha_fin_obj = datetime.combine(fecha_fin_obj, dt_time.max)
                llegadas = llegadas.filter(cita__fecha_cita__lte=fecha_fin_obj)
            except ValueError:
                pass
        
        if estado:
            llegadas = llegadas.filter(estado=estado)
        
        if orden_suministro:
            llegadas = llegadas.filter(
                Q(numero_orden_suministro__icontains=orden_suministro) |
                Q(cita__numero_orden_suministro__icontains=orden_suministro)
            )
        
        if remision:
            llegadas = llegadas.filter(
                Q(remision__icontains=remision) |
                Q(cita__numero_orden_remision__icontains=remision)
            )
        
        # Ordenar por fecha de creación descendente
        llegadas = llegadas.order_by('-fecha_creacion')
        
        # Paginación
        paginator = Paginator(llegadas, 25)  # 25 registros por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Obtener opciones de estado para el select
        estados_choices = LlegadaProveedor.ESTADO_CHOICES
        
        context = {
            'page_obj': page_obj,
            'paginator': paginator,
            'folio': folio,
            'proveedor': proveedor,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'estado': estado,
            'orden_suministro': orden_suministro,
            'remision': remision,
            'estados_choices': estados_choices,
        }
        
        return render(request, "inventario/llegadas/lista_llegadas.html", context)


def exportar_llegadas_excel(request):
    """Exporta la lista de llegadas a Excel con campos de cita y llegada"""
    
    # Obtener llegadas con los mismos filtros que la vista de lista
    llegadas = LlegadaProveedor.objects.select_related('proveedor', 'cita', 'almacen', 
                                                        'usuario_calidad', 'usuario_facturacion', 
                                                        'usuario_supervision', 'usuario_ubicacion',
                                                        'cita__usuario_autorizacion').all()
    
    # Aplicar los mismos filtros que la vista de lista
    folio = request.GET.get('folio', '').strip()
    proveedor = request.GET.get('proveedor', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    estado = request.GET.get('estado', '').strip()
    orden_suministro = request.GET.get('orden_suministro', '').strip()
    remision = request.GET.get('remision', '').strip()
    
    if folio:
        llegadas = llegadas.filter(folio__icontains=folio)
    
    if proveedor:
        llegadas = llegadas.filter(
            Q(proveedor__razon_social__icontains=proveedor) |
            Q(proveedor__rfc__icontains=proveedor)
        )
    
    # Filtro de rango de fechas
    if fecha_inicio:
        try:
            from datetime import datetime
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            llegadas = llegadas.filter(cita__fecha_cita__date__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            from datetime import datetime, time as dt_time
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Incluir todo el día
            fecha_fin_obj = datetime.combine(fecha_fin_obj, dt_time.max)
            llegadas = llegadas.filter(cita__fecha_cita__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    if estado:
        llegadas = llegadas.filter(estado=estado)
    
    if orden_suministro:
        llegadas = llegadas.filter(
            Q(numero_orden_suministro__icontains=orden_suministro) |
            Q(cita__numero_orden_suministro__icontains=orden_suministro)
        )
    
    if remision:
        llegadas = llegadas.filter(
            Q(remision__icontains=remision) |
            Q(cita__numero_orden_remision__icontains=remision)
        )
    
    # Ordenar por fecha de creación descendente
    llegadas = llegadas.order_by('-fecha_creacion')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Llegadas de Proveedores'
    
    # Estilos
    header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados - Campos de Cita y Llegada
    headers = [
        # Campos de Cita
        'Folio Cita', 'Fecha Cita', 'Estado Cita', 'Tipo Entrega',
        'Orden Suministro (Cita)', 'Orden Remisión (Cita)', 'Contrato (Cita)', 'Clave Medicamento (Cita)',
        'Usuario Autorización Cita',
        # Campos de Llegada
        'Folio Llegada', 'Estado Llegada', 'Remisión',
        'Piezas Emitidas', 'Piezas Recibidas', 'Tipo Red', 'Almacén',
        'Orden Suministro (Llegada)', 'Contrato (Llegada)', 'Procedimiento',
        # Control de Calidad
        'Estado Calidad', 'Observaciones Calidad', 'Usuario Calidad', 'Fecha Validación Calidad',
        # Facturación
        'Número Factura', 'Usuario Facturación', 'Fecha Facturación',
        # Supervisión
        'Estado Supervisión', 'Observaciones Supervisión', 'Usuario Supervisión', 'Fecha Supervisión',
        # Ubicación
        'Usuario Ubicación', 'Fecha Ubicación',
        # Observaciones
        'Observaciones Recepción',
        # Proveedor
        'Proveedor', 'RFC Proveedor'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    row_num = 2
    for llegada in llegadas:
        col = 1
        
        # Campos de Cita
        ws.cell(row=row_num, column=col).value = llegada.cita.folio if llegada.cita and llegada.cita.folio else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.cita.fecha_cita.strftime('%d/%m/%Y %H:%M') if llegada.cita and llegada.cita.fecha_cita else ''
        col += 1
        ws.cell(row=row_num, column=col).value = dict(llegada.cita.ESTADOS_CITA).get(llegada.cita.estado, llegada.cita.estado) if llegada.cita else ''
        col += 1
        ws.cell(row=row_num, column=col).value = dict([(t[0], t[1]) for t in llegada.cita.TIPOS_ENTREGA]).get(llegada.cita.tipo_entrega, llegada.cita.tipo_entrega) if llegada.cita else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.cita.numero_orden_suministro if llegada.cita else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.cita.numero_orden_remision if llegada.cita else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.cita.numero_contrato if llegada.cita else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.cita.clave_medicamento if llegada.cita else ''
        col += 1
        # (Se elimina la columna 'Fecha Autorización Cita' del Excel)
        ws.cell(row=row_num, column=col).value = f"{llegada.cita.usuario_autorizacion.first_name} {llegada.cita.usuario_autorizacion.last_name}".strip() if llegada.cita and llegada.cita.usuario_autorizacion else ''
        col += 1
        
        # Campos de Llegada
        ws.cell(row=row_num, column=col).value = llegada.folio
        col += 1
        # (Se elimina la columna 'Fecha Llegada Real' del Excel)
        ws.cell(row=row_num, column=col).value = dict(LlegadaProveedor.ESTADO_CHOICES).get(llegada.estado, llegada.estado)
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.remision
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.numero_piezas_emitidas
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.numero_piezas_recibidas
        col += 1
        if llegada.tipo_red:
            tipo_red_display = dict([('FRIA', 'Red Fría'), ('SECA', 'Red Seca')]).get(llegada.tipo_red, llegada.tipo_red)
        else:
            tipo_red_display = ''
        ws.cell(row=row_num, column=col).value = tipo_red_display
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.almacen.nombre if llegada.almacen else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.numero_orden_suministro or ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.numero_contrato or ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.numero_procedimiento or ''
        col += 1
        
        # Control de Calidad
        ws.cell(row=row_num, column=col).value = dict([('APROBADO', 'Aprobado'), ('RECHAZADO', 'Rechazado')]).get(llegada.estado_calidad, llegada.estado_calidad or '')
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.observaciones_calidad or ''
        col += 1
        ws.cell(row=row_num, column=col).value = f"{llegada.usuario_calidad.first_name} {llegada.usuario_calidad.last_name}".strip() if llegada.usuario_calidad else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.fecha_validacion_calidad.strftime('%d/%m/%Y %H:%M') if llegada.fecha_validacion_calidad else ''
        col += 1
        
        # Facturación
        ws.cell(row=row_num, column=col).value = llegada.numero_factura or ''
        col += 1
        ws.cell(row=row_num, column=col).value = f"{llegada.usuario_facturacion.first_name} {llegada.usuario_facturacion.last_name}".strip() if llegada.usuario_facturacion else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.fecha_facturacion.strftime('%d/%m/%Y %H:%M') if llegada.fecha_facturacion else ''
        col += 1
        
        # Supervisión
        ws.cell(row=row_num, column=col).value = dict([('VALIDADO', 'Validado'), ('RECHAZADO', 'Rechazado')]).get(llegada.estado_supervision, llegada.estado_supervision or '')
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.observaciones_supervision or ''
        col += 1
        ws.cell(row=row_num, column=col).value = f"{llegada.usuario_supervision.first_name} {llegada.usuario_supervision.last_name}".strip() if llegada.usuario_supervision else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.fecha_supervision.strftime('%d/%m/%Y %H:%M') if llegada.fecha_supervision else ''
        col += 1
        
        # Ubicación
        ws.cell(row=row_num, column=col).value = f"{llegada.usuario_ubicacion.first_name} {llegada.usuario_ubicacion.last_name}".strip() if llegada.usuario_ubicacion else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.fecha_ubicacion.strftime('%d/%m/%Y %H:%M') if llegada.fecha_ubicacion else ''
        col += 1
        
        # Observaciones
        ws.cell(row=row_num, column=col).value = llegada.observaciones_recepcion or ''
        col += 1
        
        # Proveedor
        ws.cell(row=row_num, column=col).value = llegada.proveedor.razon_social if llegada.proveedor else ''
        col += 1
        ws.cell(row=row_num, column=col).value = llegada.proveedor.rfc if llegada.proveedor else ''
        
        # Aplicar bordes a toda la fila
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border
        
        row_num += 1
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 15, 'B': 18, 'C': 15, 'D': 15, 'E': 20, 'F': 20, 'G': 15, 'H': 20,
        'I': 18, 'J': 20, 'K': 15, 'L': 18, 'M': 15, 'N': 15, 'O': 12, 'P': 12,
        'Q': 12, 'R': 20, 'S': 20, 'T': 15, 'U': 15, 'V': 15, 'W': 25, 'X': 20,
        'Y': 18, 'Z': 15, 'AA': 20, 'AB': 18, 'AC': 15, 'AD': 25, 'AE': 20,
        'AF': 18, 'AG': 20, 'AH': 18, 'AI': 25, 'AJ': 30, 'AK': 15
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"llegadas_proveedores_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class CrearLlegadaView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Crea una nueva llegada de proveedor"""
    permission_required = 'inventario.add_llegadaproveedor'
    
    def get(self, request):
        form = LlegadaProveedorForm()
        formset = ItemLlegadaFormSet(prefix="items")
        
        # Debug: Verificar cuántas citas están disponibles
        import logging
        logger = logging.getLogger(__name__)
        citas_disponibles = form.fields['cita'].queryset
        total_citas = citas_disponibles.count()
        logger.info(f"[DEBUG CREAR_LLEGADA] Total citas disponibles: {total_citas}")
        
        # Obtener todas las citas autorizadas para comparar
        from django.apps import apps
        CitaProveedor = apps.get_model('inventario', 'CitaProveedor')
        LlegadaProveedor = apps.get_model('inventario', 'LlegadaProveedor')
        todas_autorizadas = CitaProveedor.objects.filter(estado='autorizada').count()
        logger.info(f"[DEBUG CREAR_LLEGADA] Total citas autorizadas: {todas_autorizadas}")
        
        # Verificar llegadas en estado EN_RECEPCION
        llegadas_en_recepcion = LlegadaProveedor.objects.filter(estado='EN_RECEPCION').count()
        logger.info(f"[DEBUG CREAR_LLEGADA] Total llegadas en EN_RECEPCION: {llegadas_en_recepcion}")
        
        # Listar todas las citas disponibles con detalles
        logger.info(f"[DEBUG CREAR_LLEGADA] === LISTA DE CITAS DISPONIBLES ===")
        for cita in citas_disponibles[:10]:  # Primeras 10 para debug
            tiene_llegada = False
            estado_llegada = None
            llegada_id = None
            try:
                if hasattr(cita, 'llegada_proveedor'):
                    try:
                        llegada = cita.llegada_proveedor
                        tiene_llegada = True
                        estado_llegada = llegada.estado
                        llegada_id = str(llegada.id)
                    except:
                        tiene_llegada = False
            except Exception as e:
                logger.error(f"[DEBUG CREAR_LLEGADA] Error al acceder a llegada_proveedor: {str(e)}")
            
            logger.info(f"[DEBUG CREAR_LLEGADA] Cita ID: {cita.id}, Folio: {cita.folio or 'Sin folio'}, Estado: {cita.estado}, Fecha: {cita.fecha_cita}, Tiene llegada: {tiene_llegada}, Estado llegada: {estado_llegada}, Llegada ID: {llegada_id}")
        
        # Verificar específicamente las citas con folios IB-2026-000031 e IB-2026-000032
        try:
            cita_31 = CitaProveedor.objects.filter(folio='IB-2026-000031').first()
            if cita_31:
                logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000031 - ID: {cita_31.id}, Estado: {cita_31.estado}")
                try:
                    llegada_31 = cita_31.llegada_proveedor
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000031 - Tiene llegada: True, Estado llegada: {llegada_31.estado}, Llegada ID: {llegada_31.id}")
                    esta_en_queryset_31 = citas_disponibles.filter(id=cita_31.id).exists()
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000031 - Está en queryset: {esta_en_queryset_31}")
                except:
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000031 - Tiene llegada: False")
                    esta_en_queryset_31 = citas_disponibles.filter(id=cita_31.id).exists()
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000031 - Está en queryset: {esta_en_queryset_31}")
            
            cita_32 = CitaProveedor.objects.filter(folio='IB-2026-000032').first()
            if cita_32:
                logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000032 - ID: {cita_32.id}, Estado: {cita_32.estado}")
                try:
                    llegada_32 = cita_32.llegada_proveedor
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000032 - Tiene llegada: True, Estado llegada: {llegada_32.estado}, Llegada ID: {llegada_32.id}")
                    esta_en_queryset_32 = citas_disponibles.filter(id=cita_32.id).exists()
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000032 - Está en queryset: {esta_en_queryset_32}")
                except:
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000032 - Tiene llegada: False")
                    esta_en_queryset_32 = citas_disponibles.filter(id=cita_32.id).exists()
                    logger.info(f"[DEBUG CREAR_LLEGADA] Cita IB-2026-000032 - Está en queryset: {esta_en_queryset_32}")
        except Exception as e:
            logger.error(f"[DEBUG CREAR_LLEGADA] Error al verificar citas específicas: {str(e)}")
        
        logger.info(f"[DEBUG CREAR_LLEGADA] === FIN LISTA ===")
        
        # Debug: Mostrar el SQL generado
        try:
            sql_query = str(citas_disponibles.query)
            logger.info(f"[DEBUG CREAR_LLEGADA] SQL Query generado: {sql_query}")
        except Exception as e:
            logger.error(f"[DEBUG CREAR_LLEGADA] Error al obtener SQL: {str(e)}")
        
        # Agregar información de debug al contexto para mostrarla en el template
        context = {
            "form": form,
            "formset": formset,
            "debug_total_citas": total_citas,
            "debug_citas_list": list(citas_disponibles.values('id', 'folio', 'estado')[:10])
        }
        
        return render(request, "inventario/llegadas/crear_llegada.html", context)
    
    def post(self, request):
        form = LlegadaProveedorForm(request.POST)
        formset = ItemLlegadaFormSet(request.POST, prefix="items")
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                cita = form.cleaned_data.get('cita')
                
                # Verificar si ya existe una llegada para esta cita en estado EN_RECEPCION
                if cita and hasattr(cita, 'llegada_proveedor') and cita.llegada_proveedor:
                    llegada = cita.llegada_proveedor
                    # Actualizar los datos del formulario
                    llegada.proveedor = form.cleaned_data.get('proveedor', llegada.proveedor)
                    llegada.remision = form.cleaned_data.get('remision', llegada.remision)
                    llegada.numero_piezas_emitidas = form.cleaned_data.get('numero_piezas_emitidas', llegada.numero_piezas_emitidas)
                    llegada.numero_piezas_recibidas = form.cleaned_data.get('numero_piezas_recibidas', llegada.numero_piezas_recibidas)
                    llegada.almacen = form.cleaned_data.get('almacen', llegada.almacen)
                    llegada.tipo_red = form.cleaned_data.get('tipo_red', llegada.tipo_red)
                    llegada.numero_orden_suministro = form.cleaned_data.get('numero_orden_suministro', llegada.numero_orden_suministro)
                    llegada.numero_contrato = form.cleaned_data.get('numero_contrato', llegada.numero_contrato)
                    llegada.numero_procedimiento = form.cleaned_data.get('numero_procedimiento', llegada.numero_procedimiento)
                    llegada.observaciones_recepcion = form.cleaned_data.get('observaciones_recepcion', llegada.observaciones_recepcion)
                    llegada.creado_por = request.user
                    llegada.save()
                    
                    formset.instance = llegada
                    formset.save()
                    
                    messages.success(request, f"Llegada {llegada.folio} actualizada exitosamente")
                else:
                    # Crear nueva llegada
                    llegada = form.save(commit=False)
                    llegada.creado_por = request.user
                    llegada.save()
                    
                    formset.instance = llegada
                    formset.save()
                    
                    messages.success(request, f"Llegada {llegada.folio} creada exitosamente")
                
                return redirect('logistica:llegadas:detalle_llegada', pk=llegada.pk)
        
        return render(request, "inventario/llegadas/crear_llegada.html", {"form": form, "formset": formset})


class EditarLlegadaView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Edita una llegada existente"""
    permission_required = 'inventario.change_llegadaproveedor'
    
    def get(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        if not llegada.puede_editar_recepcion():
            messages.error(request, "No se puede editar esta llegada en su estado actual")
            return redirect('logistica:llegadas:detalle_llegada', pk=pk)
        
        form = LlegadaProveedorForm(instance=llegada)
        
        # Si hay items existentes, no mostrar formularios extra (extra=0)
        # Si no hay items, mostrar 1 formulario vacío (extra=1)
        if llegada.items.exists():
            # Hay items existentes, solo mostrar esos (sin extra)
            from django.forms import inlineformset_factory
            from .llegada_forms import ItemLlegadaForm
            formset = inlineformset_factory(
                LlegadaProveedor,
                ItemLlegada,
                form=ItemLlegadaForm,
                extra=0,  # No mostrar formularios extra si hay items existentes
                can_delete=True,
                can_delete_extra=True,
                fk_name="llegada",
                min_num=1,
                validate_min=True,
            )(instance=llegada, prefix="items")
        else:
            # No hay items, mostrar solo 1 formulario vacío (extra=1)
            from django.forms import inlineformset_factory
            from .llegada_forms import ItemLlegadaForm
            formset = inlineformset_factory(
                LlegadaProveedor,
                ItemLlegada,
                form=ItemLlegadaForm,
                extra=1,  # Mostrar 1 formulario vacío cuando no hay items
                can_delete=True,
                can_delete_extra=True,
                fk_name="llegada",
                min_num=0,  # No requerir mínimo cuando no hay items (solo mostrar extra=1)
                validate_min=False,
            )(instance=llegada, prefix="items")
        
        return render(request, "inventario/llegadas/editar_llegada.html", {
            "form": form, 
            "formset": formset, 
            "llegada": llegada
        })
    
    def post(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        if not llegada.puede_editar_recepcion():
            messages.error(request, "No se puede editar esta llegada en su estado actual")
            return redirect('logistica:llegadas:detalle_llegada', pk=pk)
        
        form = LlegadaProveedorForm(request.POST, instance=llegada)
        formset = ItemLlegadaFormSet(request.POST, instance=llegada, prefix="items")
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.save()
                if 'tipo_entrega' in form.cleaned_data and llegada.cita_id:
                    llegada.cita.tipo_entrega = form.cleaned_data['tipo_entrega'] or llegada.cita.tipo_entrega
                    llegada.cita.save()
                formset.save()
                messages.success(request, f"Llegada {llegada.folio} actualizada exitosamente")
                return redirect('logistica:llegadas:detalle_llegada', pk=llegada.pk)
        
        return render(request, "inventario/llegadas/editar_llegada.html", {"form": form, "formset": formset, "llegada": llegada})


class AprobarEntradaView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Aprueba la entrada de una llegada"""
    permission_required = 'inventario.change_llegadaproveedor'
    
    def post(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        if not llegada.puede_editar_recepcion():
            messages.error(request, "No se puede aprobar esta llegada en su estado actual")
            return redirect('logistica:llegadas:detalle_llegada', pk=pk)
        
        # Verificar si se solicita inspección visual
        solicita_inspeccion_visual = request.POST.get('solicita_inspeccion_visual') == '1'
        
        if solicita_inspeccion_visual:
            # Si solicita inspección visual, pasar a CONTROL_CALIDAD
            llegada.estado = 'CONTROL_CALIDAD'
            llegada.save()
            messages.success(request, f"Llegada {llegada.folio} aprobada. Pendiente de validación de calidad (inspección visual)")
        else:
            # Si no solicita inspección visual, saltarse CONTROL_CALIDAD y pasar directamente a UBICACION
            llegada.estado = 'UBICACION'
            llegada.save()
            messages.success(request, f"Llegada {llegada.folio} aprobada. Saltando control de calidad, lista para asignar ubicación")
        
        return redirect('logistica:llegadas:detalle_llegada', pk=pk)


class DetalleLlegadaView(LoginRequiredMixin, View):
    """Muestra el detalle de una llegada"""
    
    def get(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        return render(request, "inventario/llegadas/detalle_llegada.html", {"llegada": llegada})


class ControlCalidadView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Control de calidad de la llegada"""
    permission_required = 'inventario.change_llegadaproveedor'
    
    def get(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        form = ControlCalidadForm(instance=llegada)
        return render(request, "inventario/llegadas/control_calidad.html", {"llegada": llegada, "form": form})
    
    def post(self, request, pk):
        from django.db import transaction
        from .models import Lote
        
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        form = ControlCalidadForm(request.POST, instance=llegada)
        
        if form.is_valid():
            with transaction.atomic():
                llegada = form.save(commit=False)
                llegada.usuario_calidad = request.user
                llegada.fecha_validacion_calidad = timezone.now()
                
                # Si se aprueba la inspeccion visual, cambiar estado a UBICACION y crear lotes
                if llegada.estado_calidad == 'APROBADO':
                    llegada.estado = 'UBICACION'
                    llegada.save()
                    
                    # Crear lotes para cada item si no existen
                    for item in llegada.items.all():
                        if not item.lote_creado:
                            from .models import Institucion
                            # Obtener la institucion (usar la primera disponible)
                            institucion = Institucion.objects.first()
                            
                            lote = Lote.objects.create(
                                producto=item.producto,
                                numero_lote=item.numero_lote,
                                fecha_caducidad=item.fecha_caducidad,
                                fecha_fabricacion=item.fecha_elaboracion,
                                cantidad_inicial=item.cantidad_recibida,
                                cantidad_disponible=item.cantidad_recibida,
                                cantidad_reservada=0,
                                almacen=llegada.almacen,
                                institucion=institucion,
                                precio_unitario=item.precio_unitario_sin_iva or 0,
                                valor_total=(item.precio_unitario_sin_iva or 0) * item.cantidad_recibida,
                                fecha_recepcion=llegada.fecha_llegada_real.date(),
                                estado=1
                            )
                            item.lote_creado = lote
                            item.save()
                    
                    messages.success(request, "Inspeccion visual aprobada. Lotes creados. Pendiente de asignacion de ubicacion")
                else:
                    llegada.save()
                    messages.warning(request, "Inspeccion visual rechazada")
            
            return redirect('logistica:llegadas:detalle_llegada', pk=llegada.pk)
        
        return render(request, "inventario/llegadas/control_calidad.html", {"llegada": llegada, "form": form})


class FacturacionView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Captura de datos de facturación"""
    permission_required = 'inventario.change_llegadaproveedor'
    
    def get(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        form = FacturacionForm(instance=llegada)
        formset = ItemFacturacionFormSet(instance=llegada)
        return render(request, "inventario/llegadas/facturacion.html", {
            "llegada": llegada,
            "form": form,
            "formset": formset
        })
    
    def post(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        form = FacturacionForm(request.POST, instance=llegada)
        formset = ItemFacturacionFormSet(request.POST, instance=llegada)
        
        if form.is_valid() and formset.is_valid():
            llegada = form.save(commit=False)
            llegada.usuario_facturacion = request.user
            llegada.save()
            formset.save()
            messages.success(request, "Datos de facturación guardados")
            return redirect('logistica:llegadas:detalle_llegada', pk=llegada.pk)
        
        return render(request, "inventario/llegadas/facturacion.html", {
            "llegada": llegada,
            "form": form,
            "formset": formset
        })


class SupervisionView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Supervisión de la llegada"""
    permission_required = 'inventario.change_llegadaproveedor'
    
    def get(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        form = SupervisionForm(instance=llegada)
        return render(request, "inventario/llegadas/supervision.html", {"llegada": llegada, "form": form})
    
    def post(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        form = SupervisionForm(request.POST, instance=llegada)
        
        if form.is_valid():
            llegada = form.save(commit=False)
            llegada.usuario_supervision = request.user
            llegada.fecha_supervision = timezone.now()
            llegada.save()
            messages.success(request, "Supervisión registrada")
            return redirect('logistica:llegadas:detalle_llegada', pk=llegada.pk)
        
        return render(request, "inventario/llegadas/supervision.html", {"llegada": llegada, "form": form})


class UbicacionView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Asignación de ubicación de lotes"""
    permission_required = 'inventario.change_llegadaproveedor'
    
    def get(self, request, pk):
        from .models import Almacen
        import logging
        logger = logging.getLogger(__name__)
        
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        items = list(llegada.items.all())
        
        logger.info(f"DEBUG: Llegada {llegada.pk} - Estado: {llegada.estado}")
        logger.info(f"DEBUG: Total de items: {len(items)}")
        for idx, item in enumerate(items):
            logger.info(f"DEBUG: Item {idx} - Lote creado: {item.lote_creado}")
        
        # Obtener almacenes disponibles
        almacenes = Almacen.objects.all()
        
        # Preparar datos para el template
        items_with_ubicaciones = []
        for idx, item in enumerate(items):
            # Verificar de forma segura si existe producto
            try:
                producto = item.producto if hasattr(item, 'producto_id') and item.producto_id else None
            except Exception:
                producto = None
            
            items_with_ubicaciones.append({
                'producto': producto,
                'lote': item.lote_creado,
                'cantidad_recibida': item.cantidad_recibida,
                'index': idx
            })
        
        logger.info(f"DEBUG: Items con ubicaciones preparados: {len(items_with_ubicaciones)}")
        
        return render(request, "inventario/llegadas/ubicacion.html", {
            "llegada": llegada,
            "items_with_ubicaciones": items_with_ubicaciones,
            "almacenes": almacenes
        })
    
    def post(self, request, pk):
        from .models import Almacen, Lote, LoteUbicacion, UbicacionAlmacen, MovimientoInventario
        
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        items = list(llegada.items.all())
        
        try:
            with transaction.atomic():
                for i, item in enumerate(items):
                    # Verificar que el item tenga producto
                    try:
                        producto_desc = item.producto.descripcion if item.producto else "Item sin producto"
                    except Exception:
                        producto_desc = "Item sin producto"
                    
                    almacen_id = request.POST.get(f'ubicacion-detalle-{i}-0-almacen')
                    
                    if not almacen_id:
                        messages.error(request, f"Debe seleccionar un almacén para {producto_desc}")
                        return redirect("logistica:llegadas:ubicacion", pk=llegada.pk)
                    
                    almacen = get_object_or_404(Almacen, pk=almacen_id)
                    
                    # Procesar ubicaciones desde POST
                    ubicacion_data = []
                    j = 0
                    while True:
                        ubicacion_id_key = f'ubicacion-detalle-{i}-{j}-ubicacion'
                        cantidad_key = f'ubicacion-detalle-{i}-{j}-cantidad'
                        
                        if ubicacion_id_key not in request.POST:
                            break
                        
                        ubicacion_id = request.POST.get(ubicacion_id_key)
                        cantidad_str = request.POST.get(cantidad_key, '0')
                        
                        if ubicacion_id and cantidad_str:
                            try:
                                cantidad = int(cantidad_str)
                                if cantidad > 0:
                                    ubicacion_data.append({
                                        'ubicacion_id': ubicacion_id,
                                        'cantidad': cantidad
                                    })
                            except ValueError:
                                pass
                        j += 1
                    
                    # Validar que hay al menos una ubicación
                    if not ubicacion_data:
                        messages.error(request, f"Debe asignar al menos una ubicación para {producto_desc}")
                        return redirect("logistica:llegadas:ubicacion", pk=llegada.pk)
                    
                    # Validar que la suma de cantidades sea igual a la cantidad recibida
                    total_cantidad = sum(u['cantidad'] for u in ubicacion_data)
                    if total_cantidad != item.cantidad_recibida:
                        messages.error(
                            request,
                            f"Para {producto_desc}: La suma de cantidades ({total_cantidad}) "
                            f"debe ser igual a la cantidad recibida ({item.cantidad_recibida})"
                        )
                        return redirect("logistica:llegadas:ubicacion", pk=llegada.pk)
                    
                    # Si el lote ya existe, actualizar; si no, crear
                    if item.lote_creado:
                        lote = item.lote_creado
                        lote.almacen = almacen
                        lote.save()
                        # Eliminar ubicaciones anteriores
                        lote.ubicaciones_detalle.all().delete()
                    else:
                        # Crear nuevo lote (fallback si no se creó en Control de Calidad)
                        from .models import Institucion
                        institucion = Institucion.objects.first()
                        lote = Lote.objects.create(
                            producto=item.producto,
                            numero_lote=item.numero_lote,
                            fecha_caducidad=item.fecha_caducidad,
                            fecha_fabricacion=item.fecha_elaboracion,
                            cantidad_inicial=item.cantidad_recibida,
                            cantidad_disponible=item.cantidad_recibida,
                            cantidad_reservada=0,
                            almacen=almacen,
                            institucion=institucion,
                            precio_unitario=item.precio_unitario_sin_iva or 0,
                            valor_total=(item.precio_unitario_sin_iva or 0) * item.cantidad_recibida,
                            fecha_recepcion=llegada.fecha_llegada_real.date(),
                            estado=1
                        )
                        item.lote_creado = lote
                        item.save()
                    
                    # Crear registros de LoteUbicacion para cada ubicación
                    for ubi_data in ubicacion_data:
                        ubicacion = get_object_or_404(UbicacionAlmacen, pk=ubi_data['ubicacion_id'])
                        LoteUbicacion.objects.create(
                            lote=lote,
                            ubicacion=ubicacion,
                            cantidad=ubi_data['cantidad'],
                            usuario_asignacion=request.user
                        )
                    
                    # Crear movimiento de entrada en inventario
                    MovimientoInventario.objects.create(
                        lote=lote,
                        tipo_movimiento='ENTRADA',
                        cantidad=item.cantidad_recibida,
                        cantidad_anterior=0,
                        cantidad_nueva=item.cantidad_recibida,
                        motivo=f'Entrada de proveedor - Folio: {llegada.folio}',
                        documento_referencia=llegada.remision,
                        contrato=llegada.numero_contrato,
                        remision=llegada.remision,
                        folio=llegada.folio,
                        usuario=request.user
                    )
                
                # Marcar llegada como completada
                llegada.estado = 'APROBADA'
                llegada.usuario_ubicacion = request.user
                llegada.fecha_ubicacion = timezone.now()
                llegada.save()
                
                messages.success(request, "Ubicaciones asignadas correctamente")
                return redirect('logistica:llegadas:detalle_llegada', pk=llegada.pk)
        
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al asignar ubicaciones: {str(e)}", exc_info=True)
            messages.error(request, f"Error al asignar ubicaciones: {str(e)}")
            return redirect("logistica:llegadas:ubicacion", pk=llegada.pk)


class SubirDocumentoView(LoginRequiredMixin, View):
    """Subir documentos adjuntos"""
    
    def post(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        form = DocumentoLlegadaForm(request.POST, request.FILES)
        
        if form.is_valid():
            documento = form.save(commit=False)
            documento.llegada = llegada
            documento.save()
            messages.success(request, "Documento subido exitosamente")
        
        return redirect('logistica:llegadas:detalle_llegada', pk=llegada.pk)


# API para obtener productos en formato JSON
@require_GET
def api_productos(request):
    """API que devuelve los productos disponibles en formato JSON"""
    from django.apps import apps
    Producto = apps.get_model('inventario', 'Producto')
    
    productos = Producto.objects.all().order_by('descripcion')
    data = []
    for producto in productos:
        data.append({
            'id': producto.id,
            'clave_cnis': producto.clave_cnis,
            'descripcion': f"{producto.clave_cnis} - {producto.descripcion[:50]}...",
            'iva': float(producto.iva)
        })
    return JsonResponse(data, safe=False)


@require_GET
def api_ubicaciones_por_almacen(request):
    """
    API que devuelve las ubicaciones disponibles para un almacén específico.
    Parámetro GET: almacen_id
    """
    from django.apps import apps
    UbicacionAlmacen = apps.get_model('inventario', 'UbicacionAlmacen')
    
    almacen_id = request.GET.get('almacen_id')
    
    if not almacen_id:
        return JsonResponse({'error': 'almacen_id es requerido'}, status=400)
    
    try:
        almacen_id = int(almacen_id)
    except ValueError:
        return JsonResponse({'error': 'almacen_id debe ser un número'}, status=400)
    
    ubicaciones = UbicacionAlmacen.objects.filter(
        almacen_id=almacen_id
    ).values('id', 'codigo', 'descripcion').order_by('codigo')
    
    return JsonResponse({
        'ubicaciones': list(ubicaciones)
    })


@require_GET
def api_debug_citas_disponibles(request):
    """
    Endpoint de debug para verificar qué citas están disponibles en el formulario.
    URL: /logistica/llegadas/api/debug/citas-disponibles/
    """
    from django.apps import apps
    from django.db.models import Q, OuterRef, Exists
    CitaProveedor = apps.get_model('inventario', 'CitaProveedor')
    LlegadaProveedor = apps.get_model('inventario', 'LlegadaProveedor')
    
    # Replicar EXACTAMENTE la lógica del formulario
    queryset = CitaProveedor.objects.filter(estado='autorizada').select_related('proveedor')
    
    # Subquery para verificar si existe una llegada en estado EN_RECEPCION para esta cita
    llegada_en_recepcion = LlegadaProveedor.objects.filter(
        cita=OuterRef('pk'),
        estado='EN_RECEPCION'
    )
    
    # Filtrar: citas sin llegada_proveedor O con llegada_proveedor en estado EN_RECEPCION
    queryset = queryset.filter(
        Q(llegada_proveedor__isnull=True) | Exists(llegada_en_recepcion)
    )
    queryset = queryset.order_by('-fecha_cita')
    
    # Obtener el SQL query
    sql_query = str(queryset.query)
    
    citas_data = []
    for cita in queryset[:20]:  # Primeras 20
        tiene_llegada = False
        estado_llegada = None
        llegada_id = None
        try:
            if hasattr(cita, 'llegada_proveedor'):
                try:
                    llegada = cita.llegada_proveedor
                    tiene_llegada = True
                    estado_llegada = llegada.estado
                    llegada_id = str(llegada.id)
                except:
                    tiene_llegada = False
        except:
            pass
        
        citas_data.append({
            'id': str(cita.id),
            'folio': cita.folio or 'Sin folio',
            'proveedor': cita.proveedor.razon_social if cita.proveedor else 'Sin proveedor',
            'fecha_cita': cita.fecha_cita.isoformat() if cita.fecha_cita else None,
            'estado': cita.estado,
            'tiene_llegada': tiene_llegada,
            'estado_llegada': estado_llegada,
            'llegada_id': llegada_id,
        })
    
    # Verificar específicamente las citas IB-2026-000031 e IB-2026-000032
    cita_31 = CitaProveedor.objects.filter(folio='IB-2026-000031').first()
    cita_32 = CitaProveedor.objects.filter(folio='IB-2026-000032').first()
    
    debug_info = {
        'total': queryset.count(),
        'sql_query': sql_query,
        'citas': citas_data,
        'cita_31_info': None,
        'cita_32_info': None,
    }
    
    if cita_31:
        tiene_llegada_31 = False
        estado_llegada_31 = None
        llegada_id_31 = None
        esta_en_queryset_31 = queryset.filter(id=cita_31.id).exists()
        try:
            if hasattr(cita_31, 'llegada_proveedor'):
                llegada_31 = cita_31.llegada_proveedor
                tiene_llegada_31 = True
                estado_llegada_31 = llegada_31.estado
                llegada_id_31 = str(llegada_31.id)
        except:
            pass
        
        debug_info['cita_31_info'] = {
            'id': str(cita_31.id),
            'folio': cita_31.folio,
            'estado': cita_31.estado,
            'tiene_llegada': tiene_llegada_31,
            'estado_llegada': estado_llegada_31,
            'llegada_id': llegada_id_31,
            'esta_en_queryset': esta_en_queryset_31,
        }
    
    if cita_32:
        tiene_llegada_32 = False
        estado_llegada_32 = None
        llegada_id_32 = None
        esta_en_queryset_32 = queryset.filter(id=cita_32.id).exists()
        try:
            if hasattr(cita_32, 'llegada_proveedor'):
                llegada_32 = cita_32.llegada_proveedor
                tiene_llegada_32 = True
                estado_llegada_32 = llegada_32.estado
                llegada_id_32 = str(llegada_32.id)
        except:
            pass
        
        debug_info['cita_32_info'] = {
            'id': str(cita_32.id),
            'folio': cita_32.folio,
            'estado': cita_32.estado,
            'tiene_llegada': tiene_llegada_32,
            'estado_llegada': estado_llegada_32,
            'llegada_id': llegada_id_32,
            'esta_en_queryset': esta_en_queryset_32,
        }
    
    return JsonResponse(debug_info, json_dumps_params={'indent': 2, 'ensure_ascii': False})


@require_GET
def api_cita_folio(request, cita_id):
    """
    API que devuelve los datos de una cita específica.
    URL: /logistica/llegadas/api/cita/<cita_id>/folio/
    Devuelve: folio, remisión, orden de suministro, almacén, y datos de llegada si existe
    """
    try:
        from django.apps import apps
        CitaProveedor = apps.get_model('inventario', 'CitaProveedor')
        LlegadaProveedor = apps.get_model('inventario', 'LlegadaProveedor')
        
        cita = CitaProveedor.objects.get(id=cita_id)
        
        response_data = {
            'folio': cita.folio or f"TEMP-{cita_id}",
            'remision': cita.numero_orden_remision or '',
            'orden_suministro': cita.numero_orden_suministro or '',
            'contrato': cita.numero_contrato or '',
            'almacen_id': cita.almacen_id or '',
            'proveedor_id': cita.proveedor_id or '',
            'clave_medicamento': cita.clave_medicamento or '',
            'cita_id': cita_id
        }
        
        # Si la cita tiene una llegada en estado EN_RECEPCION, incluir sus datos
        try:
            llegada = LlegadaProveedor.objects.get(cita=cita, estado='EN_RECEPCION')
            response_data['llegada_existe'] = True
            response_data['llegada_id'] = str(llegada.id)
            response_data['remision'] = llegada.remision or response_data['remision']
            response_data['orden_suministro'] = llegada.numero_orden_suministro or response_data['orden_suministro']
            response_data['contrato'] = llegada.numero_contrato or response_data['contrato']
            response_data['almacen_id'] = llegada.almacen_id or response_data['almacen_id']
            response_data['tipo_red'] = llegada.tipo_red or ''
            response_data['numero_piezas_emitidas'] = llegada.numero_piezas_emitidas or 0
            response_data['numero_piezas_recibidas'] = llegada.numero_piezas_recibidas or 0
            response_data['numero_procedimiento'] = llegada.numero_procedimiento or ''
            response_data['observaciones_recepcion'] = llegada.observaciones_recepcion or ''
        except LlegadaProveedor.DoesNotExist:
            response_data['llegada_existe'] = False
        
        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': 'Cita no encontrada'}, status=404)



class ImprimirEPAView(LoginRequiredMixin, View):
    """Genera e imprime el documento EPA (Entrada de Existencias en Almacén)"""
    
    def get(self, request, pk):
        llegada = get_object_or_404(LlegadaProveedor, pk=pk)
        
        # Verificar que la llegada esté en UBICACION o APROBADA (se puede imprimir EPA desde que está en UBICACION)
        if llegada.estado not in ['UBICACION', 'APROBADA']:
            messages.error(request, "Solo se pueden imprimir EPA de llegadas en estado 'Asignando Ubicación' o 'Aprobada'")
            return redirect('logistica:llegadas:detalle_llegada', pk=pk)
        
        # Obtener usuarios con roles específicos para prellenar firmas
        titular_entrada = None
        titular_insumos = None
        
        try:
            from django.contrib.auth.models import Group
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            # Buscar usuario con rol "Titular entrada"
            grupo_titular_entrada = Group.objects.filter(name='Titular entrada').first()
            if grupo_titular_entrada:
                usuarios_titular_entrada = grupo_titular_entrada.user_set.filter(is_active=True).first()
                if usuarios_titular_entrada:
                    titular_entrada = usuarios_titular_entrada
            
            # Buscar usuario con rol "Titular insumos"
            grupo_titular_insumos = Group.objects.filter(name='Titular insumos').first()
            if grupo_titular_insumos:
                usuarios_titular_insumos = grupo_titular_insumos.user_set.filter(is_active=True).first()
                if usuarios_titular_insumos:
                    titular_insumos = usuarios_titular_insumos
        except Exception:
            # Si hay algún error, continuar sin prellenar
            pass
        
        # Preparar nombres para las firmas
        nombre_titular_entrada = ''
        if titular_entrada:
            nombre_titular_entrada = titular_entrada.get_full_name() if titular_entrada.get_full_name() else titular_entrada.username
        
        nombre_titular_insumos = ''
        if titular_insumos:
            nombre_titular_insumos = titular_insumos.get_full_name() if titular_insumos.get_full_name() else titular_insumos.username
        
        # Preparar datos para el template
        context = {
            'llegada': llegada,
            'items': llegada.items.all(),
            'total_items': llegada.items.aggregate(total=models.Sum('cantidad_recibida'))['total'] or 0,
            'nombre_titular_entrada': nombre_titular_entrada,
            'nombre_titular_insumos': nombre_titular_insumos,
        }
        
        return render(request, 'inventario/llegadas/imprimir_epa.html', context)
