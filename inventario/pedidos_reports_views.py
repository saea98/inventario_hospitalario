"""
Vistas para reportes de errores en carga masiva de pedidos
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q, F, Value, CharField
from django.db.models.functions import Concat
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
from .pedidos_models import LogErrorPedido, SolicitudPedido, ItemSolicitud, PropuestaPedido, ItemPropuesta, LoteAsignado
from .pedidos_utils import obtener_resumen_errores
from .models import Producto, Institucion
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@login_required
def reporte_errores_pedidos(request):
    """
    Muestra un reporte de los errores en carga masiva de pedidos.
    Permite filtrar por fecha, tipo de error, institución, etc.
    """
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_error = request.GET.get('tipo_error')
    institucion = request.GET.get('institucion')
    clave_solicitada = request.GET.get('clave_solicitada')
    filtro_pedido = request.GET.get('pedido', '').strip()
    
    # Base queryset
    errores = LogErrorPedido.objects.select_related(
        'usuario', 'institucion', 'almacen'
    ).all()
    
    # Aplicar filtros
    if fecha_inicio:
        from datetime import datetime
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        errores = errores.filter(fecha_error__gte=fecha_inicio_dt)
    
    if fecha_fin:
        from datetime import datetime
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
        fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
        errores = errores.filter(fecha_error__lte=fecha_fin_dt)
    
    if tipo_error:
        errores = errores.filter(tipo_error=tipo_error)
    
    if institucion:
        errores = errores.filter(institucion__id=institucion)
    
    if clave_solicitada:
        errores = errores.filter(clave_solicitada__icontains=clave_solicitada)
    
    # Obtener resumen
    resumen = obtener_resumen_errores(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        tipo_error=tipo_error
    )
    
    # Estadísticas generales
    total_errores = errores.count()
    errores_hoy = errores.filter(
        fecha_error__date=timezone.now().date()
    ).count()
    errores_semana = errores.filter(
        fecha_error__gte=timezone.now() - timedelta(days=7)
    ).count()
    
    # Errores sin alerta
    errores_sin_alerta = errores.filter(alerta_enviada=False).count()
    
    # Claves más frecuentes con error
    claves_frecuentes = errores.values('clave_solicitada').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Tipos de error más frecuentes
    tipos_frecuentes = errores.values('tipo_error').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Instituciones con más errores
    instituciones_frecuentes = errores.values(
        'institucion__nombre'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Usuarios con más errores
    usuarios_frecuentes = errores.values(
        'usuario__username'
    ).annotate(
        nombre_completo=Concat(
            F('usuario__first_name'), Value(' '), F('usuario__last_name'),
            output_field=CharField()
        ),
        count=Count('id')
    ).order_by('-count')[:10]

    # Obtener folio del pedido (observaciones_solicitud) para cada error
    # Si hay filtro por pedido, procesar más registros para poder filtrar
    limite = 500 if filtro_pedido else 100
    errores_lista = list(errores[:limite])
    errores_con_folio = []
    for error in errores_lista:
        folio_pedido = _obtener_folio_pedido_para_error(error)
        folio_str = folio_pedido or '-'
        errores_con_folio.append({'error': error, 'folio_pedido': folio_str})
    # Aplicar filtro por pedido (folio) si está activo
    if filtro_pedido:
        termino = filtro_pedido.lower()
        errores_con_folio = [x for x in errores_con_folio if termino in (x['folio_pedido'] or '').lower()]
    errores_con_folio = errores_con_folio[:100]
    
    # Instituciones para filtro
    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')

    context = {
        'errores': [x['error'] for x in errores_con_folio],
        'errores_con_folio': errores_con_folio,
        'total_errores': total_errores,
        'errores_hoy': errores_hoy,
        'errores_semana': errores_semana,
        'errores_sin_alerta': errores_sin_alerta,
        'claves_frecuentes': claves_frecuentes,
        'tipos_frecuentes': tipos_frecuentes,
        'instituciones_frecuentes': instituciones_frecuentes,
        'usuarios_frecuentes': usuarios_frecuentes,
        'resumen': resumen,
        'errores_por_tipo': resumen.get('errores_por_tipo', {}),
        'alertas_enviadas': errores.filter(alerta_enviada=True).count(),
        'tipo_error_choices': LogErrorPedido.TIPO_ERROR_CHOICES,
        'instituciones': instituciones,
        'filtro_pedido': filtro_pedido,
        'filtro_institucion': institucion,
        'page_title': 'Reporte de Errores en Carga Masiva de Pedidos'
    }
    
    return render(request, 'inventario/pedidos/reporte_errores_pedidos.html', context)


def _obtener_folio_pedido_para_error(error):
    """
    Busca la solicitud relacionada con el error y retorna el folio del pedido
    (observaciones_solicitud). Se busca por clave+usuario+fecha.
    """
    try:
        producto = Producto.objects.get(clave_cnis=error.clave_solicitada)
        fecha_inicio = error.fecha_error - timedelta(hours=24)
        fecha_fin = error.fecha_error + timedelta(hours=1)
        items = ItemSolicitud.objects.filter(
            producto=producto,
            solicitud__fecha_solicitud__gte=fecha_inicio,
            solicitud__fecha_solicitud__lte=fecha_fin
        ).select_related('solicitud').order_by('-solicitud__fecha_solicitud')
        if error.usuario:
            items = items.filter(solicitud__usuario_solicitante=error.usuario)
        item = items.first()
        if item and item.solicitud:
            return (item.solicitud.observaciones_solicitud or '').strip() or (item.solicitud.folio or '')
    except Producto.DoesNotExist:
        if error.usuario:
            solicitud = SolicitudPedido.objects.filter(
                usuario_solicitante=error.usuario,
                fecha_solicitud__gte=error.fecha_error - timedelta(hours=24),
                fecha_solicitud__lte=error.fecha_error + timedelta(hours=1)
            ).order_by('-fecha_solicitud').first()
            if solicitud:
                return (solicitud.observaciones_solicitud or '').strip() or (solicitud.folio or '')
    return None


def _obtener_items_no_surtidos(request):
    """
    Aplica filtros y construye la lista de items no surtidos.
    Usado por reporte_items_no_surtidos y exportar_items_no_surtidos_excel.
    """
    from django.db.models import Sum

    estado_propuesta = request.GET.get('estado', '')
    institucion_id = request.GET.get('institucion', '')
    folio = request.GET.get('folio', '').strip()
    clave_cnis = request.GET.get('clave_cnis', '').strip()

    propuestas = PropuestaPedido.objects.filter(
        solicitud__isnull=False
    ).select_related(
        'solicitud__institucion_solicitante',
        'solicitud__almacen_destino'
    ).prefetch_related(
        'items__producto',
        'items__lotes_asignados'
    ).order_by('-fecha_generacion')

    if estado_propuesta:
        propuestas = propuestas.filter(estado=estado_propuesta)
    if institucion_id:
        propuestas = propuestas.filter(solicitud__institucion_solicitante_id=institucion_id)
    if folio:
        propuestas = propuestas.filter(
            Q(solicitud__folio__icontains=folio) |
            Q(solicitud__observaciones_solicitud__icontains=folio)
        )
    if clave_cnis:
        propuestas = propuestas.filter(items__producto__clave_cnis__icontains=clave_cnis).distinct()

    items_no_surtidos = []
    for propuesta in propuestas:
        solicitud = propuesta.solicitud
        folio_pedido = (solicitud.observaciones_solicitud or '').strip() or (solicitud.folio or '')
        institucion_nombre = solicitud.institucion_solicitante.denominacion if solicitud.institucion_solicitante else '-'
        almacen_nombre = solicitud.almacen_destino.nombre if solicitud.almacen_destino else '-'

        for item in propuesta.items.all():
            if item.cantidad_propuesta <= 0:
                continue
            cantidad_surtida_real = item.lotes_asignados.filter(surtido=True).aggregate(
                total=Sum('cantidad_asignada')
            )['total'] or 0
            if cantidad_surtida_real < item.cantidad_propuesta:
                cantidad_faltante = item.cantidad_propuesta - cantidad_surtida_real
                if clave_cnis and clave_cnis.lower() not in (item.producto.clave_cnis or '').lower():
                    continue
                items_no_surtidos.append({
                    'propuesta': propuesta,
                    'item': item,
                    'producto': item.producto,
                    'folio_pedido': folio_pedido,
                    'institucion': institucion_nombre,
                    'almacen': almacen_nombre,
                    'cantidad_propuesta': item.cantidad_propuesta,
                    'cantidad_surtida': cantidad_surtida_real,
                    'cantidad_faltante': cantidad_faltante,
                    'estado_propuesta': propuesta.get_estado_display(),
                })
    return items_no_surtidos


@login_required
def reporte_items_no_surtidos(request):
    """
    Reporte de items que no se surtieron (o se surtieron parcialmente) en las propuestas de suministro.
    Muestra ItemPropuesta donde cantidad_surtida < cantidad_propuesta.
    """
    items_no_surtidos = _obtener_items_no_surtidos(request)

    # Instituciones para filtro
    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')

    # Paginación (25 items por página)
    paginator = Paginator(items_no_surtidos, 25)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Query string para paginación (preservar filtros)
    params = request.GET.copy()
    if 'page' in params:
        params.pop('page')
    query_string = params.urlencode()

    context = {
        'items_no_surtidos': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_items': paginator.count,
        'is_paginated': paginator.num_pages > 1,
        'query_string': query_string,
        'instituciones': instituciones,
        'filtro_estado': request.GET.get('estado', ''),
        'filtro_institucion': request.GET.get('institucion', ''),
        'filtro_folio': request.GET.get('folio', '').strip(),
        'filtro_clave': request.GET.get('clave_cnis', '').strip(),
        'estado_choices': PropuestaPedido.ESTADO_CHOICES,
        'page_title': 'Reporte de Items No Surtidos en Propuestas',
    }
    return render(request, 'inventario/pedidos/reporte_items_no_surtidos.html', context)


@login_required
def exportar_items_no_surtidos_excel(request):
    """Exporta el reporte de items no surtidos a Excel respetando los filtros actuales."""
    items_no_surtidos = _obtener_items_no_surtidos(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Items No Surtidos'

    header_fill = PatternFill(start_color='B45F06', end_color='B45F06', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = [
        'Folio de Pedido',
        'Institución',
        'Almacén',
        'Clave CNIS',
        'Descripción',
        'Cant. Propuesta',
        'Cant. Surtida',
        'Cant. Faltante',
        'Estado Propuesta',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(items_no_surtidos, 2):
        desc = (row['producto'].descripcion or '')[:100] if row['producto'] else ''
        ws.cell(row=row_idx, column=1).value = row['folio_pedido'] or ''
        ws.cell(row=row_idx, column=2).value = row['institucion'] or ''
        ws.cell(row=row_idx, column=3).value = row['almacen'] or ''
        ws.cell(row=row_idx, column=4).value = row['producto'].clave_cnis if row['producto'] else ''
        ws.cell(row=row_idx, column=5).value = desc
        ws.cell(row=row_idx, column=6).value = row['cantidad_propuesta'] or 0
        ws.cell(row=row_idx, column=7).value = row['cantidad_surtida'] or 0
        ws.cell(row=row_idx, column=8).value = row['cantidad_faltante'] or 0
        ws.cell(row=row_idx, column=9).value = row['estado_propuesta'] or ''
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    from datetime import date
    fecha_str = date.today().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="items_no_surtidos_{fecha_str}.xlsx"'
    wb.save(response)
    return response


def _obtener_filtros_reporte_pedidos(request):
    """Extrae y aplica filtros comunes para el reporte de pedidos (vista y Excel)."""
    from datetime import datetime
    institucion_id = request.GET.get('institucion', '').strip()
    folio = request.GET.get('folio', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    clave_cnis = request.GET.get('clave_cnis', '').strip()
    estado = request.GET.get('estado', '').strip()

    solicitudes = SolicitudPedido.objects.select_related(
        'institucion_solicitante', 'almacen_destino', 'usuario_solicitante'
    ).prefetch_related(
        'items__producto',
        'items__items_propuesta__propuesta',
        'items__items_propuesta__lotes_asignados',
    ).order_by('-fecha_solicitud')

    if institucion_id:
        solicitudes = solicitudes.filter(institucion_solicitante_id=institucion_id)
    if folio:
        solicitudes = solicitudes.filter(
            Q(observaciones_solicitud__icontains=folio) | Q(folio__icontains=folio)
        )
    if fecha_inicio:
        try:
            f_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_solicitud__date__gte=f_inicio)
        except ValueError:
            pass
    if fecha_fin:
        try:
            f_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_solicitud__date__lte=f_fin)
        except ValueError:
            pass
    if clave_cnis:
        solicitudes = solicitudes.filter(items__producto__clave_cnis__icontains=clave_cnis).distinct()
    if estado:
        solicitudes = solicitudes.filter(estado=estado)

    return solicitudes, {
        'filtro_institucion': institucion_id,
        'filtro_folio': folio,
        'filtro_fecha_inicio': fecha_inicio,
        'filtro_fecha_fin': fecha_fin,
        'filtro_clave': clave_cnis,
        'filtro_estado': estado,
    }


def _construir_filas_reporte_pedidos(solicitudes):
    """Construye lista de filas (una por item de solicitud) para el reporte de pedidos.
    Incluye cantidad suministrada (surtida) desde LoteAsignado donde surtido=True.
    """
    filas = []
    for solicitud in solicitudes:
        folio_pedido = (solicitud.observaciones_solicitud or '').strip() or (solicitud.folio or '')
        institucion_nombre = solicitud.institucion_solicitante.denominacion if solicitud.institucion_solicitante else '-'
        almacen_nombre = solicitud.almacen_destino.nombre if solicitud.almacen_destino else '-'
        for item in solicitud.items.all():
            # Cantidad suministrada: suma de LoteAsignado.cantidad_asignada donde surtido=True
            cantidad_surtida = 0
            for ip in item.items_propuesta.all():
                if ip.propuesta and ip.propuesta.solicitud_id == solicitud.id:
                    cantidad_surtida = sum(
                        la.cantidad_asignada for la in ip.lotes_asignados.all() if la.surtido
                    )
                    break
            filas.append({
                'solicitud': solicitud,
                'item': item,
                'folio_pedido': folio_pedido,
                'folio_sistema': solicitud.folio or '',
                'institucion': institucion_nombre,
                'almacen': almacen_nombre,
                'fecha_solicitud': solicitud.fecha_solicitud,
                'estado': solicitud.get_estado_display(),
                'clave_cnis': item.producto.clave_cnis if item.producto else '',
                'descripcion': (item.producto.descripcion or '')[:100] if item.producto else '',
                'cantidad_solicitada': item.cantidad_solicitada,
                'cantidad_aprobada': item.cantidad_aprobada,
                'cantidad_surtida': cantidad_surtida,
            })
    return filas


@login_required
def reporte_pedidos(request):
    """
    Reporte de pedidos (solicitudes) con filtros: institución, folio, fecha, claves, estado.
    Muestra detalle por item (una fila por producto en cada pedido).
    """
    solicitudes, filtros = _obtener_filtros_reporte_pedidos(request)
    filas = _construir_filas_reporte_pedidos(solicitudes)

    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')

    # Paginación (25 por página)
    paginator = Paginator(filas, 25)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    params = request.GET.copy()
    if 'page' in params:
        params.pop('page')
    query_string = params.urlencode()

    context = {
        'filas': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_items': paginator.count,
        'total_solicitudes': len(set(f['solicitud'].id for f in filas)),
        'is_paginated': paginator.num_pages > 1,
        'query_string': query_string,
        'instituciones': instituciones,
        'estado_choices': SolicitudPedido.ESTADO_CHOICES,
        'page_title': 'Reporte de Pedidos',
        **filtros,
    }
    return render(request, 'inventario/pedidos/reporte_pedidos.html', context)


@login_required
def exportar_reporte_pedidos_excel(request):
    """Exporta el reporte de pedidos a Excel con los mismos filtros de la vista."""
    solicitudes, _ = _obtener_filtros_reporte_pedidos(request)
    filas = _construir_filas_reporte_pedidos(solicitudes)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte de Pedidos'

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = [
        'Folio de Pedido',
        'Folio Sistema',
        'Institución',
        'Almacén',
        'Fecha Solicitud',
        'Estado',
        'Clave CNIS',
        'Descripción',
        'Cant. Solicitada',
        'Cant. Aprobada',
        'Cant. Suministrada',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_idx, f in enumerate(filas, 2):
        ws.cell(row=row_idx, column=1).value = f['folio_pedido'] or ''
        ws.cell(row=row_idx, column=2).value = f['folio_sistema'] or ''
        ws.cell(row=row_idx, column=3).value = f['institucion'] or ''
        ws.cell(row=row_idx, column=4).value = f['almacen'] or ''
        ws.cell(row=row_idx, column=5).value = f['fecha_solicitud'].strftime('%d/%m/%Y %H:%M') if f['fecha_solicitud'] else ''
        ws.cell(row=row_idx, column=6).value = f['estado'] or ''
        ws.cell(row=row_idx, column=7).value = f['clave_cnis'] or ''
        ws.cell(row=row_idx, column=8).value = f['descripcion'] or ''
        ws.cell(row=row_idx, column=9).value = f['cantidad_solicitada'] or 0
        ws.cell(row=row_idx, column=10).value = f['cantidad_aprobada'] or 0
        ws.cell(row=row_idx, column=11).value = f.get('cantidad_surtida', 0) or 0
        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 45
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    from datetime import date
    fecha_str = date.today().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="reporte_pedidos_{fecha_str}.xlsx"'
    wb.save(response)
    return response


@login_required
def reporte_claves_sin_existencia(request):
    """
    Muestra un reporte de claves solicitadas sin existencia.
    Útil para que el área médica genere oficios de solicitud.
    """
    # Obtener claves sin existencia
    errores_sin_existencia = LogErrorPedido.objects.filter(
        tipo_error='SIN_EXISTENCIA'
    ).select_related('usuario', 'institucion').order_by('-fecha_error')
    
    # Agrupar por clave
    from .models import Producto
    claves_sin_existencia = {}
    for error in errores_sin_existencia:
        if error.clave_solicitada not in claves_sin_existencia:
            descripcion = 'Producto no encontrado'
            try:
                producto = Producto.objects.get(clave_cnis=error.clave_solicitada)
                descripcion = producto.descripcion
            except Producto.DoesNotExist:
                pass
            
            claves_sin_existencia[error.clave_solicitada] = {
                'clave': error.clave_solicitada,
                'descripcion': descripcion,
                'total_solicitudes': 0,
                'cantidad_total': 0,
                'instituciones': set(),
                'ultimos_errores': []
            }
        
        claves_sin_existencia[error.clave_solicitada]['total_solicitudes'] += 1
        claves_sin_existencia[error.clave_solicitada]['cantidad_total'] += error.cantidad_solicitada or 0
        if error.institucion:
            claves_sin_existencia[error.clave_solicitada]['instituciones'].add(
                error.institucion.nombre
            )
        claves_sin_existencia[error.clave_solicitada]['ultimos_errores'].append(error)
    
    # Convertir instituciones a lista
    for clave_data in claves_sin_existencia.values():
        clave_data['instituciones'] = list(clave_data['instituciones'])
        clave_data['ultimos_errores'] = clave_data['ultimos_errores'][:5]
    
    context = {
        'claves_sin_existencia': sorted(
            claves_sin_existencia.values(),
            key=lambda x: x['total_solicitudes'],
            reverse=True
        ),
        'total_claves': len(claves_sin_existencia),
        'page_title': 'Reporte de Claves sin Existencia'
    }
    
    return render(request, 'inventario/pedidos/reporte_claves_sin_existencia.html', context)


@login_required
def reporte_pedidos_sin_existencia(request):
    """
    Muestra un reporte de pedidos con claves sin existencia agrupado por destino (institución y almacén).
    Útil para ver qué destinos tienen más problemas de disponibilidad.
    """
    # Obtener errores sin existencia
    errores_sin_existencia = LogErrorPedido.objects.filter(
        tipo_error='SIN_EXISTENCIA'
    ).select_related('usuario', 'institucion', 'almacen', 'usuario__almacen', 'usuario__almacen__institucion').order_by('-fecha_error')
    
    # Aplicar filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    institucion_id = request.GET.get('institucion')
    almacen_id = request.GET.get('almacen')
    
    if fecha_inicio:
        try:
            from datetime import datetime
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            errores_sin_existencia = errores_sin_existencia.filter(fecha_error__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            from datetime import datetime, time as dt_time
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            fecha_fin_obj = datetime.combine(fecha_fin_obj.date(), dt_time.max)
            errores_sin_existencia = errores_sin_existencia.filter(fecha_error__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    # Filtros institución/almacén se aplican después por destino del pedido
    
    # Agrupar por destino (institución + almacén del pedido)
    from .models import Producto, Institucion, Almacen
    from .pedidos_models import ItemSolicitud
    pedidos_sin_existencia = {}
    
    for error in errores_sin_existencia:
        # Buscar la solicitud relacionada para obtener institución y almacén del pedido
        institucion = None
        almacen = None
        folio_pedido = None
        solicitud_relacionada = None
        
        # Buscar solicitud que contenga un item con la clave del error
        # Esto es más preciso que buscar solo por usuario y fecha
        try:
            # Buscar el producto por clave CNIS
            producto = Producto.objects.get(clave_cnis=error.clave_solicitada)
            
            # Buscar solicitudes que tengan items con este producto
            # Buscar en un rango de tiempo más amplio para asegurar encontrar la solicitud
            fecha_inicio_busqueda = error.fecha_error - timedelta(hours=24)  # 24 horas antes
            fecha_fin_busqueda = error.fecha_error + timedelta(hours=1)  # 1 hora después
            
            # Buscar items que coincidan con el producto y la fecha
            items_relacionados = ItemSolicitud.objects.filter(
                producto=producto,
                solicitud__fecha_solicitud__gte=fecha_inicio_busqueda,
                solicitud__fecha_solicitud__lte=fecha_fin_busqueda
            ).select_related('solicitud', 'solicitud__institucion_solicitante', 'solicitud__almacen_destino').order_by('-solicitud__fecha_solicitud')
            
            # Si hay usuario, filtrar por usuario también
            if error.usuario:
                items_relacionados = items_relacionados.filter(solicitud__usuario_solicitante=error.usuario)
            
            # Tomar la primera solicitud encontrada
            item_relacionado = items_relacionados.first()
            
            if item_relacionado and item_relacionado.solicitud:
                solicitud_relacionada = item_relacionado.solicitud
                # SIEMPRE usar institución y almacén del pedido cuando se encuentra la solicitud
                # Usar los campos correctos: institucion_solicitante y almacen_destino
                institucion = solicitud_relacionada.institucion_solicitante
                almacen = solicitud_relacionada.almacen_destino
                # Usar observaciones_solicitud (folio del pedido) - este es el campo correcto
                folio_pedido = solicitud_relacionada.observaciones_solicitud.strip() if solicitud_relacionada.observaciones_solicitud else (solicitud_relacionada.folio or '')
        except Producto.DoesNotExist:
            # Si el producto no existe, intentar buscar por usuario y fecha como respaldo
            if error.usuario:
                fecha_inicio_busqueda = error.fecha_error - timedelta(hours=24)
                fecha_fin_busqueda = error.fecha_error + timedelta(hours=1)
                
                solicitud_relacionada = SolicitudPedido.objects.filter(
                    usuario_solicitante=error.usuario,
                    fecha_solicitud__gte=fecha_inicio_busqueda,
                    fecha_solicitud__lte=fecha_fin_busqueda
                ).select_related('institucion_solicitante', 'almacen_destino').order_by('-fecha_solicitud').first()
                
                if solicitud_relacionada:
                    institucion = solicitud_relacionada.institucion_solicitante
                    almacen = solicitud_relacionada.almacen_destino
                    folio_pedido = solicitud_relacionada.observaciones_solicitud.strip() if solicitud_relacionada.observaciones_solicitud else (solicitud_relacionada.folio or '')
        
        # Solo incluir en el reporte errores que están en un pedido (tienen solicitud asociada)
        if not solicitud_relacionada:
            continue
        
        # Crear clave única por destino (siempre del pedido)
        institucion_nombre = institucion.denominacion if institucion else 'Sin institución'
        almacen_nombre = almacen.nombre if almacen else 'Sin almacén'
        destino_key = f"{institucion_nombre}|{almacen_nombre}"
        
        if destino_key not in pedidos_sin_existencia:
            pedidos_sin_existencia[destino_key] = {
                'institucion': institucion_nombre,
                'almacen': almacen_nombre,
                'institucion_obj': institucion,
                'almacen_obj': almacen,
                'total_claves': 0,
                'total_solicitudes': 0,
                'cantidad_total': 0,
                'claves': set(),
                'errores': []
            }
        
        # Agregar información
        pedidos_sin_existencia[destino_key]['total_solicitudes'] += 1
        pedidos_sin_existencia[destino_key]['cantidad_total'] += error.cantidad_solicitada or 0
        pedidos_sin_existencia[destino_key]['claves'].add(error.clave_solicitada)
        
        # Agregar error con folio del pedido
        error_data = {
            'error': error,
            'folio_pedido': folio_pedido
        }
        pedidos_sin_existencia[destino_key]['errores'].append(error_data)
        
        # Agregar folio a un set para mostrar en el resumen
        if folio_pedido:
            if 'folios_pedidos' not in pedidos_sin_existencia[destino_key]:
                pedidos_sin_existencia[destino_key]['folios_pedidos'] = set()
            pedidos_sin_existencia[destino_key]['folios_pedidos'].add(folio_pedido)
    
    # Procesar datos y obtener descripciones de productos
    for destino_data in pedidos_sin_existencia.values():
        destino_data['total_claves'] = len(destino_data['claves'])
        destino_data['claves'] = sorted(list(destino_data['claves']))
        
        # Obtener descripciones de productos
        claves_con_descripcion = []
        for clave in destino_data['claves']:
            descripcion = 'Producto no encontrado'
            try:
                producto = Producto.objects.get(clave_cnis=clave)
                descripcion = producto.descripcion
            except Producto.DoesNotExist:
                pass
            claves_con_descripcion.append({
                'clave': clave,
                'descripcion': descripcion
            })
        destino_data['claves_detalle'] = claves_con_descripcion
        destino_data['errores'] = destino_data['errores'][:10]  # Limitar a 10 errores recientes
        
        # Procesar folios de pedidos
        if 'folios_pedidos' in destino_data:
            destino_data['folios_pedidos'] = sorted(list(destino_data['folios_pedidos']))
            destino_data['total_folios'] = len(destino_data['folios_pedidos'])
        else:
            destino_data['folios_pedidos'] = []
            destino_data['total_folios'] = 0
    
    # Obtener instituciones y almacenes para filtros
    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')
    almacenes = Almacen.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'pedidos_sin_existencia': sorted(
            pedidos_sin_existencia.values(),
            key=lambda x: x['total_solicitudes'],
            reverse=True
        ),
        'total_destinos': len(pedidos_sin_existencia),
        'instituciones': instituciones,
        'almacenes': almacenes,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'institucion_id': institucion_id,
        'almacen_id': almacen_id,
        'page_title': 'Reporte de Pedidos Sin Existencia por Destino'
    }
    
    return render(request, 'inventario/pedidos/reporte_pedidos_sin_existencia.html', context)


@login_required
def exportar_pedidos_sin_existencia_excel(request):
    """
    Exporta el reporte de pedidos sin existencia por destino a Excel
    """
    # Obtener errores sin existencia (mismo código que en la vista)
    errores_sin_existencia = LogErrorPedido.objects.filter(
        tipo_error='SIN_EXISTENCIA'
    ).select_related('usuario', 'institucion', 'almacen', 'usuario__almacen', 'usuario__almacen__institucion').order_by('-fecha_error')
    
    # Aplicar filtros (mismo que en la vista)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    institucion_id = request.GET.get('institucion')
    almacen_id = request.GET.get('almacen')
    
    if fecha_inicio:
        try:
            from datetime import datetime
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            errores_sin_existencia = errores_sin_existencia.filter(fecha_error__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            from datetime import datetime, time as dt_time
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            fecha_fin_obj = datetime.combine(fecha_fin_obj.date(), dt_time.max)
            errores_sin_existencia = errores_sin_existencia.filter(fecha_error__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    # Filtros institución/almacén se aplican después por destino del pedido
    
    # Agrupar por destino (mismo código que en la vista)
    from .models import Producto
    from .pedidos_models import ItemSolicitud
    pedidos_sin_existencia = {}
    
    for error in errores_sin_existencia:
        # Buscar la solicitud relacionada para obtener institución y almacén del pedido
        institucion = None
        almacen = None
        folio_pedido = None
        solicitud_relacionada = None
        
        # Buscar solicitud que contenga un item con la clave del error
        # Esto es más preciso que buscar solo por usuario y fecha
        try:
            # Buscar el producto por clave CNIS
            producto = Producto.objects.get(clave_cnis=error.clave_solicitada)
            
            # Buscar solicitudes que tengan items con este producto
            # Buscar en un rango de tiempo más amplio para asegurar encontrar la solicitud
            fecha_inicio_busqueda = error.fecha_error - timedelta(hours=24)  # 24 horas antes
            fecha_fin_busqueda = error.fecha_error + timedelta(hours=1)  # 1 hora después
            
            # Buscar items que coincidan con el producto y la fecha
            items_relacionados = ItemSolicitud.objects.filter(
                producto=producto,
                solicitud__fecha_solicitud__gte=fecha_inicio_busqueda,
                solicitud__fecha_solicitud__lte=fecha_fin_busqueda
            ).select_related('solicitud', 'solicitud__institucion_solicitante', 'solicitud__almacen_destino').order_by('-solicitud__fecha_solicitud')
            
            # Si hay usuario, filtrar por usuario también
            if error.usuario:
                items_relacionados = items_relacionados.filter(solicitud__usuario_solicitante=error.usuario)
            
            # Tomar la primera solicitud encontrada
            item_relacionado = items_relacionados.first()
            
            if item_relacionado and item_relacionado.solicitud:
                solicitud_relacionada = item_relacionado.solicitud
                # SIEMPRE usar institución y almacén del pedido cuando se encuentra la solicitud
                # Usar los campos correctos: institucion_solicitante y almacen_destino
                institucion = solicitud_relacionada.institucion_solicitante
                almacen = solicitud_relacionada.almacen_destino
                # Usar observaciones_solicitud (folio del pedido) - este es el campo correcto
                folio_pedido = solicitud_relacionada.observaciones_solicitud.strip() if solicitud_relacionada.observaciones_solicitud else (solicitud_relacionada.folio or '')
        except Producto.DoesNotExist:
            # Si el producto no existe, intentar buscar por usuario y fecha como respaldo
            if error.usuario:
                fecha_inicio_busqueda = error.fecha_error - timedelta(hours=24)
                fecha_fin_busqueda = error.fecha_error + timedelta(hours=1)
                
                solicitud_relacionada = SolicitudPedido.objects.filter(
                    usuario_solicitante=error.usuario,
                    fecha_solicitud__gte=fecha_inicio_busqueda,
                    fecha_solicitud__lte=fecha_fin_busqueda
                ).select_related('institucion_solicitante', 'almacen_destino').order_by('-fecha_solicitud').first()
                
                if solicitud_relacionada:
                    institucion = solicitud_relacionada.institucion_solicitante
                    almacen = solicitud_relacionada.almacen_destino
                    folio_pedido = solicitud_relacionada.observaciones_solicitud.strip() if solicitud_relacionada.observaciones_solicitud else (solicitud_relacionada.folio or '')
        
        # Solo incluir en el reporte errores que están en un pedido (tienen solicitud asociada)
        if not solicitud_relacionada:
            continue
        
        # Aplicar filtros por destino del pedido (institución/almacén)
        if institucion_id and (not institucion or str(institucion.id) != str(institucion_id)):
            continue
        if almacen_id and (not almacen or str(almacen.id) != str(almacen_id)):
            continue
        
        # Crear clave única por destino (siempre del pedido)
        institucion_nombre = institucion.denominacion if institucion else 'Sin institución'
        almacen_nombre = almacen.nombre if almacen else 'Sin almacén'
        destino_key = f"{institucion_nombre}|{almacen_nombre}"
        
        if destino_key not in pedidos_sin_existencia:
            pedidos_sin_existencia[destino_key] = {
                'institucion': institucion_nombre,
                'almacen': almacen_nombre,
                'total_claves': 0,
                'total_solicitudes': 0,
                'cantidad_total': 0,
                'claves': set(),
                'folios_pedidos': set(),
            }
        
        pedidos_sin_existencia[destino_key]['total_solicitudes'] += 1
        pedidos_sin_existencia[destino_key]['cantidad_total'] += error.cantidad_solicitada or 0
        pedidos_sin_existencia[destino_key]['claves'].add(error.clave_solicitada)
        
        # Agregar folio del pedido si existe
        if folio_pedido:
            pedidos_sin_existencia[destino_key]['folios_pedidos'].add(folio_pedido)
    
    # Procesar datos
    for destino_data in pedidos_sin_existencia.values():
        destino_data['total_claves'] = len(destino_data.get('claves', set()))
        destino_data['claves'] = ', '.join(sorted(list(destino_data.get('claves', set())))) if destino_data.get('claves') else ''
        destino_data['folios_pedidos'] = ', '.join(sorted(list(destino_data.get('folios_pedidos', set())))) if destino_data.get('folios_pedidos') else ''
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pedidos Sin Existencia por Destino'
    
    # Estilos
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Institución Destino', 'Almacén Destino', 'Folio(s) Pedido', 'Total Claves', 'Total Solicitudes', 'Cantidad Total Solicitada', 'Claves Afectadas']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    for row, destino_data in enumerate(sorted(
        pedidos_sin_existencia.values(),
        key=lambda x: x['total_solicitudes'],
        reverse=True
    ), 2):
        # Asegurar que todos los valores sean strings o números, nunca None
        ws.cell(row=row, column=1).value = str(destino_data.get('institucion', 'Sin institución') or 'Sin institución')
        ws.cell(row=row, column=2).value = str(destino_data.get('almacen', 'Sin almacén') or 'Sin almacén')
        ws.cell(row=row, column=3).value = str(destino_data.get('folios_pedidos', '') or '')
        ws.cell(row=row, column=4).value = destino_data.get('total_claves', 0) or 0
        ws.cell(row=row, column=5).value = destino_data.get('total_solicitudes', 0) or 0
        ws.cell(row=row, column=6).value = destino_data.get('cantidad_total', 0) or 0
        ws.cell(row=row, column=7).value = str(destino_data.get('claves', '') or '')
        
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 50
    
    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos_sin_existencia_por_destino.xlsx"'
    wb.save(response)
    
    return response


@login_required
def exportar_claves_sin_existencia_excel(request):
    """
    Exporta el reporte de claves sin existencia a Excel
    """
    # Obtener claves sin existencia (mismo código que en la vista)
    errores_sin_existencia = LogErrorPedido.objects.filter(
        tipo_error='SIN_EXISTENCIA'
    ).select_related('usuario', 'institucion').order_by('-fecha_error')
    
    # Agrupar por clave
    from .models import Producto
    claves_sin_existencia = {}
    for error in errores_sin_existencia:
        if error.clave_solicitada not in claves_sin_existencia:
            descripcion = 'Producto no encontrado'
            try:
                producto = Producto.objects.get(clave_cnis=error.clave_solicitada)
                descripcion = producto.descripcion
            except Producto.DoesNotExist:
                pass
            
            claves_sin_existencia[error.clave_solicitada] = {
                'clave': error.clave_solicitada,
                'descripcion': descripcion,
                'total_solicitudes': 0,
                'cantidad_total': 0,
                'instituciones': set(),
            }
        
        claves_sin_existencia[error.clave_solicitada]['total_solicitudes'] += 1
        claves_sin_existencia[error.clave_solicitada]['cantidad_total'] += error.cantidad_solicitada or 0
        if error.institucion:
            claves_sin_existencia[error.clave_solicitada]['instituciones'].add(
                error.institucion.nombre
            )
    
    # Convertir instituciones a lista
    for clave_data in claves_sin_existencia.values():
        clave_data['instituciones'] = ', '.join(list(clave_data['instituciones']))
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Claves Sin Existencia'
    
    # Estilos
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Clave', 'Descripción', 'Cantidad Total Solicitada', 'Total Solicitudes', 'Prioridad']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    for row, clave_data in enumerate(sorted(
        claves_sin_existencia.values(),
        key=lambda x: x['total_solicitudes'],
        reverse=True
    ), 2):
        ws.cell(row=row, column=1).value = clave_data['clave']
        ws.cell(row=row, column=2).value = clave_data['descripcion']
        ws.cell(row=row, column=3).value = clave_data['cantidad_total']
        ws.cell(row=row, column=4).value = clave_data['total_solicitudes']
        ws.cell(row=row, column=5).value = ''  # Prioridad: vacío para que lo llene quien corresponda
        
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 30
    
    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="claves_sin_existencia.xlsx"'
    wb.save(response)
    
    return response


@login_required
def reporte_claves_no_existen(request):
    """
    Muestra un reporte de claves que no existen en el catálogo.
    Útil para que el área médica genere oficios de solicitud de nuevos productos.
    """
    # Obtener claves que no existen
    errores_no_existen = LogErrorPedido.objects.filter(
        tipo_error='CLAVE_NO_EXISTE'
    ).select_related('usuario', 'institucion').order_by('-fecha_error')
    
    # Agrupar por clave
    claves_no_existen = {}
    for error in errores_no_existen:
        if error.clave_solicitada not in claves_no_existen:
            claves_no_existen[error.clave_solicitada] = {
                'clave': error.clave_solicitada,
                'total_solicitudes': 0,
                'cantidad_total': 0,
                'instituciones': set(),
                'ultimos_errores': []
            }
        
        claves_no_existen[error.clave_solicitada]['total_solicitudes'] += 1
        claves_no_existen[error.clave_solicitada]['cantidad_total'] += error.cantidad_solicitada or 0
        if error.institucion:
            claves_no_existen[error.clave_solicitada]['instituciones'].add(
                error.institucion.nombre
            )
        claves_no_existen[error.clave_solicitada]['ultimos_errores'].append(error)
    
    # Convertir instituciones a lista
    for clave_data in claves_no_existen.values():
        clave_data['instituciones'] = list(clave_data['instituciones'])
        clave_data['ultimos_errores'] = clave_data['ultimos_errores'][:5]
    
    context = {
        'claves_no_existen': sorted(
            claves_no_existen.values(),
            key=lambda x: x['total_solicitudes'],
            reverse=True
        ),
        'total_claves': len(claves_no_existen),
        'page_title': 'Reporte de Claves que no Existen en Catálogo'
    }
    
    return render(request, 'inventario/pedidos/reporte_claves_no_existen.html', context)


# ============================================================================
# DASHBOARD SURTIMIENTO POR INSTITUCIÓN
# ============================================================================

@login_required
def dashboard_surtimiento_institucion(request):
    """
    Dashboard por institución: productos surtidos completamente, no surtidos y surtidos parcialmente.
    Permite ver gráficas o tablas según selección del usuario.
    """
    from datetime import datetime

    institucion_id = request.GET.get('institucion', '').strip()
    folio_pedido = request.GET.get('folio_pedido', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    vista = request.GET.get('vista', 'graficas')  # 'graficas' | 'tablas'

    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')

    # Exigir institución O folio de pedido para no procesar todos los pedidos (mejor performance)
    tiene_filtro = bool(institucion_id or folio_pedido)
    surtido_completo = []
    no_surtido = []
    surtido_parcial = []

    if tiene_filtro:
        items = ItemPropuesta.objects.select_related(
            'propuesta__solicitud__institucion_solicitante',
            'producto',
        ).prefetch_related('lotes_asignados').filter(
            propuesta__solicitud__institucion_solicitante__activo=True,
        ).exclude(propuesta__solicitud__estado='CANCELADA')

        if institucion_id:
            items = items.filter(propuesta__solicitud__institucion_solicitante_id=institucion_id)
        if folio_pedido:
            items = items.filter(propuesta__solicitud__observaciones_solicitud__icontains=folio_pedido)
        if fecha_desde:
            try:
                fd = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                items = items.filter(propuesta__solicitud__fecha_solicitud__date__gte=fd)
            except ValueError:
                pass
        if fecha_hasta:
            try:
                fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                items = items.filter(propuesta__solicitud__fecha_solicitud__date__lte=fh)
            except ValueError:
                pass

        for item in items:
            sol = item.cantidad_solicitada or 0
            # Cantidad realmente suministrada: suma de LoteAsignado donde surtido=True (como en reporte de pedidos)
            sur = sum(la.cantidad_asignada for la in item.lotes_asignados.all() if la.surtido)
            if sur == 0:
                no_surtido.append({
                    'item': item,
                    'cantidad_solicitada': sol,
                    'cantidad_surtida': sur,
                    'diferencia': sol,
                    'folio': item.propuesta.solicitud.folio if item.propuesta and item.propuesta.solicitud else '-',
                })
            elif sur >= sol:
                surtido_completo.append({
                    'item': item,
                    'cantidad_solicitada': sol,
                    'cantidad_surtida': sur,
                    'diferencia': 0,
                    'folio': item.propuesta.solicitud.folio if item.propuesta and item.propuesta.solicitud else '-',
                })
            else:
                surtido_parcial.append({
                    'item': item,
                    'cantidad_solicitada': sol,
                    'cantidad_surtida': sur,
                    'diferencia': sol - sur,
                    'folio': item.propuesta.solicitud.folio if item.propuesta and item.propuesta.solicitud else '-',
                })

    institucion_seleccionada = None
    if institucion_id:
        institucion_seleccionada = Institucion.objects.filter(id=institucion_id).first()

    from urllib.parse import urlencode
    q = request.GET.copy()
    q['vista'] = 'graficas'
    url_graficas = '?' + q.urlencode()
    q['vista'] = 'tablas'
    url_tablas = '?' + q.urlencode()

    context = {
        'instituciones': instituciones,
        'institucion_seleccionada': institucion_seleccionada,
        'filtro_institucion': institucion_id,
        'filtro_folio_pedido': folio_pedido,
        'filtro_fecha_desde': fecha_desde,
        'filtro_fecha_hasta': fecha_hasta,
        'vista': vista,
        'tiene_filtro': tiene_filtro,
        'url_graficas': url_graficas,
        'url_tablas': url_tablas,
        'surtido_completo': surtido_completo,
        'no_surtido': no_surtido,
        'surtido_parcial': surtido_parcial,
        'total_completo': len(surtido_completo),
        'total_no_surtido': len(no_surtido),
        'total_parcial': len(surtido_parcial),
        'page_title': 'Dashboard de Surtimiento por Institución',
    }
    return render(request, 'inventario/pedidos/dashboard_surtimiento_institucion.html', context)


def _obtener_datos_dashboard_surtimiento(request):
    """
    Obtiene las listas surtido_completo, no_surtido, surtido_parcial con los mismos
    filtros que el dashboard (para reutilizar en vista y exportación Excel).
    Retorna (surtido_completo, no_surtido, surtido_parcial) o None si no hay filtro.
    """
    from datetime import datetime

    institucion_id = request.GET.get('institucion', '').strip()
    folio_pedido = request.GET.get('folio_pedido', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    if not (institucion_id or folio_pedido):
        return None

    items = ItemPropuesta.objects.select_related(
        'propuesta__solicitud__institucion_solicitante',
        'producto',
    ).prefetch_related('lotes_asignados').filter(
        propuesta__solicitud__institucion_solicitante__activo=True,
    ).exclude(propuesta__solicitud__estado='CANCELADA')

    if institucion_id:
        items = items.filter(propuesta__solicitud__institucion_solicitante_id=institucion_id)
    if folio_pedido:
        items = items.filter(propuesta__solicitud__observaciones_solicitud__icontains=folio_pedido)
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            items = items.filter(propuesta__solicitud__fecha_solicitud__date__gte=fd)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            items = items.filter(propuesta__solicitud__fecha_solicitud__date__lte=fh)
        except ValueError:
            pass

    surtido_completo = []
    no_surtido = []
    surtido_parcial = []

    for item in items:
        sol = item.cantidad_solicitada or 0
        sur = sum(la.cantidad_asignada for la in item.lotes_asignados.all() if la.surtido)
        folio = item.propuesta.solicitud.folio if item.propuesta and item.propuesta.solicitud else '-'
        clave = (item.producto.clave_cnis or '-') if item.producto else '-'
        desc = ((item.producto.descripcion or '-')[:200]) if item.producto else '-'
        row = {
            'folio': folio,
            'clave_cnis': clave,
            'descripcion': desc,
            'cantidad_solicitada': sol,
            'cantidad_surtida': sur,
            'diferencia': sol - sur if sur < sol else 0,
        }
        if sur == 0:
            row['diferencia'] = sol
            no_surtido.append(row)
        elif sur >= sol:
            surtido_completo.append(row)
        else:
            row['diferencia'] = sol - sur
            surtido_parcial.append(row)

    return (surtido_completo, no_surtido, surtido_parcial)


@login_required
def exportar_dashboard_surtimiento_excel(request):
    """Exporta el dashboard de surtimiento a Excel con los mismos filtros (institución o folio obligatorio)."""
    from datetime import date

    datos = _obtener_datos_dashboard_surtimiento(request)
    if datos is None:
        from django.shortcuts import redirect
        from django.contrib import messages
        messages.warning(request, 'Seleccione una institución o un folio de pedido para exportar.')
        return redirect('pedidos:dashboard_surtimiento_institucion')

    surtido_completo, no_surtido, surtido_parcial = datos

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def escribir_hoja(ws, titulo, filas, color_header='1F4E78'):
        headers = ['Folio', 'Clave CNIS', 'Descripción', 'Cant. solicitada', 'Cant. surtida', 'Diferencia']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col)
            c.value = h
            c.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border
        for row_idx, r in enumerate(filas, 2):
            ws.cell(row=row_idx, column=1).value = r.get('folio', '')
            ws.cell(row=row_idx, column=2).value = r.get('clave_cnis', '')
            ws.cell(row=row_idx, column=3).value = r.get('descripcion', '')
            ws.cell(row=row_idx, column=4).value = r.get('cantidad_solicitada', 0)
            ws.cell(row=row_idx, column=5).value = r.get('cantidad_surtida', 0)
            ws.cell(row=row_idx, column=6).value = r.get('diferencia', 0)
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = border
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12

    ws1 = wb.active
    ws1.title = 'Surtido completo'
    escribir_hoja(ws1, 'Surtido completo', surtido_completo, '28a745')
    ws2 = wb.create_sheet('No surtido')
    escribir_hoja(ws2, 'No surtido', no_surtido, 'dc3545')
    ws3 = wb.create_sheet('Surtido parcial')
    escribir_hoja(ws3, 'Surtido parcial', surtido_parcial, 'ffc107')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha_str = date.today().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="dashboard_surtimiento_{fecha_str}.xlsx"'
    wb.save(response)
    return response
