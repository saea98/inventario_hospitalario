"""
Vistas para el módulo de Gestión de Pedidos (Fase 2.2.1)
"""

import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone
from datetime import date, datetime, time
from django.forms import inlineformset_factory
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .pedidos_models import SolicitudPedido, ItemSolicitud, PropuestaPedido, ItemPropuesta, Producto
from .pedidos_forms import (
    SolicitudPedidoForm,
    SolicitudPedidoEdicionForm,
    ItemSolicitudForm,
    FiltroSolicitudesForm,
    ValidarSolicitudPedidoForm,
    BulkUploadForm,
    verificar_folio_pedido_duplicado,
)
from .propuesta_generator import PropuestaGenerator
from .propuesta_utils import cancelar_propuesta, eliminar_propuesta, validar_disponibilidad_para_propuesta, validar_disponibilidad_solicitud
from .pedidos_utils import registrar_error_pedido
from .pedidos_forms import _usuario_puede_duplicar_folio
from django.db import models

# ============================================================================
# VISTAS DE GESTIÓN DE PEDIDOS
# ============================================================================

@login_required
def lista_solicitudes(request):
    """
    Muestra una lista de todas las solicitudes de pedido, con filtros, paginación y exportación a Excel.
    """
    solicitudes = SolicitudPedido.objects.select_related(
        'institucion_solicitante', 'almacen_destino', 'usuario_solicitante'
    ).all().order_by('-fecha_solicitud')
    
    form = FiltroSolicitudesForm(request.GET)
    
    # Aplicar filtros directamente desde request.GET para que funcionen incluso si el form no es válido
    estado = request.GET.get('estado', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    institucion = request.GET.get('institucion', '').strip()
    folio = request.GET.get('folio', '').strip()
    
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    
    if fecha_inicio:
        try:
            from django.utils.dateparse import parse_date
            fecha_inicio_obj = parse_date(fecha_inicio)
            if fecha_inicio_obj:
                solicitudes = solicitudes.filter(fecha_solicitud__gte=fecha_inicio_obj)
        except:
            pass
    
    if fecha_fin:
        try:
            from django.utils.dateparse import parse_date
            fecha_fin_obj = parse_date(fecha_fin)
            if fecha_fin_obj:
                # Agregar 23:59:59 para incluir todo el día
                from django.utils import timezone
                from datetime import datetime, time
                fecha_fin_completa = timezone.make_aware(datetime.combine(fecha_fin_obj, time.max))
                solicitudes = solicitudes.filter(fecha_solicitud__lte=fecha_fin_completa)
        except:
            pass
    
    if institucion:
        # Buscar en denominacion (campo principal) y nombre (campo opcional)
        from django.db.models import Q
        solicitudes = solicitudes.filter(
            Q(institucion_solicitante__denominacion__icontains=institucion) |
            Q(institucion_solicitante__nombre__icontains=institucion) |
            Q(institucion_solicitante__clue__icontains=institucion)
        )
    
    if folio:
        # Buscar en el campo observaciones_solicitud y también en el folio del sistema
        from django.db.models import Q
        solicitudes = solicitudes.filter(
            Q(observaciones_solicitud__icontains=folio) |
            Q(folio__icontains=folio)
        )
    
    # Verificar si se solicita exportación a Excel
    if request.GET.get('export') == 'excel':
        return exportar_solicitudes_excel(solicitudes)
    
    # Paginación
    paginator = Paginator(solicitudes, 20)  # 20 solicitudes por página
    page = request.GET.get('page')
    
    try:
        solicitudes_pagina = paginator.page(page)
    except PageNotAnInteger:
        solicitudes_pagina = paginator.page(1)
    except EmptyPage:
        solicitudes_pagina = paginator.page(paginator.num_pages)
            
    context = {
        'solicitudes': solicitudes_pagina,
        'page_obj': solicitudes_pagina,
        'form': form,
        'page_title': 'Gestión de Pedidos'
    }
    return render(request, 'inventario/pedidos/lista_solicitudes.html', context)


def exportar_solicitudes_excel(solicitudes):
    """
    Exporta las solicitudes de pedido a Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes de Pedido"
    
    # Definir estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_num = 1
    
    # Título del reporte
    ws.merge_cells(f'A{row_num}:F{row_num}')
    ws[f'A{row_num}'] = 'REPORTE DE SOLICITUDES DE PEDIDO'
    ws[f'A{row_num}'].font = Font(bold=True, size=14)
    ws[f'A{row_num}'].alignment = Alignment(horizontal="center", vertical="center")
    row_num += 2
    
    # Encabezados
    headers = [
        'Folio (Observaciones)',
        'Institución Solicitante',
        'Almacén Destino',
        'Fecha de Solicitud',
        'Estado',
        'Folio Sistema'
    ]
    
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    row_num += 1
    
    # Datos
    for solicitud in solicitudes:
        row = [
            solicitud.observaciones_solicitud or '',
            solicitud.institucion_solicitante.denominacion if solicitud.institucion_solicitante else 'N/A',
            solicitud.almacen_destino.nombre if solicitud.almacen_destino else 'N/A',
            solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_solicitud else '',
            solicitud.get_estado_display(),
            solicitud.folio or ''
        ]
        
        ws.append(row)
        
        # Aplicar bordes
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        row_num += 1
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="solicitudes_pedido.xlsx"'
    
    wb.save(response)
    return response


@login_required
def verificar_folio_pedido(request):
    """
    API para el frontend: verifica si ya existe un pedido con el folio dado.
    GET: folio=... (folio del pedido, ej. observaciones_solicitud), solicitud_id=... (opcional, excluir al editar).
    Retorna JSON: { existe_activo: bool, existe_cancelado: bool }
    El rol administrador_local puede capturar folios duplicados: para ese perfil se retorna siempre sin duplicado.
    """
    folio = (request.GET.get('folio') or '').strip()
    solicitud_id = request.GET.get('solicitud_id', '').strip() or None
    if _usuario_puede_duplicar_folio(request.user):
        return JsonResponse({'existe_activo': False, 'existe_cancelado': False})
    existe_activo, existe_cancelado = verificar_folio_pedido_duplicado(folio, excluir_solicitud_id=solicitud_id)
    return JsonResponse({'existe_activo': existe_activo, 'existe_cancelado': existe_cancelado})


@login_required
@transaction.atomic
def crear_solicitud(request):
    """
    Permite a un usuario crear una nueva solicitud de pedido y añadirle items,
    incluyendo la opción de carga masiva por CSV.
    """
    upload_form = BulkUploadForm()
    ItemSolicitudFormSet = inlineformset_factory(SolicitudPedido, ItemSolicitud, form=ItemSolicitudForm, extra=1, can_delete=True)

    if request.method == 'POST':
        if 'upload_csv' in request.POST:
            upload_form = BulkUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                csv_file = request.FILES['csv_file']
                try:
                    items_data = []
                    # Intentar decodificar con múltiples codificaciones
                    decoded_file = None
                    codificaciones = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                    contenido_bytes = csv_file.read()
                    
                    for codificacion in codificaciones:
                        try:
                            decoded_file = contenido_bytes.decode(codificacion)
                            break
                        except (UnicodeDecodeError, AttributeError):
                            continue
                    
                    if decoded_file is None:
                        raise ValueError('No se pudo decodificar el archivo con ninguna codificación soportada')
                    
                    io_string = io.StringIO(decoded_file)
                    reader = csv.DictReader(io_string)
                    
                    for row in reader:
                        clave = row.get('CLAVE')
                        cantidad = row.get('CANTIDAD SOLICITADA')
                        
                        if clave and cantidad:
                            try:
                                cantidad_int = int(cantidad)
                            except ValueError:
                                registrar_error_pedido(
                                    usuario=request.user,
                                    tipo_error='CANTIDAD_INVALIDA',
                                    clave_solicitada=clave,
                                    cantidad_solicitada=None,
                                    descripcion_error=f"Cantidad no valida: {cantidad}",
                                    enviar_alerta=True
                                )
                                messages.warning(request, f"Cantidad invalida para clave {clave}")
                                continue
                            
                            try:
                                producto = Producto.objects.get(clave_cnis=clave)
                            except Producto.DoesNotExist:
                                registrar_error_pedido(
                                    usuario=request.user,
                                    tipo_error='CLAVE_NO_EXISTE',
                                    clave_solicitada=clave,
                                    cantidad_solicitada=cantidad_int,
                                    descripcion_error=f"Clave no existe en catalogo",
                                    enviar_alerta=True
                                )
                                messages.warning(request, f"Clave {clave} no existe")
                                continue
                            
                            # NO validar disponibilidad aquí - la validación real se hace al generar la propuesta
                            # Esto permite crear el pedido y luego en la propuesta se valida correctamente
                            # con el algoritmo completo que busca en múltiples lotes y ubicaciones
                            # resultado_validacion = validar_disponibilidad_para_propuesta(
                            #     producto.id,
                            #     cantidad_int,
                            #     None
                            # )
                            # existencia = resultado_validacion['cantidad_disponible']
                            # 
                            # if not resultado_validacion['disponible']:
                            #     registrar_error_pedido(
                            #         usuario=request.user,
                            #         tipo_error='SIN_EXISTENCIA',
                            #         clave_solicitada=clave,
                            #         cantidad_solicitada=cantidad_int,
                            #         descripcion_error=f"Insuficiente: {cantidad_int} solicitado, {existencia} disponible",
                            #         enviar_alerta=True
                            #     )
                            #     messages.warning(request, f"Sin existencia para {clave}")
                            #     continue
                            
                            items_data.append({
                                'producto': producto.id,
                                'cantidad_solicitada': cantidad_int
                            })
                    
                    ItemSolicitudFormSet = inlineformset_factory(SolicitudPedido, ItemSolicitud, form=ItemSolicitudForm, extra=len(items_data), can_delete=True)
                    formset = ItemSolicitudFormSet(initial=items_data)
                    form = SolicitudPedidoForm(user=request.user)
                    messages.success(request, f"{len(items_data)} items cargados desde el CSV.")
                    
                except Exception as e:
                    messages.error(request, f"Error al procesar el archivo CSV: {e}")
                    form = SolicitudPedidoForm(user=request.user)
                    formset = ItemSolicitudFormSet(instance=SolicitudPedido())
            else:
                messages.error(request, "Error en el formulario de carga de archivo.")
                form = SolicitudPedidoForm(user=request.user)
                formset = ItemSolicitudFormSet(instance=SolicitudPedido())
        
        else:
            form = SolicitudPedidoForm(request.POST, user=request.user)
            formset = ItemSolicitudFormSet(request.POST, instance=SolicitudPedido())
            
            if form.is_valid() and formset.is_valid():
                try:
                    if getattr(form, '_folio_existe_cancelado', False):
                        messages.warning(
                            request,
                            'Ya existe un pedido cancelado con este folio. Se permitió continuar; verifique que sea correcto.'
                        )
                    solicitud = form.save(commit=False)
                    solicitud.usuario_solicitante = request.user
                    solicitud.save()
                    
                    formset.instance = solicitud
                    formset.save()
                    messages.success(request, f"Solicitud {solicitud.folio} creada con éxito.")
                    return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
                except Exception as e:
                    messages.error(
                        request,
                        f"No se pudo guardar la solicitud. Detalle: {type(e).__name__}: {str(e)}"
                    )
            else:
                # Errores del formulario principal: mostrar con nombre legible del campo
                nombres_campos = {
                    'institucion_solicitante': 'Institución solicitante',
                    'almacen_destino': 'Almacén destino',
                    'fecha_entrega_programada': 'Fecha de entrega programada',
                    'observaciones_solicitud': 'Folio / Observaciones',
                }
                if form.errors:
                    for field, errors in form.errors.items():
                        label = nombres_campos.get(field) or (form.fields[field].label if field in form.fields else field)
                        for error in errors:
                            messages.error(request, f"Datos de la solicitud — {label}: {error}")
                # Errores del formset: indicar fila (item) y campo con error
                if formset.non_form_errors():
                    for error in formset.non_form_errors():
                        messages.error(request, f"Items: {error}")
                if formset.errors:
                    for i, item_form in enumerate(formset.forms):
                        if item_form.errors:
                            num_fila = i + 1
                            nombres_item = {
                                'producto': 'Producto',
                                'cantidad_solicitada': 'Cantidad solicitada',
                                'cantidad_aprobada': 'Cantidad aprobada',
                                'justificacion_cambio': 'Justificación',
                            }
                            for field, errors in item_form.errors.items():
                                label = nombres_item.get(field) or (item_form.fields[field].label if field in item_form.fields else field)
                                for error in errors:
                                    messages.error(request, f"Item {num_fila} — {label}: {error}")

    else:
        form = SolicitudPedidoForm(user=request.user)
        formset = ItemSolicitudFormSet(instance=SolicitudPedido())
        
    context = {
        'form': form,
        'formset': formset,
        'upload_form': upload_form,
        'page_title': 'Crear Nueva Solicitud de Pedido'
    }
    return render(request, 'inventario/pedidos/crear_solicitud.html', context)


@login_required
def detalle_solicitud(request, solicitud_id):
    """
    Muestra el detalle de una solicitud de pedido específica.
    """
    solicitud = get_object_or_404(
        SolicitudPedido.objects.select_related(
            'institucion_solicitante', 'almacen_destino', 'usuario_solicitante', 'usuario_validacion'
        ).prefetch_related('items__producto'),
        id=solicitud_id
    )
    
    propuesta = PropuestaPedido.objects.filter(solicitud=solicitud).first()

    # Lista previa: solo items con cantidad aprobada > 0 (a surtir)
    items_a_surtir = [item for item in solicitud.items.all() if item.cantidad_aprobada > 0]
    items_sin_existencia = [item for item in solicitud.items.all() if item.cantidad_aprobada == 0]

    # Validar disponibilidad para mostrar alertas
    validacion_disponibilidad = validar_disponibilidad_solicitud(solicitud_id)

    context = {
        'solicitud': solicitud,
        'propuesta': propuesta,
        'validacion_disponibilidad': validacion_disponibilidad,
        'items_a_surtir': items_a_surtir,
        'items_sin_existencia': items_sin_existencia,
        'page_title': f"Detalle de Solicitud {solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/detalle_solicitud.html', context)


@login_required
@transaction.atomic
def validar_solicitud(request, solicitud_id):
    """
    Permite a un usuario autorizado validar, modificar o rechazar los items de una solicitud.
    Genera automáticamente la propuesta de pedido si la solicitud es aprobada.
    Excluye automáticamente los items sin disponibilidad.
    """
    solicitud = get_object_or_404(SolicitudPedido, id=solicitud_id, estado='PENDIENTE')
    
    if request.method == 'POST':
        form = ValidarSolicitudPedidoForm(request.POST, solicitud=solicitud)
        if form.is_valid():
            solicitud.usuario_validacion = request.user
            solicitud.fecha_validacion = timezone.now()
            
            for item in solicitud.items.all():
                cantidad_aprobada = form.cleaned_data.get(f'item_{item.id}_cantidad_aprobada')
                justificacion = form.cleaned_data.get(f'item_{item.id}_justificacion')
                
                item.cantidad_aprobada = cantidad_aprobada
                item.justificacion_cambio = justificacion
                item.save()
            
            total_aprobado = sum(item.cantidad_aprobada for item in solicitud.items.all())
            if total_aprobado == 0:
                solicitud.estado = 'RECHAZADA'
                messages.warning(request, f"Solicitud {solicitud.folio} ha sido rechazada.")
                solicitud.save()
            else:
                # Validar disponibilidad y excluir items sin disponibilidad
                items_sin_disponibilidad = []
                items_con_disponibilidad = []
                
                for item in solicitud.items.all():
                    if item.cantidad_aprobada > 0:
                        resultado = validar_disponibilidad_para_propuesta(
                            item.producto.id,
                            item.cantidad_aprobada,
                            solicitud.institucion_solicitante.id
                        )
                        
                        # Solo excluir si NO hay disponibilidad en absoluto (cantidad_disponible == 0)
                        # Si hay disponibilidad parcial, mantener cantidad_aprobada para que el algoritmo
                        # pueda asignar lo disponible y buscar otros lotes
                        if resultado['cantidad_disponible'] == 0:
                            # No hay disponibilidad en absoluto - excluir y registrar en log
                            cantidad_solicitada_log = item.cantidad_aprobada
                            items_sin_disponibilidad.append({
                                'clave': item.producto.clave_cnis,
                                'descripcion': item.producto.descripcion,
                                'solicitado': cantidad_solicitada_log,
                                'disponible': resultado['cantidad_disponible']
                            })
                            item.cantidad_aprobada = 0
                            item.save()
                            registrar_error_pedido(
                                usuario=request.user,
                                tipo_error='SIN_EXISTENCIA',
                                clave_solicitada=item.producto.clave_cnis,
                                cantidad_solicitada=cantidad_solicitada_log,
                                descripcion_error=(
                                    f"Sin existencia al validar solicitud {solicitud.folio}. "
                                    f"Solicitado: {cantidad_solicitada_log}, disponible: 0."
                                ),
                                institucion=solicitud.institucion_solicitante,
                                almacen=solicitud.almacen_destino,
                                enviar_alerta=False
                            )
                        elif not resultado['disponible']:
                            # Hay disponibilidad parcial - mantener cantidad_aprobada
                            # El algoritmo de generación asignará lo disponible y buscará otros lotes
                            items_con_disponibilidad.append(item)
                            messages.info(
                                request,
                                f"Producto {item.producto.clave_cnis}: Se solicitaron {item.cantidad_aprobada} pero hay {resultado['cantidad_disponible']} disponibles. "
                                f"El sistema asignará lo disponible y buscará otros lotes para cubrir el resto."
                            )
                        else:
                            # Disponibilidad completa
                            items_con_disponibilidad.append(item)
                
                # Mostrar advertencia si hay items excluidos
                if items_sin_disponibilidad:
                    messages.warning(
                        request, 
                        f"Se excluyeron {len(items_sin_disponibilidad)} producto(s) de la propuesta por falta de disponibilidad. Puedes editarlos después si es necesario."
                    )
                    for item in items_sin_disponibilidad:
                        messages.warning(
                            request,
                            f"  - {item['clave']}: Se solicitaron {item['solicitado']} pero solo hay {item['disponible']} disponibles"
                        )
                
                # Marcar solicitud como VALIDADA siempre (queda registrado; items con 0 ya están en LogErrorPedido)
                solicitud.estado = 'VALIDADA'
                solicitud.save()

                # Generar propuesta solo si hay items con disponibilidad
                if items_con_disponibilidad:
                    try:
                        generator = PropuestaGenerator(solicitud.id, request.user)
                        propuesta = generator.generate()
                        cantidad_items = propuesta.items.count()
                        messages.success(
                            request,
                            f"Solicitud {solicitud.folio} validada. Propuesta generada con {cantidad_items} producto(s)."
                        )
                    except Exception as e:
                        messages.error(request, f"Error al generar la propuesta: {str(e)}")
                else:
                    # Todos los items quedaron sin existencia; ya registrados en LogErrorPedido
                    messages.success(
                        request,
                        "Solicitud validada. No se generó propuesta porque no hubo existencia para ninguno de los insumos; "
                        "quedó registrado en el log de errores de pedidos."
                    )
            
            return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
    else:
        form = ValidarSolicitudPedidoForm(solicitud=solicitud)
        
    context = {
        'solicitud': solicitud,
        'form': form,
        'page_title': f"Validar Solicitud {solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/validar_solicitud.html', context)


# ============================================================================
# VISTAS DE PROPUESTA DE PEDIDO (Para personal de almacén)
# ============================================================================

@login_required
def lista_propuestas(request):
    """
    Muestra una lista de propuestas de pedido para que el almacén las revise y surta.
    Incluye indicador de porcentaje surtido y botón de impresión para propuestas surtidas.
    Las fechas se muestran en zona horaria America/Mexico_City (UTC-6).
    """
    import pytz
    tz_mexico = pytz.timezone('America/Mexico_City')
    timezone.activate(tz_mexico)

    from django.urls import reverse
    from django.db.models import Sum
    from .models import Institucion

    propuestas = PropuestaPedido.objects.select_related(
        'solicitud__institucion_solicitante',
        'solicitud__almacen_destino',
        'solicitud__usuario_solicitante'
    ).prefetch_related('items').order_by('-fecha_generacion')

    # Filtros
    estado = request.GET.get('estado')
    if estado:
        propuestas = propuestas.filter(estado=estado)

    folio_pedido = request.GET.get('folio_pedido', '').strip()
    if folio_pedido:
        propuestas = propuestas.filter(solicitud__observaciones_solicitud__icontains=folio_pedido)

    institucion_id = request.GET.get('institucion')
    if institucion_id:
        propuestas = propuestas.filter(solicitud__institucion_solicitante_id=institucion_id)

    # Filtro por fecha de pedido (solicitud)
    fecha_pedido_desde = request.GET.get('fecha_pedido_desde', '').strip()
    fecha_pedido_hasta = request.GET.get('fecha_pedido_hasta', '').strip()
    if fecha_pedido_desde:
        try:
            d = datetime.strptime(fecha_pedido_desde, '%Y-%m-%d').date()
            start_dt = tz_mexico.localize(datetime.combine(d, time.min))
            propuestas = propuestas.filter(solicitud__fecha_solicitud__gte=start_dt.astimezone(pytz.UTC))
        except ValueError:
            pass
    if fecha_pedido_hasta:
        try:
            d = datetime.strptime(fecha_pedido_hasta, '%Y-%m-%d').date()
            end_dt = tz_mexico.localize(datetime.combine(d, time.max))
            propuestas = propuestas.filter(solicitud__fecha_solicitud__lte=end_dt.astimezone(pytz.UTC))
        except ValueError:
            pass

    # Paginación (aplicar después de filtros, antes de construir la lista)
    paginator = Paginator(propuestas, 25)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    propuestas_con_info = []
    for propuesta in page_obj:
        if not propuesta.id:
            continue

        total_solicitado = propuesta.items.aggregate(
            total=Sum('cantidad_solicitada')
        )['total'] or 0

        total_surtido = propuesta.items.aggregate(
            total=Sum('cantidad_surtida')
        )['total'] or 0

        porcentaje_surtido = 0
        if total_solicitado > 0:
            porcentaje_surtido = round((total_surtido / total_solicitado) * 100, 2)

        url_detalle = f'/logistica/propuestas/{propuesta.id}/'
        url_pdf = f'/logistica/propuestas/{propuesta.id}/acuse-pdf/'

        puede_imprimir = propuesta.estado == 'SURTIDA'

        propuestas_con_info.append({
            'propuesta': propuesta,
            'porcentaje_surtido': porcentaje_surtido,
            'total_solicitado': total_solicitado,
            'total_surtido': total_surtido,
            'puede_imprimir': puede_imprimir,
            'url_detalle': url_detalle,
            'url_pdf': url_pdf
        })

    # Query string para mantener filtros en la paginación
    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    base_query = get_copy.urlencode()

    # Obtener instituciones para el filtro
    instituciones = Institucion.objects.all().order_by('denominacion')

    context = {
        'propuestas': propuestas_con_info,
        'page_obj': page_obj,
        'paginator': paginator,
        'base_query': base_query,
        'estados': PropuestaPedido.ESTADO_CHOICES,
        'instituciones': instituciones,
        'filtro_estado': estado or '',
        'filtro_folio_pedido': folio_pedido,
        'filtro_institucion': institucion_id or '',
        'filtro_fecha_pedido_desde': fecha_pedido_desde,
        'filtro_fecha_pedido_hasta': fecha_pedido_hasta,
        'page_title': 'Propuestas de Pedido para Surtimiento'
    }
    return render(request, 'inventario/pedidos/lista_propuestas.html', context)


@login_required
def detalle_propuesta(request, propuesta_id):
    """
    Muestra el detalle de una propuesta de pedido con los lotes asignados.
    """
    propuesta = get_object_or_404(
        PropuestaPedido.objects.select_related(
            'solicitud__institucion_solicitante',
            'solicitud__almacen_destino'
        ).prefetch_related('items__lotes_asignados__lote_ubicacion__lote', 'items__lotes_asignados__lote_ubicacion__ubicacion__almacen'),
        id=propuesta_id
    )
    
    context = {
        'propuesta': propuesta,
        'page_title': f"Propuesta {propuesta.solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/detalle_propuesta.html', context)


@login_required
def auditar_surtido_documento(request):
    """
    Herramienta para comparar lo anotado en un documento físico (lista de surtido)
    con lo registrado en la propuesta en el sistema. El usuario selecciona una propuesta,
    opcionalmente adjunta imagen del documento, e ingresa las cantidades que aparecen
    en el papel (CANTIDAD SURTIDA); se muestra una comparación sistema vs documento.
    """
    propuesta = None
    propuesta_id = request.GET.get('propuesta_id') or request.POST.get('propuesta_id')
    if propuesta_id:
        propuesta = PropuestaPedido.objects.filter(id=propuesta_id).select_related(
            'solicitud__institucion_solicitante', 'solicitud__almacen_destino'
        ).prefetch_related(
            'items__producto',
            'items__lotes_asignados__lote_ubicacion__lote',
        ).first()
        if not propuesta:
            messages.warning(request, 'Propuesta no encontrada.')
            return redirect('logistica:auditar_surtido_documento')

    # Lista de propuestas para el selector: por búsqueda (folio/pedido) o últimas 100
    busqueda_folio = (request.GET.get('busqueda') or request.GET.get('q') or '').strip()
    qs_propuestas = PropuestaPedido.objects.filter(
        solicitud__isnull=False
    ).select_related('solicitud__institucion_solicitante').order_by('-fecha_generacion')
    if busqueda_folio:
        qs_propuestas = qs_propuestas.filter(
            models.Q(solicitud__observaciones_solicitud__icontains=busqueda_folio)
            | models.Q(solicitud__folio__icontains=busqueda_folio)
        )[:200]
    else:
        qs_propuestas = qs_propuestas[:100]
    propuestas_recientes = list(qs_propuestas)

    def _cantidad_surtida_real(item):
        """Cantidad realmente surtida: suma de LoteAsignado.cantidad_asignada donde surtido=True (igual que propuesta de surtimiento/dashboard)."""
        return sum(la.cantidad_asignada for la in item.lotes_asignados.all() if la.surtido)

    def _orden_surtido(item):
        """Orden: 0=completo, 1=parcial, 2=no surtido (para cuadrar con lo físico)."""
        sur = _cantidad_surtida_real(item)
        sol = item.cantidad_solicitada
        if sur >= sol:
            return 0
        return 1 if sur > 0 else 2

    def _tipo_surtido(item):
        sur = _cantidad_surtida_real(item)
        sol = item.cantidad_solicitada
        if sur >= sol:
            return 'completo'
        return 'parcial' if sur > 0 else 'no_surtido'

    def _lotes_display(item):
        parts = []
        for la in item.lotes_asignados.all():
            if getattr(la, 'lote_ubicacion', None) and getattr(la.lote_ubicacion, 'lote', None):
                parts.append(la.lote_ubicacion.lote.numero_lote)
        return ', '.join(parts) if parts else '—'

    if request.method == 'POST' and propuesta:
        # Recoger cantidades "según documento" por item
        items_con_doc = []
        for item in propuesta.items.all():
            key = f'item_{item.id}_doc'
            raw = request.POST.get(key, '').strip()
            try:
                cant_doc = int(raw) if raw else None
            except ValueError:
                cant_doc = None
            cant_sistema = _cantidad_surtida_real(item)
            coincide = (cant_doc is not None and cant_doc == cant_sistema)
            items_con_doc.append({
                'item': item,
                'cant_sistema': cant_sistema,
                'cant_documento': cant_doc,
                'coincide': coincide,
                'diferencia': (cant_doc - cant_sistema) if cant_doc is not None and cant_sistema is not None else None,
                'tipo_surtido': _tipo_surtido(item),
                'orden_surtido': _orden_surtido(item),
                'lotes_display': _lotes_display(item),
            })
        items_con_doc.sort(key=lambda r: (r['orden_surtido'], (r['item'].producto.clave_cnis or '')))
        imagen = request.FILES.get('imagen_documento')
        context = {
            'propuesta': propuesta,
            'items_con_doc': items_con_doc,
            'mostrar_resultado': True,
            'imagen_adjunta': imagen.name if imagen else None,
            'propuestas_recientes': propuestas_recientes,
            'busqueda_folio': busqueda_folio,
            'page_title': 'Auditoría: Surtido documento vs sistema',
        }
        return render(request, 'inventario/pedidos/auditar_surtido_documento.html', context)

    # Items ordenados: surtido completo, luego parcial, luego no surtido (para cuadrar con lo físico)
    # Usar cantidad surtida real (desde LoteAsignado) para estado y columna "En sistema"
    items_ordenados = []
    if propuesta:
        raw_items = list(propuesta.items.all())
        items_ordenados = [
            {
                'item': i,
                'cant_sistema': _cantidad_surtida_real(i),
                'tipo_surtido': _tipo_surtido(i),
                'lotes_display': _lotes_display(i),
            }
            for i in raw_items
        ]
        items_ordenados.sort(key=lambda r: (_orden_surtido(r['item']), (r['item'].producto.clave_cnis or '')))
    context = {
        'propuesta': propuesta,
        'items_ordenados': items_ordenados,
        'propuestas_recientes': propuestas_recientes,
        'busqueda_folio': busqueda_folio,
        'mostrar_resultado': False,
        'page_title': 'Auditar surtido con documento',
    }
    return render(request, 'inventario/pedidos/auditar_surtido_documento.html', context)


def _build_items_con_doc_auditoria(propuesta, post_data):
    """Construye la lista items_con_doc igual que en auditar_surtido_documento (para export Excel)."""
    def _cantidad_surtida_real(item):
        return sum(la.cantidad_asignada for la in item.lotes_asignados.all() if la.surtido)

    def _orden_surtido(item):
        sur, sol = _cantidad_surtida_real(item), item.cantidad_solicitada
        if sur >= sol:
            return 0
        return 1 if sur > 0 else 2

    def _tipo_surtido(item):
        sur, sol = _cantidad_surtida_real(item), item.cantidad_solicitada
        if sur >= sol:
            return 'completo'
        return 'parcial' if sur > 0 else 'no_surtido'

    def _lotes_display(item):
        parts = []
        for la in item.lotes_asignados.all():
            if getattr(la, 'lote_ubicacion', None) and getattr(la.lote_ubicacion, 'lote', None):
                parts.append(la.lote_ubicacion.lote.numero_lote)
        return ', '.join(parts) if parts else '—'

    items_con_doc = []
    for item in propuesta.items.all():
        key = f'item_{item.id}_doc'
        raw = (post_data.get(key) or '').strip()
        try:
            cant_doc = int(raw) if raw else None
        except ValueError:
            cant_doc = None
        cant_sistema = _cantidad_surtida_real(item)
        coincide = (cant_doc is not None and cant_doc == cant_sistema)
        items_con_doc.append({
            'item': item,
            'cant_sistema': cant_sistema,
            'cant_documento': cant_doc,
            'coincide': coincide,
            'diferencia': (cant_doc - cant_sistema) if cant_doc is not None and cant_sistema is not None else None,
            'tipo_surtido': _tipo_surtido(item),
            'orden_surtido': _orden_surtido(item),
            'lotes_display': _lotes_display(item),
        })
    items_con_doc.sort(key=lambda r: (r['orden_surtido'], (r['item'].producto.clave_cnis or '')))
    return items_con_doc


@login_required
def exportar_auditoria_surtido_excel(request):
    """
    Exporta a Excel el resultado de la comparación (auditar surtido con documento).
    Espera POST con propuesta_id y item_{id}_doc (mismas claves que el formulario de comparación).
    """
    if request.method != 'POST':
        return redirect('logistica:auditar_surtido_documento')
    propuesta_id = request.POST.get('propuesta_id')
    if not propuesta_id:
        messages.warning(request, 'Falta propuesta.')
        return redirect('logistica:auditar_surtido_documento')
    propuesta = PropuestaPedido.objects.filter(id=propuesta_id).select_related(
        'solicitud__institucion_solicitante', 'solicitud__almacen_destino'
    ).prefetch_related(
        'items__producto',
        'items__lotes_asignados__lote_ubicacion__lote',
    ).first()
    if not propuesta:
        messages.warning(request, 'Propuesta no encontrada.')
        return redirect('logistica:auditar_surtido_documento')
    items_con_doc = _build_items_con_doc_auditoria(propuesta, request.POST)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Auditoría surtido'
    header_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    headers = [
        'Estado', 'Clave', 'Descripción', 'Lote(s)',
        'Solicitada', 'En sistema', 'Documento dice', '¿Coincide?', 'Diferencia'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(items_con_doc, 2):
        desc = (row['item'].producto.descripcion or '')[:200]
        estado = {'completo': 'Completo', 'parcial': 'Parcial', 'no_surtido': 'No surtido'}.get(row['tipo_surtido'], row['tipo_surtido'])
        coincide_txt = '—'
        if row['cant_documento'] is not None:
            coincide_txt = 'Sí' if row['coincide'] else 'No'
        dif_txt = '' if row['diferencia'] is None else row['diferencia']
        doc_txt = '' if row['cant_documento'] is None else row['cant_documento']
        ws.cell(row=row_idx, column=1).value = estado
        ws.cell(row=row_idx, column=2).value = row['item'].producto.clave_cnis or ''
        ws.cell(row=row_idx, column=3).value = desc
        ws.cell(row=row_idx, column=4).value = row['lotes_display'] or ''
        ws.cell(row=row_idx, column=5).value = row['item'].cantidad_solicitada
        ws.cell(row=row_idx, column=6).value = row['cant_sistema']
        ws.cell(row=row_idx, column=7).value = doc_txt
        ws.cell(row=row_idx, column=8).value = coincide_txt
        ws.cell(row=row_idx, column=9).value = dif_txt
        for c in range(1, 10):
            ws.cell(row=row_idx, column=c).border = border
            ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal='left', vertical='center')

    for col, width in enumerate([14, 16, 42, 24, 12, 12, 14, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    folio_safe = (propuesta.solicitud.folio or 'propuesta').replace('/', '-')[:30]
    from datetime import date
    fecha_str = date.today().strftime('%Y%m%d')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="auditoria_surtido_{folio_safe}_{fecha_str}.xlsx"'
    wb.save(response)
    return response


@login_required
def revisar_propuesta(request, propuesta_id):
    """
    Permite al personal de almacén revisar una propuesta generada.
    """
    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id, estado='GENERADA')
    
    if request.method == 'POST':
        propuesta.estado = 'REVISADA'
        propuesta.fecha_revision = timezone.now()
        propuesta.usuario_revision = request.user
        propuesta.save()
        messages.success(request, "Propuesta revisada correctamente.")
        return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    context = {
        'propuesta': propuesta,
        'page_title': f"Revisar Propuesta {propuesta.solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/revisar_propuesta.html', context)


@login_required
@transaction.atomic
def surtir_propuesta(request, propuesta_id):
    """
    Permite al personal de almacén confirmar el surtimiento de una propuesta.
    FASE 5: Genera automáticamente movimientos de inventario.
    """
    from .fase5_utils import generar_movimientos_suministro
    
    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id, estado='REVISADA')  # 'REVISADA' = 'Revisada por Almacén'
    
    if request.method == 'POST':
        # IMPORTANTE: Generar movimientos PRIMERO. Si falla, hacemos rollback y NO marcamos como SURTIDA.
        # Así evitamos falsos positivos (propuestas SURTIDA sin movimientos de inventario).
        resultado = generar_movimientos_suministro(propuesta.id, request.user)
        if not resultado.get('exito', False):
            messages.error(
                request,
                f"No se pudo surtir la propuesta: {resultado.get('mensaje', 'Error desconocido')}. "
                "No se aplicaron cambios. Corrija el problema (ej. cantidad insuficiente) e intente de nuevo."
            )
            return redirect('logistica:surtir_propuesta', propuesta_id=propuesta.id)

        # Movimientos creados OK. Ahora marcar propuesta y lotes como surtidos.
        propuesta.estado = 'EN_SURTIMIENTO'
        propuesta.fecha_surtimiento = timezone.now()
        propuesta.usuario_surtimiento = request.user
        propuesta.save()

        for item in propuesta.items.all():
            for lote_asignado in item.lotes_asignados.all():
                lote_asignado.surtido = True
                lote_asignado.fecha_surtimiento = timezone.now()
                lote_asignado.save()

        propuesta.estado = 'SURTIDA'
        propuesta.save()

        messages.success(
            request,
            f"Propuesta surtida exitosamente. {resultado.get('mensaje', '')}"
        )
        return redirect('logistica:lista_propuestas')
    
    context = {
        'propuesta': propuesta,
        'page_title': f"Surtir Propuesta {propuesta.solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/surtir_propuesta.html', context)


@login_required
@transaction.atomic
def editar_propuesta(request, propuesta_id):
    """
    Permite al personal de almacén editar los lotes y cantidades de la propuesta.
    """
    from .pedidos_models import LoteAsignado
    from .models import LoteUbicacion
    from .propuesta_utils import reservar_cantidad_lote, liberar_cantidad_lote

    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id, estado='GENERADA')

    def _disponible_lu(lu):
        return max(0, lu.cantidad - getattr(lu, 'cantidad_reservada', 0))

    if request.method == 'POST':
        for item in propuesta.items.all():
            nueva_cantidad = request.POST.get(f'item_{item.id}_cantidad_propuesta')
            if nueva_cantidad:
                item.cantidad_propuesta = int(nueva_cantidad)
                item.save()

            lotes_actuales = list(item.lotes_asignados.select_related('lote_ubicacion__lote', 'lote_ubicacion__ubicacion').all())
            for lote_asignado in lotes_actuales:
                key_cant = f'lote_{lote_asignado.id}_cantidad'
                key_elim = f'lote_{lote_asignado.id}_eliminar'
                raw_cant = request.POST.get(key_cant)
                quiere_eliminar = request.POST.get(key_elim)
                cant_num = None
                if raw_cant is not None and str(raw_cant).strip() != '':
                    try:
                        cant_num = int(raw_cant)
                    except ValueError:
                        pass
                if quiere_eliminar or cant_num is not None and cant_num == 0:
                    lu = lote_asignado.lote_ubicacion
                    liberar_cantidad_lote(lu, lote_asignado.cantidad_asignada)
                    lote_asignado.delete()
                elif cant_num is not None and cant_num > 0:
                    lu = lote_asignado.lote_ubicacion
                    max_permitido = _disponible_lu(lu) + lote_asignado.cantidad_asignada
                    if cant_num > max_permitido:
                        cant_num = max_permitido
                        messages.warning(
                            request,
                            f"Lote {lu.lote.numero_lote} / {lu.ubicacion.codigo}: la cantidad se ajustó a {cant_num} "
                            f"(disponible restando reservas de otros pedidos)."
                        )
                    ant = lote_asignado.cantidad_asignada
                    if cant_num < ant:
                        liberar_cantidad_lote(lu, ant - cant_num)
                    elif cant_num > ant:
                        if not reservar_cantidad_lote(lu, cant_num - ant):
                            messages.error(
                                request,
                                f"No hay cantidad suficiente en {lu.lote.numero_lote} / {lu.ubicacion.codigo}. "
                                f"Se mantuvo la cantidad anterior."
                            )
                            continue
                    lote_asignado.cantidad_asignada = cant_num
                    lote_asignado.save()

            nueva_ubicacion_id = request.POST.get(f'item_{item.id}_nueva_ubicacion')
            if nueva_ubicacion_id:
                lote_ubicacion = LoteUbicacion.objects.select_related('lote', 'ubicacion').get(id=nueva_ubicacion_id)
                cantidad_nuevo = int(request.POST.get(f'item_{item.id}_cantidad_nueva_ubicacion', 0))
                if cantidad_nuevo > 0:
                    disponible = _disponible_lu(lote_ubicacion)
                    if cantidad_nuevo > disponible:
                        messages.warning(
                            request,
                            f"Lote {lote_ubicacion.lote.numero_lote} / {lote_ubicacion.ubicacion.codigo}: "
                            f"se solicitaban {cantidad_nuevo} pero solo hay {disponible} disponibles (restando reservas). "
                            f"Se asignó {disponible}."
                        )
                        cantidad_nuevo = disponible
                    if cantidad_nuevo <= 0:
                        continue
                    lote_asignado_existente = item.lotes_asignados.filter(lote_ubicacion=lote_ubicacion).first()
                    if lote_asignado_existente:
                        ant = lote_asignado_existente.cantidad_asignada
                        if cantidad_nuevo < ant:
                            liberar_cantidad_lote(lote_ubicacion, ant - cantidad_nuevo)
                        elif cantidad_nuevo > ant:
                            if not reservar_cantidad_lote(lote_ubicacion, cantidad_nuevo - ant):
                                messages.error(request, "No hay cantidad suficiente en la ubicación seleccionada. No se aplicó el cambio.")
                                continue
                        lote_asignado_existente.cantidad_asignada = cantidad_nuevo
                        lote_asignado_existente.save()
                    else:
                        if not reservar_cantidad_lote(lote_ubicacion, cantidad_nuevo):
                            messages.error(
                                request,
                                f"No hay cantidad suficiente en {lote_ubicacion.lote.numero_lote} / {lote_ubicacion.ubicacion.codigo}. "
                                f"No se agregó la asignación."
                            )
                            continue
                        LoteAsignado.objects.create(
                            item_propuesta=item,
                            lote_ubicacion=lote_ubicacion,
                            cantidad_asignada=cantidad_nuevo
                        )
        
        # Manejar nuevo item agregado
        nuevo_producto_id = request.POST.get('nuevo_item_producto')
        if nuevo_producto_id:
            from .models import Producto
            from .pedidos_models import ItemPropuesta, ItemSolicitud
            
            try:
                producto = Producto.objects.get(id=nuevo_producto_id, activo=True)
                cantidad_propuesta = int(request.POST.get('nuevo_item_cantidad', 0))
                nueva_ubicacion_id = request.POST.get('nuevo_item_ubicacion')
                cantidad_ubicacion = int(request.POST.get('nuevo_item_cantidad_ubicacion', 0))
                
                if cantidad_propuesta > 0:
                    # Buscar si ya existe un ItemSolicitud para este producto en la solicitud
                    # Si no existe, crear uno temporal (o usar el primero disponible)
                    item_solicitud = propuesta.solicitud.items.filter(producto=producto).first()
                    
                    # Si no hay item_solicitud, crear uno temporal
                    if not item_solicitud:
                        item_solicitud = ItemSolicitud.objects.create(
                            solicitud=propuesta.solicitud,
                            producto=producto,
                            cantidad_solicitada=cantidad_propuesta,
                            cantidad_aprobada=cantidad_propuesta
                        )
                    
                    # Calcular cantidad disponible
                    from django.db.models import Sum
                    cantidad_disponible = producto.lote_set.filter(estado=1).aggregate(
                        total=Sum('cantidad_disponible')
                    )['total'] or 0
                    
                    # Crear el ItemPropuesta
                    item_propuesta = ItemPropuesta.objects.create(
                        propuesta=propuesta,
                        item_solicitud=item_solicitud,
                        producto=producto,
                        cantidad_solicitada=cantidad_propuesta,
                        cantidad_propuesta=cantidad_propuesta,
                        cantidad_disponible=cantidad_disponible,
                        estado='DISPONIBLE' if cantidad_propuesta > 0 else 'NO_DISPONIBLE'
                    )
                    
                    # Si se seleccionó una ubicación, crear el LoteAsignado (validar disponible restando reservas y reservar)
                    if nueva_ubicacion_id and cantidad_ubicacion > 0:
                        lote_ubicacion = LoteUbicacion.objects.select_related('lote', 'ubicacion').get(id=nueva_ubicacion_id)
                        disponible = _disponible_lu(lote_ubicacion)
                        if cantidad_ubicacion > disponible:
                            messages.warning(
                                request,
                                f"Ubicación {lote_ubicacion.lote.numero_lote} / {lote_ubicacion.ubicacion.codigo}: "
                                f"se solicitaban {cantidad_ubicacion}, solo hay {disponible} disponibles (restando reservas). Se asignó {disponible}."
                            )
                            cantidad_ubicacion = disponible
                        if cantidad_ubicacion > 0 and reservar_cantidad_lote(lote_ubicacion, cantidad_ubicacion):
                            LoteAsignado.objects.create(
                                item_propuesta=item_propuesta,
                                lote_ubicacion=lote_ubicacion,
                                cantidad_asignada=cantidad_ubicacion
                            )
                        elif cantidad_ubicacion > 0:
                            messages.error(request, "No hay cantidad suficiente en la ubicación seleccionada (restando reservas). No se asignó lote.")
                    
                    messages.success(request, f"Item {producto.clave_cnis} agregado a la propuesta.")
            except Producto.DoesNotExist:
                messages.error(request, "El producto seleccionado no existe o no está activo.")
            except Exception as e:
                messages.error(request, f"Error al agregar el nuevo item: {str(e)}")
        
        propuesta.total_propuesto = sum(item.cantidad_propuesta for item in propuesta.items.all())
        propuesta.save()

        # Guardar borrador: quedarse en la página para seguir editando (y soporte AJAX para auto-guardado)
        is_borrador = request.POST.get('guardar_borrador')
        if is_borrador:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.utils.dateformat import DateFormat
                from django.utils.timezone import now
                return JsonResponse({
                    'success': True,
                    'message': 'Borrador guardado.',
                    'timestamp': DateFormat(now()).format('H:i'),
                })
            messages.success(request, "Borrador guardado. Puede seguir editando.")
            propuesta = get_object_or_404(
                PropuestaPedido.objects.prefetch_related(
                    'items__lotes_asignados__lote_ubicacion__lote',
                    'items__lotes_asignados__lote_ubicacion__ubicacion',
                    'items__producto',
                ).select_related('solicitud'),
                id=propuesta_id,
                estado='GENERADA',
            )
            from .models import Producto, Lote
            from datetime import date, timedelta
            productos_disponibles = Producto.objects.filter(activo=True).order_by('clave_cnis')
            fecha_minima = date.today() + timedelta(days=60)
            almacen_central_id = 1
            ubicaciones_por_item = {}
            for item in propuesta.items.select_related('producto').all():
                lotes = Lote.objects.filter(
                    producto=item.producto,
                    fecha_caducidad__gte=fecha_minima,
                    estado=1
                ).order_by('fecha_caducidad', 'numero_lote')
                lista = []
                for lote in lotes:
                    for lu in lote.ubicaciones_detalle.filter(
                        cantidad__gt=0,
                        ubicacion__almacen_id=almacen_central_id
                    ).select_related('ubicacion').order_by('ubicacion__codigo'):
                        if (lu.cantidad - getattr(lu, 'cantidad_reservada', 0)) <= 0:
                            continue
                        lista.append({'lote': lote, 'ubicacion': lu})
                ubicaciones_por_item[item.id] = lista
            for item in propuesta.items.all():
                item.ubicaciones_ordenadas_para_editar = ubicaciones_por_item.get(item.id, [])
            context = {
                'propuesta': propuesta,
                'productos_disponibles': productos_disponibles,
                'page_title': f"Editar Propuesta {propuesta.solicitud.folio}",
            }
            return render(request, 'inventario/pedidos/editar_propuesta.html', context)

        messages.success(request, "Propuesta actualizada correctamente.")
        return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    # Obtener todos los productos disponibles para el selector
    from .models import Producto, Lote, LoteUbicacion
    from datetime import date, timedelta

    productos_disponibles = Producto.objects.filter(activo=True).order_by('clave_cnis')

    # Misma regla que generación de propuestas: más próximo a caducar, >= 60 días, por disponibilidad, por ubicación
    fecha_minima = date.today() + timedelta(days=60)
    almacen_central_id = 1  # mismo criterio que PropuestaGenerator
    ubicaciones_por_item = {}
    for item in propuesta.items.select_related('producto').all():
        lotes = Lote.objects.filter(
            producto=item.producto,
            fecha_caducidad__gte=fecha_minima,
            estado=1
        ).order_by('fecha_caducidad', 'numero_lote')
        lista = []
        for lote in lotes:
            for lu in lote.ubicaciones_detalle.filter(
                cantidad__gt=0,
                ubicacion__almacen_id=almacen_central_id
            ).select_related('ubicacion').order_by('ubicacion__codigo'):
                if (lu.cantidad - getattr(lu, 'cantidad_reservada', 0)) <= 0:
                    continue
                lista.append({'lote': lote, 'ubicacion': lu})
        ubicaciones_por_item[item.id] = lista

    # Adjuntar lista a cada item para el template
    for item in propuesta.items.all():
        item.ubicaciones_ordenadas_para_editar = ubicaciones_por_item.get(item.id, [])

    context = {
        'propuesta': propuesta,
        'productos_disponibles': productos_disponibles,
        'page_title': f"Editar Propuesta {propuesta.solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/editar_propuesta.html', context)


@login_required
def obtener_ubicaciones_producto(request):
    """
    Endpoint AJAX para obtener las ubicaciones disponibles de un producto.
    """
    from .models import Lote, LoteUbicacion
    from datetime import date, timedelta
    
    producto_id = request.GET.get('producto_id')
    
    if not producto_id:
        return JsonResponse({'error': 'producto_id es requerido'}, status=400)
    
    try:
        from .models import Producto
        producto = Producto.objects.get(id=producto_id, activo=True)
        
        # Misma regla que generación de propuestas: ≥60 días por caducar, más próximo a caducar, por ubicación
        fecha_minima = date.today() + timedelta(days=60)
        lotes = Lote.objects.filter(
            producto=producto,
            fecha_caducidad__gte=fecha_minima,
            estado=1
        ).select_related('producto').order_by('fecha_caducidad', 'numero_lote')

        ubicaciones = []
        almacen_central_id = 1  # mismo criterio que PropuestaGenerator
        for lote in lotes:
            lote_ubicaciones = LoteUbicacion.objects.filter(
                lote=lote,
                cantidad__gt=0,
                ubicacion__almacen_id=almacen_central_id
            ).select_related('ubicacion').order_by('ubicacion__codigo')

            for lote_ubicacion in lote_ubicaciones:
                cantidad_disponible = lote_ubicacion.cantidad - lote_ubicacion.cantidad_reservada
                if cantidad_disponible > 0:
                    ubicaciones.append({
                        'id': lote_ubicacion.id,
                        'lote': lote.numero_lote,
                        'codigo': lote_ubicacion.ubicacion.codigo,
                        'cantidad': cantidad_disponible,
                        'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y')
                    })
        
        return JsonResponse({
            'ubicaciones': ubicaciones
        })
    
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def corregir_dato_lote(request):
    """
    Corrige el dato del lote en inventario (ej. numero_lote mal capturado) desde la edición de propuesta.
    Actualiza el registro Lote y crea un MovimientoInventario tipo AJUSTE_DATOS_LOTE.
    POST JSON: { lote_id: <int>, nuevo_numero_lote: "<string>" }
    """
    from .models import Lote, MovimientoInventario
    import json

    if request.method != 'POST':
        return JsonResponse({'exito': False, 'mensaje': 'Método no permitido'}, status=405)

    try:
        body = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        body = {}
    lote_id = body.get('lote_id')
    nuevo_numero = (body.get('nuevo_numero_lote') or '').strip()
    if not lote_id:
        return JsonResponse({'exito': False, 'mensaje': 'Falta lote_id'}, status=400)
    if not nuevo_numero:
        return JsonResponse({'exito': False, 'mensaje': 'El nuevo número de lote no puede estar vacío'}, status=400)

    try:
        lote = Lote.objects.select_related('producto', 'institucion').get(pk=lote_id)
    except Lote.DoesNotExist:
        return JsonResponse({'exito': False, 'mensaje': 'Lote no encontrado'}, status=404)

    numero_anterior = (lote.numero_lote or '').strip()
    if numero_anterior == nuevo_numero:
        return JsonResponse({'exito': False, 'mensaje': 'El nuevo número de lote es igual al actual'}, status=400)

    # unique_together = ['numero_lote', 'producto', 'institucion']
    if Lote.objects.filter(producto=lote.producto, institucion=lote.institucion, numero_lote=nuevo_numero).exclude(pk=lote_id).exists():
        return JsonResponse({
            'exito': False,
            'mensaje': f'Ya existe otro lote con el número "{nuevo_numero}" para el mismo producto e institución.'
        }, status=400)

    with transaction.atomic():
        lote.numero_lote = nuevo_numero
        lote.save(update_fields=['numero_lote'])
        cantidad_actual = lote.cantidad_disponible
        motivo = (
            f"Corrección de dato de lote en inventario (desde edición de propuesta). "
            f"Número de lote corregido de '{numero_anterior}' a '{nuevo_numero}'. "
            f"Producto: {lote.producto.clave_cnis or 'N/A'}. Sin cambio en cantidades."
        )
        MovimientoInventario.objects.create(
            lote=lote,
            tipo_movimiento='AJUSTE_DATOS_LOTE',
            cantidad=0,
            cantidad_anterior=cantidad_actual,
            cantidad_nueva=cantidad_actual,
            motivo=motivo,
            documento_referencia='CORRECCION-DATO-LOTE',
            usuario=request.user,
        )
    return JsonResponse({
        'exito': True,
        'mensaje': 'Número de lote corregido en inventario. Se generó el registro de movimiento "Ajuste a datos de lote".',
        'nuevo_numero_lote': nuevo_numero,
    })


@login_required
def cancelar_propuesta_view(request, propuesta_id):
    """
    Libera todas las cantidades reservadas y devuelve la propuesta al estado GENERADA (editable).
    Permite cambiar los ítems de suministro.
    Requiere permisos de staff o grupo Almacenero.
    """
    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id)
    
    # Validar permisos
    if not (request.user.is_staff or request.user.is_superuser or request.user.groups.filter(name='Almacenero').exists()):
        messages.error(request, "No tienes permiso para liberar propuestas.")
        return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    # Validar que la propuesta este en un estado que permita liberación
    estados_liberables = ['GENERADA', 'REVISADA', 'EN_SURTIMIENTO']
    if propuesta.estado not in estados_liberables:
        messages.error(
            request, 
            f"No se puede liberar una propuesta en estado {propuesta.get_estado_display()}."
        )
        return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    if request.method == 'POST':
        try:
            resultado = cancelar_propuesta(propuesta_id, usuario=request.user)
            
            if resultado['exito']:
                cantidad_liberada = resultado.get('cantidad_liberada', 0)
                messages.success(
                    request, 
                    f"Propuesta liberada exitosamente. Se liberaron {cantidad_liberada} unidades. "
                    f"La propuesta ha regresado al estado GENERADA (pendiente de validación) "
                    f"y la solicitud ha regresado al estado PENDIENTE para nueva aprobación."
                )
                # Redirigir a la solicitud para que pueda ser validada nuevamente
                if propuesta.solicitud:
                    return redirect('logistica:detalle_pedido', solicitud_id=propuesta.solicitud.id)
                else:
                    return redirect('logistica:lista_propuestas')
            else:
                messages.error(request, resultado['mensaje'])
                return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
        except Exception as e:
            messages.error(request, f"Error al liberar la propuesta: {str(e)}")
            return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    # GET: Redirigir al detalle de la propuesta (el modal maneja la confirmación)
    return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)


@login_required
def eliminar_propuesta_view(request, propuesta_id):
    """
    Elimina completamente una propuesta haciendo rollback de todas las reservas,
    registrando movimientos de inventario y eliminando la propuesta.
    Requiere permisos de staff o grupo Almacenero.
    """
    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id)
    
    # Validar permisos
    if not (request.user.is_staff or request.user.is_superuser or request.user.groups.filter(name='Almacenero').exists()):
        messages.error(request, "No tienes permiso para eliminar propuestas.")
        return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    # Validar que la propuesta esté en un estado que permita eliminación
    estados_eliminables = ['GENERADA', 'REVISADA', 'EN_SURTIMIENTO', 'CANCELADA']
    if propuesta.estado not in estados_eliminables:
        messages.error(
            request, 
            f"No se puede eliminar una propuesta en estado {propuesta.get_estado_display()}."
        )
        return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    if request.method == 'POST':
        try:
            resultado = eliminar_propuesta(propuesta_id, usuario=request.user)
            
            if resultado['exito']:
                messages.success(
                    request,
                    f"Propuesta eliminada exitosamente. {resultado['mensaje']}"
                )
                # Redirigir a la lista de propuestas o a la solicitud
                if propuesta.solicitud:
                    return redirect('logistica:detalle_pedido', solicitud_id=propuesta.solicitud.id)
                else:
                    return redirect('logistica:lista_propuestas')
            else:
                messages.error(request, f"Error al eliminar propuesta: {resultado['mensaje']}")
                return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
        except Exception as e:
            messages.error(request, f"Error inesperado al eliminar propuesta: {str(e)}")
            return redirect('logistica:detalle_propuesta', propuesta_id=propuesta.id)
    
    # GET: Mostrar confirmación
    cantidad_total_reservada = 0
    for item in propuesta.items.all():
        for lote_asignado in item.lotes_asignados.all():
            cantidad_total_reservada += lote_asignado.cantidad_asignada
    
    context = {
        'propuesta': propuesta,
        'cantidad_total_reservada': cantidad_total_reservada,
        'page_title': f"Eliminar Propuesta {propuesta.solicitud.folio if propuesta.solicitud else propuesta.id}"
    }
    
    return render(request, 'inventario/pedidos/eliminar_propuesta.html', context)


# ============================================================================
# VISTAS PARA EDITAR Y CANCELAR SOLICITUDES
# ============================================================================

@login_required
@transaction.atomic
def editar_solicitud(request, solicitud_id):
    """
    Permite editar los items y campos del encabezado de una solicitud PENDIENTE o VALIDADA.
    - Si PENDIENTE: simplemente edita los items y encabezado
    - Si VALIDADA: cancela la propuesta actual, recalcula y genera una nueva
    """
    solicitud = get_object_or_404(
        SolicitudPedido.objects.prefetch_related('items__producto'),
        id=solicitud_id,
        estado__in=['PENDIENTE', 'VALIDADA']
    )
    
    propuesta = PropuestaPedido.objects.filter(solicitud=solicitud).first()
    # Permitir editar aunque no haya propuesta (caso de falta de disponibilidad)
    
    ItemSolicitudFormSet = inlineformset_factory(
        SolicitudPedido, 
        ItemSolicitud, 
        form=ItemSolicitudForm, 
        extra=3,  # Permitir agregar hasta 3 items nuevos
        can_delete=True
    )
    
    if request.method == 'POST':
        # Procesar formulario del encabezado
        form_encabezado = SolicitudPedidoEdicionForm(request.POST, instance=solicitud, user=request.user)
        # Procesar formset de items
        formset = ItemSolicitudFormSet(request.POST, instance=solicitud)
        
        # Validar ambos formularios
        if form_encabezado.is_valid() and formset.is_valid():
            if getattr(form_encabezado, '_folio_existe_cancelado', False):
                messages.warning(
                    request,
                    'Ya existe un pedido cancelado con este folio. Se permitió continuar; verifique que sea correcto.'
                )
            # Si hay propuesta, cancelarla primero (libera reservas)
            if propuesta:
                resultado_cancelacion = cancelar_propuesta(propuesta.id, usuario=request.user)
                
                if not resultado_cancelacion['exito']:
                    messages.error(request, f"Error al cancelar propuesta anterior: {resultado_cancelacion['mensaje']}")
                    return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
            
            # Guardar los cambios del encabezado
            form_encabezado.save()
            
            # Guardar los cambios en los items
            instances = formset.save(commit=False)
            
            # Para items nuevos, establecer cantidad_aprobada = cantidad_solicitada si la solicitud está PENDIENTE
            for instance in instances:
                if not instance.pk:  # Item nuevo
                    if solicitud.estado == 'PENDIENTE':
                        # Si está pendiente, cantidad_aprobada = cantidad_solicitada por defecto
                        if instance.cantidad_aprobada == 0:
                            instance.cantidad_aprobada = instance.cantidad_solicitada
                instance.save()
            
            # Eliminar items marcados para eliminar
            for form in formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()
            
            # Si es PENDIENTE, solo guardar cambios
            if solicitud.estado == 'PENDIENTE':
                messages.success(request, f"Solicitud {solicitud.folio} actualizada exitosamente.")
                return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
            
            # Si es VALIDADA, validar disponibilidad y generar propuesta
            errores_disponibilidad = []
            for item in solicitud.items.all():
                if item.cantidad_aprobada > 0:
                    resultado = validar_disponibilidad_para_propuesta(
                        item.producto.id,
                        item.cantidad_aprobada,
                        solicitud.institucion_solicitante.id
                    )
                    if not resultado['disponible']:
                        errores_disponibilidad.append(
                            f"Producto {item.producto.clave_cnis}: Se requieren {item.cantidad_aprobada} pero solo hay {resultado['cantidad_disponible']} disponibles."
                        )
            
            if errores_disponibilidad:
                messages.error(request, "No se puede generar la propuesta por falta de disponibilidad:")
                for error in errores_disponibilidad:
                    messages.error(request, f"  - {error}")
                # La solicitud permanece en VALIDADA pero sin propuesta
                return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
            
            # Generar nueva propuesta
            try:
                generator = PropuestaGenerator(solicitud.id, request.user)
                nueva_propuesta = generator.generate()
                messages.success(request, f"Solicitud actualizada y nueva propuesta generada exitosamente.")
            except Exception as e:
                messages.error(request, f"Error al generar la nueva propuesta: {str(e)}")
            
            return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
        else:
            # Mostrar errores de validación
            if not form_encabezado.is_valid():
                messages.error(request, "Por favor, corrige los errores en los datos del encabezado.")
            if not formset.is_valid():
                messages.error(request, "Por favor, corrige los errores en los items.")
    else:
        form_encabezado = SolicitudPedidoEdicionForm(instance=solicitud, user=request.user)
        formset = ItemSolicitudFormSet(instance=solicitud)
        
        # Asegurar que los formsets nuevos tengan el queryset de productos
        for form in formset.forms:
            if not form.instance.pk:  # Solo para forms nuevos
                form.fields['producto'].queryset = Producto.objects.filter(activo=True)
    
    # Obtener productos para el contexto (por si acaso)
    from .models import Producto
    productos_disponibles = Producto.objects.filter(activo=True).order_by('clave_cnis')
    
    context = {
        'solicitud': solicitud,
        'propuesta': propuesta,
        'form_encabezado': form_encabezado,
        'formset': formset,
        'productos_disponibles': productos_disponibles,
        'page_title': f"Editar Solicitud {solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/editar_solicitud.html', context)


@login_required
@transaction.atomic
def cancelar_solicitud(request, solicitud_id):
    """
    Cancela una solicitud PENDIENTE o VALIDADA.
    Si tiene propuesta, libera todas las reservas antes de cancelar.
    """
    solicitud = get_object_or_404(
        SolicitudPedido,
        id=solicitud_id,
        estado__in=['PENDIENTE', 'VALIDADA']
    )
    
    propuesta = PropuestaPedido.objects.filter(solicitud=solicitud).first()
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Si hay propuesta, cancelarla primero (libera reservas)
                if propuesta:
                    resultado_cancelacion = cancelar_propuesta(propuesta.id, usuario=request.user)
                    
                    if not resultado_cancelacion['exito']:
                        messages.error(
                            request, 
                            f"Error al liberar propuesta: {resultado_cancelacion['mensaje']}"
                        )
                        return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
                    
                    # Eliminar la propuesta después de liberar
                    propuesta.delete()
                
                # Cambiar estado de solicitud a CANCELADA
                solicitud.estado = 'CANCELADA'
                solicitud.save()
                
                messages.success(
                    request, 
                    f"Solicitud {solicitud.folio} cancelada exitosamente. Todas las reservas han sido liberadas."
                )
        except Exception as e:
            messages.error(request, f"Error al cancelar la solicitud: {str(e)}")
            return redirect('logistica:detalle_pedido', solicitud_id=solicitud.id)
        
        return redirect('logistica:lista_pedidos')
    
    # GET: Mostrar confirmación
    cantidad_total_reservada = 0
    if propuesta:
        for item in propuesta.items.all():
            for lote_asignado in item.lotes_asignados.all():
                cantidad_total_reservada += lote_asignado.cantidad_asignada
    
    context = {
        'solicitud': solicitud,
        'propuesta': propuesta,
        'cantidad_total_reservada': cantidad_total_reservada,
        'page_title': f"Cancelar Solicitud {solicitud.folio}"
    }
    return render(request, 'inventario/pedidos/cancelar_solicitud.html', context)
