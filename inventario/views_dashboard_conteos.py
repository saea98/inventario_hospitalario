"""
Dashboard de Reportes de Conteos Físicos

Permite visualizar el progreso de los conteos realizados con:
- Filtros por fecha, almacén, estado, usuario
- Gráficos de resumen
- Tabla detallada de conteos
- Exportación a Excel y PDF
"""

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Count, Case, When, Value, CharField, F
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json

from .models import RegistroConteoFisico, LoteUbicacion, Lote, Almacen, Producto
from .access_control import requiere_rol


@requiere_rol('Almacenero', 'Administrador', 'Gestor de Inventario', 'Supervisión')
def dashboard_conteos(request):
    """
    Dashboard principal de conteos físicos.
    
    Muestra:
    - Tarjetas de resumen (completados, en progreso, pendientes)
    - Gráficos de progreso
    - Tabla filtrable de conteos
    """
    
    # Obtener parámetros de filtro
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    almacen_id = request.GET.get('almacen')
    estado = request.GET.get('estado')  # 'completado', 'en_progreso', 'todos'
    usuario_id = request.GET.get('usuario')
    
    # Valores por defecto
    if not fecha_desde:
        fecha_desde = timezone.now().date()
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
    
    if not fecha_hasta:
        fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    if not estado:
        estado = 'todos'
    
    # Construir query base
    query = RegistroConteoFisico.objects.select_related(
        'lote_ubicacion__lote__producto',
        'lote_ubicacion__ubicacion__almacen',
        'usuario_creacion'
    )
    
    # Filtrar por fecha (fecha_creacion)
    query = query.filter(
        fecha_creacion__date__gte=fecha_desde,
        fecha_creacion__date__lte=fecha_hasta
    )
    
    # Filtrar por almacén
    if almacen_id:
        query = query.filter(lote_ubicacion__ubicacion__almacen_id=almacen_id)
    
    # Filtrar por estado
    if estado == 'completado':
        query = query.filter(completado=True)
    elif estado == 'en_progreso':
        query = query.filter(completado=False)
    
    # Filtrar por usuario
    if usuario_id:
        query = query.filter(usuario_creacion_id=usuario_id)
    
    # Obtener conteos
    conteos = query.order_by('-fecha_actualizacion')
    
    # Calcular estadísticas
    total_conteos = conteos.count()
    conteos_completados = conteos.filter(completado=True).count()
    conteos_en_progreso = conteos.filter(completado=False).count()
    
    # Conteos por progreso
    conteos_1_3 = conteos.filter(
        primer_conteo__isnull=False,
        segundo_conteo__isnull=True,
        tercer_conteo__isnull=True
    ).count()
    
    conteos_2_3 = conteos.filter(
        primer_conteo__isnull=False,
        segundo_conteo__isnull=False,
        tercer_conteo__isnull=True
    ).count()
    
    conteos_3_3 = conteos.filter(
        tercer_conteo__isnull=False
    ).count()
    
    # Obtener almacenes para filtro
    almacenes = Almacen.objects.all().order_by('nombre')
    
    # Obtener usuarios para filtro
    from django.contrib.auth import get_user_model
    User = get_user_model()
    usuarios = User.objects.filter(
        registros_conteo_creados__isnull=False
    ).distinct().order_by('first_name', 'last_name')
    
    # Preparar datos para gráficos
    datos_grafico = {
        'completados': conteos_completados,
        'en_progreso': conteos_en_progreso,
        '1_3': conteos_1_3,
        '2_3': conteos_2_3,
        '3_3': conteos_3_3,
    }
    
    # Conteos por almacén (para gráfico de barras)
    conteos_por_almacen = {}
    for almacen in almacenes:
        count = conteos.filter(lote_ubicacion__ubicacion__almacen=almacen).count()
        if count > 0:
            conteos_por_almacen[almacen.nombre] = count
    
    # Conteos por usuario (para gráfico de barras)
    conteos_por_usuario = {}
    for usuario in usuarios:
        count = conteos.filter(usuario_creacion=usuario).count()
        if count > 0:
            nombre_usuario = f"{usuario.first_name} {usuario.last_name}".strip() or usuario.username
            conteos_por_usuario[nombre_usuario] = count
    
    # Preparar tabla con detalles
    tabla_conteos = []
    for conteo in conteos[:100]:  # Limitar a 100 registros en la tabla
        tabla_conteos.append({
            'id': conteo.id,
            'clave': conteo.lote_ubicacion.lote.producto.clave_cnis,
            'producto': conteo.lote_ubicacion.lote.producto.descripcion,
            'lote': conteo.lote_ubicacion.lote.numero_lote,
            'ubicacion': f"{conteo.lote_ubicacion.ubicacion.codigo} - {conteo.lote_ubicacion.ubicacion.almacen.nombre}",
            'progreso': conteo.progreso,
            'completado': 'Sí' if conteo.completado else 'No',
            'usuario': f"{conteo.usuario_creacion.first_name} {conteo.usuario_creacion.last_name}".strip() or conteo.usuario_creacion.username,
            'fecha': conteo.fecha_actualizacion.strftime('%d/%m/%Y %H:%M'),
            'primer_conteo': conteo.primer_conteo or '-',
            'segundo_conteo': conteo.segundo_conteo or '-',
            'tercer_conteo': conteo.tercer_conteo or '-',
        })
    
    contexto = {
        'total_conteos': total_conteos,
        'conteos_completados': conteos_completados,
        'conteos_en_progreso': conteos_en_progreso,
        'conteos_1_3': conteos_1_3,
        'conteos_2_3': conteos_2_3,
        'conteos_3_3': conteos_3_3,
        'datos_grafico': json.dumps(datos_grafico),
        'conteos_por_almacen': json.dumps(conteos_por_almacen),
        'conteos_por_usuario': json.dumps(conteos_por_usuario),
        'tabla_conteos': tabla_conteos,
        'almacenes': almacenes,
        'usuarios': usuarios,
        'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
        'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d'),
        'almacen_id': almacen_id,
        'estado': estado,
        'usuario_id': usuario_id,
    }
    
    return render(request, 'inventario/dashboard_conteos.html', contexto)


@requiere_rol('Almacenero', 'Administrador', 'Gestor de Inventario', 'Supervisión')
def exportar_conteos_excel(request):
    """
    Exportar conteos a Excel.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('Error: openpyxl no está instalado', status=500)
    
    # Obtener parámetros de filtro
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    almacen_id = request.GET.get('almacen')
    estado = request.GET.get('estado', 'todos')
    usuario_id = request.GET.get('usuario')
    
    # Valores por defecto
    if not fecha_desde:
        fecha_desde = timezone.now().date()
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
    
    if not fecha_hasta:
        fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Construir query
    query = RegistroConteoFisico.objects.select_related(
        'lote_ubicacion__lote__producto',
        'lote_ubicacion__ubicacion__almacen',
        'usuario_creacion'
    )
    
    query = query.filter(
        fecha_creacion__date__gte=fecha_desde,
        fecha_creacion__date__lte=fecha_hasta
    )
    
    if almacen_id:
        query = query.filter(lote_ubicacion__ubicacion__almacen_id=almacen_id)
    
    if estado == 'completado':
        query = query.filter(completado=True)
    elif estado == 'en_progreso':
        query = query.filter(completado=False)
    
    if usuario_id:
        query = query.filter(usuario_creacion_id=usuario_id)
    
    conteos = query.order_by('-fecha_actualizacion')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conteos"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'CLAVE (CNIS)',
        'PRODUCTO',
        'LOTE',
        'UBICACIÓN',
        'PROGRESO',
        'COMPLETADO',
        'USUARIO',
        'FECHA',
        'PRIMER CONTEO',
        'SEGUNDO CONTEO',
        'TERCER CONTEO',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row, conteo in enumerate(conteos, 2):
        datos = [
            conteo.lote_ubicacion.lote.producto.clave_cnis,
            conteo.lote_ubicacion.lote.producto.descripcion,
            conteo.lote_ubicacion.lote.numero_lote,
            f"{conteo.lote_ubicacion.ubicacion.codigo} - {conteo.lote_ubicacion.ubicacion.almacen.nombre}",
            conteo.progreso,
            'Sí' if conteo.completado else 'No',
            f"{conteo.usuario_creacion.first_name} {conteo.usuario_creacion.last_name}".strip() or conteo.usuario_creacion.username,
            conteo.fecha_actualizacion.strftime('%d/%m/%Y %H:%M'),
            conteo.primer_conteo or '',
            conteo.segundo_conteo or '',
            conteo.tercer_conteo or '',
        ]
        
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = valor
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="conteos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@requiere_rol('Almacenero', 'Administrador', 'Gestor de Inventario', 'Supervisión')
def exportar_conteos_pdf(request):
    """
    Exportar conteos a PDF.
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return HttpResponse('Error: reportlab no está instalado', status=500)
    
    # Obtener parámetros de filtro
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    almacen_id = request.GET.get('almacen')
    estado = request.GET.get('estado', 'todos')
    usuario_id = request.GET.get('usuario')
    
    # Valores por defecto
    if not fecha_desde:
        fecha_desde = timezone.now().date()
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
    
    if not fecha_hasta:
        fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Construir query
    query = RegistroConteoFisico.objects.select_related(
        'lote_ubicacion__lote__producto',
        'lote_ubicacion__ubicacion__almacen',
        'usuario_creacion'
    )
    
    query = query.filter(
        fecha_creacion__date__gte=fecha_desde,
        fecha_creacion__date__lte=fecha_hasta
    )
    
    if almacen_id:
        query = query.filter(lote_ubicacion__ubicacion__almacen_id=almacen_id)
    
    if estado == 'completado':
        query = query.filter(completado=True)
    elif estado == 'en_progreso':
        query = query.filter(completado=False)
    
    if usuario_id:
        query = query.filter(usuario_creacion_id=usuario_id)
    
    conteos = query.order_by('-fecha_actualizacion')
    
    # Crear PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="conteos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Título
    titulo = Paragraph("REPORTE DE CONTEOS FÍSICOS", title_style)
    elements.append(titulo)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Información del reporte
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=6
    )
    
    elements.append(Paragraph(f"<b>Período:</b> {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}", info_style))
    elements.append(Paragraph(f"<b>Generado:</b> {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}", info_style))
    elements.append(Paragraph(f"<b>Total de Registros:</b> {conteos.count()}", info_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Tabla
    data = [
        [
            'CLAVE',
            'PRODUCTO',
            'LOTE',
            'UBICACIÓN',
            'PROGRESO',
            'COMPLETADO',
            'USUARIO',
            'FECHA',
        ]
    ]
    
    for conteo in conteos[:50]:  # Limitar a 50 registros por PDF
        data.append([
            conteo.lote_ubicacion.lote.producto.clave_cnis[:12],
            conteo.lote_ubicacion.lote.producto.nombre[:20],
            conteo.lote_ubicacion.lote.numero_lote[:12],
            f"{conteo.lote_ubicacion.ubicacion.codigo}",
            conteo.progreso,
            'Sí' if conteo.completado else 'No',
            conteo.usuario_creacion.username[:15],
            conteo.fecha_actualizacion.strftime('%d/%m %H:%M'),
        ])
    
    table = Table(data, colWidths=[0.8*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.8*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    return response
