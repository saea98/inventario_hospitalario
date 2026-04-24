"""
Vistas para reporte de lotes y pedidos asociados.
Muestra todos los lotes que están asignados en pedidos, con paginación y filtros.

Relaciones:
- Producto (clave_cnis) -> Lote (numero_lote) -> LoteUbicacion -> LoteAsignado -> ItemPropuesta -> PropuestaPedido -> SolicitudPedido
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import date, datetime
import logging

from .pedidos_models import LoteAsignado, PropuestaPedido
from .propuesta_utils import (
    cantidad_existencia_fisica_lote_como_reporte_existencias,
    totales_reserva_activa_por_lote_ids,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
logger = logging.getLogger(__name__)


def _agrupar_lotes_pedidos(lotes_asignados_query):
    """
    Agrupa asignaciones por lote y por propuesta/pedido.

    Reserva del lote: suma de LoteAsignado con surtido=False (misma regla que
    editar propuesta / reportes de reservas), vía totales_reserva_activa_por_lote_ids.
    No usa el campo Lote.cantidad_reservada. Existencia física: Lote.cantidad_disponible
    (mismo criterio que el reporte de existencias, no la suma directa de ubicaciones).
    """
    lote_ids = list(
        lotes_asignados_query.values_list('lote_ubicacion__lote_id', flat=True).distinct()
    )
    reservas_por_lote = totales_reserva_activa_por_lote_ids(lote_ids)
    lotes_dict = {}

    for lote_asignado in lotes_asignados_query:
        try:
            lote = lote_asignado.lote_ubicacion.lote
            item_prop = lote_asignado.item_propuesta
            propuesta = item_prop.propuesta
            solicitud = propuesta.solicitud

            lote_key = lote.id

            if lote_key not in lotes_dict:
                reserva_desde_pedidos = reservas_por_lote.get(lote.id, 0)
                disp = cantidad_existencia_fisica_lote_como_reporte_existencias(lote)
                cantidad_neta = max(0, disp - reserva_desde_pedidos)
                lotes_dict[lote_key] = {
                    'lote_id': lote.id,
                    'clave': lote.producto.clave_cnis,
                    'descripcion': lote.producto.descripcion,
                    'numero_lote': lote.numero_lote,
                    'institucion': lote.institucion.denominacion if lote.institucion else 'N/A',
                    'cantidad_disponible': disp,
                    'cantidad_reservada': reserva_desde_pedidos,
                    'cantidad_neta': cantidad_neta,
                    'sobre_reserva': reserva_desde_pedidos > disp,
                    'deficit_unidades': max(0, reserva_desde_pedidos - disp),
                    'fecha_caducidad': lote.fecha_caducidad,
                    'precio_unitario': lote.precio_unitario,
                    'valor_total': lote.valor_total,
                    'pedidos': {},
                }

            pedido_key = propuesta.id

            if pedido_key not in lotes_dict[lote_key]['pedidos']:
                lotes_dict[lote_key]['pedidos'][pedido_key] = {
                    'propuesta_id': propuesta.id,
                    'solicitud_folio': solicitud.observaciones_solicitud or solicitud.folio,
                    'institucion_solicitante': solicitud.institucion_solicitante.denominacion,
                    'estado_propuesta': propuesta.get_estado_display(),
                    'fecha_generacion': propuesta.fecha_generacion,
                    'cantidad_total_asignada': 0,
                    'cantidad_total_surtida': 0,
                    'cantidad_pendiente_surtir': 0,
                }

            lotes_dict[lote_key]['pedidos'][pedido_key]['cantidad_total_asignada'] += (
                lote_asignado.cantidad_asignada
            )
            if lote_asignado.surtido:
                lotes_dict[lote_key]['pedidos'][pedido_key]['cantidad_total_surtida'] += (
                    lote_asignado.cantidad_asignada
                )

            lotes_dict[lote_key]['pedidos'][pedido_key]['cantidad_pendiente_surtir'] = max(
                0,
                lotes_dict[lote_key]['pedidos'][pedido_key]['cantidad_total_asignada']
                - lotes_dict[lote_key]['pedidos'][pedido_key]['cantidad_total_surtida'],
            )

        except Exception as e:
            logger.error(f"Error procesando lote asignado: {str(e)}")

    datos_lotes = []
    for lote_key, lote_data in lotes_dict.items():
        lote_data['pedidos'] = list(lote_data['pedidos'].values())
        lote_data['total_pedidos'] = len(lote_data['pedidos'])
        lote_data['total_cantidad_asignada'] = sum(
            p['cantidad_total_asignada'] for p in lote_data['pedidos']
        )
        lote_data['total_cantidad_surtida'] = sum(
            p['cantidad_total_surtida'] for p in lote_data['pedidos']
        )
        datos_lotes.append(lote_data)

    datos_lotes.sort(key=lambda x: x['numero_lote'])
    return datos_lotes


@login_required
def reporte_lote_pedidos(request):
    """
    Reporte de lotes y pedidos asociados.
    Muestra todos los lotes que están asignados en pedidos, con paginación y filtros.

    Reservado / neto inventario: suma de LoteAsignado (surtido=False), no el campo
    Lote.cantidad_reservada. Disponible: Lote.cantidad_disponible (misma base que el reporte de existencias).

    En la tabla de pedidos, «Pendiente de surtir» = asignado − surtido (no es saldo neto de stock).
    """
    
    # Obtener parámetros de filtro
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_folio = request.GET.get('folio', '').strip()
    filtro_fecha_desde = request.GET.get('fecha_desde', '').strip()
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    filtro_estado = request.GET.get('estado', '').strip()
    
    # Obtener todos los lotes asignados en pedidos (excluir los que ya fueron surtidos)
    lotes_asignados_query = LoteAsignado.objects.select_related(
        'lote_ubicacion__lote__producto',
        'lote_ubicacion__lote__institucion',
        'item_propuesta__propuesta__solicitud',
        'item_propuesta__propuesta',
        'item_propuesta__producto'
    ).filter(surtido=False)  # Excluir asignaciones ya surtidas
    
    # Aplicar filtros
    if filtro_clave:
        lotes_asignados_query = lotes_asignados_query.filter(
            Q(lote_ubicacion__lote__producto__clave_cnis__icontains=filtro_clave) |
            Q(lote_ubicacion__lote__producto__descripcion__icontains=filtro_clave)
        )
    
    if filtro_lote:
        lotes_asignados_query = lotes_asignados_query.filter(
            lote_ubicacion__lote__numero_lote__icontains=filtro_lote
        )
    
    if filtro_folio:
        lotes_asignados_query = lotes_asignados_query.filter(
            Q(item_propuesta__propuesta__solicitud__observaciones_solicitud__icontains=filtro_folio) |
            Q(item_propuesta__propuesta__solicitud__folio__icontains=filtro_folio)
        )
    
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            lotes_asignados_query = lotes_asignados_query.filter(
                item_propuesta__propuesta__fecha_generacion__date__gte=fecha_desde
            )
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            lotes_asignados_query = lotes_asignados_query.filter(
                item_propuesta__propuesta__fecha_generacion__date__lte=fecha_hasta
            )
        except ValueError:
            pass
    
    if filtro_estado:
        lotes_asignados_query = lotes_asignados_query.filter(
            item_propuesta__propuesta__estado=filtro_estado
        )

    datos_lotes = _agrupar_lotes_pedidos(lotes_asignados_query)

    # Paginación
    paginator = Paginator(datos_lotes, 20)  # 20 lotes por página
    page = request.GET.get('page')
    
    try:
        lotes_pagina = paginator.page(page)
    except PageNotAnInteger:
        lotes_pagina = paginator.page(1)
    except EmptyPage:
        lotes_pagina = paginator.page(paginator.num_pages)
    
    # Estados de propuesta para el filtro
    estados_propuesta = PropuestaPedido.ESTADO_CHOICES
    
    # Contexto
    context = {
        'page_title': 'Reporte de Lotes y Pedidos Asociados',
        'filtro_clave': filtro_clave,
        'filtro_lote': filtro_lote,
        'filtro_folio': filtro_folio,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
        'filtro_estado': filtro_estado,
        'estados_propuesta': estados_propuesta,
        'page_obj': lotes_pagina,
        'paginator': paginator,
        'total_lotes': paginator.count,
    }
    
    return render(request, 'inventario/reportes/reporte_lote_pedidos.html', context)


@login_required
def exportar_lote_pedidos_excel(request):
    """
    Exporta el reporte de lotes y pedidos asociados a Excel.
    Usa la misma lógica de filtrado que la vista principal.
    """
    
    # Obtener parámetros de filtro (GET o POST)
    filtro_clave = request.POST.get('clave', request.GET.get('clave', '')).strip()
    filtro_lote = request.POST.get('lote', request.GET.get('lote', '')).strip()
    filtro_folio = request.POST.get('folio', request.GET.get('folio', '')).strip()
    filtro_fecha_desde = request.POST.get('fecha_desde', request.GET.get('fecha_desde', '')).strip()
    filtro_fecha_hasta = request.POST.get('fecha_hasta', request.GET.get('fecha_hasta', '')).strip()
    filtro_estado = request.POST.get('estado', request.GET.get('estado', '')).strip()
    
    # Obtener todos los lotes asignados en pedidos (misma lógica que la vista)
    # Excluir los que ya fueron surtidos
    lotes_asignados_query = LoteAsignado.objects.select_related(
        'lote_ubicacion__lote__producto',
        'lote_ubicacion__lote__institucion',
        'item_propuesta__propuesta__solicitud',
        'item_propuesta__propuesta',
        'item_propuesta__producto'
    ).filter(surtido=False)  # Excluir asignaciones ya surtidas
    
    # Aplicar filtros (misma lógica que la vista)
    if filtro_clave:
        lotes_asignados_query = lotes_asignados_query.filter(
            Q(lote_ubicacion__lote__producto__clave_cnis__icontains=filtro_clave) |
            Q(lote_ubicacion__lote__producto__descripcion__icontains=filtro_clave)
        )
    
    if filtro_lote:
        lotes_asignados_query = lotes_asignados_query.filter(
            lote_ubicacion__lote__numero_lote__icontains=filtro_lote
        )
    
    if filtro_folio:
        lotes_asignados_query = lotes_asignados_query.filter(
            Q(item_propuesta__propuesta__solicitud__observaciones_solicitud__icontains=filtro_folio) |
            Q(item_propuesta__propuesta__solicitud__folio__icontains=filtro_folio)
        )
    
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            lotes_asignados_query = lotes_asignados_query.filter(
                item_propuesta__propuesta__fecha_generacion__date__gte=fecha_desde
            )
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            lotes_asignados_query = lotes_asignados_query.filter(
                item_propuesta__propuesta__fecha_generacion__date__lte=fecha_hasta
            )
        except ValueError:
            pass
    
    if filtro_estado:
        lotes_asignados_query = lotes_asignados_query.filter(
            item_propuesta__propuesta__estado=filtro_estado
        )

    datos_lotes = _agrupar_lotes_pedidos(lotes_asignados_query)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lotes y Pedidos"
    
    # Definir estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    info_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    info_font = Font(bold=True, size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_num = 1
    
    # Título del reporte
    ws.merge_cells(f'A{row_num}:L{row_num}')
    ws[f'A{row_num}'] = 'REPORTE DE LOTES Y PEDIDOS ASOCIADOS'
    ws[f'A{row_num}'].font = Font(bold=True, size=14)
    ws[f'A{row_num}'].alignment = Alignment(horizontal="center", vertical="center")
    row_num += 2
    
    # Resumen general
    ws.merge_cells(f'A{row_num}:L{row_num}')
    ws[f'A{row_num}'] = f'RESUMEN: {len(datos_lotes)} lotes encontrados'
    ws[f'A{row_num}'].font = info_font
    ws[f'A{row_num}'].fill = info_fill
    ws[f'A{row_num}'].alignment = Alignment(horizontal="center", vertical="center")
    row_num += 2
    
    # Encabezados de la tabla principal
    headers = [
        'Clave CNIS',
        'Descripción',
        'Número Lote',
        'Institución',
        'Cantidad Disponible (Lote, como Existencias)',
        'Cantidad Reservada (LoteAsignado activo)',
        'Cantidad Neta inventario',
        'Exceso reserva vs físico (uds)',
        'Total Pedidos',
        'Cantidad Asignada',
        'Cantidad Surtida',
        'Fecha Caducidad',
    ]
    
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    row_num += 1
    
    # Datos de lotes
    for lote in datos_lotes:
        row = [
            lote['clave'],
            lote['descripcion'][:50],  # Limitar longitud
            lote['numero_lote'],
            lote['institucion'],
            lote['cantidad_disponible'],
            lote['cantidad_reservada'],
            lote['cantidad_neta'],
            lote['deficit_unidades'],
            lote['total_pedidos'],
            lote['total_cantidad_asignada'],
            lote['total_cantidad_surtida'],
            lote['fecha_caducidad'].strftime('%d/%m/%Y') if lote['fecha_caducidad'] else 'N/A',
        ]
        
        ws.append(row)
        
        # Aplicar bordes
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        row_num += 1
        
        # Agregar detalles de pedidos del lote
        if lote['pedidos']:
            # Encabezados de pedidos
            ws.merge_cells(f'A{row_num}:L{row_num}')
            ws[f'A{row_num}'] = f'  Pedidos del Lote {lote["numero_lote"]} ({lote["clave"]})'
            ws[f'A{row_num}'].font = Font(bold=True, italic=True, size=10)
            ws[f'A{row_num}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row_num += 1
            
            headers_pedidos = [
                'Folio Solicitud',
                'Institución Solicitante',
                'Estado Propuesta',
                'Fecha Generación',
                'Cantidad Asignada',
                'Cantidad Surtida',
                'Pendiente de surtir (no es saldo de inventario)',
            ]
            
            ws.append(headers_pedidos)
            
            # Aplicar estilos a encabezados de pedidos
            for col in range(1, len(headers_pedidos) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.font = Font(bold=True, size=10)
                cell.alignment = header_alignment
                cell.border = border
            
            row_num += 1
            
            # Datos de pedidos
            for pedido in lote['pedidos']:
                row = [
                    pedido['solicitud_folio'],
                    pedido['institucion_solicitante'],
                    pedido['estado_propuesta'],
                    pedido['fecha_generacion'].strftime('%d/%m/%Y %H:%M') if pedido['fecha_generacion'] else 'N/A',
                    pedido['cantidad_total_asignada'],
                    pedido['cantidad_total_surtida'],
                    pedido['cantidad_pendiente_surtir'],
                ]
                
                ws.append(row)
                
                # Aplicar bordes
                for col in range(1, len(row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                row_num += 1
            
            row_num += 1  # Espacio entre lotes
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 14
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo
    filename = f"reporte_lotes_pedidos_{date.today().strftime('%Y%m%d')}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
