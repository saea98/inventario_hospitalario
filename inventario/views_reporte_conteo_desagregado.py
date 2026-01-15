"""
Reporte de Conteo de Almacén Desagregado por Lote y Caducidad
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Q
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

from .models import Lote, LoteUbicacion, RegistroConteoFisico, Producto, Almacen


@login_required
def reporte_conteo_desagregado(request):
    """
    Reporte de conteo de almacén desagregado por producto, lote y caducidad.
    """
    
    # Obtener filtros
    filtro_fecha_desde = request.GET.get("fecha_desde", "")
    filtro_fecha_hasta = request.GET.get("fecha_hasta", "")
    filtro_almacen = request.GET.get("almacen", "")
    filtro_solo_con_ubicacion = request.GET.get("solo_con_ubicacion", "")
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        "producto",
        "institucion",
        "almacen"
    ).all()
    
    # Aplicar filtro de almacén
    if filtro_almacen:
        lotes = lotes.filter(almacen__nombre=filtro_almacen)
    
    # Obtener conteos
    conteos = RegistroConteoFisico.objects.select_related(
        "lote_ubicacion",
        "lote_ubicacion__lote",
        "lote_ubicacion__lote__producto"
    ).all()
    
    # Aplicar filtro de fechas a conteos
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, "%Y-%m-%d").date()
            conteos = conteos.filter(fecha_creacion__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, "%Y-%m-%d").date()
            fecha_hasta = fecha_hasta + timedelta(days=1)
            conteos = conteos.filter(fecha_creacion__date__lt=fecha_hasta)
        except ValueError:
            pass
            
    # Agrupar conteos por lote
    conteos_dict = defaultdict(lambda: {
        "primer_conteo": 0,
        "segundo_conteo": 0,
        "tercer_conteo": 0
    })
    
    for conteo in conteos:
        if not conteo.lote_ubicacion or not conteo.lote_ubicacion.lote:
            continue
        
        lote_id = conteo.lote_ubicacion.lote.id
        
        if conteo.primer_conteo:
            conteos_dict[lote_id]["primer_conteo"] += conteo.primer_conteo
        if conteo.segundo_conteo:
            conteos_dict[lote_id]["segundo_conteo"] += conteo.segundo_conteo
        if conteo.tercer_conteo:
            conteos_dict[lote_id]["tercer_conteo"] += conteo.tercer_conteo
            
    # Combinar datos
    reporte_data = []
    consecutivo = 1
    
    for lote in lotes:
        if not lote.producto:
            continue
        
        # Aplicar filtro: solo lotes con ubicación asignada
        if filtro_solo_con_ubicacion:
            ubicaciones = LoteUbicacion.objects.filter(lote=lote).exists()
            if not ubicaciones:
                continue
            
        conteo_info = conteos_dict.get(lote.id, {
            "primer_conteo": 0,
            "segundo_conteo": 0,
            "tercer_conteo": 0
        })
        
        reporte_data.append({
            "consecutivo": consecutivo,
            "fuente_financiamiento": "U013",
            "clave_cnis": lote.producto.clave_cnis,
            "descripcion": lote.producto.descripcion,
            "unidad_medida": lote.producto.unidad_medida or "PIEZA",
            "lote": lote.numero_lote,
            "fecha_caducidad": lote.fecha_caducidad.strftime("%d/%m/%Y") if lote.fecha_caducidad else "N/A",
            "cantidad": lote.cantidad_disponible,
            "importe": lote.cantidad_disponible * (lote.precio_unitario or 0),
            "cifra_primer_conteo": conteo_info.get("primer_conteo", 0),
            "cifra_segundo_conteo": conteo_info.get("segundo_conteo", 0),
            "cifra_tercer_conteo": conteo_info.get("tercer_conteo", 0),
        })
        
        consecutivo += 1
        
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(reporte_data, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    # Obtener almacenes
    almacenes = Almacen.objects.all().order_by("nombre")
    
    # Calcular totales
    total_cantidad = sum(r["cantidad"] for r in reporte_data)
    total_importe = sum(r["importe"] for r in reporte_data)
    total_primer_conteo = sum(r["cifra_primer_conteo"] for r in reporte_data)
    total_segundo_conteo = sum(r["cifra_segundo_conteo"] for r in reporte_data)
    total_tercer_conteo = sum(r["cifra_tercer_conteo"] for r in reporte_data)
    
    context = {
        "page_obj": page_obj,
        "total_registros": len(reporte_data),
        "total_cantidad": total_cantidad,
        "total_importe": total_importe,
        "total_primer_conteo": total_primer_conteo,
        "total_segundo_conteo": total_segundo_conteo,
        "total_tercer_conteo": total_tercer_conteo,
        "almacenes": almacenes,
        "filtro_fecha_desde": filtro_fecha_desde,
        "filtro_fecha_hasta": filtro_fecha_hasta,
        "filtro_almacen": filtro_almacen,
        "filtro_solo_con_ubicacion": filtro_solo_con_ubicacion,
    }
    
    return render(request, "inventario/reporte_conteo_desagregado.html", context)


@login_required
def exportar_conteo_desagregado_excel(request):
    """
    Exporta el reporte de conteo desagregado a Excel.
    """
    
    # Obtener filtros
    filtro_fecha_desde = request.GET.get("fecha_desde", "")
    filtro_fecha_hasta = request.GET.get("fecha_hasta", "")
    filtro_almacen = request.GET.get("almacen", "")
    filtro_solo_con_ubicacion = request.GET.get("solo_con_ubicacion", "")
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        "producto",
        "institucion",
        "almacen"
    ).all()
    
    # Aplicar filtro de almacén
    if filtro_almacen:
        lotes = lotes.filter(almacen__nombre=filtro_almacen)
    
    # Obtener conteos
    conteos = RegistroConteoFisico.objects.select_related(
        "lote_ubicacion",
        "lote_ubicacion__lote",
        "lote_ubicacion__lote__producto"
    ).all()
    
    # Aplicar filtro de fechas a conteos
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, "%Y-%m-%d").date()
            conteos = conteos.filter(fecha_creacion__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, "%Y-%m-%d").date()
            fecha_hasta = fecha_hasta + timedelta(days=1)
            conteos = conteos.filter(fecha_creacion__date__lt=fecha_hasta)
        except ValueError:
            pass
            
    # Agrupar conteos por lote
    conteos_dict = defaultdict(lambda: {
        "primer_conteo": 0,
        "segundo_conteo": 0,
        "tercer_conteo": 0
    })
    
    for conteo in conteos:
        if not conteo.lote_ubicacion or not conteo.lote_ubicacion.lote:
            continue
        
        lote_id = conteo.lote_ubicacion.lote.id
        
        if conteo.primer_conteo:
            conteos_dict[lote_id]["primer_conteo"] += conteo.primer_conteo
        if conteo.segundo_conteo:
            conteos_dict[lote_id]["segundo_conteo"] += conteo.segundo_conteo
        if conteo.tercer_conteo:
            conteos_dict[lote_id]["tercer_conteo"] += conteo.tercer_conteo
            
    # Combinar datos
    reporte_data = []
    consecutivo = 1
    
    for lote in lotes:
        if not lote.producto:
            continue
        
        # Aplicar filtro: solo lotes con ubicación asignada
        if filtro_solo_con_ubicacion:
            ubicaciones = LoteUbicacion.objects.filter(lote=lote).exists()
            if not ubicaciones:
                continue
            
        conteo_info = conteos_dict.get(lote.id, {
            "primer_conteo": 0,
            "segundo_conteo": 0,
            "tercer_conteo": 0
        })
        
        reporte_data.append({
            "consecutivo": consecutivo,
            "fuente_financiamiento": "U013",
            "clave_cnis": lote.producto.clave_cnis,
            "descripcion": lote.producto.descripcion,
            "unidad_medida": lote.producto.unidad_medida or "PIEZA",
            "lote": lote.numero_lote,
            "fecha_caducidad": lote.fecha_caducidad.strftime("%d/%m/%Y") if lote.fecha_caducidad else "N/A",
            "cantidad": lote.cantidad_disponible,
            "importe": lote.cantidad_disponible * (lote.precio_unitario or 0),
            "cifra_primer_conteo": conteo_info.get("primer_conteo", 0),
            "cifra_segundo_conteo": conteo_info.get("segundo_conteo", 0),
            "cifra_tercer_conteo": conteo_info.get("tercer_conteo", 0),
        })
        
        consecutivo += 1
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo Desagregado"
    
    # Estilos
    header_fill = PatternFill(start_color="8B1538", end_color="8B1538", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Encabezados
    headers = [
        "#",
        "Fuente Financiamiento",
        "Clave CNIS",
        "Descripción",
        "U.M.",
        "Lote",
        "Caducidad",
        "Cantidad",
        "Importe",
        "1er Conteo",
        "2do Conteo",
        "3er Conteo"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    for row_num, item in enumerate(reporte_data, 2):
        ws.cell(row=row_num, column=1).value = item["consecutivo"]
        ws.cell(row=row_num, column=2).value = item["fuente_financiamiento"]
        ws.cell(row=row_num, column=3).value = item["clave_cnis"]
        ws.cell(row=row_num, column=4).value = item["descripcion"]
        ws.cell(row=row_num, column=5).value = item["unidad_medida"]
        ws.cell(row=row_num, column=6).value = item["lote"]
        ws.cell(row=row_num, column=7).value = item["fecha_caducidad"]
        ws.cell(row=row_num, column=8).value = item["cantidad"]
        ws.cell(row=row_num, column=9).value = round(item["importe"], 2)
        ws.cell(row=row_num, column=10).value = item["cifra_primer_conteo"]
        ws.cell(row=row_num, column=11).value = item["cifra_segundo_conteo"]
        ws.cell(row=row_num, column=12).value = item["cifra_tercer_conteo"]
        
        # Aplicar bordes
        for col_num in range(1, 13):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Fila de totales
    total_row = len(reporte_data) + 2
    ws.cell(row=total_row, column=1).value = "TOTALES"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    total_cantidad = sum(r["cantidad"] for r in reporte_data)
    total_importe = sum(r["importe"] for r in reporte_data)
    total_primer_conteo = sum(r["cifra_primer_conteo"] for r in reporte_data)
    total_segundo_conteo = sum(r["cifra_segundo_conteo"] for r in reporte_data)
    total_tercer_conteo = sum(r["cifra_tercer_conteo"] for r in reporte_data)
    
    ws.cell(row=total_row, column=8).value = total_cantidad
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=9).value = round(total_importe, 2)
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    ws.cell(row=total_row, column=10).value = total_primer_conteo
    ws.cell(row=total_row, column=10).font = Font(bold=True)
    ws.cell(row=total_row, column=11).value = total_segundo_conteo
    ws.cell(row=total_row, column=11).font = Font(bold=True)
    ws.cell(row=total_row, column=12).value = total_tercer_conteo
    ws.cell(row=total_row, column=12).font = Font(bold=True)
    
    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12
    
    # Generar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=reporte_conteo_desagregado.xlsx"
    wb.save(response)
    
    return response
