"""
Reporte de Entradas al Inventario

Muestra todas las entradas (MovimientoInventario con tipo ENTRADA)
con el layout oficial: RFC, PROVEEDOR, PARTIDA, Clave CNIS, Producto, etc. (32 columnas).
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Subquery, OuterRef
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import MovimientoInventario, Institucion, Almacen, Proveedor, UbicacionAlmacen
from .llegada_models import LlegadaProveedor

# Layout del reporte de entradas (34 columnas, incl. estado llegada y ubicación)
ENTRADAS_LAYOUT_HEADERS = [
    'RFC',
    'PROVEEDOR',
    'PARTIDA',
    'Clave CNIS',
    'Producto',
    'UNIDAD DE MEDIDA',
    'CANTIDAD RECIBIDA',
    'FECHA DE ENTREGA',
    'LUGAR DE ENTREGA',
    'CONTRATO',
    'REMISION',
    'ORDEN DE SUMINISTRO',
    'LOTE',
    'CADUCIDAD',
    'FOLIO',
    'ESTADO LLEGADA',
    'UBIC. ASIGNADA',
    'PRECIO UNIT. CON IVA / SIN IVA',
    'SUBTOTAL',
    'IVA',
    'IMPORTE TOTAL',
    'MARCA',
    'FABRICANTE',
    'FECHA DE FABRICACIÓN',
    'CADUCIDAD CONFORME A CERTIFICADO',
    'TIPO DE ENTREGA',
    'LICITACION/PROCEDIMIENTO',
    'RESPONSABLE',
    'REVISO',
    'TIPO DE RED',
    'FECHA DE CAPTURA',
    'OBSERVACION',
    'FECHA DE EMISIÓN',
    'FUENTE DE FMTO',
    'USUARIO MOVIMIENTO',
]

# Etiquetas de estado de llegada para el reporte
ESTADO_LLEGADA_LABELS = dict(LlegadaProveedor.ESTADO_CHOICES)


def _valor(o, default=''):
    if o is None:
        return default
    return o


def _fecha(d, fmt='%d/%m/%Y'):
    if not d:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime(fmt)
    return str(d)


def _decimal(d, default=''):
    if d is None:
        return default
    if isinstance(d, Decimal):
        return float(d)
    return d


@login_required
def reporte_entradas(request):
    """
    Reporte de entradas al inventario con layout oficial (32 columnas).
    Muestra todos los MovimientoInventario con tipo ENTRADA.
    """
    subq_estado = LlegadaProveedor.objects.filter(folio=OuterRef('folio')).values('estado')[:1]
    subq_ubicacion = LlegadaProveedor.objects.filter(folio=OuterRef('folio')).values('usuario_ubicacion')[:1]
    entradas = MovimientoInventario.objects.filter(
        tipo_movimiento='ENTRADA', anulado=False
    ).annotate(
        estado_llegada=Subquery(subq_estado),
        llegada_tiene_ubicacion=Subquery(subq_ubicacion),
    ).select_related(
        'lote',
        'lote__producto',
        'lote__institucion',
        'lote__almacen',
        'lote__ubicacion',
        'lote__orden_suministro__proveedor',
        'lote__orden_suministro__fuente_financiamiento',
        'usuario'
    ).order_by('-fecha_movimiento')
    
    # Filtros
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_institucion = request.GET.get('institucion', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_ubicacion = request.GET.get('ubicacion', '')
    filtro_proveedor = request.GET.get('proveedor', '')
    filtro_estado_llegada = request.GET.get('estado_llegada', '').strip()
    filtro_con_ubicacion = request.GET.get('con_ubicacion', '') == '1'
    
    # Aplicar filtro de fecha desde
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            entradas = entradas.filter(fecha_movimiento__date__gte=fecha_desde)
        except ValueError:
            pass
    
    # Aplicar filtro de fecha hasta
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            # Sumar 1 día para incluir todo el día hasta
            fecha_hasta = fecha_hasta + timedelta(days=1)
            entradas = entradas.filter(fecha_movimiento__date__lt=fecha_hasta)
        except ValueError:
            pass
    
    # Aplicar filtro de clave
    if filtro_clave:
        entradas = entradas.filter(lote__producto__clave_cnis__icontains=filtro_clave)
    
    # Aplicar filtro de lote
    if filtro_lote:
        entradas = entradas.filter(lote__numero_lote__icontains=filtro_lote)
    
    # Aplicar filtro de institución
    if filtro_institucion:
        entradas = entradas.filter(lote__institucion__denominacion=filtro_institucion)
    
    # Aplicar filtro de almacén
    if filtro_almacen:
        entradas = entradas.filter(lote__almacen__nombre=filtro_almacen)
    
    # Aplicar filtro de ubicación
    if filtro_ubicacion:
        entradas = entradas.filter(lote__ubicacion_id=filtro_ubicacion)
    
    # Aplicar filtro de proveedor (buscar en documento_referencia o motivo)
    if filtro_proveedor:
        entradas = entradas.filter(
            Q(documento_referencia__icontains=filtro_proveedor) |
            Q(motivo__icontains=filtro_proveedor)
        )
    
    # Filtros por estado de llegada (solo entradas vinculadas a LlegadaProveedor por folio)
    if filtro_estado_llegada:
        entradas = entradas.filter(estado_llegada=filtro_estado_llegada)
    if filtro_con_ubicacion:
        entradas = entradas.filter(llegada_tiene_ubicacion__isnull=False)
    
    # Construir lista con el layout de 34 columnas
    entradas_lista = []
    total_cantidad = 0
    total_valor = 0

    for m in entradas:
        lote = m.lote
        prod = lote.producto if lote else None
        os = lote.orden_suministro if lote else None
        prov = os.proveedor if os else None

        precio = (lote and lote.precio_unitario) or Decimal('0')
        subtotal_val = (m.subtotal or (lote and lote.subtotal)) or (m.cantidad * precio)
        iva_val = (m.iva or (lote and lote.iva)) or Decimal('0')
        importe_val = (m.importe_total or (lote and lote.importe_total)) or (m.cantidad * precio)
        total_cantidad += m.cantidad
        total_valor += float(importe_val) if importe_val else 0

        lugar_entrega = ''
        if lote:
            if lote.almacen:
                lugar_entrega = lote.almacen.nombre
            if lote.institucion and not lugar_entrega:
                lugar_entrega = lote.institucion.denominacion or ''

        estado_llegada_val = getattr(m, 'estado_llegada', None)
        ubicacion_asignada = getattr(m, 'llegada_tiene_ubicacion', None)
        estado_llegada_label = ESTADO_LLEGADA_LABELS.get(estado_llegada_val, '') if estado_llegada_val else ''
        ubicacion_label = 'Sí' if ubicacion_asignada else ('—' if estado_llegada_val is None else 'No')

        row = [
            _valor(prov and prov.rfc or (lote and lote.rfc_proveedor)),
            _valor(prov and prov.razon_social or (lote and lote.proveedor)),
            _valor(lote and lote.partida or (os and os.partida_presupuestal)),
            _valor(prod and prod.clave_cnis),
            _valor(prod and prod.descripcion),
            _valor(prod and prod.unidad_medida) or 'PIEZA',
            m.cantidad,
            _fecha(lote and lote.fecha_recepcion or (m.fecha_movimiento.date() if m.fecha_movimiento else None)),
            lugar_entrega,
            _valor(m.contrato or (lote and lote.contrato)),
            _valor(m.remision or (lote and lote.remision)),
            _valor(os and os.numero_orden),
            _valor(lote and lote.numero_lote),
            _fecha(lote and lote.fecha_caducidad),
            _valor(m.folio or (lote and lote.folio)),
            _valor(estado_llegada_label),
            _valor(ubicacion_label),
            _decimal(precio),
            _decimal(m.subtotal or (lote and lote.subtotal)) if (m.subtotal or (lote and lote.subtotal)) is not None else _decimal(subtotal_val),
            _decimal(m.iva or (lote and lote.iva)) if (m.iva or (lote and lote.iva)) is not None else _decimal(iva_val),
            _decimal(m.importe_total or (lote and lote.importe_total)) if (m.importe_total or (lote and lote.importe_total)) is not None else _decimal(importe_val),
            _valor(prod and prod.marca),
            _valor(prod and prod.fabricante),
            _fecha(lote and lote.fecha_fabricacion),
            _fecha(lote and lote.fecha_caducidad),
            _valor(m.tipo_entrega or (lote and lote.tipo_entrega)),
            _valor(m.licitacion or (lote and lote.licitacion)),
            _valor(m.responsable or (lote and lote.responsable)),
            _valor(m.reviso or (lote and lote.reviso)),
            _valor(m.tipo_red or (lote and lote.tipo_red)),
            _fecha(m.fecha_movimiento, '%d/%m/%Y %H:%M') if m.fecha_movimiento else '',
            _valor(lote and lote.observaciones),
            _fecha(os and os.fecha_orden) or _fecha(lote and lote.fecha_recepcion),
            _valor((os and os.fuente_financiamiento and os.fuente_financiamiento.nombre) or (lote and lote.fuente_datos)),
            _valor((m.usuario.get_full_name() or m.usuario.username) if getattr(m, 'usuario', None) else ''),
        ]
        entradas_lista.append({'row': row, 'id': m.id})
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(entradas_lista, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener opciones de filtro
    instituciones = Institucion.objects.all().order_by('denominacion')
    almacenes = Almacen.objects.all().order_by('nombre')
    ubicaciones = UbicacionAlmacen.objects.filter(activo=True).select_related('almacen').order_by('almacen__nombre', 'codigo')
    
    context = {
        'page_obj': page_obj,
        'headers': ENTRADAS_LAYOUT_HEADERS,
        'total_registros': len(entradas_lista),
        'total_cantidad': total_cantidad,
        'total_valor': total_valor,
        'instituciones': instituciones,
        'almacenes': almacenes,
        'ubicaciones': ubicaciones,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
        'filtro_clave': filtro_clave,
        'filtro_lote': filtro_lote,
        'filtro_institucion': filtro_institucion,
        'filtro_almacen': filtro_almacen,
        'filtro_ubicacion': filtro_ubicacion,
        'filtro_proveedor': filtro_proveedor,
        'filtro_estado_llegada': filtro_estado_llegada,
        'filtro_con_ubicacion': filtro_con_ubicacion,
        'estado_llegada_choices': LlegadaProveedor.ESTADO_CHOICES,
    }
    return render(request, 'inventario/reporte_entradas.html', context)


def _construir_fila_entrada(m):
    """Construye la fila de 34 valores para un MovimientoInventario ENTRADA (mismo orden que ENTRADAS_LAYOUT_HEADERS)."""
    lote = m.lote
    prod = lote.producto if lote else None
    os = lote.orden_suministro if lote else None
    prov = os.proveedor if os else None
    precio = (lote and lote.precio_unitario) or Decimal('0')
    subtotal_val = (m.subtotal or (lote and lote.subtotal)) or (m.cantidad * precio)
    iva_val = (m.iva or (lote and lote.iva)) or Decimal('0')
    importe_val = (m.importe_total or (lote and lote.importe_total)) or (m.cantidad * precio)
    lugar_entrega = ''
    if lote:
        lugar_entrega = (lote.almacen and lote.almacen.nombre) or (lote.institucion and lote.institucion.denominacion) or ''
    estado_llegada_val = getattr(m, 'estado_llegada', None)
    ubicacion_asignada = getattr(m, 'llegada_tiene_ubicacion', None)
    estado_llegada_label = ESTADO_LLEGADA_LABELS.get(estado_llegada_val, '') if estado_llegada_val else ''
    ubicacion_label = 'Sí' if ubicacion_asignada else ('—' if estado_llegada_val is None else 'No')
    return [
        _valor(prov and prov.rfc or (lote and lote.rfc_proveedor)),
        _valor(prov and prov.razon_social or (lote and lote.proveedor)),
        _valor(lote and lote.partida or (os and os.partida_presupuestal)),
        _valor(prod and prod.clave_cnis),
        _valor(prod and prod.descripcion),
        _valor(prod and prod.unidad_medida) or 'PIEZA',
        m.cantidad,
        _fecha(lote and lote.fecha_recepcion or (m.fecha_movimiento.date() if m.fecha_movimiento else None)),
        lugar_entrega,
        _valor(m.contrato or (lote and lote.contrato)),
        _valor(m.remision or (lote and lote.remision)),
        _valor(os and os.numero_orden),
        _valor(lote and lote.numero_lote),
        _fecha(lote and lote.fecha_caducidad),
        _valor(m.folio or (lote and lote.folio)),
        _valor(estado_llegada_label),
        _valor(ubicacion_label),
        _decimal(precio),
        _decimal(m.subtotal or (lote and lote.subtotal)) if (m.subtotal or (lote and lote.subtotal)) is not None else _decimal(subtotal_val),
        _decimal(m.iva or (lote and lote.iva)) if (m.iva or (lote and lote.iva)) is not None else _decimal(iva_val),
        _decimal(m.importe_total or (lote and lote.importe_total)) if (m.importe_total or (lote and lote.importe_total)) is not None else _decimal(importe_val),
        _valor(prod and prod.marca),
        _valor(prod and prod.fabricante),
        _fecha(lote and lote.fecha_fabricacion),
        _fecha(lote and lote.fecha_caducidad),
        _valor(m.tipo_entrega or (lote and lote.tipo_entrega)),
        _valor(m.licitacion or (lote and lote.licitacion)),
        _valor(m.responsable or (lote and lote.responsable)),
        _valor(m.reviso or (lote and lote.reviso)),
        _valor(m.tipo_red or (lote and lote.tipo_red)),
        _fecha(m.fecha_movimiento, '%d/%m/%Y %H:%M') if m.fecha_movimiento else '',
        _valor(lote and lote.observaciones),
        _fecha(os and os.fecha_orden) or _fecha(lote and lote.fecha_recepcion),
        _valor((os and os.fuente_financiamiento and os.fuente_financiamiento.nombre) or (lote and lote.fuente_datos)),
        _valor((m.usuario.get_full_name() or m.usuario.username) if getattr(m, 'usuario', None) else ''),
    ]


@login_required
def exportar_entradas_excel(request):
    """Exporta el reporte de entradas a Excel con el layout oficial (34 columnas)."""
    subq_estado = LlegadaProveedor.objects.filter(folio=OuterRef('folio')).values('estado')[:1]
    subq_ubicacion = LlegadaProveedor.objects.filter(folio=OuterRef('folio')).values('usuario_ubicacion')[:1]
    entradas = MovimientoInventario.objects.filter(
        tipo_movimiento='ENTRADA', anulado=False
    ).annotate(
        estado_llegada=Subquery(subq_estado),
        llegada_tiene_ubicacion=Subquery(subq_ubicacion),
    ).select_related(
        'lote',
        'lote__producto',
        'lote__institucion',
        'lote__almacen',
        'lote__orden_suministro__proveedor',
        'lote__orden_suministro__fuente_financiamiento',
        'usuario'
    ).order_by('-fecha_movimiento')

    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_institucion = request.GET.get('institucion', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_ubicacion = request.GET.get('ubicacion', '')
    filtro_proveedor = request.GET.get('proveedor', '')
    filtro_estado_llegada = request.GET.get('estado_llegada', '').strip()
    filtro_con_ubicacion = request.GET.get('con_ubicacion', '') == '1'

    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            entradas = entradas.filter(fecha_movimiento__date__gte=fecha_desde)
        except ValueError:
            pass
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date() + timedelta(days=1)
            entradas = entradas.filter(fecha_movimiento__date__lt=fecha_hasta)
        except ValueError:
            pass
    if filtro_clave:
        entradas = entradas.filter(lote__producto__clave_cnis__icontains=filtro_clave)
    if filtro_lote:
        entradas = entradas.filter(lote__numero_lote__icontains=filtro_lote)
    if filtro_institucion:
        entradas = entradas.filter(lote__institucion__denominacion=filtro_institucion)
    if filtro_almacen:
        entradas = entradas.filter(lote__almacen__nombre=filtro_almacen)
    if filtro_ubicacion:
        entradas = entradas.filter(lote__ubicacion_id=filtro_ubicacion)
    if filtro_proveedor:
        entradas = entradas.filter(
            Q(lote__rfc_proveedor__icontains=filtro_proveedor) |
            Q(lote__proveedor__icontains=filtro_proveedor) |
            Q(lote__orden_suministro__proveedor__razon_social__icontains=filtro_proveedor) |
            Q(documento_referencia__icontains=filtro_proveedor) | Q(motivo__icontains=filtro_proveedor)
        )
    if filtro_estado_llegada:
        entradas = entradas.filter(estado_llegada=filtro_estado_llegada)
    if filtro_con_ubicacion:
        entradas = entradas.filter(llegada_tiene_ubicacion__isnull=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Entradas"
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="E8F5E9", size=10)
    total_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    total_font = Font(bold=True, size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(ENTRADAS_LAYOUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    total_cantidad = 0
    total_importe = 0
    listado = []
    for m in entradas:
        fila = _construir_fila_entrada(m)
        listado.append(fila)
        total_cantidad += fila[6] or 0
        total_importe += float(fila[20] or 0)  # IMPORTE TOTAL (índice 20 con nuevas columnas)

    for row_num, fila in enumerate(listado, 2):
        for col_num, val in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.border = border
            if col_num in (7, 17, 18, 19, 20, 21):
                cell.alignment = Alignment(horizontal='right')

    total_row = len(listado) + 2
    ws.cell(row=total_row, column=1).value = "TOTALES"
    ws.cell(row=total_row, column=1).font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=7).value = total_cantidad
    ws.cell(row=total_row, column=7).font = total_font
    ws.cell(row=total_row, column=7).fill = total_fill
    ws.cell(row=total_row, column=21).value = total_importe
    ws.cell(row=total_row, column=21).font = total_font
    ws.cell(row=total_row, column=21).fill = total_fill
    for col in range(1, len(ENTRADAS_LAYOUT_HEADERS) + 1):
        ws.cell(row=total_row, column=col).border = border

    for col in range(1, len(ENTRADAS_LAYOUT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_entradas.xlsx"'
    wb.save(response)
    return response
