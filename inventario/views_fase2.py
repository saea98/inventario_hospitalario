"""
VISTAS PARA FASE 2: Gestión Logística
Incluye: Citas, Traslados y Conteo Físico
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from datetime import datetime

from .models import (
    CitaProveedor, OrdenTraslado, ItemTraslado, 
    ConteoFisico, ItemConteoFisico, Folio, TipoEntrega,
    Lote, Almacen, Proveedor, ListaRevision, ItemRevision
)
from .forms import (
    CitaProveedorForm, OrdenTrasladoForm, LogisticaTrasladoForm,
    CargaMasivaCitasForm, CitaProveedorEditForm, ValidarEntradaForm, RechazarEntradaForm
)
from .servicio_lista_revision import ServicioListaRevision
from .servicio_folio import ServicioFolio
from .servicios_notificaciones import notificaciones
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse


# ============================================================================
# VISTAS PARA CITAS DE PROVEEDORES
# ============================================================================

@login_required
def lista_citas(request):
    """Lista todas las citas programadas con paginación y filtros avanzados"""
    from django.db.models import Q
    
    citas = CitaProveedor.objects.all().order_by('-fecha_cita')
    
    # Filtros
    estado = request.GET.get('estado')
    proveedor = request.GET.get('proveedor')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    numero_orden = request.GET.get('numero_orden')
    numero_remision = request.GET.get('numero_remision')
    numero_contrato = request.GET.get('numero_contrato')
    clave_medicamento = request.GET.get('clave_medicamento')
    
    # Aplicar filtros
    if estado:
        citas = citas.filter(estado=estado)
    
    if proveedor:
        citas = citas.filter(proveedor__razon_social__icontains=proveedor)
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            citas = citas.filter(fecha_cita__gte=fecha_obj)
        except:
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime, timedelta
            fecha_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_obj = fecha_obj + timedelta(days=1)  # Incluir todo el día
            citas = citas.filter(fecha_cita__lt=fecha_obj)
        except:
            pass
    
    if numero_orden:
        citas = citas.filter(numero_orden_suministro__icontains=numero_orden)
    
    if numero_remision:
        citas = citas.filter(numero_orden_remision__icontains=numero_remision)
    
    if numero_contrato:
        citas = citas.filter(numero_contrato__icontains=numero_contrato)
    
    if clave_medicamento:
        citas = citas.filter(clave_medicamento__icontains=clave_medicamento)
    
    # Contar por estado (antes de paginar)
    estados_count = {
        'programada': CitaProveedor.objects.filter(estado='programada').count(),
        'autorizada': CitaProveedor.objects.filter(estado='autorizada').count(),
        'completada': CitaProveedor.objects.filter(estado='completada').count(),
        'cancelada': CitaProveedor.objects.filter(estado='cancelada').count(),
    }
    
    # Paginación
    paginator = Paginator(citas, 15)  # 15 citas por página
    page = request.GET.get('page')
    
    try:
        citas_page = paginator.page(page)
    except PageNotAnInteger:
        citas_page = paginator.page(1)
    except EmptyPage:
        citas_page = paginator.page(paginator.num_pages)
    
    context = {
        'citas': citas_page,
        'paginator': paginator,
        'page_obj': citas_page,
        'estados': CitaProveedor.ESTADOS_CITA,
        'estados_count': estados_count,
        'estado_seleccionado': estado,
        'proveedor_seleccionado': proveedor,
        'fecha_desde_seleccionada': fecha_desde,
        'fecha_hasta_seleccionada': fecha_hasta,
        'numero_orden_seleccionado': numero_orden,
        'numero_remision_seleccionado': numero_remision,
        'numero_contrato_seleccionado': numero_contrato,
        'clave_medicamento_seleccionada': clave_medicamento,
        'total_citas': paginator.count,
    }
    return render(request, 'inventario/citas/lista.html', context)

@login_required
def crear_cita(request):
    """Crear una nueva cita o cargar citas masivas"""
    
    # Determinar qué tipo de operación se está realizando
    tipo_operacion = request.POST.get('tipo_operacion', 'manual') if request.method == 'POST' else 'manual'
    
    if request.method == 'POST':
        if tipo_operacion == 'masiva':
            # Procesar carga masiva
            form_masiva = CargaMasivaCitasForm(request.POST, request.FILES)
            form_manual = CitaProveedorForm()
            
            if form_masiva.is_valid():
                try:
                    from .citas_masivas import CargaMasivaCitas
                    
                    archivo = request.FILES['archivo']
                    cargador = CargaMasivaCitas()
                    resultado = cargador.procesar_archivo(archivo)
                    
                    if resultado['exito']:
                        messages.success(
                            request,
                            f"✓ Carga completada: {resultado['citas_creadas']} citas creadas"
                        )
                        
                        if resultado['advertencias']:
                            for adv in resultado['advertencias'][:5]:  # Mostrar primeras 5
                                messages.warning(request, f"⚠️ {adv}")
                            if len(resultado['advertencias']) > 5:
                                messages.warning(request, f"⚠️ ...y {len(resultado['advertencias']) - 5} advertencias más")
                        
                        return redirect('logistica:lista_citas')
                    else:
                        for error in resultado['errores']:
                            messages.error(request, f"❌ {error}")
                        
                        for adv in resultado['advertencias'][:5]:
                            messages.warning(request, f"⚠️ {adv}")
                
                except Exception as e:
                    messages.error(request, f"Error al procesar archivo: {str(e)}")
            else:
                messages.error(request, "Verifica el archivo seleccionado")
        
        else:
            # Procesar captura manual
            form_manual = CitaProveedorForm(request.POST)
            form_masiva = CargaMasivaCitasForm()
            
            if form_manual.is_valid():
                cita = form_manual.save(commit=False)
                cita.usuario_creacion = request.user
                cita.save()
                
                # Generar folio si no existe
                ServicioFolio.asignar_folio_a_cita(cita)
                
                # Enviar notificación
                notificaciones.notificar_cita_creada(cita)
                
                messages.success(request, f'✓ Cita creada exitosamente con {cita.proveedor.razon_social}')
                return redirect('logistica:lista_citas')
            else:
                messages.error(request, 'Error al crear la cita. Verifica los datos.')
    
    else:
        form_manual = CitaProveedorForm()
        form_masiva = CargaMasivaCitasForm()
    
    context = {
        'form_manual': form_manual,
        'form_masiva': form_masiva,
        'tipo_operacion': tipo_operacion,
    }
    return render(request, 'inventario/citas/crear.html', context)

@login_required
def detalle_cita(request, pk):
    """Ver detalles de una cita"""
    cita = get_object_or_404(CitaProveedor, pk=pk)
    
    # Verificar si el usuario tiene permiso para validar entrada de cita
    puede_validar = request.user.has_perm('inventario.validar_entrada_cita') or request.user.is_superuser
    
    # Obtener lista de revisión si existe
    lista_revision = None
    try:
        lista_revision = cita.lista_revision
    except ListaRevision.DoesNotExist:
        pass
    
    context = {
        'cita': cita,
        'lista_revision': lista_revision,
        'puede_autorizar': cita.estado == 'programada' and puede_validar,
        'puede_completar': cita.estado == 'autorizada',
        'puede_validar': puede_validar,
    }
    return render(request, 'inventario/citas/detalle.html', context)


@login_required
@require_http_methods(["POST"])
def toggle_no_material_medico(request, pk):
    """Alterna la bandera 'No es material médico' de la cita y redirige al detalle."""
    cita = get_object_or_404(CitaProveedor, pk=pk)
    cita.no_es_material_medico = not cita.no_es_material_medico
    cita.save(update_fields=['no_es_material_medico'])
    estado = "marcado como no material médico" if cita.no_es_material_medico else "marcado como material médico"
    messages.success(request, f"La cita se actualizó: {estado}.")
    return redirect('logistica:detalle_cita', pk=pk)


@login_required
def editar_cita(request, pk):
    """Editar una cita existente"""
    cita = get_object_or_404(CitaProveedor, pk=pk)
    
    # Solo permitir editar si está en estado 'programada'
    if cita.estado != 'programada':
        messages.warning(request, f'No se puede editar una cita en estado {cita.get_estado_display()}')
        return redirect('logistica:lista_citas')
    
    if request.method == 'POST':
        form = CitaProveedorEditForm(request.POST, instance=cita)
        if form.is_valid():
            form.save()
            messages.success(request, '✓ Cita actualizada exitosamente')
            return redirect('logistica:lista_citas')
    else:
        form = CitaProveedorEditForm(instance=cita)
    
    return render(request, 'inventario/citas/editar.html', {
        'form': form,
        'cita': cita
    })


@login_required
def validar_entrada(request, pk):
    """Validar entrada de una cita (nueva lista de revisión)"""
    cita = get_object_or_404(CitaProveedor, pk=pk)
    
    # Generar folio si no existe
    if not cita.folio:
        ServicioFolio.asignar_folio_a_cita(cita)
    
    # Obtener lista de revisión si existe
    lista_revision = None
    try:
        lista_revision = ListaRevision.objects.get(cita=cita)
    except ListaRevision.DoesNotExist:
        pass
    
    # Validar acceso según el estado de la cita
    # - Si está 'programada': permitir validar (crear/editar lista de revisión)
    # - Si está 'autorizada' y tiene lista aprobada: permitir solo lectura (imprimir)
    # - Cualquier otro caso: denegar acceso
    if cita.estado == 'programada':
        # Crear lista de revisión si no existe
        if not lista_revision:
            lista_revision = ServicioListaRevision.crear_lista_revision(cita, request.user)
    elif cita.estado == 'autorizada' and lista_revision and lista_revision.estado == 'aprobada':
        # Permitir acceso de solo lectura para imprimir
        pass
    else:
        messages.warning(request, f'Solo se pueden validar citas en estado "Programada"')
        return redirect('logistica:detalle_cita', pk=pk)
    
    # Solo permitir POST (validar/aprobar/rechazar) cuando la cita está en 'programada'
    if request.method == 'POST' and cita.estado != 'programada':
        messages.warning(request, f'Solo se pueden validar citas en estado "Programada"')
        return redirect('logistica:detalle_cita', pk=pk)
    
    if request.method == 'POST':
        # Procesar items de revisión
        for item in lista_revision.items.all():
            resultado_key = f'item_{item.id}_resultado'
            observaciones_key = f'item_{item.id}_observaciones'
            
            if resultado_key in request.POST:
                item.resultado = request.POST[resultado_key]
                item.observaciones = request.POST.get(observaciones_key, '')
                item.save()
        
        # Determinar si es aprobación o rechazo
        if 'aprobar' in request.POST:
            form = ValidarEntradaForm(request.POST)
            if form.is_valid():
                # Guardar tipo_red si se proporcionó
                tipo_red = form.cleaned_data.get('tipo_red')
                if tipo_red:
                    # Obtener o crear LlegadaProveedor
                    from .llegada_models import LlegadaProveedor
                    try:
                        llegada = LlegadaProveedor.objects.get(cita=cita)
                        llegada.tipo_red = tipo_red
                        llegada.save()
                    except LlegadaProveedor.DoesNotExist:
                        # Si no existe la llegada, crearla con los datos básicos
                        # Esto puede suceder si la validación se hace antes de crear la llegada
                        # IMPORTANTE: Copiar todos los datos relevantes de la cita
                        llegada = LlegadaProveedor.objects.create(
                            cita=cita,
                            proveedor=cita.proveedor,
                            almacen=cita.almacen,
                            folio=cita.folio or '',
                            remision=cita.numero_orden_remision or '',  # Copiar remisión de la cita
                            numero_piezas_emitidas=0,
                            numero_piezas_recibidas=0,
                            tipo_red=tipo_red,
                            estado='EN_RECEPCION',
                            # Copiar datos de la cita que son relevantes para la llegada
                            numero_orden_suministro=cita.numero_orden_suministro or '',
                            numero_contrato=cita.numero_contrato or '',
                            numero_procedimiento='',  # Este campo no existe en CitaProveedor
                        )
                
                ServicioListaRevision.validar_entrada(
                    lista_revision,
                    request.user,
                    form.cleaned_data.get('observaciones', '')
                )
                messages.success(request, '✓ Entrada validada exitosamente. La cita está autorizada.')
                return redirect('logistica:lista_citas')
        
        elif 'rechazar' in request.POST:
            form = RechazarEntradaForm(request.POST)
            if form.is_valid():
                ServicioListaRevision.rechazar_entrada(
                    lista_revision,
                    request.user,
                    form.cleaned_data['justificacion']
                )
                messages.warning(request, '✗ Entrada rechazada. La cita ha sido cancelada.')
                return redirect('logistica:lista_citas')
    
    # Preparar contexto con valor inicial de tipo_red si existe
    form_validar_initial = {}
    try:
        llegada = cita.llegada_proveedor
        if llegada and llegada.tipo_red:
            form_validar_initial['tipo_red'] = llegada.tipo_red
    except:
        pass
    
    # Determinar si es modo solo lectura (cita autorizada)
    es_solo_lectura = cita.estado == 'autorizada'
    
    # Preparar contexto
    context = {
        'cita': cita,
        'lista_revision': lista_revision,
        'items': lista_revision.items.all() if lista_revision else [],
        'form_validar': ValidarEntradaForm(initial=form_validar_initial),
        'form_rechazar': RechazarEntradaForm(),
        'es_solo_lectura': es_solo_lectura,
    }
    
    return render(request, 'inventario/citas/validar_entrada.html', context)


@login_required
def cancelar_cita(request, pk):
    """Cancelar una cita"""
    cita = get_object_or_404(CitaProveedor, pk=pk)
    
    if cita.estado == 'cancelada':
        messages.warning(request, 'Esta cita ya está cancelada')
        return redirect('logistica:lista_citas')
    
    if request.method == 'POST':
        cita.estado = 'cancelada'
        cita.save()
        
        # Enviar notificación
        notificaciones.notificar_cita_cancelada(cita)
        
        messages.success(request, f'✓ Cita cancelada')
        return redirect('logistica:lista_citas')
    
    return render(request, 'inventario/citas/cancelar.html', {'cita': cita})


# ============================================================================
# VISTAS PARA TRASLADOS
# ============================================================================

@login_required
def lista_traslados(request):
    """Lista todas las órdenes de traslado"""
    traslados = OrdenTraslado.objects.all().order_by('-fecha_creacion')
    
    # Filtros
    estado = request.GET.get('estado')
    almacen_origen = request.GET.get('almacen_origen')
    
    if estado:
        traslados = traslados.filter(estado=estado)
    
    if almacen_origen:
        traslados = traslados.filter(almacen_origen__id=almacen_origen)
    
    # Contar por estado
    estados_count = {
        'creada': traslados.filter(estado='creada').count(),
        'logistica_asignada': traslados.filter(estado='logistica_asignada').count(),
        'en_transito': traslados.filter(estado='en_transito').count(),
        'recibida': traslados.filter(estado='recibida').count(),
        'completada': traslados.filter(estado='completada').count(),
    }
    
    almacenes = Almacen.objects.all()
    
    context = {
        'traslados': traslados,
        'estados': OrdenTraslado.ESTADOS_TRASLADO,
        'almacenes': almacenes,
        'estados_count': estados_count,
        'estado_seleccionado': estado,
        'almacen_seleccionado': almacen_origen,
    }
    return render(request, 'inventario/traslados/lista.html', context)


@login_required
def crear_traslado(request):
    """Crear una nueva orden de traslado"""
    if request.method == 'POST':
        form = OrdenTrasladoForm(request.POST)
        if form.is_valid():
            orden = form.save(commit=False)
            orden.usuario_creacion = request.user
            
            # Generar folio automáticamente
            try:
                tipo_entrega = TipoEntrega.objects.get(codigo='TRA')
                folio_obj = Folio.objects.get(tipo_entrega=tipo_entrega)
                orden.folio = folio_obj.generar_folio()
            except:
                orden.folio = f"TRA-{timezone.now().strftime('%Y%m%d%H%M%S')}"
            
            orden.save()
            messages.success(request, f'✓ Orden de traslado creada: {orden.folio}')
            return redirect('detalle_traslado', pk=orden.pk)
    else:
        form = OrdenTrasladoForm()
    
    return render(request, 'inventario/traslados/crear.html', {'form': form})


@login_required
def detalle_traslado(request, pk):
    """Ver detalles de una orden de traslado"""
    orden = get_object_or_404(OrdenTraslado, pk=pk)
    items = orden.items.all()
    
    context = {
        'orden': orden,
        'items': items,
        'puede_asignar_logistica': orden.estado == 'creada',
        'puede_iniciar_transito': orden.estado == 'logistica_asignada',
    }
    return render(request, 'inventario/traslados/detalle.html', context)


@login_required
def asignar_logistica_traslado(request, pk):
    """Asignar vehículo, chofer y ruta a una orden de traslado"""
    orden = get_object_or_404(OrdenTraslado, pk=pk)
    
    if orden.estado != 'creada':
        messages.warning(request, 'Solo se puede asignar logística a órdenes en estado "Creada"')
        return redirect('detalle_traslado', pk=pk)
    
    if request.method == 'POST':
        form = LogisticaTrasladoForm(request.POST, instance=orden)
        if form.is_valid():
            orden = form.save(commit=False)
            orden.estado = 'logistica_asignada'
            orden.save()
            messages.success(request, '✓ Logística asignada exitosamente')
            return redirect('detalle_traslado', pk=orden.pk)
    else:
        form = LogisticaTrasladoForm(instance=orden)
    
    return render(request, 'inventario/traslados/asignar_logistica.html', {
        'form': form,
        'orden': orden
    })


# ============================================================================
# VISTAS PARA CONTEO FÍSICO
# ============================================================================

@login_required
def lista_conteos(request):
    """Lista todos los conteos físicos"""
    conteos = ConteoFisico.objects.all().order_by('-fecha_inicio')
    
    # Filtros
    estado = request.GET.get('estado')
    almacen = request.GET.get('almacen')
    
    if estado:
        conteos = conteos.filter(estado=estado)
    
    if almacen:
        conteos = conteos.filter(almacen__id=almacen)
    
    # Contar por estado
    estados_count = {
        'iniciado': conteos.filter(estado='iniciado').count(),
        'en_progreso': conteos.filter(estado='en_progreso').count(),
        'completado': conteos.filter(estado='completado').count(),
    }
    
    almacenes = Almacen.objects.all()
    
    context = {
        'conteos': conteos,
        'almacenes': almacenes,
        'estados_count': estados_count,
        'estado_seleccionado': estado,
        'almacen_seleccionado': almacen,
    }
    return render(request, 'inventario/conteo_fisico/lista.html', context)


@login_required
def iniciar_conteo(request):
    """Iniciar una nueva sesión de conteo físico"""
    if request.method == 'POST':
        almacen_id = request.POST.get('almacen')
        observaciones = request.POST.get('observaciones', '')
        
        try:
            almacen = Almacen.objects.get(pk=almacen_id)
            
            conteo = ConteoFisico.objects.create(
                almacen=almacen,
                observaciones=observaciones,
                usuario_creacion=request.user,
                estado='iniciado'
            )
            
            # Generar folio automáticamente
            try:
                tipo_entrega = TipoEntrega.objects.get(codigo='CNT')
                folio_obj = Folio.objects.get(tipo_entrega=tipo_entrega)
                conteo.folio = folio_obj.generar_folio()
                conteo.save()
            except:
                conteo.folio = f"CNT-{timezone.now().strftime('%Y%m%d%H%M%S')}"
                conteo.save()
            
            messages.success(request, f'✓ Conteo iniciado: {conteo.folio}')
            return redirect('capturar_conteo', pk=conteo.pk)
        except Almacen.DoesNotExist:
            messages.error(request, 'Almacén no encontrado')
    
    almacenes = Almacen.objects.all()
    context = {'almacenes': almacenes}
    return render(request, 'inventario/conteo_fisico/iniciar.html', context)


@login_required
def capturar_conteo(request, pk):
    """Capturar datos de conteo físico"""
    conteo = get_object_or_404(ConteoFisico, pk=pk)
    items = conteo.items.all()
    
    # Obtener lotes disponibles del almacén
    lotes_disponibles = Lote.objects.filter(
        institucion__almacen=conteo.almacen,
        cantidad_disponible__gt=0
    ).order_by('numero_lote')
    
    context = {
        'conteo': conteo,
        'items': items,
        'lotes_disponibles': lotes_disponibles,
    }
    return render(request, 'inventario/conteo_fisico/capturar.html', context)


@login_required
def detalle_conteo(request, pk):
    """Ver detalles de un conteo físico"""
    conteo = get_object_or_404(ConteoFisico, pk=pk)
    items = conteo.items.all()
    
    # Calcular diferencias
    total_teorico = sum(item.cantidad_teorica for item in items)
    total_fisico = sum(item.cantidad_fisica for item in items)
    diferencia_total = total_fisico - total_teorico
    
    context = {
        'conteo': conteo,
        'items': items,
        'total_teorico': total_teorico,
        'total_fisico': total_fisico,
        'diferencia_total': diferencia_total,
    }
    return render(request, 'inventario/conteo_fisico/detalle.html', context)

@login_required
def exportar_citas_excel(request):
    """Exporta la lista de citas a Excel"""
    
    # Obtener todas las citas
    citas = CitaProveedor.objects.all().order_by('-fecha_cita')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Citas'
    
    # Estilos
    header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Proveedor', 'RFC', 'Fecha y Hora', 'Almacén', 'Orden Suministro', 'Orden Remisión', 'Clave', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row_num = 2
    for cita in citas:
        ws.cell(row=row_num, column=1).value = cita.proveedor.razon_social if cita.proveedor else ''
        ws.cell(row=row_num, column=2).value = cita.proveedor.rfc if cita.proveedor else ''
        ws.cell(row=row_num, column=3).value = cita.fecha_cita.strftime('%d/%m/%Y %H:%M') if cita.fecha_cita else ''
        ws.cell(row=row_num, column=4).value = cita.almacen.nombre if cita.almacen else ''
        ws.cell(row=row_num, column=5).value = cita.numero_orden_suministro or ''
        ws.cell(row=row_num, column=6).value = cita.numero_orden_remision or ''
        ws.cell(row=row_num, column=7).value = cita.clave_medicamento or ''
        ws.cell(row=row_num, column=8).value = dict(CitaProveedor.ESTADOS_CITA).get(cita.estado, cita.estado)
        
        # Aplicar bordes
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = border
        
        row_num += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    
    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="citas_proveedores.xlsx"'
    wb.save(response)
    return response

