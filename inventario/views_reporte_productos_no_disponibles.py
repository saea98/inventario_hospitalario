"""
Reporte de Productos No Disponibles en Almacén Destino
Muestra productos que están disponibles en otros almacenes pero no en el almacén destino.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Sum
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

from .pedidos_models import ProductoNoDisponibleAlmacen, PropuestaPedido
from .models import Almacen


@login_required
def reporte_productos_no_disponibles(request):
    """
    Reporte de productos no disponibles en el almacén destino pero disponibles en otros almacenes.
    """
    # Filtros
    almacen_id = request.GET.get('almacen')
    notificado = request.GET.get('notificado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    # Query base
    productos = ProductoNoDisponibleAlmacen.objects.select_related(
        'propuesta__solicitud__institucion_solicitante',
        'propuesta__solicitud__almacen_destino',
        'producto',
        'almacen_destino'
    ).order_by('-fecha_registro')
    
    # Aplicar filtros
    if almacen_id:
        productos = productos.filter(almacen_destino_id=almacen_id)
    
    if notificado == 'si':
        productos = productos.filter(notificado_telegram=True)
    elif notificado == 'no':
        productos = productos.filter(notificado_telegram=False)
    
    if fecha_desde:
        productos = productos.filter(fecha_registro__gte=fecha_desde)
    
    if fecha_hasta:
        productos = productos.filter(fecha_registro__lte=fecha_hasta)
    
    # Preparar datos con almacenes parseados
    productos_con_almacenes = []
    for producto in productos:
        try:
            almacenes_info = json.loads(producto.almacenes_con_disponibilidad) if producto.almacenes_con_disponibilidad else []
        except:
            almacenes_info = []
        productos_con_almacenes.append({
            'registro': producto,
            'almacenes': almacenes_info
        })
    
    # Paginación
    paginator = Paginator(productos_con_almacenes, 25)
    page = request.GET.get('page')
    try:
        productos_paginados = paginator.page(page)
    except PageNotAnInteger:
        productos_paginados = paginator.page(1)
    except EmptyPage:
        productos_paginados = paginator.page(paginator.num_pages)
    
    # Obtener almacenes para el filtro
    almacenes = Almacen.objects.all().order_by('nombre')
    
    # Resumen
    resumen = {
        'total': productos.count(),
        'notificados': productos.filter(notificado_telegram=True).count(),
        'no_notificados': productos.filter(notificado_telegram=False).count(),
    }
    
    context = {
        'page_title': 'Productos No Disponibles en Almacén Destino',
        'productos': productos_paginados,
        'almacenes': almacenes,
        'resumen': resumen,
        'filtro_almacen': almacen_id or '',
        'filtro_notificado': notificado or '',
        'filtro_fecha_desde': fecha_desde or '',
        'filtro_fecha_hasta': fecha_hasta or '',
    }
    
    return render(request, 'inventario/reportes/reporte_productos_no_disponibles.html', context)


@login_required
def exportar_productos_no_disponibles_excel(request):
    """
    Exporta el reporte de productos no disponibles a Excel.
    """
    # Filtros (mismos que en la vista)
    almacen_id = request.GET.get('almacen') or request.POST.get('almacen')
    notificado = request.GET.get('notificado') or request.POST.get('notificado')
    fecha_desde = request.GET.get('fecha_desde') or request.POST.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta') or request.POST.get('fecha_hasta')
    
    # Query base
    productos = ProductoNoDisponibleAlmacen.objects.select_related(
        'propuesta__solicitud__institucion_solicitante',
        'propuesta__solicitud__almacen_destino',
        'producto',
        'almacen_destino'
    ).order_by('-fecha_registro')
    
    # Aplicar filtros
    if almacen_id:
        productos = productos.filter(almacen_destino_id=almacen_id)
    
    if notificado == 'si':
        productos = productos.filter(notificado_telegram=True)
    elif notificado == 'no':
        productos = productos.filter(notificado_telegram=False)
    
    if fecha_desde:
        productos = productos.filter(fecha_registro__gte=fecha_desde)
    
    if fecha_hasta:
        productos = productos.filter(fecha_registro__lte=fecha_hasta)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos No Disponibles"
    
    # Estilos
    header_fill = PatternFill(start_color="8B1538", end_color="8B1538", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'Fecha Registro',
        'Folio Propuesta',
        'Institución',
        'Almacén Destino',
        'Clave CNIS',
        'Descripción',
        'Cantidad Requerida',
        'Disponible en Destino',
        'Disponible en Otros',
        'Almacenes con Disponibilidad',
        'Notificado Telegram'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row_num, registro in enumerate(productos, 2):
        try:
            almacenes_info = json.loads(registro.almacenes_con_disponibilidad) if registro.almacenes_con_disponibilidad else []
        except:
            almacenes_info = []
        almacenes_texto = ", ".join([f"{a['almacen_nombre']} ({a['cantidad']})" for a in almacenes_info])
        
        data = [
            registro.fecha_registro.strftime('%d/%m/%Y %H:%M'),
            registro.propuesta.solicitud.folio,
            registro.propuesta.solicitud.institucion_solicitante.denominacion,
            registro.almacen_destino.nombre,
            registro.producto.clave_cnis,
            registro.producto.descripcion,
            registro.cantidad_requerida,
            registro.cantidad_disponible_destino,
            registro.cantidad_disponible_otros,
            almacenes_texto,
            'Sí' if registro.notificado_telegram else 'No'
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Ajustar anchos de columna
    column_widths = [18, 20, 30, 25, 15, 50, 18, 20, 20, 40, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos_no_disponibles_almacen.xlsx"'
    
    wb.save(response)
    return response
