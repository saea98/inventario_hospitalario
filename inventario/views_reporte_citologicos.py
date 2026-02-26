"""
Reporte de productos citológicos.
Por ahora basado en un listado fijo de claves CNIS; después se puede generalizar
(categoría, archivo de claves, etc.).
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Sum, Avg, Max
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime

from .models import Lote, MovimientoInventario
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Listado fijo de claves citológicas (único por clave)
CLAVES_CITOLOGICOS = frozenset({
    "010.000.1754.00",
    "010.000.1760.00",
    "010.000.1764.00",
    "010.000.1768.01",
    "010.000.1774.00",
    "010.000.2152.00",
    "010.000.2192.00",
    "010.000.2195.00",
    "010.000.3003.00",
    "010.000.3047.00",
    "010.000.3049.00",
    "010.000.4225.00",
    "010.000.4226.00",
    "010.000.4229.00",
    "010.000.4233.00",
    "010.000.4323.00",
    "010.000.4429.00",
    "010.000.4431.00",
    "010.000.4434.00",
    "010.000.4437.00",
    "010.000.4444.00",
    "010.000.5140.01",
    "010.000.5307.00",
    "010.000.5421.00",
    "010.000.5423.00",
    "010.000.5432.00",
    "010.000.5436.00",
    "010.000.5437.00",
    "010.000.5438.00",
    "010.000.5440.01",
    "010.000.5445.00",
    "010.000.5445.01",
    "010.000.5449.00",
    "010.000.5450.00",
    "010.000.5457.00",
    "010.000.5458.00",
    "010.000.5465.02",
    "010.000.5466.00",
    "010.000.5468.00",
    "010.000.5470.00",
    "010.000.5480.00",
    "010.000.5541.00",
    "010.000.5650.00",
    "010.000.5654.00",
    "010.000.5771.00",
    "010.000.5880.00",
    "010.000.5887.00",
    "010.000.5972.00",
    "010.000.6016.00",
    "010.000.6018.00",
    "010.000.6023.00",
    "010.000.6023.01",
    "010.000.6024.00",
    "010.000.6037.00",
    "010.000.6085.00",
    "010.000.6086.00",
    "010.000.6093.00",
    "010.000.6095.00",
    "010.000.6120.00",
    "010.000.6153.00",
    "010.000.6165.00",
    "010.000.6174.00",
    "010.000.6211.00",
    "010.000.6214.00",
    "010.000.6226.00",
    "010.000.6227.00",
    "010.000.6282.00",
    "010.000.6293.00",
    "010.000.6358.00",
    "010.000.6359.00",
    "010.000.7179.00",
})

HEADERS_EXCEL = [
    "Clave CNIS", "Producto", "Unidad de medida", "Almacén", "Existencia", "Precio", "Importe",
    "Cant. Reservada", "ENTIDAD FEDERATIVA", "CLUES", "FUENTE DE FINANCIAMIENTO", "PARTIDA PRESUPUESTAL",
    "LUGAR DE ENTREGA", "ENTRADAS", "SALIDAS",
]
KEY_ORDER = [
    "clave_cnis", "producto", "unidad_medida", "almacen", "existencia", "precio", "importe",
    "cant_reservada", "entidad_federativa", "clues", "fuente_financiamiento", "partida_presupuestal",
    "lugar_entrega", "entradas", "salidas",
]


def _parse_fecha(fecha_str):
    if not fecha_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(fecha_str).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _obtener_datos_citologicos(request):
    """
    Obtiene existencias agrupadas por (producto, institución, almacén)
    solo para las claves del listado citológico. Aplica filtros opcionales desde GET.
    """
    filtro_entidad = (request.GET.get("entidad") or "").strip()
    filtro_clues = (request.GET.get("clues") or "").strip()
    filtro_almacen = (request.GET.get("almacen") or "").strip()
    filtro_orden = (request.GET.get("orden") or "").strip()
    filtro_estado = (request.GET.get("estado") or "").strip()
    excluir_sin_orden = (request.GET.get("excluir_sin_orden") or "") == "si"
    fecha_rec_desde = _parse_fecha((request.GET.get("fecha_rec_desde") or "").strip())
    fecha_rec_hasta = _parse_fecha((request.GET.get("fecha_rec_hasta") or "").strip())
    cad_desde = _parse_fecha((request.GET.get("cad_desde") or "").strip())
    cad_hasta = _parse_fecha((request.GET.get("cad_hasta") or "").strip())

    lotes_base = (
        Lote.objects.filter(producto__clave_cnis__in=CLAVES_CITOLOGICOS)
        .select_related("producto", "institucion", "orden_suministro__proveedor", "almacen")
    )

    if excluir_sin_orden:
        lotes_base = lotes_base.exclude(orden_suministro__isnull=True)
    if filtro_entidad:
        lotes_base = lotes_base.filter(institucion__denominacion__icontains=filtro_entidad)
    if filtro_clues:
        lotes_base = lotes_base.filter(institucion__clue__icontains=filtro_clues)
    if filtro_almacen:
        lotes_base = lotes_base.filter(almacen__nombre__icontains=filtro_almacen)
    if filtro_orden:
        lotes_base = lotes_base.filter(orden_suministro__numero_orden__icontains=filtro_orden)
    if filtro_estado:
        lotes_base = lotes_base.filter(estado=filtro_estado)
    if fecha_rec_desde:
        lotes_base = lotes_base.filter(fecha_recepcion__gte=fecha_rec_desde)
    if fecha_rec_hasta:
        lotes_base = lotes_base.filter(fecha_recepcion__lte=fecha_rec_hasta)
    if cad_desde:
        lotes_base = lotes_base.filter(fecha_caducidad__gte=cad_desde)
    if cad_hasta:
        lotes_base = lotes_base.filter(fecha_caducidad__lte=cad_hasta)

    qs = (
        lotes_base.values(
            "producto_id",
            "institucion_id",
            "almacen_id",
            "producto__clave_cnis",
            "producto__descripcion",
            "producto__unidad_medida",
            "almacen__nombre",
            "institucion__clue",
            "institucion__denominacion",
        )
        .annotate(
            existencia=Sum("cantidad_disponible"),
            importe=Sum("valor_total"),
            cant_reservada=Sum("cantidad_reservada"),
            precio_promedio=Avg("precio_unitario"),
            partida_presup=Max("orden_suministro__partida_presupuestal"),
        )
        .order_by("producto__clave_cnis", "institucion__denominacion", "almacen__nombre")
    )

    entradas_map = {}
    for row in MovimientoInventario.objects.filter(
        tipo_movimiento__in=["ENTRADA", "TRANSFERENCIA_ENTRADA", "AJUSTE_POSITIVO"],
        anulado=False,
    ).values("lote__producto_id", "lote__almacen_id", "lote__institucion_id").annotate(total=Sum("cantidad")):
        key = (row["lote__producto_id"], row["lote__almacen_id"], row["lote__institucion_id"])
        entradas_map[key] = row["total"]

    salidas_map = {}
    for row in MovimientoInventario.objects.filter(
        tipo_movimiento__in=["SALIDA", "TRANSFERENCIA_SALIDA", "AJUSTE_NEGATIVO", "CADUCIDAD", "DETERIORO"],
        anulado=False,
    ).values("lote__producto_id", "lote__almacen_id", "lote__institucion_id").annotate(total=Sum("cantidad")):
        key = (row["lote__producto_id"], row["lote__almacen_id"], row["lote__institucion_id"])
        salidas_map[key] = row["total"]

    fuente_map = {}
    for row in lotes_base.filter(orden_suministro__isnull=False).values(
        "producto_id", "institucion_id", "almacen_id"
    ).distinct():
        key = (row["producto_id"], row["institucion_id"], row["almacen_id"])
        if key not in fuente_map:
            lo = lotes_base.filter(
                producto_id=row["producto_id"],
                institucion_id=row["institucion_id"],
                almacen_id=row["almacen_id"],
            ).select_related("orden_suministro__fuente_financiamiento").first()
            if lo and lo.orden_suministro and getattr(lo.orden_suministro, "fuente_financiamiento", None):
                fuente_map[key] = lo.orden_suministro.fuente_financiamiento.nombre
            else:
                fuente_map[key] = ""

    rows = []
    for g in qs:
        key = (g["producto_id"], g["institucion_id"], g["almacen_id"])
        entradas = entradas_map.get(key, 0)
        salidas = salidas_map.get(key, 0)
        fuente = fuente_map.get(key, "")
        partida = g.get("partida_presup") or ""
        if not partida:
            lote_sample = lotes_base.filter(
                producto_id=g["producto_id"],
                institucion_id=g["institucion_id"],
                almacen_id=g["almacen_id"],
            ).first()
            if lote_sample:
                partida = getattr(lote_sample, "partida", "") or ""
        lugar_entrega = g.get("almacen__nombre") or g.get("institucion__denominacion") or ""
        rows.append({
            "clave_cnis": g.get("producto__clave_cnis") or "",
            "producto": (g.get("producto__descripcion") or "")[:500],
            "unidad_medida": g.get("producto__unidad_medida") or "PIEZA",
            "almacen": g.get("almacen__nombre") or "",
            "existencia": g["existencia"] or 0,
            "precio": g.get("precio_promedio") or 0,
            "importe": g["importe"] or 0,
            "cant_reservada": g["cant_reservada"] or 0,
            "entidad_federativa": "CIUDAD DE MÉXICO",
            "clues": g.get("institucion__clue") or "",
            "fuente_financiamiento": fuente,
            "partida_presupuestal": partida,
            "lugar_entrega": lugar_entrega,
            "entradas": entradas,
            "salidas": salidas,
        })
    return rows


@login_required
def reporte_citologicos(request):
    """Vista del reporte de productos citológicos (listado fijo de claves)."""
    datos = _obtener_datos_citologicos(request)
    paginator = Paginator(datos, 50)
    page = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    get_copy = request.GET.copy()
    get_copy.pop("page", None)
    base_query = get_copy.urlencode()

    context = {
        "datos_reporte": page_obj,
        "paginator": paginator,
        "base_query": base_query,
        "filtro_entidad": request.GET.get("entidad", "").strip(),
        "filtro_clues": request.GET.get("clues", "").strip(),
        "filtro_almacen": request.GET.get("almacen", "").strip(),
        "filtro_orden": request.GET.get("orden", "").strip(),
        "filtro_estado": request.GET.get("estado", "").strip(),
        "excluir_sin_orden": (request.GET.get("excluir_sin_orden") or "") == "si",
        "fecha_rec_desde": request.GET.get("fecha_rec_desde", "").strip(),
        "fecha_rec_hasta": request.GET.get("fecha_rec_hasta", "").strip(),
        "cad_desde": request.GET.get("cad_desde", "").strip(),
        "cad_hasta": request.GET.get("cad_hasta", "").strip(),
        "estados_lote": [(1, "Disponible"), (4, "Suspendido"), (5, "Deteriorado"), (6, "Caducado")],
        "total_claves_listado": len(CLAVES_CITOLOGICOS),
    }
    return render(request, "inventario/reportes/reporte_citologicos.html", context)


@login_required
def exportar_citologicos_excel(request):
    """Exporta el reporte de productos citológicos a Excel."""
    datos = _obtener_datos_citologicos(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos citológicos"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(HEADERS_EXCEL, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row_num, row in enumerate(datos, 2):
        for col_num, key in enumerate(KEY_ORDER, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = row.get(key, "")
            cell.border = border
            if key in ("existencia", "precio", "importe", "cant_reservada", "entradas", "salidas"):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_num in range(1, len(HEADERS_EXCEL) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_citologicos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response
