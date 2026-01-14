"""
Reporte de Registros No Afectados por Movimientos

Muestra todos los registros (Lotes, LoteUbicacion, Productos) que NO han sido
afectados por movimientos de conteo o asignación de ubicación desde el 28 de diciembre hacia atrás.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Exists, OuterRef
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    Lote,
    LoteUbicacion,
    Producto,
    RegistroConteoFisico,
    MovimientoInventario,
    Institucion,
    Almacen
)


@login_required
def reporte_no_afectados(request):
    """
    Reporte de registros no afectados por movimientos de conteo y asignación de ubicación
    desde el 28 de diciembre hacia atrás.
    """
    
    # Fecha de corte: 28 de diciembre
    fecha_corte = datetime(2025, 12, 28).date()
    fecha_corte_aware = timezone.make_aware(datetime.combine(fecha_corte, datetime.min.time()))
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        'producto',
        'institucion'
    ).all()
    
    # Registros no afectados
    registros_no_afectados = []
    
    for lote in lotes:
        # Verificar si el lote tiene movimientos de conteo o asignación de ubicación después del 28 de diciembre
        tiene_conteo = RegistroConteoFisico.objects.filter(
            lote_ubicacion__lote=lote,
            fecha_creacion__gte=fecha_corte_aware
        ).exists()
        
        tiene_asignacion = LoteUbicacion.objects.filter(
            lote=lote,
            fecha_asignacion__gte=fecha_corte_aware
        ).exists()
        
        # Si no tiene conteo ni asignación, es un registro no afectado
        if not tiene_conteo and not tiene_asignacion:
            # Obtener ubicaciones del lote (si las hay)
            ubicaciones = LoteUbicacion.objects.filter(lote=lote).select_related(
                'ubicacion',
                'ubicacion__almacen'
            )
            
            if ubicaciones.exists():
                # Si tiene ubicaciones, mostrar cada una
                for ub in ubicaciones:
                    registros_no_afectados.append({
                        'tipo': 'LoteUbicacion',
                        'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                        'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                        'lote': lote.numero_lote,
                        'institucion': lote.institucion.denominacion if lote.institucion else '-',
                        'almacen': ub.ubicacion.almacen.nombre if ub.ubicacion and ub.ubicacion.almacen else '-',
                        'ubicacion': ub.ubicacion.codigo if ub.ubicacion else '-',
                        'cantidad': ub.cantidad,
                        'fecha_asignacion': ub.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if ub.fecha_asignacion else '-',
                        'lote_id': lote.id,
                        'ubicacion_id': ub.id,
                    })
            else:
                # Si no tiene ubicaciones, mostrar el lote sin ubicación
                registros_no_afectados.append({
                    'tipo': 'Lote',
                    'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                    'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                    'lote': lote.numero_lote,
                    'institucion': lote.institucion.denominacion if lote.institucion else '-',
                    'almacen': '-',
                    'ubicacion': 'Sin ubicación asignada',
                    'cantidad': lote.cantidad_disponible,
                    'fecha_asignacion': '-',
                    'lote_id': lote.id,
                    'ubicacion_id': None,
                })
    
    # Filtros opcionales
    filtro_institucion = request.GET.get('institucion', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_clave = request.GET.get('clave', '').strip()
    
    # Aplicar filtros
    if filtro_institucion:
        registros_no_afectados = [r for r in registros_no_afectados if r['institucion'] == filtro_institucion]
    
    if filtro_almacen:
        registros_no_afectados = [r for r in registros_no_afectados if r['almacen'] == filtro_almacen]
    
    if filtro_clave:
        registros_no_afectados = [r for r in registros_no_afectados if filtro_clave.lower() in r['clave_cnis'].lower()]
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(registros_no_afectados, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener opciones de filtro
    instituciones = Institucion.objects.all().order_by('denominacion')
    almacenes = Almacen.objects.all().order_by('nombre')
    
    context = {
        'page_obj': page_obj,
        'total_registros': len(registros_no_afectados),
        'fecha_corte': fecha_corte.strftime('%d/%m/%Y'),
        'instituciones': instituciones,
        'almacenes': almacenes,
        'filtro_institucion': filtro_institucion,
        'filtro_almacen': filtro_almacen,
        'filtro_clave': filtro_clave,
    }
    
    return render(request, 'inventario/reporte_no_afectados.html', context)


@login_required
def exportar_no_afectados_excel(request):
    """
    Exporta el reporte de registros no afectados a Excel.
    """
    
    # Fecha de corte: 28 de diciembre
    fecha_corte = datetime(2025, 12, 28).date()
    fecha_corte_aware = timezone.make_aware(datetime.combine(fecha_corte, datetime.min.time()))
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        'producto',
        'institucion'
    ).all()
    
    # Registros no afectados
    registros_no_afectados = []
    
    for lote in lotes:
        # Verificar si el lote tiene movimientos de conteo o asignación de ubicación después del 28 de diciembre
        tiene_conteo = RegistroConteoFisico.objects.filter(
            lote_ubicacion__lote=lote,
            fecha_creacion__gte=fecha_corte_aware
        ).exists()
        
        tiene_asignacion = LoteUbicacion.objects.filter(
            lote=lote,
            fecha_asignacion__gte=fecha_corte_aware
        ).exists()
        
        # Si no tiene conteo ni asignación, es un registro no afectado
        if not tiene_conteo and not tiene_asignacion:
            # Obtener ubicaciones del lote (si las hay)
            ubicaciones = LoteUbicacion.objects.filter(lote=lote).select_related(
                'ubicacion',
                'ubicacion__almacen'
            )
            
            if ubicaciones.exists():
                # Si tiene ubicaciones, mostrar cada una
                for ub in ubicaciones:
                    registros_no_afectados.append({
                        'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                        'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                        'lote': lote.numero_lote,
                        'institucion': lote.institucion.denominacion if lote.institucion else '-',
                        'almacen': ub.ubicacion.almacen.nombre if ub.ubicacion and ub.ubicacion.almacen else '-',
                        'ubicacion': ub.ubicacion.codigo if ub.ubicacion else '-',
                        'cantidad': ub.cantidad,
                        'fecha_asignacion': ub.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if ub.fecha_asignacion else '-',
                    })
            else:
                # Si no tiene ubicaciones, mostrar el lote sin ubicación
                registros_no_afectados.append({
                    'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                    'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                    'lote': lote.numero_lote,
                    'institucion': lote.institucion.denominacion if lote.institucion else '-',
                    'almacen': '-',
                    'ubicacion': 'Sin ubicación asignada',
                    'cantidad': lote.cantidad_disponible,
                    'fecha_asignacion': '-',
                })
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "No Afectados"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'CLAVE CNIS',
        'DESCRIPCIÓN PRODUCTO',
        'LOTE',
        'INSTITUCIÓN',
        'ALMACÉN',
        'UBICACIÓN',
        'CANTIDAD',
        'FECHA ASIGNACIÓN'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, registro in enumerate(registros_no_afectados, 2):
        ws.cell(row=row, column=1).value = registro['clave_cnis']
        ws.cell(row=row, column=2).value = registro['descripcion_producto']
        ws.cell(row=row, column=3).value = registro['lote']
        ws.cell(row=row, column=4).value = registro['institucion']
        ws.cell(row=row, column=5).value = registro['almacen']
        ws.cell(row=row, column=6).value = registro['ubicacion']
        ws.cell(row=row, column=7).value = registro['cantidad']
        ws.cell(row=row, column=8).value = registro['fecha_asignacion']
        
        # Aplicar bordes
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_no_afectados.xlsx"'
    wb.save(response)
    
    return response


@login_required
def eliminar_registro_no_afectado(request, lote_id, ubicacion_id=None):
    """
    Elimina un registro no afectado (Lote o LoteUbicacion).
    """
    from django.shortcuts import redirect
    from django.contrib import messages
    
    try:
        if ubicacion_id:
            # Eliminar LoteUbicacion
            lote_ubicacion = LoteUbicacion.objects.get(id=ubicacion_id)
            lote_ubicacion.delete()
            messages.success(request, f"Ubicación eliminada correctamente.")
        else:
            # Eliminar Lote
            lote = Lote.objects.get(id=lote_id)
            lote.delete()
            messages.success(request, f"Lote {lote.numero_lote} eliminado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar el registro: {str(e)}")
    
    # Redirigir al reporte
    return redirect('reporte_no_afectados')
