"""
Vistas para reporte de disponibilidad vs reservas por clave y lote.
Permite tomar decisiones sobre el inventario considerando las reservas en propuestas.

Misma base que lote-pedidos y reporte de existencias:
- Existencia: Lote.cantidad_disponible (alineado con reportes/existencias).
- Reserva: suma de LoteAsignado con surtido=False (totales_reserva_activa_por_lote_ids),
  no el campo Lote.cantidad_reservada (puede estar desactualizado).
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import date
import logging

from .models import Lote, Institucion
from .propuesta_utils import (
    cantidad_existencia_fisica_lote_como_reporte_existencias,
    totales_reserva_activa_por_lote_ids,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

logger = logging.getLogger(__name__)


@login_required
def reporte_disponibilidad_lotes(request):
    """
    Reporte de disponibilidad vs reservas por clave y lote.
    Existencia = Lote.cantidad_disponible; reserva = suma LoteAsignado (surtido=False).
    """

    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_institucion = request.GET.get('institucion', '')
    filtro_estado = request.GET.get('estado', '')

    lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'almacen',
    ).filter(estado=1)

    if filtro_clave:
        lotes = lotes.filter(
            Q(producto__clave_cnis__icontains=filtro_clave)
            | Q(producto__descripcion__icontains=filtro_clave)
        )

    if filtro_lote:
        lotes = lotes.filter(numero_lote__icontains=filtro_lote)

    if filtro_institucion:
        lotes = lotes.filter(institucion_id=filtro_institucion)

    if filtro_estado:
        lotes = lotes.filter(estado=int(filtro_estado))

    lotes = lotes.order_by('producto__clave_cnis', 'fecha_caducidad')
    lotes_list = list(lotes.prefetch_related('ubicaciones_detalle'))
    lote_ids = [lo.id for lo in lotes_list]
    reservas_map = totales_reserva_activa_por_lote_ids(lote_ids)

    total_disponible = 0
    total_reservado = 0
    total_neto = 0
    for lote in lotes_list:
        disp = cantidad_existencia_fisica_lote_como_reporte_existencias(lote)
        res = reservas_map.get(lote.id, 0)
        total_disponible += disp
        total_reservado += res
        total_neto += max(0, disp - res)

    paginator = Paginator(lotes_list, 20)
    page = request.GET.get('page')

    try:
        lotes_pagina = paginator.page(page)
    except PageNotAnInteger:
        lotes_pagina = paginator.page(1)
    except EmptyPage:
        lotes_pagina = paginator.page(paginator.num_pages)

    datos_reporte = []

    for lote in lotes_pagina:
        cantidad_disponible = cantidad_existencia_fisica_lote_como_reporte_existencias(lote)
        cantidad_reservada = reservas_map.get(lote.id, 0)
        cantidad_neta = max(0, cantidad_disponible - cantidad_reservada)
        porcentaje_reserva = (
            (cantidad_reservada / cantidad_disponible * 100) if cantidad_disponible > 0 else 0
        )

        if cantidad_neta <= 0:
            estado_reserva = 'AGOTADO'
            clase_alerta = 'danger'
        elif porcentaje_reserva >= 80:
            estado_reserva = 'CRÍTICO'
            clase_alerta = 'danger'
        elif porcentaje_reserva >= 50:
            estado_reserva = 'ALTO'
            clase_alerta = 'warning'
        else:
            estado_reserva = 'NORMAL'
            clase_alerta = 'success'

        dias_caducidad = None
        if lote.fecha_caducidad:
            dias_caducidad = (lote.fecha_caducidad - date.today()).days

        datos_reporte.append({
            'lote_id': lote.id,
            'clave': lote.producto.clave_cnis,
            'descripcion': lote.producto.descripcion[:60],
            'numero_lote': lote.numero_lote,
            'institucion': lote.institucion.denominacion if lote.institucion else 'N/A',
            'cantidad_disponible': cantidad_disponible,
            'cantidad_reservada': cantidad_reservada,
            'cantidad_neta': cantidad_neta,
            'porcentaje_reserva': round(porcentaje_reserva, 2),
            'fecha_caducidad': lote.fecha_caducidad,
            'dias_caducidad': dias_caducidad,
            'estado_reserva': estado_reserva,
            'clase_alerta': clase_alerta,
            'precio_unitario': lote.precio_unitario,
            'valor_total': lote.valor_total,
        })

    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')
    
    # Contexto
    context = {
        'page_title': 'Reporte de Disponibilidad vs Reservas',
        'datos_reporte': datos_reporte,
        'instituciones': instituciones,
        'filtro_clave': filtro_clave,
        'filtro_lote': filtro_lote,
        'filtro_institucion': filtro_institucion,
        'filtro_estado': filtro_estado,
        'total_disponible': total_disponible,
        'total_reservado': total_reservado,
        'total_neto': total_neto,
        'total_lotes': paginator.count,
        'porcentaje_reserva_total': (
            (total_reservado / total_disponible * 100)
            if total_disponible > 0 else 0
        ),
        'page_obj': lotes_pagina,
        'paginator': paginator,
    }
    
    logger.warning(f"[REPORTE_DISPONIBILIDAD] Total lotes: {paginator.count} | Disponible: {total_disponible} | Reservado: {total_reservado} | Neto: {total_neto}")
    
    return render(request, 'inventario/reportes/reporte_disponibilidad_lotes.html', context)


@login_required
def exportar_disponibilidad_excel(request):
    """
    Exporta el reporte de disponibilidad vs reservas a Excel.
    """
    
    # Obtener parámetros de filtro (GET o POST)
    filtro_clave = request.POST.get('clave', request.GET.get('clave', '')).strip()
    filtro_lote = request.POST.get('lote', request.GET.get('lote', '')).strip()
    filtro_institucion = request.POST.get('institucion', request.GET.get('institucion', ''))
    
    # Query base
    lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'almacen'
    ).filter(
        estado=1
    )
    
    # Aplicar filtros
    if filtro_clave:
        lotes = lotes.filter(
            Q(producto__clave_cnis__icontains=filtro_clave) |
            Q(producto__descripcion__icontains=filtro_clave)
        )
    
    if filtro_lote:
        lotes = lotes.filter(numero_lote__icontains=filtro_lote)
    
    if filtro_institucion:
        lotes = lotes.filter(institucion_id=filtro_institucion)
    
    # Si no hay filtros, mostrar mensaje
    if not filtro_clave and not filtro_lote and not filtro_institucion:
        logger.info(f"[EXPORTAR_DISPONIBILIDAD] Exportando sin filtros")
    else:
        logger.info(f"[EXPORTAR_DISPONIBILIDAD] Filtros: clave={filtro_clave}, lote={filtro_lote}, institucion={filtro_institucion}")
    
    lotes = lotes.order_by('producto__clave_cnis', 'fecha_caducidad')
    lotes_list = list(lotes.prefetch_related('ubicaciones_detalle'))
    lote_ids = [lo.id for lo in lotes_list]
    reservas_map = totales_reserva_activa_por_lote_ids(lote_ids)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Disponibilidad"
    
    # Definir estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        'Clave CNIS',
        'Descripción',
        'Número Lote',
        'Institución',
        'Cantidad Disponible (Σ ubicaciones)',
        'Cantidad Reservada (pedidos activos)',
        'Cantidad Neta',
        '% Reserva',
        'Fecha Caducidad',
        'Días para Caducar',
        'Precio Unitario',
        'Valor Total',
    ]
    
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    
    # Llenar datos (todos los lotes, no solo la página actual)
    total_disponible = 0
    total_reservado = 0
    total_neto = 0
    
    for lote in lotes_list:
        cantidad_disponible = cantidad_existencia_fisica_lote_como_reporte_existencias(lote)
        cantidad_reservada = reservas_map.get(lote.id, 0)
        cantidad_neta = max(0, cantidad_disponible - cantidad_reservada)
        porcentaje_reserva = (
            (cantidad_reservada / cantidad_disponible * 100) if cantidad_disponible > 0 else 0
        )

        dias_caducidad = None
        if lote.fecha_caducidad:
            dias_caducidad = (lote.fecha_caducidad - date.today()).days

        row = [
            lote.producto.clave_cnis,
            lote.producto.descripcion[:60],
            lote.numero_lote,
            lote.institucion.denominacion if lote.institucion else 'N/A',
            cantidad_disponible,
            cantidad_reservada,
            cantidad_neta,
            f"{porcentaje_reserva:.2f}%",
            lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A',
            dias_caducidad if dias_caducidad is not None else 'N/A',
            f"${lote.precio_unitario:.2f}",
            f"${lote.valor_total:.2f}",
        ]
        
        ws.append(row)
        
        # Aplicar bordes y alineación
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Colorear según nivel de reserva
        if cantidad_neta <= 0:
            color = "FFC7CE"  # Rojo claro
        elif porcentaje_reserva >= 80:
            color = "FFC7CE"  # Rojo claro
        elif porcentaje_reserva >= 50:
            color = "FFEB9C"  # Amarillo claro
        else:
            color = "C6EFCE"  # Verde claro
        
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        total_disponible += cantidad_disponible
        total_reservado += cantidad_reservada
        total_neto += cantidad_neta
    
    # Agregar fila de totales
    ws.append([])
    total_row = ws.max_row
    ws[f'A{total_row}'] = 'TOTAL'
    ws[f'E{total_row}'] = total_disponible
    ws[f'F{total_row}'] = total_reservado
    ws[f'G{total_row}'] = total_neto
    
    # Aplicar estilos a fila de totales
    for col in ['A', 'E', 'F', 'G']:
        cell = ws[f'{col}{total_row}']
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_disponibilidad_lotes.xlsx"'
    
    wb.save(response)
    return response
