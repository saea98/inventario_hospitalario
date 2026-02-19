"""
Reporte de Salidas al Inventario (Inventario de Salidas)

Para auditorías: muestra todas las salidas (MovimientoInventario tipo SALIDA)
con el layout oficial: Clave CNIS, Producto, LOTE, CANTIDAD SURTIDA,
CLUES DEL ALMACÉN, ALMACÉN, P.P., RFC, Proveedor, FOLIO DE SALIDA, FECHA DE ENTREGA, UNIDAD HOSPITALARIA,
CLUES DESTINO SSA/IMB, CONTRATO, REMISIÓN, ORDEN DE SUMINISTRO, LICITACIÓN, Precio, Importe.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, OuterRef, Subquery, UUIDField
from django.db.models.functions import Cast
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import MovimientoInventario, Institucion, Almacen, UbicacionAlmacen
from .pedidos_models import PropuestaPedido

# Layout del reporte de salidas para auditorías
SALIDAS_LAYOUT_HEADERS = [
    'Clave CNIS',
    'Producto',
    'UNIDAD DE MEDIDA',
    'LOTE',
    'CADUCIDAD',
    'CANTIDAD SURTIDA',
    'CLUES DEL ALMACÉN',
    'ALMACÉN',
    'P.P.',
    'RFC',
    'Proveedor',
    'FOLIO DE SALIDA',
    'FOLIO DE PEDIDO',
    'FECHA DE ENTREGA',
    'UNIDAD HOSPITALARIA',
    'CLUES DESTINO SSA',
    'CLUES DESTINO IMB',
    'CONTRATO',
    'REMISION',
    'ORDEN DE SUMINISTRO',
    'LICITACIÓN/PROCEDIMIENTO',
    'Precio',
    'Importe',
    'USUARIO MOVIMIENTO',
]


def _valor(o, default=''):
    if o is None:
        return default
    return o


def _fecha(d, fmt='%d/%m/%Y'):
    if not d:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime(fmt)
    return str(d)


def _decimal(d, default=''):
    if d is None:
        return default
    if isinstance(d, Decimal):
        return float(d)
    return d


def _construir_fila_salida(m):
    """Construye la fila para un MovimientoInventario SALIDA."""
    lote = m.lote
    prod = lote.producto if lote else None
    os = lote.orden_suministro if lote else None
    prov = os.proveedor if os else None
    almacen = lote.almacen if lote else None
    inst_origen = almacen.institucion if almacen else None
    inst_destino = m.institucion_destino

    precio = (lote and lote.precio_unitario) or Decimal('0')
    importe_val = (m.importe_total or (lote and lote.importe_total))
    if importe_val is None:
        importe_val = m.cantidad * precio

    partida_presupuestal = (lote and lote.partida) or (os and os.partida_presupuestal)
    rfc = _valor(prov and prov.rfc or (lote and getattr(lote, 'rfc_proveedor', None)))
    proveedor_nombre = _valor(prov and prov.razon_social or (lote and getattr(lote, 'proveedor', None)))

    return [
        _valor(prod and prod.clave_cnis),
        _valor(prod and prod.descripcion),
        _valor(prod and prod.unidad_medida) or 'PIEZA',
        _valor(lote and lote.numero_lote),
        _fecha(lote and lote.fecha_caducidad),
        m.cantidad,
        _valor(inst_origen and inst_origen.clue),
        _valor(almacen and almacen.nombre),
        _valor(partida_presupuestal),
        rfc,
        proveedor_nombre,
        _valor(m.folio),
        _valor(getattr(m, 'folio_pedido_solicitud', None) or (lote and lote.observaciones)),
        _fecha(m.fecha_movimiento, '%d/%m/%Y %H:%M') if m.fecha_movimiento else '',
        _valor(inst_destino and inst_destino.denominacion),
        _valor(inst_destino and inst_destino.clue),
        _valor(inst_destino and inst_destino.ib_clue),
        _valor(m.contrato or (lote and lote.contrato)),
        _valor(m.remision or (lote and lote.remision)),
        _valor(os and os.numero_orden),
        _valor(m.licitacion or (lote and lote.licitacion)),
        _decimal(precio),
        _decimal(importe_val),
        _valor((m.usuario.get_full_name() or m.usuario.username) if getattr(m, 'usuario', None) else ''),
    ]


@login_required
def reporte_salidas(request):
    """
    Reporte de salidas al inventario (Inventario de Salidas) para auditorías.
    Layout oficial de 20 columnas.
    """
    subq_folio_pedido = PropuestaPedido.objects.filter(
        id=Cast(OuterRef('folio'), UUIDField())
    ).values('solicitud__observaciones_solicitud')[:1]
    salidas = MovimientoInventario.objects.filter(
        tipo_movimiento='SALIDA', anulado=False
    ).annotate(
        folio_pedido_solicitud=Subquery(subq_folio_pedido)
    ).select_related(
        'lote',
        'lote__producto',
        'lote__institucion',
        'lote__almacen',
        'lote__orden_suministro',
        'lote__orden_suministro__proveedor',
        'institucion_destino',
        'usuario'
    ).order_by('-fecha_movimiento')

    # Filtros
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_almacen = request.GET.get('almacen', '')
    filtro_destino = request.GET.get('destino', '').strip()
    filtro_folio = request.GET.get('folio', '').strip()

    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            salidas = salidas.filter(fecha_movimiento__date__gte=fecha_desde)
        except ValueError:
            pass
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date() + timedelta(days=1)
            salidas = salidas.filter(fecha_movimiento__date__lt=fecha_hasta)
        except ValueError:
            pass
    if filtro_clave:
        salidas = salidas.filter(lote__producto__clave_cnis__icontains=filtro_clave)
    if filtro_lote:
        salidas = salidas.filter(lote__numero_lote__icontains=filtro_lote)
    if filtro_almacen:
        salidas = salidas.filter(lote__almacen__nombre=filtro_almacen)
    if filtro_destino:
        salidas = salidas.filter(
            Q(institucion_destino__denominacion__icontains=filtro_destino) |
            Q(institucion_destino__clue__icontains=filtro_destino)
        )
    if filtro_folio:
        salidas = salidas.filter(folio__icontains=filtro_folio)

    # Construir lista de filas
    salidas_lista = []
    total_cantidad = 0
    total_importe = 0.0
    for m in salidas:
        row = _construir_fila_salida(m)
        salidas_lista.append({'row': row, 'id': m.id})
        total_cantidad += row[5]  # CANTIDAD SURTIDA
        total_importe += float(row[22] or 0)  # Importe

    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(salidas_lista, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    instituciones = Institucion.objects.all().order_by('denominacion')
    almacenes = Almacen.objects.all().order_by('nombre')

    context = {
        'page_obj': page_obj,
        'headers': SALIDAS_LAYOUT_HEADERS,
        'total_registros': len(salidas_lista),
        'total_cantidad': total_cantidad,
        'total_importe': total_importe,
        'almacenes': almacenes,
        'instituciones': instituciones,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
        'filtro_clave': filtro_clave,
        'filtro_lote': filtro_lote,
        'filtro_almacen': filtro_almacen,
        'filtro_destino': filtro_destino,
        'filtro_folio': filtro_folio,
    }
    return render(request, 'inventario/reporte_salidas.html', context)


@login_required
def exportar_salidas_excel(request):
    """Exporta el reporte de salidas a Excel con el layout oficial (incluye RFC y Proveedor)."""
    subq_folio_pedido_exp = PropuestaPedido.objects.filter(
        id=Cast(OuterRef('folio'), UUIDField())
    ).values('solicitud__observaciones_solicitud')[:1]
    salidas = MovimientoInventario.objects.filter(
        tipo_movimiento='SALIDA', anulado=False
    ).annotate(
        folio_pedido_solicitud=Subquery(subq_folio_pedido_exp)
    ).select_related(
        'lote',
        'lote__producto',
        'lote__institucion',
        'lote__almacen',
        'lote__orden_suministro',
        'lote__orden_suministro__proveedor',
        'institucion_destino',
        'usuario'
    ).order_by('-fecha_movimiento')

    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_almacen = request.GET.get('almacen', '')
    filtro_destino = request.GET.get('destino', '').strip()
    filtro_folio = request.GET.get('folio', '').strip()

    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            salidas = salidas.filter(fecha_movimiento__date__gte=fecha_desde)
        except ValueError:
            pass
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date() + timedelta(days=1)
            salidas = salidas.filter(fecha_movimiento__date__lt=fecha_hasta)
        except ValueError:
            pass
    if filtro_clave:
        salidas = salidas.filter(lote__producto__clave_cnis__icontains=filtro_clave)
    if filtro_lote:
        salidas = salidas.filter(lote__numero_lote__icontains=filtro_lote)
    if filtro_almacen:
        salidas = salidas.filter(lote__almacen__nombre=filtro_almacen)
    if filtro_destino:
        salidas = salidas.filter(
            Q(institucion_destino__denominacion__icontains=filtro_destino) |
            Q(institucion_destino__clue__icontains=filtro_destino)
        )
    if filtro_folio:
        salidas = salidas.filter(folio__icontains=filtro_folio)

    wb = Workbook()
    ws = wb.active
    ws.title = "Salidas"
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="E3F2FD", size=10)
    total_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    total_font = Font(bold=True, size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(SALIDAS_LAYOUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    total_cantidad = 0
    total_importe_val = 0.0
    listado = []
    for m in salidas:
        fila = _construir_fila_salida(m)
        listado.append(fila)
        total_cantidad += fila[5] or 0
        total_importe_val += float(fila[22] or 0)

    for row_num, fila in enumerate(listado, 2):
        for col_num, val in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.border = border
            if col_num in (6, 22, 23):
                cell.alignment = Alignment(horizontal='right')

    total_row = len(listado) + 2
    ws.cell(row=total_row, column=1).value = "TOTALES"
    ws.cell(row=total_row, column=1).font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=6).value = total_cantidad
    ws.cell(row=total_row, column=6).font = total_font
    ws.cell(row=total_row, column=6).fill = total_fill
    ws.cell(row=total_row, column=23).value = total_importe_val
    ws.cell(row=total_row, column=23).font = total_font
    ws.cell(row=total_row, column=23).fill = total_fill
    for col in range(1, len(SALIDAS_LAYOUT_HEADERS) + 1):
        ws.cell(row=total_row, column=col).border = border

    for col in range(1, len(SALIDAS_LAYOUT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_salidas.xlsx"'
    wb.save(response)
    return response
