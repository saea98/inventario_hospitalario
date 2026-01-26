"""
Módulo para convertir Excel de Acuse de Entrega a PDF
Usa openpyxl para leer el Excel y reportlab para generar el PDF
"""

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.units import mm, inch
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageTemplate, Frame, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from datetime import datetime
import os
from django.conf import settings


def convertir_acuse_excel_a_pdf(excel_buffer):
    """
    Convierte un buffer de Excel de Acuse a PDF usando openpyxl y reportlab.
    
    Args:
        excel_buffer: BytesIO con contenido del archivo Excel
        
    Returns:
        BytesIO: Buffer con contenido del PDF
        
    Raises:
        Exception: Si falla la conversión
    """
    
    try:
        # Leer el Excel
        excel_buffer.seek(0)
        workbook = load_workbook(excel_buffer)
        worksheet = workbook.active
        
        # Extraer datos del Excel
        folio = _obtener_valor_celda(worksheet['I1']).replace('#FOLIO: ', '')
        folio_pedido = _obtener_valor_celda(worksheet['I3']).replace('FOLIO DE PEDIDO: ', '')
        fecha = _obtener_valor_celda(worksheet['I4']).replace('FECHA: ', '')
        institucion = _obtener_valor_celda(worksheet['A10']).replace('INSTITUCIÓN: ', '')
        
        # Extraer datos de la tabla (a partir de fila 18)
        datos_tabla = []
        for row_idx in range(18, worksheet.max_row + 1):
            fila = []
            for col_idx in range(1, 12):
                celda = worksheet.cell(row=row_idx, column=col_idx)
                valor = _obtener_valor_celda(celda)
                
                # Para la columna DESCRIPCIÓN (3), usar Paragraph para mejor ajuste
                if col_idx == 3 and valor:
                    # Crear un Paragraph con el texto
                    style = ParagraphStyle(
                        'DescripcionStyle',
                        fontSize=7,
                        leading=9,
                        alignment=TA_LEFT,
                        fontName='Helvetica'
                    )
                    fila.append(Paragraph(valor, style))
                else:
                    fila.append(valor if valor else "")
            
            # Si la fila tiene contenido, agregarla
            if any(fila):
                datos_tabla.append(fila)
        
        # Crear PDF
        pdf_buffer = BytesIO()
        
        # Crear documento con tamaño letter landscape
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=1.8*inch,
            bottomMargin=0.8*inch
        )
        
        # Contenido del documento
        story = []
        
        # Crear tabla de encabezado con logotipo y información
        # Buscar logotipo
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'templates', 'inventario', 'images', 'logo_imss.jpg'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_imss.jpg'),
        ]
        
        logo_img = None
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                try:
                    logo_img = Image(logo_path, width=1.5*inch, height=0.4*inch)
                    break
                except:
                    pass
        
        # Crear tabla de encabezado: logo + título a la izquierda, información a la derecha
        # Estilo para el título del sistema
        titulo_sistema_style = ParagraphStyle(
            'TituloSistema',
            fontSize=9,
            textColor=colors.HexColor('#8B1538'),
            alignment=TA_LEFT,
            fontName='Helvetica-Bold',
            leading=11,
            spaceAfter=0
        )
        
        # Crear contenido izquierdo (logo + título)
        left_content = []
        if logo_img:
            left_content.append(logo_img)
        left_content.append(Paragraph('Sistema de Abasto, Inventarios y Control de<br/>Almacenes', titulo_sistema_style))
        
        # Información a la derecha
        info_style = ParagraphStyle(
            'InfoStyle',
            fontSize=8,
            alignment=TA_RIGHT,
            fontName='Helvetica',
            leading=10
        )
        info_right = Paragraph(
            f'#FOLIO: {folio}<br/>FECHA: {fecha}<br/>FOLIO DE PEDIDO: {folio_pedido}',
            info_style
        )
        
        # Tabla de encabezado
        header_data = [[left_content, info_right]]
        header_table = Table(header_data, colWidths=[5.5*inch, 4.5*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (0, 0), 0),
            ('RIGHTPADDING', (0, 0), (0, 0), 0),
            ('LEFTPADDING', (1, 0), (1, 0), 10),
            ('RIGHTPADDING', (1, 0), (1, 0), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        
        story.append(header_table)
        story.append(Spacer(1, 0.05*inch))
        
        # Información adicional (TRANSFERENCIA y TIPO) - alineada a la derecha
        info_adicional_style = ParagraphStyle(
            'InfoAdicional',
            fontSize=8,
            alignment=TA_RIGHT,
            fontName='Helvetica',
            leading=10
        )
        story.append(Paragraph('TRANSFERENCIA: prueba', info_adicional_style))
        story.append(Paragraph('TIPO: TRANSFERENCIA (SURTIMIENTO)', info_adicional_style))
        story.append(Spacer(1, 0.15*inch))
        
        # Título de Acuse de Entrega
        titulo_style = ParagraphStyle(
            'TituloAcuse',
            fontSize=11,
            textColor=colors.HexColor('#8B1538'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph('ACUSE DE ENTREGA', titulo_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Tabla de firmas
        # Crear Paragraph para la institución para que se ajuste al ancho
        # El ancho de la columna es 2.5*inch, menos padding (10 puntos = ~0.14 inch)
        # Entonces el ancho disponible es aproximadamente 2.36*inch
        institucion_style = ParagraphStyle(
            'InstitucionStyle',
            fontSize=7,
            leading=9,
            alignment=TA_LEFT,
            fontName='Helvetica',
            wordWrap='CJK',  # Permite word wrap
            leftIndent=0,
            rightIndent=0,
        )
        
        # Preparar texto de institución con word wrap manual si es muy largo
        texto_institucion = f'<b>INSTITUCIÓN:</b><br/>{institucion}'
        
        # Reordenar columnas según el formato deseado: UNIDAD DE DESTINO, AUTORIZA (ALMACEN), RECIBE (UNIDAD DE DESTINO), ENTREGA (ALMACEN)
        firma_data = [
            ['UNIDAD DE DESTINO', 'AUTORIZA (ALMACEN)', 'RECIBE (UNIDAD DE DESTINO)', 'ENTREGA (ALMACEN)'],
            [
                Paragraph(texto_institucion, institucion_style),
                'NOMBRE:\n\nPUESTO:\n\nFIRMA: __________________',
                'NOMBRE: __________________\n\nPUESTO: __________________\n\nFIRMA: __________________',
                'NOMBRE: __________________\n\nPUESTO: __________________\n\nFIRMA: __________________'
            ]
        ]
        
        firma_table = Table(firma_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch, 2.5*inch])
        firma_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B1538')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 40),
            # Especial para columna UNIDAD DE DESTINO - más padding para el texto
            ('LEFTPADDING', (0, 1), (0, 1), 8),
            ('RIGHTPADDING', (0, 1), (0, 1), 8),
            ('TOPPADDING', (0, 1), (0, 1), 8),
            ('BOTTOMPADDING', (0, 1), (0, 1), 8),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            # Asegurar que el texto se ajuste al ancho de la columna
            ('ALIGN', (0, 1), (0, 1), 'LEFT'),
        ]))
        
        story.append(firma_table)
        story.append(Spacer(1, 0.1*inch))
        
        # Tabla de items
        encabezados = ['#', 'CLAVE CNIS', 'DESCRIPCIÓN', 'U.M.', 'TIPO', 'LOTE', 'CADUCIDAD', 'CLASIFICACIÓN', 'UBICACIÓN', 'CANTIDAD', 'FOLIO PEDIDO']
        datos_completos = [encabezados] + datos_tabla
        
        # Definir anchos de columnas
        col_widths = [0.35*inch, 0.85*inch, 1.9*inch, 0.75*inch, 0.75*inch, 0.85*inch, 0.85*inch, 0.85*inch, 0.85*inch, 0.75*inch, 1.5*inch]
        
        items_table = Table(datos_completos, colWidths=col_widths)
        
        items_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B1538')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            
            # Datos - Alineación
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),
            ('ALIGN', (3, 1), (-1, -1), 'CENTER'),
            
            # Datos - Alineación vertical
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            
            # Datos - Estilos de fuente
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            
            # Datos - Colores de fondo
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            
            # Especial para columna DESCRIPCIÓN - permitir wrap
            ('LEFTPADDING', (2, 1), (2, -1), 5),
            ('RIGHTPADDING', (2, 1), (2, -1), 5),
            ('TOPPADDING', (2, 1), (2, -1), 5),
            ('BOTTOMPADDING', (2, 1), (2, -1), 5),
            
            # Altura de filas para mejor legibilidad
            ('ROWHEIGHT', (0, 1), (-1, -1), 18*mm),
        ]))
        
        story.append(items_table)
        
        # Construir el PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return pdf_buffer
        
    except Exception as e:
        raise Exception(f"Error al convertir Acuse Excel a PDF: {str(e)}")


def _obtener_valor_celda(celda):
    """Obtiene el valor de una celda."""
    
    if celda is None:
        return ""
    
    value = celda.value if hasattr(celda, 'value') else celda
    
    if value is None:
        return ""
    
    # Formatear fechas
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y')
    
    return str(value)
