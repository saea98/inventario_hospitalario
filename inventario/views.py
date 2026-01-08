# ==========================================
# SISTEMA DE INVENTARIO HOSPITALARIO - VISTAS
# ==========================================
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import date, timedelta
from django.http import HttpResponse
import json
import os
import pandas as pd
from django.db.models import Q
from django.db.models import Count
from django.db.models import Case, When, Value, BooleanField

# Control de acceso por roles
from .access_control import requiere_rol


from .forms import CargaLotesForm
from .carga_datos import carga_lotes_desde_excel

# Modelos
from .models import (
    Institucion, Producto, Proveedor, Lote, OrdenSuministro,
    CategoriaProducto, Alcaldia, TipoInstitucion, FuenteFinanciamiento,
    MovimientoInventario, AlertaCaducidad, CargaInventario, SolicitudInventario, EstadoInsumo, User,
    LoteUbicacion
)

# Formularios
from .forms import (
    InstitucionForm, ProductoForm, ProveedorForm, LoteForm,
    MovimientoInventarioForm, CargaInventarioForm, FiltroInventarioForm,
    CustomUserCreationForm, LoteUbicacionForm, LoteUbicacionFormSet
)

# Utilidades
from .utils import ExcelProcessor
from .reports import ReportGenerator
import openpyxl
from .models import Institucion
from .forms import CargaMasivaInstitucionForm
from .models import SolicitudInventario, EstadoInsumo
from inventario import models
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from django.core.exceptions import FieldDoesNotExist


# ==========================================
# 1. DASHBOARD PRINCIPAL
# ==========================================
@login_required
def dashboard(request):
    """Vista principal del panel de control"""
    total_instituciones = Institucion.objects.filter(activo=True).count()
    total_productos = Producto.objects.filter(activo=True).count()
    total_lotes = Lote.objects.filter(estado=1).count()

    valor_total_inventario = Lote.objects.filter(estado=1).aggregate(
        total=Sum('valor_total')
    )['total'] or 0

    hoy = date.today()
    alertas_30_dias = Lote.objects.filter(
        estado=1, fecha_caducidad__lte=hoy + timedelta(days=30), fecha_caducidad__gt=hoy
    ).count()
    productos_caducados = Lote.objects.filter(fecha_caducidad__lt=hoy).count()
    productos_bajo_stock = Lote.objects.filter(
        estado=1, cantidad_disponible__lt=10
    ).count()

    ultimos_movimientos = MovimientoInventario.objects.select_related(
        'lote__producto', 'lote__institucion', 'usuario'
    ).order_by('-fecha_movimiento')[:10]

    instituciones_top = Institucion.objects.annotate(
        total_lotes=Count('lote'), valor_inventario=Sum('lote__valor_total')
    ).filter(activo=True).order_by('-valor_inventario')[:5]

    context = {
        'total_instituciones': total_instituciones,
        'total_productos': total_productos,
        'total_lotes': total_lotes,
        'valor_total_inventario': valor_total_inventario,
        'alertas_30_dias': alertas_30_dias,
        'productos_caducados': productos_caducados,
        'productos_bajo_stock': productos_bajo_stock,
        'ultimos_movimientos': ultimos_movimientos,
        'instituciones_top': instituciones_top,
    }
    return render(request, 'inventario/dashboard.html', context)


@login_required
@require_http_methods(["GET"])
def api_estadisticas_dashboard(request):
    """API JSON para gráficas del dashboard"""
    hoy = date.today()
    estadisticas_mensuales = []
    for i in range(6):
        fecha = hoy.replace(day=1) - timedelta(days=30 * i)
        mes_siguiente = (fecha.replace(month=fecha.month + 1)
                         if fecha.month < 12 else fecha.replace(year=fecha.year + 1, month=1))

        lotes_mes = Lote.objects.filter(
            fecha_recepcion__gte=fecha, fecha_recepcion__lt=mes_siguiente
        ).count()

        valor_mes = Lote.objects.filter(
            fecha_recepcion__gte=fecha, fecha_recepcion__lt=mes_siguiente
        ).aggregate(total=Sum('valor_total'))['total'] or 0

        estadisticas_mensuales.append({
            'mes': fecha.strftime('%Y-%m'),
            'lotes': lotes_mes,
            'valor': float(valor_mes)
        })

    categorias = CategoriaProducto.objects.annotate(
        total_lotes=Count('producto__lote'),
        valor_total=Sum('producto__lote__valor_total')
    ).values('nombre', 'total_lotes', 'valor_total')

    return JsonResponse({
        'estadisticas_mensuales': list(reversed(estadisticas_mensuales)),
        'categorias': list(categorias)
    })


# ==========================================
# 2. INSTITUCIONES (CLUES)
# ==========================================
@login_required
@requiere_rol('Administrador')
def lista_instituciones(request):
    instituciones = Institucion.objects.select_related('alcaldia', 'tipo_institucion')
    search = request.GET.get('search')
    alcaldia_id = request.GET.get('alcaldia')
    tipo_id = request.GET.get('tipo')

    if search:
        instituciones = instituciones.filter(
            Q(clue__icontains=search) | Q(ib_clue__icontains=search) | Q(denominacion__icontains=search)
        )
    if alcaldia_id:
        instituciones = instituciones.filter(alcaldia_id=alcaldia_id)
    if tipo_id:
        instituciones = instituciones.filter(tipo_institucion_id=tipo_id)

    paginator = Paginator(instituciones, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'inventario/instituciones/lista.html', {
        'page_obj': page_obj,
        'alcaldias': Alcaldia.objects.all(),
        'tipos_institucion': TipoInstitucion.objects.all(),
        'search': search,
        'alcaldia_selected': alcaldia_id,
        'tipo_selected': tipo_id,
    })


@login_required
def detalle_institucion(request, pk):
    institucion = get_object_or_404(Institucion, pk=pk)

    # 🔹 Productos que tienen al menos un lote asociado a esta institución
    productos = (
        Producto.objects.filter(lote__institucion=institucion)
        .distinct()
        .annotate(total_lotes=Count('lote'))
    )

    # 🔹 Totales
    total_productos = productos.count()
    total_lotes = Lote.objects.filter(institucion=institucion).count()

    contexto = {
        'institucion': institucion,
        'productos': productos,
        'total_productos': total_productos,
        'total_lotes': total_lotes,
    }
    return render(request, 'inventario/instituciones/detalle.html', contexto)


@login_required
def crear_institucion(request):
    form = InstitucionForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Institución creada exitosamente.')
        return redirect('lista_instituciones')
    return render(request, 'inventario/instituciones/form.html', {'form': form, 'titulo': 'Crear Institución'})


@login_required
def editar_institucion(request, pk):
    institucion = get_object_or_404(Institucion, pk=pk)
    form = InstitucionForm(request.POST or None, instance=institucion)
    if form.is_valid():
        form.save()
        messages.success(request, 'Institución actualizada exitosamente.')
        return redirect('detalle_institucion', pk=pk)
    return render(request, 'inventario/instituciones/form.html', {'form': form, 'titulo': 'Editar Institución'})



@login_required
def reporte_personalizado(request):
    if request.method == "POST":
        try:
            # 1️⃣ Recuperar campos seleccionados y ordenados
            campos = request.POST.getlist("columnas")  # checkboxes seleccionados
            orden_columnas = request.POST.get("orden_columnas", "")
            print("🔍 Campos seleccionados:", campos)
            print("🔍 Orden columnas:", orden_columnas)

            # Si hay un orden personalizado, lo respetamos
            if orden_columnas:
                orden = [c for c in orden_columnas.split(",") if c in campos]
                if orden:
                    campos = orden

            if not campos:
                return JsonResponse({"error": "No se seleccionaron columnas válidas"}, status=400)

            # 2️⃣ Consultar los datos (solo esos campos)
            datos = (
                Lote.objects.select_related("producto", "institucion", "creado_por")
                .values(*campos)
            )

            datos_lista = list(datos)
            if not datos_lista:
                print("⚠️ No hay datos para exportar")
                return JsonResponse({"error": "No hay datos para exportar"}, status=404)

            print("🔍 Primer registro:", datos_lista[0])

            # 🧩 Diccionario de estados legibles
            ESTADOS = {
                0: "Inactivo",
                1: "Activo",
                2: "Bloqueado",
                3: "Caducado",
            }

            # 3️⃣ Procesar campos legibles
            for registro in datos_lista:
                # Producto
                if "producto_id" in registro:
                    producto = Producto.objects.filter(id=registro["producto_id"]).first()
                    registro["producto"] = (
                        f"{producto.descripcion} ({producto.clave_cnis})" if producto else ""
                    )
                    registro.pop("producto_id", None)

                # Institución
                if "institucion_id" in registro:
                    inst = Institucion.objects.filter(id=registro["institucion_id"]).first()
                    registro["institucion"] = (
                        f"{inst.clue} - {inst.denominacion}" if inst else ""
                    )
                    registro.pop("institucion_id", None)

                # Usuario creador
                if "creado_por_id" in registro:
                    user = User.objects.filter(id=registro["creado_por_id"]).first()
                    registro["creado_por"] = user.username if user else ""
                    registro.pop("creado_por_id", None)

                # Estado legible
                if "estado" in registro:
                    registro["estado"] = ESTADOS.get(registro["estado"], f"Desconocido ({registro['estado']})")

                # Fechas y decimales legibles
                for k, v in registro.items():
                    if isinstance(v, Decimal):
                        registro[k] = float(v)
                    elif isinstance(v, (date, datetime)):
                        registro[k] = v.strftime("%Y-%m-%d")
                    elif v is None:
                        registro[k] = ""

            # 4️⃣ Exportar a Excel respetando el orden de columnas
            df = pd.DataFrame(datos_lista)

            # Si hay orden definido, reordenamos las columnas
            columnas_finales = [col for col in campos if col in df.columns]
            df = df[columnas_finales]

            # Generar Excel
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="reporte_personalizado.xlsx"'
            df.to_excel(response, index=False)

            return response

        except Exception as e:
            print("❌ Error en sonalizado:", str(e))
            return JsonResponse({"error": f"Error generando el reporte: {str(e)}"}, status=500)

    

# ==========================================
# 3. PRODUCTOS
# ==========================================
@login_required
def lista_productos(request):
    productos = Producto.objects.select_related('categoria')
    search = request.GET.get('search')
    categoria_id = request.GET.get('categoria')
    es_cpm = request.GET.get('es_cpm')

    if search:
        productos = productos.filter(Q(clave_cnis__icontains=search) | Q(descripcion__icontains=search))
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    if es_cpm:
        productos = productos.filter(es_insumo_cpm=True)

    paginator = Paginator(productos, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'inventario/productos/lista.html', {
        'page_obj': page_obj,
        'categorias': CategoriaProducto.objects.all(),
        'search': search,
        'categoria_selected': categoria_id,
        'es_cpm': es_cpm,
    })


@login_required
def detalle_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    lotes = Lote.objects.filter(producto=producto).select_related('institucion', 'orden_suministro').order_by('-fecha_recepcion')

    total_stock = lotes.filter(estado=1).aggregate(total=Sum('cantidad_disponible'))['total'] or 0
    valor_total = lotes.filter(estado=1).aggregate(total=Sum('valor_total'))['total'] or 0
    distribucion = lotes.filter(estado=1).values('institucion__denominacion').annotate(
        cantidad=Sum('cantidad_disponible'), valor=Sum('valor_total')
    ).order_by('-cantidad')[:10]

    return render(request, 'inventario/productos/detalle.html', {
        'producto': producto,
        'lotes': lotes[:20],
        'total_stock': total_stock,
        'valor_total': valor_total,
        'distribucion': distribucion,
    })


@login_required
def crear_producto(request):
    form = ProductoForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Producto creado exitosamente.')
        return redirect('lista_productos')
    return render(request, 'inventario/productos/form.html', {'form': form, 'titulo': 'Crear Producto'})


@login_required
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    form = ProductoForm(request.POST or None, instance=producto)
    if form.is_valid():
        form.save()
        messages.success(request, 'Producto actualizado exitosamente.')
        return redirect('detalle_producto', pk=pk)
    return render(request, 'inventario/productos/form.html', {'form': form, 'titulo': 'Editar Producto'})


@login_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        if Lote.objects.filter(producto=producto).exists():
            messages.error(request, 'No se puede eliminar el producto porque tiene lotes asociados.')
            return redirect('detalle_producto', pk=producto.pk)
        producto.delete()
        messages.success(request, 'Producto eliminado exitosamente.')
        return redirect('lista_productos')
    return render(request, 'inventario/productos/confirmar_eliminar.html', {'producto': producto})

# ============================================================
# Lotes (Gestión de Inventario)
# ============================================================

@login_required
def lista_lotes(request):
    """Lista de lotes con filtros y resumen de caducidad"""
    form = FiltroInventarioForm(request.GET or None)
    hoy = date.today()

    lotes = (
        Lote.objects
        .select_related('producto', 'institucion', 'orden_suministro')
        .annotate(
            prox_caducar=Case(
                When(
                    fecha_caducidad__lte=hoy + timedelta(days=90),
                    fecha_caducidad__gte=hoy,
                    then=Value(True)
                ),
                default=Value(False),
                output_field=BooleanField()
            ),
            caducado=Case(
                When(fecha_caducidad__lt=hoy, then=Value(True)),
                default=Value(False),
                output_field=BooleanField()
            )
        )
    )

    # 🔹 Aplicar filtros
    if form.is_valid():
        institucion = form.cleaned_data.get('institucion')
        producto = form.cleaned_data.get('producto')
        categoria = form.cleaned_data.get('categoria')
        estado = form.cleaned_data.get('estado')

        if institucion:
            lotes = lotes.filter(institucion=institucion)
        if producto:
            lotes = lotes.filter(producto=producto)
        if categoria:
            lotes = lotes.filter(producto__categoria=categoria)
        if estado:
            lotes = lotes.filter(estado=estado)

    # 🔹 Filtro caducidad
    caducidad = request.GET.get('caducidad', '')
    if caducidad:
        if caducidad == 'caducados':
            lotes = lotes.filter(fecha_caducidad__lt=hoy)
        elif caducidad in ['30', '60', '90']:
            dias = int(caducidad)
            lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=dias))

    # 🔹 Filtro de búsqueda libre (lote, CNIS o producto)
    search = request.GET.get('search', '').strip()
    if search:
        lotes = lotes.filter(
            Q(numero_lote__icontains=search) |
            Q(producto__clave_cnis__icontains=search) |
            Q(producto__descripcion__icontains=search)
        )

    # 🔹 Resumen
    resumen = {
        'valor_total': lotes.aggregate(total=Sum('valor_total'))['total'] or 0,
        'cantidad_total': lotes.aggregate(total=Sum('cantidad_disponible'))['total'] or 0,
        'proximos_caducar': lotes.filter(prox_caducar=True).count(),
        'caducados': lotes.filter(caducado=True).count(),
    }

    # Ordenar para asegurar que el queryset se procesa correctamente
    lotes = lotes.order_by('-fecha_recepcion')
    
    paginator = Paginator(lotes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # DEBUG: Escribir en archivo
    with open('/tmp/debug_lotes.log', 'w') as f:
        f.write(f'Total lotes: {lotes.count()}\n')
        f.write(f'Lotes en pagina: {len(page_obj)}\n')
        f.write(f'Total paginas: {paginator.num_pages}\n')
        f.write(f'page_obj type: {type(page_obj)}\n')
        f.write(f'page_obj class: {page_obj.__class__.__name__}\n')
        f.write(f'lotes type: {type(lotes)}\n')

    instituciones = Institucion.objects.filter(activo=True).order_by('clue')
    alertas_caducidad = resumen['proximos_caducar'] + resumen['caducados']
    
    columnas_disponibles = [
        {"value": "numero_lote", "label": "Número de Lote"},
        {"value": "producto", "label": "Producto"},
        {"value": "institucion", "label": "Institución"},
        {"value": "cantidad_inicial", "label": "Cantidad Inicial"},
        {"value": "cantidad_disponible", "label": "Cantidad Disponible"},
        {"value": "precio_unitario", "label": "Precio Unitario"},
        {"value": "valor_total", "label": "Valor Total"},
        {"value": "fecha_fabricacion", "label": "Fecha de Fabricación"},
        {"value": "fecha_caducidad", "label": "Fecha de Caducidad"},
        {"value": "fecha_recepcion", "label": "Fecha de Recepción"},
        {"value": "estado", "label": "Estado"},
        {"value": "observaciones", "label": "Observaciones"},
        # --- Nuevos campos ---
        {"value": "rfc_proveedor", "label": "RFC Proveedor"},
        {"value": "proveedor", "label": "Proveedor"},
        {"value": "partida", "label": "Partida"},
        {"value": "clave_saica", "label": "Clave SAICA"},
        {"value": "descripcion_saica", "label": "Descripción SAICA"},
        {"value": "unidad_saica", "label": "Unidad SAICA"},
        {"value": "fuente_datos", "label": "Fuente de Datos"},
        {"value": "contrato", "label": "Contrato"},
        {"value": "folio", "label": "Folio"},
        {"value": "subtotal", "label": "Subtotal"},
        {"value": "iva", "label": "IVA"},
        {"value": "importe_total", "label": "Importe Total"},
        {"value": "licitacion", "label": "Licitación / Procedimiento"},
        {"value": "pedido", "label": "Pedido"},
        {"value": "remision", "label": "Remisión"},
        {"value": "responsable", "label": "Responsable"},
        {"value": "reviso", "label": "Revisó"},
        {"value": "tipo_entrega", "label": "Tipo de Entrega"},
        {"value": "tipo_red", "label": "Tipo de Red"},
        {"value": "epa", "label": "EPA"},
    ]

    context = {
        'form': form,
        'lotes': lotes,
        'instituciones': instituciones,
        'resumen': resumen,
        'institucion_selected': request.GET.get('institucion', ''),
        'estado_selected': request.GET.get('estado', ''),
        'caducidad_selected': caducidad,
        'search': search,
        'page_obj': page_obj,
        'alertas_caducidad': alertas_caducidad,
        "columnas_disponibles": columnas_disponibles,
    }
    return render(request, 'inventario/lotes/lista_lotes.html', context)


from django.shortcuts import get_object_or_404
from django.db.models import Sum, Q
from .models import Lote, MovimientoInventario

@login_required
def detalle_lote(request, pk):
    """Detalle de un lote con historial de movimientos"""
    lote = get_object_or_404(
        Lote.objects.select_related('producto', 'institucion', 'orden_suministro'),
        pk=pk
    )

    # Historial de movimientos asociados al lote
    movimientos = (
        MovimientoInventario.objects
        .filter(lote=lote)
        .order_by('-fecha_movimiento')
    )

    # Resumen de totales del lote
    resumen = {
        'entradas': movimientos.filter(tipo_movimiento='ENTRADA').aggregate(total=Sum('cantidad'))['total'] or 0,
        'salidas': movimientos.filter(tipo_movimiento='SALIDA').aggregate(total=Sum('cantidad'))['total'] or 0,
        'ajustes': movimientos.filter(tipo_movimiento='AJUSTE').aggregate(total=Sum('cantidad'))['total'] or 0,
        'transferencias': movimientos.filter(tipo_movimiento='TRANSFERENCIA').aggregate(total=Sum('cantidad'))['total'] or 0,
    }

    saldo_actual = lote.cantidad_disponible
    total_movimientos = movimientos.count()

    context = {
        'lote': lote,
        'movimientos': movimientos,
        'resumen': resumen,
        'saldo_actual': saldo_actual,
        'total_movimientos': total_movimientos,
    }
    return render(request, 'inventario/lotes/detalle_lote.html', context)



@login_required
def crear_lote(request):
    if request.method == 'POST':
        form = LoteForm(request.POST)
        if form.is_valid():
            lote = form.save(commit=False)
            lote.creado_por = request.user
            lote.save()
            messages.success(request, "Lote creado correctamente.")
            return redirect('lista_lotes')
    else:
        form = LoteForm()
    return render(request, 'inventario/lotes/form_lote.html', {'form': form, 'accion': 'Crear'})


@login_required
def editar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if request.method == 'POST':
        form = LoteForm(request.POST, instance=lote)
        if form.is_valid():
            lote = form.save(commit=False)
            lote.usuario_cambio_estado = request.user
            lote.fecha_cambio_estado = timezone.now()
            lote.save()
            messages.success(request, "✅ Lote actualizado correctamente.")
            return redirect('detalle_lote', pk=lote.pk)
    else:
        form = LoteForm(instance=lote)

    return render(request, 'inventario/lotes/form_lote.html', {
        'form': form,
        'accion': 'Editar',
    })



@login_required
def eliminar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if request.method == 'POST':
        lote.delete()
        messages.success(request, "Lote eliminado correctamente.")
        return redirect('lista_lotes')
    return render(request, 'inventario/lotes/confirmar_eliminar.html', {'objeto': lote})


# ============================================================
# Movimientos de Inventario
# ============================================================

@login_required
def lista_movimientos(request):
    search = request.GET.get('search', '')
    fecha_desde = request.GET.get('fecha_desde', '')

    movimientos = (
        MovimientoInventario.objects
        .select_related('lote__producto', 'lote__institucion', 'usuario')
        .order_by('-fecha_movimiento')
    )

    # 🔍 Filtros de búsqueda
    if search:
        movimientos = movimientos.filter(
            Q(lote__numero_lote__icontains=search) |
            Q(lote__producto__clave_cnis__icontains=search) |
            Q(motivo__icontains=search)
        )

    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__gte=fecha_desde)

    # 📄 Paginación
    paginator = Paginator(movimientos, 25)  # 25 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search': search,
        'fecha_desde': fecha_desde,
    }
    return render(request, 'inventario/movimientos/lista_movimientos.html', context)


@login_required
def editar_movimiento(request, pk):
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)

    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST, instance=movimiento)
        if form.is_valid():
            form.save()
            messages.success(request, "Movimiento actualizado correctamente.")
            return redirect('lista_movimientos')
    else:
        form = MovimientoInventarioForm(instance=movimiento)

    return render(request, 'inventario/movimientos/form_movimiento.html', {'form': form, 'editar': True})

@login_required
def detalle_movimiento(request, pk):
    """
    Muestra el detalle de un movimiento de inventario.
    """
    movimiento = get_object_or_404(MovimientoInventario.objects.select_related('lote', 'usuario', 'lote__producto', 'lote__institucion'), pk=pk)

    return render(
        request,
        'inventario/movimientos/detalle_movimiento.html',
        {'movimiento': movimiento}
    )


@login_required
def anular_movimiento(request, pk):
    """Anula un movimiento de inventario y revierte el stock asociado"""
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)

    # Evitar doble anulación
    if movimiento.anulado:
        messages.warning(request, "Este movimiento ya fue anulado anteriormente.")
        return redirect('lista_movimientos')

    lote = movimiento.lote

    # Revertir la cantidad según el tipo de movimiento
    if movimiento.tipo_movimiento in ['SALIDA', 'AJUSTE_NEGATIVO', 'TRANSFERENCIA_SALIDA']:
        lote.cantidad_disponible += movimiento.cantidad
    elif movimiento.tipo_movimiento in ['ENTRADA', 'AJUSTE_POSITIVO', 'TRANSFERENCIA_ENTRADA']:
        lote.cantidad_disponible -= movimiento.cantidad

    # Evitar valores negativos tras reversión
    if lote.cantidad_disponible < 0:
        messages.error(request, "No se puede anular: la reversión dejaría el stock negativo.")
        return redirect('lista_movimientos')

    # Guardar cambios
    lote.save()
    movimiento.anulado = True
    movimiento.fecha_anulacion = timezone.now()
    movimiento.usuario_anulacion = request.user
    movimiento.save()

    messages.success(request, f"Movimiento {movimiento.id} anulado correctamente y stock revertido.")
    return redirect('lista_movimientos')


@login_required
def crear_movimiento(request):
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            lote = movimiento.lote
            producto = lote.producto  # 👈 Relación con el producto
            cantidad_anterior = lote.cantidad_disponible

            # --- Calcular nueva cantidad del lote ---
            if movimiento.tipo_movimiento in ['SALIDA', 'AJUSTE_NEGATIVO', 'TRANSFERENCIA_SALIDA']:
                movimiento.cantidad_nueva = cantidad_anterior - movimiento.cantidad
            else:
                movimiento.cantidad_nueva = cantidad_anterior + movimiento.cantidad

            # --- Validación: no permitir cantidades negativas ---
            if movimiento.cantidad_nueva < 0:
                messages.error(request, "La cantidad resultante no puede ser negativa.")
                return redirect('crear_movimiento')

            # --- Guardar cambios en el lote ---
            lote.cantidad_disponible = movimiento.cantidad_nueva
            lote.save()

            # --- Actualizar cantidad global del producto ---
            # Se suma la cantidad disponible de todos los lotes del producto
            producto.existencia_total = sum(
                l.cantidad_disponible for l in producto.lote_set.all()
            )
            producto.save()

            # --- Guardar datos del movimiento ---
            movimiento.cantidad_anterior = cantidad_anterior
            movimiento.usuario = request.user
            movimiento.save()

            messages.success(request, "Movimiento registrado correctamente.")
            return redirect('lista_movimientos')
    else:
        form = MovimientoInventarioForm()

    return render(request, 'inventario/movimientos/form_movimiento.html', {'form': form})


# ============================================================
# Carga de Inventario desde Excel
# ============================================================

@login_required
def lista_cargas(request):
    cargas = CargaInventario.objects.select_related('usuario').order_by('-fecha_carga')
    return render(request, 'inventario/cargas/lista_cargas.html', {'cargas': cargas})


@login_required
def cargar_archivo_excel(request):
    if request.method == 'POST':
        form = CargaInventarioForm(request.POST, request.FILES)
        if form.is_valid():
            carga = form.save(commit=False)
            carga.nombre_archivo = request.FILES['archivo'].name
            carga.usuario = request.user
            carga.estado = 'PROCESANDO'
            carga.save()

            # Lógica de procesamiento del archivo Excel
            import pandas as pd
            try:
                df = pd.read_excel(carga.archivo.path)
                carga.total_registros = len(df)
                procesados = 0

                for _, row in df.iterrows():
                    try:
                        producto = Producto.objects.get(clave_cnis=row['clave_cnis'])
                        institucion = Institucion.objects.get(clue=row['clue'])
                        orden = OrdenSuministro.objects.first()

                        Lote.objects.create(
                            numero_lote=row['numero_lote'],
                            producto=producto,
                            institucion=institucion,
                            orden_suministro=orden,
                            cantidad_inicial=row['cantidad'],
                            cantidad_disponible=row['cantidad'],
                            precio_unitario=row['precio_unitario'],
                            fecha_fabricacion=row['fecha_fabricacion'],
                            fecha_caducidad=row['fecha_caducidad'],
                            fecha_recepcion=row['fecha_recepcion'],
                            creado_por=request.user
                        )
                        procesados += 1
                    except Exception as e:
                        carga.log_errores = (carga.log_errores or '') + f"\nError en registro: {e}"

                carga.registros_procesados = procesados
                carga.registros_exitosos = procesados
                carga.estado = 'COMPLETADA'
                messages.success(request, f"Carga completada ({procesados} registros procesados).")

            except Exception as e:
                carga.estado = 'ERROR'
                carga.log_errores = str(e)
                messages.error(request, f"Error al procesar archivo: {e}")

            carga.save()
            return redirect('lista_cargas')
    else:
        form = CargaInventarioForm()

    return render(request, 'inventario/cargas/form_carga.html', {'form': form})


@login_required
def detalle_carga(request, pk):
    carga = get_object_or_404(CargaInventario, pk=pk)
    return render(request, 'inventario/cargas/detalle_carga.html', {'carga': carga})


@login_required
def reporte_lotes_excel(request):
    """Exporta los lotes filtrados a Excel con layout personalizado"""

    hoy = date.today()
    lotes = Lote.objects.select_related('producto', 'institucion', 'orden_suministro')

    # 🔹 Obtener filtros desde GET (igual que en lista_lotes)
    search = request.GET.get('search', '').strip()
    institucion = request.GET.get('institucion')
    estado = request.GET.get('estado')
    caducidad = request.GET.get('caducidad')

    if search:
        lotes = lotes.filter(
            Q(numero_lote__icontains=search) |
            Q(producto__clave_cnis__icontains=search) |
            Q(producto__descripcion__icontains=search)
        )
    if institucion:
        lotes = lotes.filter(institucion_id=institucion)
    if estado:
        lotes = lotes.filter(estado=estado)
    if caducidad:
        if caducidad == 'caducados':
            lotes = lotes.filter(fecha_caducidad__lt=hoy)
        elif caducidad in ['30', '60', '90']:
            dias = int(caducidad)
            lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=dias))

    # 🔹 Crear workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Lotes"

    headers = [
        "ENTIDAD", "CLUES", "ORDEN DE SUMINISTRO", "RFC", "CLAVE",
        "DESCRIPCIÓN",  # Nuevo campo agregado
        "ESTADO DEL INSUMO", "INVENTARIO DISPONIBLE", "LOTE",
        "F_CAD", "F_FAB", "F_REC"
    ]
    ws.append(headers)

    for lote in lotes:
        row = [
            getattr(lote.institucion, 'entidad', ''),  # ENTIDAD
            lote.institucion.clue,                     # CLUES
            getattr(lote.orden_suministro, 'numero_orden', ''),  # ORDEN DE SUMINISTRO
            getattr(lote.producto, 'rfc', ''),        # RFC
            lote.producto.clave_cnis,                  # CLAVE
            lote.producto.descripcion,                 # DESCRIPCIÓN - Nuevo campo
            lote.get_estado_display(),                 # ESTADO DEL INSUMO
            lote.cantidad_disponible,                  # INVENTARIO DISPONIBLE
            lote.numero_lote,                          # LOTE
            lote.fecha_caducidad.strftime("%d/%m/%Y") if lote.fecha_caducidad else "",  # F_CAD
            lote.fecha_fabricacion.strftime("%d/%m/%Y") if lote.fecha_fabricacion else "", # F_FAB
            lote.fecha_recepcion.strftime("%d/%m/%Y") if lote.fecha_recepcion else ""      # F_REC
        ]
        ws.append(row)

    # Ajustar ancho de columnas
    for i, column_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value)) for cell in column_cells)
        # Ajuste especial para la columna de descripción (posición 6)
        if i == 6:  # Columna de DESCRIPCIÓN
            ws.column_dimensions[get_column_letter(i)].width = min(length + 2, 50)  # Máximo 50 caracteres
        else:
            ws.column_dimensions[get_column_letter(i)].width = length + 2

    # Devolver respuesta
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=lotes_inventario.xlsx'
    wb.save(response)
    return response

# ============================================================
# Reportes (Excel)
# ============================================================

@login_required
def descargar_reporte_inventario(request):
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append(['CLUE', 'Producto', 'Cantidad', 'Caducidad'])

    for lote in Lote.objects.all():
        ws.append([
            lote.institucion.clue,
            lote.producto.descripcion,
            lote.cantidad_disponible,
            lote.fecha_caducidad.strftime("%d/%m/%Y")
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'
    wb.save(response)
    return response


@login_required
def descargar_reporte_movimientos(request):
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws.append(['Fecha', 'Tipo', 'Producto', 'Cantidad', 'Usuario'])

    for mov in MovimientoInventario.objects.all():
        ws.append([
            mov.fecha_movimiento.strftime("%d/%m/%Y %H:%M"),
            mov.tipo_movimiento,
            mov.lote.producto.descripcion,
            mov.cantidad,
            mov.usuario.username
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_movimientos.xlsx"'
    wb.save(response)
    return response


@login_required
def descargar_reporte_caducidades(request):
    from openpyxl import Workbook
    from django.http import HttpResponse
    from datetime import date

    wb = Workbook()
    ws = wb.active
    ws.title = "Caducidades"

    ws.append(['CLUE', 'Producto', 'Fecha Caducidad', 'Días Restantes'])

    for lote in Lote.objects.all():
        dias = lote.dias_para_caducidad
        ws.append([
            lote.institucion.clue,
            lote.producto.descripcion,
            lote.fecha_caducidad.strftime("%d/%m/%Y"),
            dias
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_caducidades.xlsx"'
    wb.save(response)
    return response


# ============================================================
# Configuración y Ayuda
# ============================================================

@login_required
def configuracion_sistema(request):
    return render(request, 'inventario/configuracion.html')


@login_required
def ayuda_sistema(request):
    return render(request, 'inventario/ayuda.html')


@login_required
def manual_usuario(request):
    return render(request, 'inventario/manual_usuario.html')


# ============================================================
# Registro de Usuario
# ============================================================

def registro_usuario(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado correctamente. Ya puedes iniciar sesión.")
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/registro.html', {'form': form})

# ============================================================
# Acciones sobre Lotes (AJAX / utilidades)
# ============================================================

@login_required
def marcar_lote_caducado(request, pk):
    """Marca un lote como caducado"""
    lote = get_object_or_404(Lote, pk=pk)
    lote.estado = 6  # 6 = Caducado según tus choices
    lote.motivo_cambio_estado = "Marcado como caducado manualmente"
    lote.fecha_cambio_estado = timezone.now()
    lote.usuario_cambio_estado = request.user
    lote.save()

    messages.warning(request, f"El lote {lote.numero_lote} fue marcado como caducado.")
    return redirect('detalle_lote', pk=lote.pk)

@login_required
def crear_alerta_lote(request, pk):
    """Crea una alerta manual para un lote"""
    lote = get_object_or_404(Lote, pk=pk)
    AlertaCaducidad.objects.get_or_create(
        lote=lote,
        tipo_alerta='CADUCADO',
        defaults={'usuario_vista': request.user}
    )
    messages.info(request, f"Se generó una alerta para el lote {lote.numero_lote}.")
    return redirect('detalle_lote', pk=lote.pk)

# ============================================================
# Reportes - Panel principal de reportes
# ============================================================

@login_required
def reportes_dashboard(request):
    """Panel general de reportes con resumen de datos"""
    total_lotes = Lote.objects.count()
    total_movimientos = MovimientoInventario.objects.count()
    total_alertas = AlertaCaducidad.objects.count()

    proximos_a_caducar = Lote.objects.filter(
        fecha_caducidad__gte=timezone.now(),
        fecha_caducidad__lte=timezone.now() + timezone.timedelta(days=90)
    ).count()

    context = {
        'total_lotes': total_lotes,
        'total_movimientos': total_movimientos,
        'total_alertas': total_alertas,
        'proximos_a_caducar': proximos_a_caducar,
    }

    return render(request, 'inventario/reportes/dashboard.html', context)

@login_required
def alertas_caducidad(request):
    """Muestra las alertas de caducidad filtradas por prioridad e institución"""
    hoy = timezone.now().date()

    # Filtros GET
    prioridad_selected = request.GET.get('prioridad', '')
    institucion_selected = request.GET.get('institucion', '')
    search = request.GET.get('search', '')

    lotes = Lote.objects.filter(estado=1)  # Solo lotes activos

    # Filtrar por institución
    if institucion_selected:
        lotes = lotes.filter(institucion_id=institucion_selected)

    # Filtrar por búsqueda
    if search:
        lotes = lotes.filter(
            Q(producto__descripcion__icontains=search) |
            Q(producto__clave_cnis__icontains=search)
        )

    # Filtrar por prioridad
    if prioridad_selected == 'critica':
        lotes = lotes.filter(fecha_caducidad__lt=hoy)
    elif prioridad_selected == 'alta':
        lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=30))
    elif prioridad_selected == 'media':
        lotes = lotes.filter(fecha_caducidad__gte=hoy + timedelta(days=31), fecha_caducidad__lte=hoy + timedelta(days=60))
    elif prioridad_selected == 'baja':
        lotes = lotes.filter(fecha_caducidad__gte=hoy + timedelta(days=61), fecha_caducidad__lte=hoy + timedelta(days=90))
    else:
        # Si no hay filtro de prioridad, mostrar todos los lótes próximos a caducar (próximos 90 días)
        lotes = lotes.filter(fecha_caducidad__lte=hoy + timedelta(days=90))

    # Estadísticas generales para los cards
    stats = {
        'caducados': Lote.objects.filter(estado=1, fecha_caducidad__lt=hoy).count(),
        'proximos_30': Lote.objects.filter(estado=1, fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=30)).count(),
        'proximos_60': Lote.objects.filter(estado=1, fecha_caducidad__gte=hoy + timedelta(days=31), fecha_caducidad__lte=hoy + timedelta(days=60)).count(),
        'proximos_90': Lote.objects.filter(estado=1, fecha_caducidad__gte=hoy + timedelta(days=61), fecha_caducidad__lte=hoy + timedelta(days=90)).count(),
    }

    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')

    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(lotes.order_by('fecha_caducidad'), 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'instituciones': instituciones,
        'prioridad_selected': prioridad_selected,
        'institucion_selected': institucion_selected,
        'search': search,
    }
    return render(request, 'inventario/alertas/caducidad.html', context)

# inventario/views.py


@login_required
def carga_masiva_instituciones(request):
    if request.method == "POST" and request.FILES.get('archivo_excel'):
        archivo_excel = request.FILES['archivo_excel']

        try:
            df = pd.read_excel(archivo_excel)
        except Exception as e:
            messages.error(request, f"❌ Error al leer el archivo Excel: {str(e)}")
            return redirect('carga_masiva_instituciones')

        # Validar columnas mínimas
        columnas_necesarias = ['Clue', 'IB CLUE', 'Denominación', 'Dirección']
        for col in columnas_necesarias:
            if col not in df.columns:
                messages.error(request, f"❌ Falta la columna obligatoria: {col}")
                return redirect('carga_masiva_instituciones')

        registros_guardados = 0
        registros_omitidos = 0

        for _, row in df.iterrows():
            clue = str(row.get('Clue', '')).strip()
            ib_clue = str(row.get('IB CLUE', '')).strip()
            denominacion = str(row.get('Denominación', '')).strip()
            direccion = str(row.get('Dirección', '')).strip()

            if not clue:
                registros_omitidos += 1
                continue

            # update_or_create usando clue como clave, actualizando ib_clue
            institucion, created = Institucion.objects.update_or_create(
                clue=clue,
                defaults={
                    'ib_clue': ib_clue,
                    'denominacion': denominacion,
                    'direccion': direccion,
                }
            )
            registros_guardados += 1

        messages.success(request,
                         f"✅ Carga finalizada: {registros_guardados} guardados, {registros_omitidos} omitidos.")
        return redirect('lista_instituciones')

    return render(request, 'inventario/instituciones/carga_masiva_instituciones.html')


@login_required
def borrar_instituciones(request):
    if request.method == "POST":
        Institucion.objects.all().delete()
        messages.success(request, "✅ Todos los registros de instituciones han sido eliminados.")
        return redirect('lista_instituciones')
    return render(request, 'inventario/instituciones/instituciones_confirm_delete.html')


from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import pandas as pd
from datetime import datetime

from .models import SolicitudInventario, Producto, EstadoInsumo, Institucion

@login_required
def carga_masiva_solicitud(request):
    import pandas as pd
    from datetime import datetime
    from decimal import Decimal
    from django.db import transaction
    from django.contrib import messages
    from django.shortcuts import redirect, render
    from .models import Producto, CategoriaProducto, Lote, Institucion

    fecha_actual = datetime.now().date()
    errores = []
    registros_creados = 0

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo_excel = request.FILES['archivo_excel']

        # Leer archivo Excel
        try:
            df = pd.read_excel(archivo_excel)
            df.columns = [col.strip().upper() for col in df.columns]
        except Exception as e:
            messages.error(request, f"❌ Error al leer el archivo Excel: {str(e)}")
            return redirect('carga_masiva_solicitud')

        # Columnas requeridas
        columnas_esperadas = [
            'CLAVE/CNIS', 'DESCRIPCIÓN', 'UNIDAD DE MEDIDA',
            'LOTE', 'INVENTARIO DISPONIBLE', 'FECHA DE CADUCIDAD', 'CLUES'
        ]
        for col in columnas_esperadas:
            if col not in df.columns:
                messages.error(request, f"❌ Falta la columna obligatoria: '{col}'")
                return redirect('carga_masiva_solicitud')

        # Categoría por defecto
        categoria_default = CategoriaProducto.objects.filter(id=1).first()
        if not categoria_default:
            messages.error(
                request,
                "❌ No existe la categoría por defecto (id=1). "
                "Crea una llamada 'Asignar Categoría'."
            )
            return redirect('carga_masiva_solicitud')

        usuario = request.user

        # Procesar filas del Excel
        for index, row in df.iterrows():
            try:
                clave_cnis = str(row.get('CLAVE/CNIS', '')).strip()
                descripcion = str(row.get('DESCRIPCIÓN', '')).strip()
                unidad_medida = str(row.get('UNIDAD DE MEDIDA', '')).strip()
                lote_numero = str(row.get('LOTE', '')).strip()
                cantidad = row.get('INVENTARIO DISPONIBLE', 0) or 0
                fecha_caducidad = row.get('FECHA DE CADUCIDAD')
                if fecha_caducidad:
                    try:
                        # Si viene como texto, verificamos si es “S/C” o similar
                        if isinstance(fecha_caducidad, str):
                            texto = fecha_caducidad.strip().upper()
                            if texto in ["S/C", "SC", "SIN CADUCIDAD", "NA", "N/A"]:
                                fecha_caducidad = None
                            else:
                                fecha_caducidad = datetime.fromisoformat(fecha_caducidad).date()
                        elif hasattr(fecha_caducidad, 'date'):
                            fecha_caducidad = fecha_caducidad.date()
                    except Exception:
                        errores.append(f"Fila {index + 2}: Fecha de caducidad no válida, se marcó como S/C.")
                        fecha_caducidad = None
                else:
                    # Si viene vacío, también lo tratamos como sin caducidad
                    fecha_caducidad = None
                ib_clue = str(row.get('CLUES', '')).strip()

                # Validar fecha de caducidad
                if fecha_caducidad:
                    try:
                        if isinstance(fecha_caducidad, str):
                            fecha_caducidad = datetime.fromisoformat(fecha_caducidad).date()
                        elif hasattr(fecha_caducidad, 'date'):
                            fecha_caducidad = fecha_caducidad.date()
                    except Exception:
                        errores.append(f"Fila {index + 2}: Fecha de caducidad no válida.")
                        fecha_caducidad = None

                # Validar datos mínimos
                if not clave_cnis or not descripcion:
                    errores.append(f"Fila {index + 2}: Falta CLAVE/CNIS o descripción.")
                    continue
                if not ib_clue:
                    errores.append(f"Fila {index + 2}: Falta IB CLUE para asociar institución.")
                    continue

                # Buscar la institución correspondiente al IB CLUE
                #institucion = Institucion.objects.filter(ib_clue=ib_clue).first()
                institucion = Institucion.objects.filter(
                    Q(ib_clue=ib_clue) | Q(clue=ib_clue)
                ).first()

                if not institucion:
                    errores.append(f"Fila {index + 2}: No se encontró institución con IB CLUE '{ib_clue}'.")
                    continue

                with transaction.atomic():
                    # Crear o actualizar producto
                    producto, _ = Producto.objects.update_or_create(
                        clave_cnis=clave_cnis,
                        defaults={
                            'descripcion': descripcion,
                            'unidad_medida': unidad_medida,
                            'categoria': categoria_default,
                            'activo': True,
                        }
                    )

                    # Crear o actualizar lote
                    if lote_numero:
                        lote, created = Lote.objects.update_or_create(
                            producto=producto,
                            institucion=institucion,
                            numero_lote=lote_numero,
                            defaults={
                                'cantidad_inicial': cantidad,
                                'cantidad_disponible': cantidad,
                                'precio_unitario': Decimal('0.01'),
                                'valor_total': Decimal(cantidad) * Decimal('0.01'),
                                'fecha_caducidad': fecha_caducidad,
                                'fecha_fabricacion': fecha_actual,
                                'fecha_recepcion': fecha_actual,
                                'creado_por': usuario,
                            }
                        )
                        registros_creados += 1
                    else:
                        errores.append(f"Fila {index + 2}: No se proporcionó número de lote.")
                        continue

            except Exception as e:
                errores.append(f"Fila {index + 2}: Error inesperado ({str(e)})")

        # Mensajes finales
        if errores:
            for err in errores:
                messages.warning(request, err)

        messages.success(request, f"✅ Carga completada. Registros procesados: {registros_creados}")
        return redirect('carga_masiva_solicitud')

    # GET
    contexto = {
        'fecha_actual': fecha_actual,
        'errores': errores,
    }
    return render(request, 'inventario/solicitud/carga_masiva_solicitud.html', contexto)



@login_required
def complemento_carga_masiva(request):
    import pandas as pd
    from datetime import datetime
    from decimal import Decimal
    from django.db import transaction
    from django.contrib import messages
    from django.shortcuts import redirect, render
    from .models import Producto, CategoriaProducto, Lote, Institucion, Proveedor
    from django.db.models import Q

    fecha_actual = datetime.now().date()
    errores = []
    registros_creados = 0
    registros_actualizados = 0

    if request.method == 'POST' and request.FILES.get('archivo_csv'):
        archivo_csv = request.FILES['archivo_csv']

        try:
            df = pd.read_csv(archivo_csv)
            df.columns = [col.strip().replace('\n', ' ').replace('\r', '').upper() for col in df.columns]
        except Exception as e:
            messages.error(request, f"❌ Error al leer el archivo CSV: {str(e)}")
            return redirect('complemento_carga_masiva')

        columnas_minimas = ['DESCRIPCIÓN', 'UNIDAD DE MEDIDA']
        clave_col = 'CLAVE' if 'CLAVE' in df.columns else 'CLAVE (SAICA)' if 'CLAVE (SAICA)' in df.columns else None

        if not clave_col:
            messages.error(request, "❌ Falta la columna obligatoria de clave: 'CLAVE' o 'CLAVE (SAICA)'")
            return redirect('complemento_carga_masiva')

        for col in columnas_minimas:
            if col not in df.columns:
                messages.error(request, f"❌ Falta la columna obligatoria: '{col}'")
                return redirect('complemento_carga_masiva')

        categoria_default = CategoriaProducto.objects.filter(id=1).first()
        if not categoria_default:
            messages.error(request, "❌ No existe la categoría por defecto (id=1).")
            return redirect('complemento_carga_masiva')

        institucion_default = Institucion.objects.first()
        if not institucion_default:
            messages.error(request, "❌ No existe institución por defecto.")
            return redirect('complemento_carga_masiva')

        usuario = request.user

        # === Procesamiento de filas ===
        for index, row in df.iterrows():
            try:
                clave = str(row.get(clave_col, '')).strip()
                descripcion = str(row.get('DESCRIPCIÓN', '')).strip()
                unidad_medida = str(row.get('UNIDAD DE MEDIDA', 'PIEZA')).strip()
                cantidad = Decimal(str(row.get('CANT', row.get('CANTIDAD RECIBIDA', 0)) or 0))

                # Datos complementarios
                proveedor_nombre = str(row.get('PROVEEDOR', '')).strip()
                rfc = str(row.get('RFC', '')).strip().upper()
                partida = str(row.get('PARTIDA', '')).strip()
                clave_saica = str(row.get('CLAVE (SAICA)', '')).strip()
                descripcion_saica = str(row.get('DESCRIPCIÓN (SAICA)', '')).strip()
                unidad_saica = str(row.get('UNIDAD DE MEDIDA (SAICA)', '')).strip()
                marca = str(row.get('MARCA', '')).strip()
                fabricante = str(row.get('FABRICANTE', '')).strip()
                cns = str(row.get('CNS', '')).strip() or None
                fuente_datos = str(row.get('FUENTE DE DATOS', '')).strip() or None

                # Campos logísticos y de costos
                contrato = str(row.get('CONTRATO', '')).strip()
                remision = str(row.get('REMISION', '')).strip()
                pedido = str(row.get('PEDIDO', '')).strip()
                lote_num = str(row.get('LOTE', '')).strip()
                caducidad = row.get('CADUCIDAD', None)
                folio = str(row.get('FOLIO', '')).strip()
                observaciones = str(row.get('OBSERVACIONES', '')).strip()
                costo_unitario = Decimal(str(row.get('COSTO UNITARIO PESOS', 0) or 0))
                subtotal = Decimal(str(row.get('SUBTOTAL', 0) or 0))
                iva = Decimal(str(row.get('IVA', row.get('IVA.1', 0)) or 0))
                importe_total = Decimal(str(row.get('IMPORTE TOTAL', 0) or 0))
                tipo_entrega = str(row.get('TIPO DE ENTREGA', '')).strip()
                licitacion = str(row.get('LICITACION/PROCEDIMIENTO', '')).strip()
                tipo_red = str(row.get('TIPO DE RED', '')).strip()
                responsable = str(row.get('RESPONSABLE', '')).strip()
                reviso = str(row.get('REVISO', '')).strip()

                # Validación mínima
                if not clave or not descripcion:
                    errores.append(f"Fila {index + 2}: Falta CLAVE o DESCRIPCIÓN.")
                    continue

                with transaction.atomic():
                    # === Proveedor ===
                    proveedor_obj = None
                    if rfc:
                        proveedor_obj, _ = Proveedor.objects.update_or_create(
                            rfc=rfc,
                            defaults={
                                'razon_social': proveedor_nombre or "Proveedor sin nombre",
                                'activo': True,
                                'fecha_actualizacion': datetime.now()
                            }
                        )

                    # === Producto ===
                    producto, creado = Producto.objects.update_or_create(
                        clave_cnis=clave,
                        defaults={
                            'descripcion': descripcion,
                            'unidad_medida': unidad_medida,
                            'categoria': categoria_default,
                            'activo': True,
                            'clave_saica': clave_saica or None,
                            'descripcion_saica': descripcion_saica or None,
                            'unidad_medida_saica': unidad_saica or None,
                            'proveedor': proveedor_obj.razon_social if proveedor_obj else proveedor_nombre or None,
                            'rfc_proveedor': proveedor_obj.rfc if proveedor_obj else rfc or None,
                            'partida_presupuestal': partida or None,
                            'marca': marca or None,
                            'fabricante': fabricante or None,
                            'fecha_actualizacion': datetime.now(),
                        }
                    )

                    if creado:
                        registros_creados += 1
                    else:
                        registros_actualizados += 1

                    # === Lote ===
                    if cantidad > 0:
                        numero_lote = lote_num or f"L-{clave}-{fecha_actual.strftime('%Y%m%d')}"
                        fecha_fabricacion_lote = fecha_actual
                        if row.get('FECHA_FABRICACION'):
                            try:
                                fecha_fabricacion_lote = pd.to_datetime(row['FECHA_FABRICACION']).date()
                            except:
                                pass

                        Lote.objects.update_or_create(
                            producto=producto,
                            institucion=institucion_default,
                            numero_lote=numero_lote,
                            defaults={
                                'cantidad_inicial': cantidad,
                                'cantidad_disponible': cantidad,
                                'precio_unitario': costo_unitario,
                                'valor_total': importe_total or (cantidad * costo_unitario),
                                'fecha_fabricacion': fecha_fabricacion_lote,
                                'fecha_recepcion': fecha_actual,
                                'observaciones': observaciones or None,
                                'contrato': contrato or None,
                                'remision': remision or None,
                                'pedido': pedido or None,
                                'folio': folio or None,
                                'tipo_entrega': tipo_entrega or None,
                                'licitacion': licitacion or None,
                                'tipo_red': tipo_red or None,
                                'responsable': responsable or None,
                                'reviso': reviso or None,
                                'subtotal': subtotal or None,
                                'iva': iva or None,
                                'importe_total': importe_total or None,
                                'creado_por': usuario,
                                'cns': cns,
                                'fuente_datos': fuente_datos,
                                'estado': 1,
                                # Campos SAICA y proveedor
                                'proveedor': proveedor_obj.razon_social if proveedor_obj else proveedor_nombre or None,
                                'rfc_proveedor': proveedor_obj.rfc if proveedor_obj else rfc or None,
                                'partida': partida or None,
                                'clave_saica': clave_saica or None,
                                'descripcion_saica': descripcion_saica or None,
                                'unidad_saica': unidad_saica or None,
                            }
                        )

            except Exception as e:
                import traceback
                errores.append(f"Fila {index + 2}: {str(e)}\nValores: {row.to_dict()}\n{traceback.format_exc()}")

        # --- Mensajes finales ---
        for err in errores:
            messages.warning(request, err)

        messages.success(
            request,
            f"✅ Complemento de carga completado: {registros_creados} creados, {registros_actualizados} actualizados."
        )
        return redirect('complemento_carga_masiva')

    # GET
    contexto = {
        'fecha_actual': fecha_actual,
        'errores': errores,
    }
    return render(request, 'inventario/solicitud/complemento_carga_masiva.html', contexto)


def ajax_ubicaciones_por_almacen(request):
    almacen_id = request.GET.get('almacen')
    ubicaciones = []
    if almacen_id:
        ubicaciones = list(UbicacionAlmacen.objects.filter(
            almacen_id=almacen_id,
            activo=True
        ).values('id', 'codigo', 'descripcion'))
    return JsonResponse(ubicaciones, safe=False)


@login_required
def borrar_solicitud(request, fecha=None):
    """
    Borrar solicitudes por fecha de generación.
    Si fecha=None, borra TODAS las solicitudes.
    """
    if request.method == "POST":
        if fecha:
            SolicitudInventario.objects.filter(fecha_generacion=fecha).delete()
            messages.success(request, f"✅ Solicitudes de {fecha} eliminadas.")
        else:
            SolicitudInventario.objects.all().delete()
            messages.success(request, "✅ Todas las solicitudes eliminadas.")
        return redirect('lista_solicitud_inventario')

    context = {
        'fecha': fecha
    }
    return render(request, 'inventario/solicitud/solicitud_confirm_delete.html', context)

@login_required
def lista_solicitudes_semanales(request):
    # Obtener todas las fechas disponibles
    fechas = SolicitudInventario.objects.values_list('fecha_generacion', flat=True).distinct().order_by('-fecha_generacion')
    
    # Obtener la fecha del parámetro GET o usar la más reciente
    fecha_param = request.GET.get('fecha', '')
    
    if fecha_param:
        try:
            # Convertir parámetro a fecha
            fecha_seleccionada = datetime.strptime(fecha_param, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            # Si el formato es inválido, usar la fecha más reciente
            fecha_seleccionada = fechas.first() if fechas else timezone.now().date()
    else:
        # Si no hay parámetro, usar la fecha más reciente
        fecha_seleccionada = fechas.first() if fechas else timezone.now().date()
    
    # Filtrar solicitudes por la fecha seleccionada
    solicitudes = SolicitudInventario.objects.filter(
        fecha_generacion=fecha_seleccionada
    ).select_related('estado_insumo').order_by('clues', 'clave_cnis')
    
    # CALCULAR LOS TOTALES EN LA VISTA (EVITAR HACERLO EN EL TEMPLATE)
    total_solicitudes = solicitudes.count()
    total_valor = solicitudes.aggregate(total=Sum('valor_total'))['total'] or 0
    total_inventario = solicitudes.aggregate(total=Sum('inventario_disponible'))['total'] or 0
    
    # Pasar la fecha en formato string para la exportación
    fecha_exportacion = fecha_seleccionada.isoformat()
    
    context = {
        'fechas': fechas,
        'solicitudes': solicitudes,
        'fecha_seleccionada': fecha_seleccionada,
        'fecha_actual': fecha_exportacion,
        'fecha_formateada': fecha_seleccionada.strftime('%d/%m/%Y'),
        'total_solicitudes': total_solicitudes,
        'total_valor': total_valor,
        'total_inventario': total_inventario,
    }
    
    return render(request, 'inventario/solicitud/lista_solicitudes.html', context)


@login_required
def detalle_solicitud(request, fecha):
    solicitudes = SolicitudInventario.objects.filter(fecha_generacion=fecha)
    return render(request, 'inventario/solicitud/detalle_solicitud.html', {'solicitudes': solicitudes, 'fecha': fecha})


@login_required
def exportar_solicitud_excel(request, fecha):
    try:
        # Convertir el string de fecha a objeto date
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Formato de fecha inválido.")
        return redirect('lista_solicitudes')
    
    # Filtrar por la fecha convertida
    solicitudes = SolicitudInventario.objects.filter(fecha_generacion=fecha_obj)
    
    if not solicitudes.exists():
        messages.error(request, "No hay solicitudes para exportar.")
        return redirect('lista_solicitudes')

    # Crear DataFrame
    data = []
    for s in solicitudes:
        data.append({
            'Entidad Federativa': s.entidad_federativa or '',
            'CLUES': s.clues or '',
            'Orden de Suministro': s.orden_suministro or '',
            'RFC Proveedor': s.rfc_proveedor or '',
            'Fuente de Financiamiento': s.fuente_financiamiento or '',
            'Partida Presupuestal': s.partida_presupuestal or '',
            'Concatenar': s.concatenar or '',
            'Clave/CNIS': s.clave_cnis or '',
            'Descripción': s.descripcion or '',
            'Precio Unitario': float(s.precio_unitario) if s.precio_unitario else 0.0,
            'Valor Total': float(s.valor_total) if s.valor_total else 0.0,
            'Insumo en CPM': s.insumo_en_cpm or '',
            'Estado del Insumo': s.estado_insumo.descripcion if s.estado_insumo else '',
            'Inventario Disponible': s.inventario_disponible or 0,
            'Unidad de Medida': s.unidad_medida or '',
            'Lote': s.lote or '',
            'Fecha Caducidad': s.fecha_caducidad.strftime('%d/%m/%Y') if s.fecha_caducidad else '',
            'Fecha Fabricación': s.fecha_fabricacion.strftime('%d/%m/%Y') if s.fecha_fabricacion else '',
            'Fecha Recepción': s.fecha_recepcion.strftime('%d/%m/%Y') if s.fecha_recepcion else '',
        })
    
    df = pd.DataFrame(data)

    # Crear respuesta Excel con formato mejorado
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="solicitudes_{fecha}.xlsx"'
    
    # Usar ExcelWriter para mejor control del formato
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Solicitudes', index=False)
        
        # Autoajustar el ancho de las columnas
        worksheet = writer.sheets['Solicitudes']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return response

@login_required
def carga_lotes_desde_excel_view(request):
    """
    Vista para cargar lotes desde archivo Excel
    """
    if request.method == 'POST':
        form = CargaLotesForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            institucion_id = form.cleaned_data['institucion'].id
            
            # Guardar archivo temporalmente
            file_path = f"/tmp/{archivo.name}"
            with open(file_path, 'wb+') as destination:
                for chunk in archivo.chunks():
                    destination.write(chunk)
            
            # Procesar archivo
            resultado = carga_lotes_desde_excel(
                archivo_excel=file_path,
                institucion_id=institucion_id,
                usuario=request.user
            )
            
            # Limpiar archivo temporal
            if os.path.exists(file_path):
                os.remove(file_path)
            
            if resultado['success']:
                messages.success(
                    request, 
                    f"✅ Carga completada: {resultado['exitosos']} creados, "
                    f"{resultado['actualizados']} actualizados, {resultado['errores']} errores"
                )
                if resultado['ubicaciones_no_encontradas']:
                    messages.warning(
                        request, 
                        f"⚠️ {len(resultado['ubicaciones_no_encontradas'])} ubicaciones no encontradas"
                    )
            else:
                messages.error(request, resultado['message'])
            
            return redirect('carga_lotes_excel')
    
    else:
        form = CargaLotesForm()
    
    return render(request, 'inventario/carga_lotes_excel.html', {'form': form})


# Función mejorada para editar_ubicaciones_lote
# Reemplazar la función existente en views.py (líneas 1788-1882)

@login_required
def editar_ubicaciones_lote(request, pk):
    """
    Vista para editar las ubicaciones asignadas a un lote.
    Permite agregar, editar y eliminar ubicaciones.
    Registra movimientos de inventario para cada cambio.
    """
    from .models import UbicacionAlmacen, Almacen, MovimientoInventario
    
    lote = get_object_or_404(Lote, pk=pk)
    ubicaciones_actuales = LoteUbicacion.objects.filter(lote=lote).order_by('ubicacion__codigo')
    
    if request.method == 'POST':
        # Procesar formulario POST
        try:
            # Obtener las ubicaciones enviadas
            ubicaciones_data = {}
            for key, value in request.POST.items():
                if key.startswith('ubicacion_'):
                    parts = key.split('_')
                    if len(parts) >= 3:
                        ubicacion_id = parts[1]
                        field_name = '_'.join(parts[2:])
                        
                        if ubicacion_id not in ubicaciones_data:
                            ubicaciones_data[ubicacion_id] = {}
                        ubicaciones_data[ubicacion_id][field_name] = value
            
            # Actualizar ubicaciones existentes
            for ubicacion_id, data in ubicaciones_data.items():
                try:
                    ubicacion = LoteUbicacion.objects.get(id=ubicacion_id, lote=lote)
                    if 'cantidad' in data and data['cantidad']:
                        cantidad_anterior = ubicacion.cantidad
                        cantidad_nueva = int(data['cantidad'])
                        
                        # Registrar movimiento si hay cambio
                        if cantidad_anterior != cantidad_nueva:
                            diferencia = cantidad_nueva - cantidad_anterior
                            tipo_movimiento = 'AJUSTE_POSITIVO' if diferencia > 0 else 'AJUSTE_NEGATIVO'
                            
                            MovimientoInventario.objects.create(
                                lote=lote,
                                tipo_movimiento=tipo_movimiento,
                                cantidad=abs(diferencia),
                                cantidad_anterior=cantidad_anterior,
                                cantidad_nueva=cantidad_nueva,
                                motivo=f'Ajuste de cantidad en ubicación {ubicacion.ubicacion.codigo} (Edición de ubicaciones)',
                                usuario=request.user
                            )
                        
                        ubicacion.cantidad = cantidad_nueva
                        ubicacion.save()
                except (ValueError, LoteUbicacion.DoesNotExist):
                    pass
            
            # Agregar nueva ubicación si se proporciona
            nueva_ubicacion_id = request.POST.get('nueva_ubicacion')
            nueva_cantidad = request.POST.get('nueva_cantidad')
            
            if nueva_ubicacion_id and nueva_cantidad:
                try:
                    nueva_cantidad = int(nueva_cantidad)
                    if nueva_cantidad > 0:
                        ubicacion_almacen = UbicacionAlmacen.objects.get(id=nueva_ubicacion_id)
                        
                        # Verificar que no exista ya
                        if not LoteUbicacion.objects.filter(lote=lote, ubicacion=ubicacion_almacen).exists():
                            lote_ubicacion = LoteUbicacion.objects.create(
                                lote=lote,
                                ubicacion=ubicacion_almacen,
                                cantidad=nueva_cantidad,
                                usuario_asignacion=request.user
                            )
                            
                            # Registrar movimiento de entrada
                            MovimientoInventario.objects.create(
                                lote=lote,
                                tipo_movimiento='AJUSTE_POSITIVO',
                                cantidad=nueva_cantidad,
                                cantidad_anterior=0,
                                cantidad_nueva=nueva_cantidad,
                                motivo=f'Nueva ubicación {ubicacion_almacen.codigo} agregada (Edición de ubicaciones)',
                                usuario=request.user
                            )
                            
                            messages.success(request, f"✅ Ubicación {ubicacion_almacen.codigo} agregada correctamente.")
                        else:
                            messages.warning(request, f"⚠️ La ubicación {ubicacion_almacen.codigo} ya está asignada a este lote.")
                except (ValueError, UbicacionAlmacen.DoesNotExist):
                    messages.error(request, "❌ Error al agregar la ubicación.")
            
            # Eliminar ubicaciones marcadas
            eliminar_ids = request.POST.getlist('eliminar_ubicacion')
            for ubicacion_id in eliminar_ids:
                try:
                    ubicacion = LoteUbicacion.objects.get(id=ubicacion_id, lote=lote)
                    cantidad_eliminada = ubicacion.cantidad
                    
                    # Registrar movimiento de eliminación
                    MovimientoInventario.objects.create(
                        lote=lote,
                        tipo_movimiento='AJUSTE_NEGATIVO',
                        cantidad=cantidad_eliminada,
                        cantidad_anterior=cantidad_eliminada,
                        cantidad_nueva=0,
                        motivo=f'Ubicación {ubicacion.ubicacion.codigo} eliminada (Edición de ubicaciones)',
                        usuario=request.user
                    )
                    
                    ubicacion.delete()
                    messages.success(request, f"✅ Ubicación eliminada correctamente.")
                except LoteUbicacion.DoesNotExist:
                    pass
            
            # Sincronizar cantidad del lote después de cambios
            lote.sincronizar_cantidad_disponible()
            
            messages.success(request, "✅ Ubicaciones actualizadas correctamente.")
            return redirect('editar_ubicaciones_lote', pk=lote.pk)
        
        except Exception as e:
            messages.error(request, f"❌ Error al guardar: {str(e)}")
    
    # Obtener ubicaciones disponibles del almacén
    ubicaciones_disponibles = UbicacionAlmacen.objects.filter(
        almacen=lote.almacen,
        activo=True
    ).exclude(
        id__in=ubicaciones_actuales.values_list('ubicacion_id', flat=True)
    ).order_by('codigo')
    
    context = {
        'lote': lote,
        'ubicaciones_actuales': ubicaciones_actuales,
        'ubicaciones_disponibles': ubicaciones_disponibles,
        'total_cantidad': sum(u.cantidad for u in ubicaciones_actuales),
    }
    
    return render(request, 'inventario/lotes/editar_ubicaciones_lote.html', context)


