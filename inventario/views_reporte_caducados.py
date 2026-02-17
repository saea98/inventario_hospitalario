"""
Reporte de Caducados y Próximos a Caducar

Muestra lotes caducados y los que caducan en 90, 60 o 30 días,
con el mismo layout de columnas que el reporte de entradas (34 columnas + DÍAS PARA CADUCAR).
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Lote, Institucion, Almacen, UbicacionAlmacen

# Mismo layout que reporte de entradas (34 columnas) + columna de estado caducidad
CADUCADOS_LAYOUT_HEADERS = [
    'RFC',
    'PROVEEDOR',
    'PARTIDA',
    'Clave CNIS',
    'Producto',
    'UNIDAD DE MEDIDA',
    'CANTIDAD RECIBIDA',
    'FECHA DE ENTREGA',
    'LUGAR DE ENTREGA',
    'CONTRATO',
    'REMISION',
    'ORDEN DE SUMINISTRO',
    'LOTE',
    'CADUCIDAD',
    'FOLIO',
    'ESTADO LLEGADA',
    'UBIC. ASIGNADA',
    'PRECIO UNIT. CON IVA / SIN IVA',
    'SUBTOTAL',
    'IVA',
    'IMPORTE TOTAL',
    'MARCA',
    'FABRICANTE',
    'FECHA DE FABRICACIÓN',
    'CADUCIDAD CONFORME A CERTIFICADO',
    'TIPO DE ENTREGA',
    'LICITACION/PROCEDIMIENTO',
    'RESPONSABLE',
    'REVISO',
    'TIPO DE RED',
    'FECHA DE CAPTURA',
    'OBSERVACION',
    'FECHA DE EMISIÓN',
    'FUENTE DE FMTO',
    'DÍAS PARA CADUCAR',
]

RANGO_CADUCIDAD = [
    ('', 'Todos'),
    ('caducado', 'Caducado'),
    ('30', '≤ 30 días'),
    ('60', '≤ 60 días'),
    ('90', '≤ 90 días'),
]


def _valor(o, default=''):
    if o is None:
        return default
    return o


def _fecha(d, fmt='%d/%m/%Y'):
    if not d:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime(fmt)
    return str(d)


def _decimal(d, default=''):
    if d is None:
        return default
    if isinstance(d, Decimal):
        return float(d)
    return d


def _dias_para_caducar(lote):
    if not lote or not lote.fecha_caducidad:
        return None
    return (lote.fecha_caducidad - timezone.now().date()).days


def _etiqueta_rango(dias):
    if dias is None:
        return '—'
    if dias < 0:
        return 'Caducado'
    if dias <= 30:
        return '≤ 30 días'
    if dias <= 60:
        return '≤ 60 días'
    if dias <= 90:
        return '≤ 90 días'
    return '> 90 días'


def _construir_fila_caducado(lote):
    """Construye la fila con el mismo orden que el reporte de entradas (34 cols) + DÍAS PARA CADUCAR."""
    prod = lote.producto if lote else None
    os = lote.orden_suministro if lote else None
    prov = os.proveedor if os else None
    precio = (lote and lote.precio_unitario) or Decimal('0')
    cantidad = (lote and lote.cantidad_disponible) or 0
    subtotal_val = (lote and lote.subtotal) or (cantidad * precio)
    iva_val = (lote and lote.iva) or Decimal('0')
    importe_val = (lote and lote.importe_total) or (cantidad * precio)

    lugar_entrega = ''
    if lote:
        lugar_entrega = (lote.almacen and lote.almacen.nombre) or (lote.institucion and lote.institucion.denominacion) or ''

    ubicacion_label = (lote.ubicacion.codigo if lote and lote.ubicacion else '') or '—'

    dias = _dias_para_caducar(lote)
    if dias is None:
        etiqueta_dias = '—'
    elif dias < 0:
        etiqueta_dias = f"Caducado ({-dias} días)"
    else:
        etiqueta_dias = f"{dias} días"

    row = [
        _valor(prov and prov.rfc or (lote and lote.rfc_proveedor)),
        _valor(prov and prov.razon_social or (lote and lote.proveedor)),
        _valor(lote and lote.partida or (os and os.partida_presupuestal)),
        _valor(prod and prod.clave_cnis),
        _valor(prod and prod.descripcion),
        _valor(prod and prod.unidad_medida) or 'PIEZA',
        cantidad,
        _fecha(lote and lote.fecha_recepcion),
        lugar_entrega,
        _valor(lote and lote.contrato),
        _valor(lote and lote.remision),
        _valor(os and os.numero_orden),
        _valor(lote and lote.numero_lote),
        _fecha(lote and lote.fecha_caducidad),
        _valor(lote and lote.folio),
        '—',  # ESTADO LLEGADA (no aplica en caducados)
        _valor(ubicacion_label),
        _decimal(precio),
        _decimal(lote and lote.subtotal) if (lote and lote.subtotal) is not None else _decimal(subtotal_val),
        _decimal(lote and lote.iva) if (lote and lote.iva) is not None else _decimal(iva_val),
        _decimal(lote and lote.importe_total) if (lote and lote.importe_total) is not None else _decimal(importe_val),
        _valor(prod and prod.marca),
        _valor(prod and prod.fabricante),
        _fecha(lote and lote.fecha_fabricacion),
        _fecha(lote and lote.fecha_caducidad),
        _valor(lote and lote.tipo_entrega),
        _valor(lote and lote.licitacion),
        _valor(lote and lote.responsable),
        _valor(lote and lote.reviso),
        _valor(lote and lote.tipo_red),
        _fecha(lote and getattr(lote, 'fecha_creacion', None), '%d/%m/%Y %H:%M') if lote and getattr(lote, 'fecha_creacion', None) else '',
        _valor(lote and lote.observaciones),
        _fecha(os and os.fecha_orden) or _fecha(lote and lote.fecha_recepcion),
        _valor((os and os.fuente_financiamiento and os.fuente_financiamiento.nombre) or (lote and lote.fuente_datos)),
        etiqueta_dias if dias is not None else '—',
    ]
    return row


@login_required
def reporte_caducados(request):
    """
    Reporte de lotes caducados y próximos a caducar (≤ 90 días).
    Mismo layout de columnas que el reporte de entradas + DÍAS PARA CADUCAR.
    """
    hoy = timezone.now().date()
    limite_90 = hoy + timedelta(days=90)

    lotes = Lote.objects.filter(
        fecha_caducidad__isnull=False,
        fecha_caducidad__lte=limite_90,
        cantidad_disponible__gt=0,
    ).select_related(
        'producto',
        'institucion',
        'almacen',
        'ubicacion',
        'orden_suministro',
        'orden_suministro__proveedor',
        'orden_suministro__fuente_financiamiento',
    ).order_by('fecha_caducidad', 'numero_lote')

    # Filtros
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_institucion = request.GET.get('institucion', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_ubicacion = request.GET.get('ubicacion', '')
    filtro_rango = request.GET.get('rango', '').strip()

    if filtro_clave:
        lotes = lotes.filter(producto__clave_cnis__icontains=filtro_clave)
    if filtro_lote:
        lotes = lotes.filter(numero_lote__icontains=filtro_lote)
    if filtro_institucion:
        lotes = lotes.filter(institucion_id=filtro_institucion)
    if filtro_almacen:
        lotes = lotes.filter(almacen__nombre=filtro_almacen)
    if filtro_ubicacion:
        lotes = lotes.filter(ubicacion_id=filtro_ubicacion)

    # Construir lista de filas
    lista_filas = []
    total_cantidad = 0
    total_valor = 0

    for lote in lotes:
        dias = _dias_para_caducar(lote)
        # Filtro por rango: caducado, ≤30, ≤60, ≤90
        if filtro_rango == 'caducado':
            if dias is None or dias >= 0:
                continue
        elif filtro_rango == '30':
            if dias is None or dias < 0 or dias > 30:
                continue
        elif filtro_rango == '60':
            if dias is None or dias < 0 or dias > 60:
                continue
        elif filtro_rango == '90':
            if dias is None or dias < 0 or dias > 90:
                continue

        row = _construir_fila_caducado(lote)
        lista_filas.append({'row': row, 'id': lote.id, 'dias': dias})
        total_cantidad += row[6] or 0
        total_valor += float(row[20] or 0)

    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(lista_filas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    instituciones = Institucion.objects.all().order_by('denominacion')
    almacenes = Almacen.objects.all().order_by('nombre')
    ubicaciones = UbicacionAlmacen.objects.filter(activo=True).select_related('almacen').order_by('almacen__nombre', 'codigo')

    # Conteos por rango para tarjetas
    lotes_sin_filtro_rango = Lote.objects.filter(
        fecha_caducidad__isnull=False,
        fecha_caducidad__lte=limite_90,
        cantidad_disponible__gt=0,
    )
    if filtro_clave:
        lotes_sin_filtro_rango = lotes_sin_filtro_rango.filter(producto__clave_cnis__icontains=filtro_clave)
    if filtro_lote:
        lotes_sin_filtro_rango = lotes_sin_filtro_rango.filter(numero_lote__icontains=filtro_lote)
    if filtro_institucion:
        lotes_sin_filtro_rango = lotes_sin_filtro_rango.filter(institucion_id=filtro_institucion)
    if filtro_almacen:
        lotes_sin_filtro_rango = lotes_sin_filtro_rango.filter(almacen__nombre=filtro_almacen)
    if filtro_ubicacion:
        lotes_sin_filtro_rango = lotes_sin_filtro_rango.filter(ubicacion_id=filtro_ubicacion)

    conteo_caducado = lotes_sin_filtro_rango.filter(fecha_caducidad__lt=hoy).count()
    conteo_30 = lotes_sin_filtro_rango.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=30)).count()
    conteo_60_total = lotes_sin_filtro_rango.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=60)).count()
    conteo_90_extra = lotes_sin_filtro_rango.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=90)).exclude(fecha_caducidad__lte=hoy + timedelta(days=60)).count()
    conteo_90_total = conteo_60_total + conteo_90_extra  # total próximos 0-90 días

    context = {
        'page_obj': page_obj,
        'headers': CADUCADOS_LAYOUT_HEADERS,
        'total_registros': len(lista_filas),
        'total_cantidad': total_cantidad,
        'total_valor': total_valor,
        'instituciones': instituciones,
        'almacenes': almacenes,
        'ubicaciones': ubicaciones,
        'filtro_clave': filtro_clave,
        'filtro_lote': filtro_lote,
        'filtro_institucion': filtro_institucion,
        'filtro_almacen': filtro_almacen,
        'filtro_ubicacion': filtro_ubicacion,
        'filtro_rango': filtro_rango,
        'rango_choices': RANGO_CADUCIDAD,
        'conteo_caducado': conteo_caducado,
        'conteo_30': conteo_30,
        'conteo_60': conteo_60_total,
        'conteo_90': conteo_90_total,
    }
    return render(request, 'inventario/reporte_caducados.html', context)


@login_required
def exportar_caducados_excel(request):
    """Exporta el reporte de caducados a Excel con el mismo layout que entradas + DÍAS PARA CADUCAR."""
    hoy = timezone.now().date()
    limite_90 = hoy + timedelta(days=90)

    lotes = Lote.objects.filter(
        fecha_caducidad__isnull=False,
        fecha_caducidad__lte=limite_90,
        cantidad_disponible__gt=0,
    ).select_related(
        'producto',
        'institucion',
        'almacen',
        'ubicacion',
        'orden_suministro',
        'orden_suministro__proveedor',
        'orden_suministro__fuente_financiamiento',
    ).order_by('fecha_caducidad', 'numero_lote')

    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_institucion = request.GET.get('institucion', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_ubicacion = request.GET.get('ubicacion', '')
    filtro_rango = request.GET.get('rango', '').strip()

    if filtro_clave:
        lotes = lotes.filter(producto__clave_cnis__icontains=filtro_clave)
    if filtro_lote:
        lotes = lotes.filter(numero_lote__icontains=filtro_lote)
    if filtro_institucion:
        lotes = lotes.filter(institucion_id=filtro_institucion)
    if filtro_almacen:
        lotes = lotes.filter(almacen__nombre=filtro_almacen)
    if filtro_ubicacion:
        lotes = lotes.filter(ubicacion_id=filtro_ubicacion)

    listado = []
    total_cantidad = 0
    total_importe = 0
    for lote in lotes:
        dias = _dias_para_caducar(lote)
        if filtro_rango == 'caducado' and (dias is None or dias >= 0):
            continue
        if filtro_rango == '30' and (dias is None or dias < 0 or dias > 30):
            continue
        if filtro_rango == '60' and (dias is None or dias < 0 or dias > 60):
            continue
        if filtro_rango == '90' and (dias is None or dias < 0 or dias > 90):
            continue
        fila = _construir_fila_caducado(lote)
        listado.append(fila)
        total_cantidad += fila[6] or 0
        total_importe += float(fila[20] or 0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Caducados y próximos"
    header_fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
    header_font = Font(bold=True, color="FFEBEE", size=10)
    total_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    total_font = Font(bold=True, size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(CADUCADOS_LAYOUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_num, fila in enumerate(listado, 2):
        for col_num, val in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = border
            if col_num in (7, 17, 18, 19, 20, 21):
                cell.alignment = Alignment(horizontal='right')

    total_row = len(listado) + 2
    ws.cell(row=total_row, column=1).value = "TOTALES"
    ws.cell(row=total_row, column=1).font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=7).value = total_cantidad
    ws.cell(row=total_row, column=7).font = total_font
    ws.cell(row=total_row, column=7).fill = total_fill
    ws.cell(row=total_row, column=21).value = total_importe
    ws.cell(row=total_row, column=21).font = total_font
    ws.cell(row=total_row, column=21).fill = total_fill
    for col in range(1, len(CADUCADOS_LAYOUT_HEADERS) + 1):
        ws.cell(row=total_row, column=col).border = border

    for col in range(1, len(CADUCADOS_LAYOUT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_caducados.xlsx"'
    wb.save(response)
    return response
