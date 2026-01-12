"""
Reporte de Lotes sin Fecha de Caducidad Válida

Muestra todos los lotes que no tienen fecha de caducidad o tienen una fecha inválida.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Lote, Institucion, Almacen


@login_required
def reporte_sin_caducidad(request):
    """
    Reporte de lotes sin fecha de caducidad válida.
    Muestra lotes donde:
    - No tienen fecha de caducidad (NULL)
    - Tienen fecha de caducidad inválida (en el pasado o formato incorrecto)
    """
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        'producto',
        'institucion'
    ).all()
    
    # Registros sin caducidad válida
    registros_sin_caducidad = []
    
    for lote in lotes:
        es_invalido = False
        razon_invalido = ""
        
        # Verificar si no tiene fecha de caducidad
        if not lote.fecha_caducidad:
            es_invalido = True
            razon_invalido = "Sin fecha de caducidad"
        else:
            # Verificar si la fecha es válida (no en el pasado)
            fecha_hoy = timezone.now().date()
            if lote.fecha_caducidad < fecha_hoy:
                es_invalido = True
                razon_invalido = f"Fecha vencida ({lote.fecha_caducidad.strftime('%d/%m/%Y')})"
        
        if es_invalido:
            # Obtener ubicaciones del lote
            ubicaciones = lote.ubicaciones_detalle.select_related(
                'ubicacion',
                'ubicacion__almacen'
            )
            
            if ubicaciones.exists():
                # Si tiene ubicaciones, mostrar cada una
                for ub in ubicaciones:
                    registros_sin_caducidad.append({
                        'tipo': 'LoteUbicacion',
                        'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                        'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                        'lote': lote.numero_lote,
                        'institucion': lote.institucion.denominacion if lote.institucion else '-',
                        'almacen': ub.ubicacion.almacen.nombre if ub.ubicacion and ub.ubicacion.almacen else '-',
                        'ubicacion': ub.ubicacion.codigo if ub.ubicacion else '-',
                        'cantidad': ub.cantidad,
                        'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'SIN FECHA',
                        'razon': razon_invalido,
                        'fecha_fabricacion': lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else '-',
                        'lote_id': lote.id,
                        'ubicacion_id': ub.id,
                    })
            else:
                # Si no tiene ubicaciones, mostrar el lote sin ubicación
                registros_sin_caducidad.append({
                    'tipo': 'Lote',
                    'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                    'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                    'lote': lote.numero_lote,
                    'institucion': lote.institucion.denominacion if lote.institucion else '-',
                    'almacen': '-',
                    'ubicacion': 'Sin ubicación',
                    'cantidad': lote.cantidad_disponible,
                    'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'SIN FECHA',
                    'razon': razon_invalido,
                    'fecha_fabricacion': lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else '-',
                    'lote_id': lote.id,
                    'ubicacion_id': None,
                })
    
    # Filtros opcionales
    filtro_institucion = request.GET.get('institucion', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_razon = request.GET.get('razon', '')  # 'sin_fecha' o 'vencida'
    
    # Aplicar filtros
    if filtro_institucion:
        registros_sin_caducidad = [r for r in registros_sin_caducidad if r['institucion'] == filtro_institucion]
    
    if filtro_almacen:
        registros_sin_caducidad = [r for r in registros_sin_caducidad if r['almacen'] == filtro_almacen]
    
    if filtro_clave:
        registros_sin_caducidad = [r for r in registros_sin_caducidad if filtro_clave.lower() in r['clave_cnis'].lower()]
    
    if filtro_razon == 'sin_fecha':
        registros_sin_caducidad = [r for r in registros_sin_caducidad if 'Sin fecha' in r['razon']]
    elif filtro_razon == 'vencida':
        registros_sin_caducidad = [r for r in registros_sin_caducidad if 'vencida' in r['razon'].lower()]
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(registros_sin_caducidad, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener opciones de filtro
    instituciones = Institucion.objects.all().order_by('denominacion')
    almacenes = Almacen.objects.all().order_by('nombre')
    
    # Estadísticas
    total_sin_fecha = len([r for r in registros_sin_caducidad if 'Sin fecha' in r['razon']])
    total_vencida = len([r for r in registros_sin_caducidad if 'vencida' in r['razon'].lower()])
    
    context = {
        'page_obj': page_obj,
        'total_registros': len(registros_sin_caducidad),
        'total_sin_fecha': total_sin_fecha,
        'total_vencida': total_vencida,
        'instituciones': instituciones,
        'almacenes': almacenes,
        'filtro_institucion': filtro_institucion,
        'filtro_almacen': filtro_almacen,
        'filtro_clave': filtro_clave,
        'filtro_razon': filtro_razon,
    }
    
    return render(request, 'inventario/reporte_sin_caducidad.html', context)


@login_required
def exportar_sin_caducidad_excel(request):
    """
    Exporta el reporte de lotes sin fecha de caducidad válida a Excel.
    """
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        'producto',
        'institucion'
    ).all()
    
    # Registros sin caducidad válida
    registros_sin_caducidad = []
    
    for lote in lotes:
        es_invalido = False
        razon_invalido = ""
        
        # Verificar si no tiene fecha de caducidad
        if not lote.fecha_caducidad:
            es_invalido = True
            razon_invalido = "Sin fecha de caducidad"
        else:
            # Verificar si la fecha es válida (no en el pasado)
            fecha_hoy = timezone.now().date()
            if lote.fecha_caducidad < fecha_hoy:
                es_invalido = True
                razon_invalido = f"Fecha vencida ({lote.fecha_caducidad.strftime('%d/%m/%Y')})"
        
        if es_invalido:
            # Obtener ubicaciones del lote
            ubicaciones = lote.ubicaciones_detalle.select_related(
                'ubicacion',
                'ubicacion__almacen'
            )
            
            if ubicaciones.exists():
                # Si tiene ubicaciones, mostrar cada una
                for ub in ubicaciones:
                    registros_sin_caducidad.append({
                        'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                        'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                        'lote': lote.numero_lote,
                        'institucion': lote.institucion.denominacion if lote.institucion else '-',
                        'almacen': ub.ubicacion.almacen.nombre if ub.ubicacion and ub.ubicacion.almacen else '-',
                        'ubicacion': ub.ubicacion.codigo if ub.ubicacion else '-',
                        'cantidad': ub.cantidad,
                        'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'SIN FECHA',
                        'fecha_fabricacion': lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else '-',
                        'razon': razon_invalido,
                    })
            else:
                # Si no tiene ubicaciones, mostrar el lote sin ubicación
                registros_sin_caducidad.append({
                    'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                    'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                    'lote': lote.numero_lote,
                    'institucion': lote.institucion.denominacion if lote.institucion else '-',
                    'almacen': '-',
                    'ubicacion': 'Sin ubicación',
                    'cantidad': lote.cantidad_disponible,
                    'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'SIN FECHA',
                    'fecha_fabricacion': lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else '-',
                    'razon': razon_invalido,
                })
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sin Caducidad"
    
    # Estilos
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'CLAVE CNIS',
        'DESCRIPCIÓN PRODUCTO',
        'LOTE',
        'INSTITUCIÓN',
        'ALMACÉN',
        'UBICACIÓN',
        'CANTIDAD',
        'FECHA CADUCIDAD',
        'FECHA FABRICACIÓN',
        'RAZÓN'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, registro in enumerate(registros_sin_caducidad, 2):
        ws.cell(row=row, column=1).value = registro['clave_cnis']
        ws.cell(row=row, column=2).value = registro['descripcion_producto']
        ws.cell(row=row, column=3).value = registro['lote']
        ws.cell(row=row, column=4).value = registro['institucion']
        ws.cell(row=row, column=5).value = registro['almacen']
        ws.cell(row=row, column=6).value = registro['ubicacion']
        ws.cell(row=row, column=7).value = registro['cantidad']
        ws.cell(row=row, column=8).value = registro['fecha_caducidad']
        ws.cell(row=row, column=9).value = registro['fecha_fabricacion']
        ws.cell(row=row, column=10).value = registro['razon']
        
        # Aplicar bordes
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 30
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_sin_caducidad.xlsx"'
    wb.save(response)
    
    return response
