"""
Reporte Kardex / Libro mayor: movimientos por lote con saldo acumulado.
"""

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .kardex_utils import construir_kardex_desde_request
from .models import MovimientoInventario


@login_required
def reporte_kardex(request):
    kardexes = []
    error_msg = None
    filtros = {
        'clave': request.GET.get('clave', '') or request.GET.get('busqueda_clave', ''),
        'lote': request.GET.get('lote', '') or request.GET.get('busqueda_lote', ''),
        'clues': request.GET.get('clues', ''),
        'almacen': request.GET.get('almacen', ''),
        'producto': request.GET.get('producto', '') or request.GET.get('busqueda_producto', ''),
        'fecha_desde': request.GET.get('fecha_desde', ''),
        'fecha_hasta': request.GET.get('fecha_hasta', ''),
        'tipo': request.GET.get('tipo', ''),
        'incluir_anulados': request.GET.get('incluir_anulados', '') == 'si',
    }

    if any(
        [
            filtros['clave'],
            filtros['lote'],
            filtros['fecha_desde'],
            filtros['fecha_hasta'],
        ]
    ):
        kardexes, error_msg, filtros = construir_kardex_desde_request(request)
        if error_msg:
            messages.warning(request, error_msg)

    context = {
        'kardexes': kardexes,
        'filtros': filtros,
        'tipos_movimiento': MovimientoInventario.TIPOS_MOVIMIENTO,
        'tiene_resultados': bool(kardexes),
        'query_export': request.GET.urlencode(),
    }
    return render(request, 'inventario/reportes/reporte_kardex.html', context)


@login_required
def exportar_kardex_excel(request):
    kardexes, error_msg, _ = construir_kardex_desde_request(request, max_lotes=50)
    if error_msg or not kardexes:
        messages.warning(request, error_msg or 'Sin datos para exportar.')
        return redirect('reportes:reporte_kardex')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kardex'
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    row_num = 1

    headers = [
        'Clave CNIS',
        'Lote',
        'CLUES',
        'Almacén',
        'Fecha',
        'Tipo',
        'Documento',
        'Folio',
        'Entrada (+)',
        'Salida (−)',
        'Saldo',
        'Motivo',
        'Usuario',
        'Anulado',
    ]

    for k in kardexes:
        ws.cell(row=row_num, column=1, value=f"LOTE: {k['numero_lote']} | {k['clave_cnis']} | {k['clues']}")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        ws.cell(
            row=row_num,
            column=1,
            value=f"Saldo inicial: {k['saldo_inicial']} | Saldo final: {k['saldo_final']} | Existencia actual: {k['existencia_actual']}",
        )
        row_num += 1

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
        row_num += 1

        if k['saldo_inicial'] and not k['movimientos']:
            ws.cell(row=row_num, column=8, value=k['saldo_inicial'])
            ws.cell(row=row_num, column=11, value=k['saldo_inicial'])
            row_num += 1

        for f in k['movimientos']:
            ws.cell(row=row_num, column=1, value=k['clave_cnis'])
            ws.cell(row=row_num, column=2, value=k['numero_lote'])
            ws.cell(row=row_num, column=3, value=k['clues'])
            ws.cell(row=row_num, column=4, value=k['almacen'])
            ws.cell(
                row=row_num,
                column=5,
                value=f['fecha'].strftime('%d/%m/%Y %H:%M') if f['fecha'] else '',
            )
            ws.cell(row=row_num, column=6, value=f['tipo_display'])
            ws.cell(row=row_num, column=7, value=f['documento'])
            ws.cell(row=row_num, column=8, value=f['folio'] or f['pedido'])
            ws.cell(row=row_num, column=9, value=f['entrada'] or '')
            ws.cell(row=row_num, column=10, value=f['salida'] or '')
            ws.cell(row=row_num, column=11, value=f['saldo'])
            ws.cell(row=row_num, column=12, value=f['motivo'][:500])
            ws.cell(row=row_num, column=13, value=f['usuario'])
            ws.cell(row=row_num, column=14, value='Sí' if f['anulado'] else '')
            row_num += 1

        row_num += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="kardex_inventario.xlsx"'
    wb.save(response)
    return response
