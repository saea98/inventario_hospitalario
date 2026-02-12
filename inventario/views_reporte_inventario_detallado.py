"""
Vista para reporte detallado de inventario con las columnas:
ENTIDAD, CLUES, ORDEN DE SUMINISTRO, RFC, CLAVE, ESTADO DEL INSUMO,
INVENTARIO DISPONIBLE, LOTE, F_CAD, F_FAB, F_REC
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from datetime import date, datetime

from .models import Lote, Producto, Institucion, OrdenSuministro, Proveedor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required
def reporte_inventario_detallado(request):
    """
    Reporte detallado de inventario con información de lotes, instituciones y órdenes de suministro.
    """
    
    # Obtener parámetros de filtro
    filtro_entidad = request.GET.get('entidad', '').strip()
    filtro_clues = request.GET.get('clues', '').strip()
    filtro_orden = request.GET.get('orden', '').strip()
    filtro_rfc = request.GET.get('rfc', '').strip()
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_estado = request.GET.get('estado', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    excluir_sin_orden = request.GET.get('excluir_sin_orden', '') == 'si'

    # Query base con todas las relaciones necesarias
    lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'orden_suministro__proveedor'
    ).all()

    # Excluir lotes sin orden de suministro (opción del usuario)
    if excluir_sin_orden:
        lotes = lotes.exclude(orden_suministro__isnull=True)

    # Aplicar filtros
    if filtro_entidad:
        lotes = lotes.filter(institucion__denominacion__icontains=filtro_entidad)
    
    if filtro_clues:
        lotes = lotes.filter(institucion__clue__icontains=filtro_clues)
    
    if filtro_orden:
        lotes = lotes.filter(orden_suministro__numero_orden__icontains=filtro_orden)
    
    if filtro_rfc:
        lotes = lotes.filter(
            Q(orden_suministro__proveedor__rfc__icontains=filtro_rfc) |
            Q(rfc_proveedor__icontains=filtro_rfc)
        )

    if filtro_clave:
        lotes = lotes.filter(producto__clave_cnis__icontains=filtro_clave)
    
    if filtro_estado:
        lotes = lotes.filter(estado=filtro_estado)
    
    if filtro_lote:
        lotes = lotes.filter(numero_lote__icontains=filtro_lote)

    # Ordenación por clic en encabezado (sort=columna&order=asc|desc)
    sort_column = request.GET.get('sort', '').strip()
    sort_order = request.GET.get('order', 'asc').strip().lower()
    if sort_order not in ('asc', 'desc'):
        sort_order = 'asc'
    # Mapeo: parámetro GET -> campo(s) order_by (prefijo - para desc)
    sort_map = {
        'entidad': 'institucion__denominacion',
        'clues': 'institucion__clue',
        'orden': 'orden_suministro__numero_orden',
        'rfc': 'orden_suministro__proveedor__rfc',
        'clave': 'producto__clave_cnis',
        'descripcion': 'producto__descripcion',
        'estado': 'estado',
        'inventario': 'cantidad_disponible',
        'lote': 'numero_lote',
        'f_cad': 'fecha_caducidad',
        'f_fab': 'fecha_fabricacion',
        'f_rec': 'fecha_recepcion',
    }
    if sort_column and sort_column in sort_map:
        order_field = sort_map[sort_column]
        if sort_order == 'desc':
            order_field = f'-{order_field}'
        lotes = lotes.order_by(order_field)
    else:
        # Orden por defecto
        lotes = lotes.order_by('institucion__denominacion', '-fecha_recepcion', 'producto__clave_cnis')

    # Paginación
    paginator = Paginator(lotes, 50)  # 50 items por página
    page = request.GET.get('page', 1)
    
    try:
        lotes_paginados = paginator.page(page)
    except PageNotAnInteger:
        lotes_paginados = paginator.page(1)
    except EmptyPage:
        lotes_paginados = paginator.page(paginator.num_pages)
    
    # Preparar datos para el template (f_cad = fecha caducidad, f_rec = fecha recepción; campos distintos del modelo)
    datos_reporte = []
    for lote in lotes_paginados:
        # Estado del insumo: leyenda (get_estado_display), no el dígito
        estado_texto = lote.get_estado_display() if lote.estado is not None else ''
        if not estado_texto and lote.estado is not None:
            estado_texto = str(lote.estado)

        # Fechas: usar explícitamente el campo correcto para cada columna
        fecha_caducidad_str = lote.fecha_caducidad.strftime('%d/%m/%Y') if getattr(lote, 'fecha_caducidad', None) else ''
        fecha_fabricacion_str = lote.fecha_fabricacion.strftime('%d/%m/%Y') if getattr(lote, 'fecha_fabricacion', None) else ''
        fecha_recepcion_str = lote.fecha_recepcion.strftime('%d/%m/%Y') if getattr(lote, 'fecha_recepcion', None) else ''

        # RFC: prioridad orden/proveedor; si no hay, usar el guardado en el lote (carga masiva)
        rfc_display = ''
        if lote.orden_suministro and lote.orden_suministro.proveedor:
            rfc_display = (lote.orden_suministro.proveedor.rfc or '').strip()
        if not rfc_display:
            rfc_display = (getattr(lote, 'rfc_proveedor', None) or '').strip()

        # Descripción: lote.descripcion_saica o producto.descripcion
        descripcion_val = (getattr(lote, 'descripcion_saica', None) or '').strip()
        if not descripcion_val and lote.producto:
            descripcion_val = (getattr(lote.producto, 'descripcion', None) or '').strip()
        descripcion_val = descripcion_val or ''

        datos_reporte.append({
            'entidad': 'CIUDAD DE MÉXICO',  # Leyenda fija (no almacén/institucion)
            'clues': lote.institucion.clue if lote.institucion else '',
            'orden_suministro': lote.orden_suministro.numero_orden if lote.orden_suministro else '',
            'rfc': rfc_display,
            'clave': lote.producto.clave_cnis if lote.producto else '',
            'descripcion': descripcion_val,
            'estado_insumo': estado_texto,
            'inventario_disponible': lote.cantidad_disponible,
            'lote': lote.numero_lote,
            'f_cad': fecha_caducidad_str,
            'f_fab': fecha_fabricacion_str,
            'f_rec': fecha_recepcion_str,
        })
    
    # Query string base para enlaces de ordenación (conserva filtros, quita sort/order/page)
    get_copy = request.GET.copy()
    get_copy.pop('sort', None)
    get_copy.pop('order', None)
    get_copy.pop('page', None)
    sort_base_query = get_copy.urlencode()

    # Obtener listas para filtros
    instituciones = Institucion.objects.filter(activo=True).order_by('denominacion')[:100]

    # Columnas disponibles para exportar a Excel (selector en modal)
    columnas_disponibles = [
        {'value': 'entidad', 'label': 'ENTIDAD'},
        {'value': 'clues', 'label': 'CLUES'},
        {'value': 'orden_suministro', 'label': 'ORDEN DE SUMINISTRO'},
        {'value': 'rfc', 'label': 'RFC'},
        {'value': 'clave', 'label': 'CLAVE'},
        {'value': 'descripcion', 'label': 'DESCRIPCIÓN'},
        {'value': 'estado_insumo', 'label': 'ESTADO DEL INSUMO'},
        {'value': 'inventario_disponible', 'label': 'INVENTARIO DISPONIBLE'},
        {'value': 'lote', 'label': 'LOTE'},
        {'value': 'f_cad', 'label': 'F_CAD'},
        {'value': 'f_fab', 'label': 'F_FAB'},
        {'value': 'f_rec', 'label': 'F_REC'},
    ]

    context = {
        'datos_reporte': datos_reporte,
        'lotes_paginados': lotes_paginados,
        'filtro_entidad': filtro_entidad,
        'filtro_clues': filtro_clues,
        'filtro_orden': filtro_orden,
        'filtro_rfc': filtro_rfc,
        'filtro_clave': filtro_clave,
        'filtro_estado': filtro_estado,
        'filtro_lote': filtro_lote,
        'excluir_sin_orden': excluir_sin_orden,
        'sort_column': sort_column,
        'sort_order': sort_order,
        'sort_base_query': sort_base_query,
        'columnas_disponibles': columnas_disponibles,
        'estados_lote': [
            (1, 'Disponible'),
            (4, 'Suspendido'),
            (5, 'Deteriorado'),
            (6, 'Caducado'),
        ],
    }
    
    return render(request, 'inventario/reportes/reporte_inventario_detallado.html', context)


# Mapeo columna id -> etiqueta para Excel
COLUMNAS_EXCEL_LABELS = {
    'entidad': 'ENTIDAD',
    'clues': 'CLUES',
    'orden_suministro': 'ORDEN DE SUMINISTRO',
    'rfc': 'RFC',
    'clave': 'CLAVE',
    'descripcion': 'DESCRIPCIÓN',
    'estado_insumo': 'ESTADO DEL INSUMO',
    'inventario_disponible': 'INVENTARIO DISPONIBLE',
    'lote': 'LOTE',
    'f_cad': 'F_CAD',
    'f_fab': 'F_FAB',
    'f_rec': 'F_REC',
}


def _obtener_lotes_filtrados(request, from_post=False):
    """Aplica filtros desde GET o POST y devuelve el queryset de lotes."""
    get_param = request.POST.get if from_post else request.GET.get
    filtro_entidad = (get_param('entidad') or '').strip()
    filtro_clues = (get_param('clues') or '').strip()
    filtro_orden = (get_param('orden') or '').strip()
    filtro_rfc = (get_param('rfc') or '').strip()
    filtro_clave = (get_param('clave') or '').strip()
    filtro_estado = (get_param('estado') or '').strip()
    filtro_lote = (get_param('lote') or '').strip()
    excluir_sin_orden = (get_param('excluir_sin_orden') or '') == 'si'
    filtro_almacen = (get_param('almacen') or '').strip()
    filtro_proveedor = (get_param('proveedor') or '').strip()
    filtro_marca = (get_param('marca') or '').strip()
    filtro_fabricante = (get_param('fabricante') or '').strip()
    fecha_rec_desde_str = (get_param('fecha_rec_desde') or '').strip()
    fecha_rec_hasta_str = (get_param('fecha_rec_hasta') or '').strip()
    cad_desde_str = (get_param('cad_desde') or '').strip()
    cad_hasta_str = (get_param('cad_hasta') or '').strip()

    lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'orden_suministro__proveedor',
        'almacen',
        'ubicacion',
    ).all()

    if excluir_sin_orden:
        lotes = lotes.exclude(orden_suministro__isnull=True)
    if filtro_entidad:
        lotes = lotes.filter(institucion__denominacion__icontains=filtro_entidad)
    if filtro_clues:
        lotes = lotes.filter(institucion__clue__icontains=filtro_clues)
    if filtro_orden:
        lotes = lotes.filter(orden_suministro__numero_orden__icontains=filtro_orden)
    if filtro_rfc:
        lotes = lotes.filter(
            Q(orden_suministro__proveedor__rfc__icontains=filtro_rfc) |
            Q(rfc_proveedor__icontains=filtro_rfc)
        )
    if filtro_clave:
        lotes = lotes.filter(producto__clave_cnis__icontains=filtro_clave)
    if filtro_estado:
        lotes = lotes.filter(estado=filtro_estado)
    if filtro_lote:
        lotes = lotes.filter(numero_lote__icontains=filtro_lote)

    if filtro_almacen:
        lotes = lotes.filter(almacen__nombre__icontains=filtro_almacen)
    if filtro_proveedor:
        lotes = lotes.filter(
            Q(orden_suministro__proveedor__razon_social__icontains=filtro_proveedor)
            | Q(proveedor__icontains=filtro_proveedor)
        )
    if filtro_marca:
        lotes = lotes.filter(producto__marca__icontains=filtro_marca)
    if filtro_fabricante:
        lotes = lotes.filter(producto__fabricante__icontains=filtro_fabricante)

    # Filtros de rango de fechas (coherentes) para recepción y caducidad
    def _parse_fecha(fecha_str):
        if not fecha_str:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(fecha_str, fmt).date()
            except ValueError:
                continue
        return None

    fecha_rec_desde = _parse_fecha(fecha_rec_desde_str)
    fecha_rec_hasta = _parse_fecha(fecha_rec_hasta_str)
    cad_desde = _parse_fecha(cad_desde_str)
    cad_hasta = _parse_fecha(cad_hasta_str)

    if fecha_rec_desde:
        lotes = lotes.filter(fecha_recepcion__gte=fecha_rec_desde)
    if fecha_rec_hasta:
        lotes = lotes.filter(fecha_recepcion__lte=fecha_rec_hasta)
    if cad_desde:
        lotes = lotes.filter(fecha_caducidad__gte=cad_desde)
    if cad_hasta:
        lotes = lotes.filter(fecha_caducidad__lte=cad_hasta)

    return lotes.order_by('institucion__denominacion', '-fecha_recepcion', 'producto__clave_cnis')


def _fila_lote_a_dict(lote):
    """Convierte un lote a diccionario con todas las columnas del reporte."""
    estado_texto = lote.get_estado_display() if lote.estado is not None else ''
    if not estado_texto and lote.estado is not None:
        estado_texto = str(lote.estado)
    rfc_val = ''
    if lote.orden_suministro and lote.orden_suministro.proveedor:
        rfc_val = (lote.orden_suministro.proveedor.rfc or '').strip()
    if not rfc_val:
        rfc_val = (getattr(lote, 'rfc_proveedor', None) or '').strip()
    descripcion_val = (getattr(lote, 'descripcion_saica', None) or '').strip() or (lote.producto.descripcion if lote.producto else '') or ''
    return {
        'entidad': 'CIUDAD DE MÉXICO',
        'clues': lote.institucion.clue if lote.institucion else '',
        'orden_suministro': lote.orden_suministro.numero_orden if lote.orden_suministro else '',
        'rfc': rfc_val,
        'clave': lote.producto.clave_cnis if lote.producto else '',
        'descripcion': descripcion_val,
        'estado_insumo': estado_texto,
        'inventario_disponible': lote.cantidad_disponible,
        'lote': lote.numero_lote,
        'f_cad': lote.fecha_caducidad.strftime('%d/%m/%Y') if getattr(lote, 'fecha_caducidad', None) else '',
        'f_fab': lote.fecha_fabricacion.strftime('%d/%m/%Y') if getattr(lote, 'fecha_fabricacion', None) else '',
        'f_rec': lote.fecha_recepcion.strftime('%d/%m/%Y') if getattr(lote, 'fecha_recepcion', None) else '',
    }


# =========================
# REPORTE DE EXISTENCIAS
# =========================

EXISTENCIAS_COLUMNAS_LABELS = {
    'lote': 'Lote',
    'clave_cnis': 'Clave CNIS',
    'producto': 'Producto',
    'unidad_medida': 'Unidad de medida',
    'almacen': 'Almacén',
    'ubicacion': 'Ubicación',
    'existencia': 'Existencia',
    'precio': 'Precio',
    'importe': 'Importe',
    'contrato': 'Contrato',
    'orden_suministro': 'Orden de Suministro',
    'remision': 'Remisión',
    'proveedor': 'Proveedor',
    'marca': 'Marca',
    'fabricante': 'Fabricante',
    'clues': 'CLUES',
    'institucion': 'Institución',
    'f_cad': 'Fecha Caducidad',
    'f_fab': 'Fecha Fabricación',
    'f_rec': 'Fecha Recepción',
}


def _fila_lote_existencias_a_dict(lote):
    """Convierte un lote a diccionario con columnas para el reporte de existencias."""
    producto = getattr(lote, 'producto', None)
    institucion = getattr(lote, 'institucion', None)
    orden = getattr(lote, 'orden_suministro', None)
    proveedor_os = getattr(orden, 'proveedor', None) if orden else None

    proveedor_nombre = ''
    if proveedor_os and getattr(proveedor_os, 'razon_social', None):
        proveedor_nombre = proveedor_os.razon_social
    elif getattr(lote, 'proveedor', None):
        proveedor_nombre = lote.proveedor

    return {
        'lote': lote.numero_lote,
        'clave_cnis': producto.clave_cnis if producto else '',
        'producto': producto.descripcion if producto else '',
        'unidad_medida': productounidad if (productounidad := getattr(producto, 'unidad_medida', None)) else '',
        'almacen': lote.almacen.nombre if getattr(lote, 'almacen', None) else '',
        'ubicacion': lote.ubicacion.codigo if getattr(lote, 'ubicacion', None) else '',
        'existencia': lote.cantidad_disponible,
        'precio': lote.precio_unitario,
        'importe': lote.valor_total,
        'contrato': getattr(lote, 'contrato', '') or '',
        'orden_suministro': orden.numero_orden if orden else '',
        'remision': getattr(lote, 'remision', '') or '',
        'proveedor': proveedor_nombre,
        'marca': getattr(producto, 'marca', '') or '' if producto else '',
        'fabricante': getattr(producto, 'fabricante', '') or '' if producto else '',
        'clues': institucion.clue if institucion else '',
        'institucion': institucion.denominacion if institucion else '',
        'f_cad': lote.fecha_caducidad.strftime('%d/%m/%Y') if getattr(lote, 'fecha_caducidad', None) else '',
        'f_fab': lote.fecha_fabricacion.strftime('%d/%m/%Y') if getattr(lote, 'fecha_fabricacion', None) else '',
        'f_rec': lote.fecha_recepcion.strftime('%d/%m/%Y') if getattr(lote, 'fecha_recepcion', None) else '',
    }


@login_required
def reporte_existencias(request):
    """
    Reporte de existencias (por lote) con filtros por CLUES, almacén, proveedor,
    marca, fechas de recepción/caducidad, etc.
    """
    lotes = _obtener_lotes_filtrados(request, from_post=False)

    paginator = Paginator(lotes, 50)
    page = request.GET.get('page', 1)
    try:
        lotes_paginados = paginator.page(page)
    except PageNotAnInteger:
        lotes_paginados = paginator.page(1)
    except EmptyPage:
        lotes_paginados = paginator.page(paginator.num_pages)

    datos_reporte = [_fila_lote_existencias_a_dict(l) for l in lotes_paginados]

    # Para mantener filtros en paginación y orden, reutilizamos los parámetros de GET
    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    base_query = get_copy.urlencode()

    context = {
        'datos_reporte': datos_reporte,
        'lotes_paginados': lotes_paginados,
        'base_query': base_query,
        # Filtros actuales (para rellenar el formulario)
        'filtro_entidad': request.GET.get('entidad', '').strip(),
        'filtro_clues': request.GET.get('clues', '').strip(),
        'filtro_orden': request.GET.get('orden', '').strip(),
        'filtro_rfc': request.GET.get('rfc', '').strip(),
        'filtro_clave': request.GET.get('clave', '').strip(),
        'filtro_estado': request.GET.get('estado', '').strip(),
        'filtro_lote': request.GET.get('lote', '').strip(),
        'excluir_sin_orden': (request.GET.get('excluir_sin_orden') or '') == 'si',
        'filtro_almacen': request.GET.get('almacen', '').strip(),
        'filtro_proveedor': request.GET.get('proveedor', '').strip(),
        'filtro_marca': request.GET.get('marca', '').strip(),
        'filtro_fabricante': request.GET.get('fabricante', '').strip(),
        'fecha_rec_desde': request.GET.get('fecha_rec_desde', '').strip(),
        'fecha_rec_hasta': request.GET.get('fecha_rec_hasta', '').strip(),
        'cad_desde': request.GET.get('cad_desde', '').strip(),
        'cad_hasta': request.GET.get('cad_hasta', '').strip(),
        'estados_lote': [
            (1, 'Disponible'),
            (4, 'Suspendido'),
            (5, 'Deteriorado'),
            (6, 'Caducado'),
        ],
        'columnas_existencias': EXISTENCIAS_COLUMNAS_LABELS,
    }

    return render(request, 'inventario/reportes/reporte_existencias.html', context)


@login_required
def exportar_existencias_excel(request):
    """Exporta el reporte de existencias a Excel con las columnas definidas."""
    lotes = _obtener_lotes_filtrados(request, from_post=False)

    columnas = list(EXISTENCIAS_COLUMNAS_LABELS.keys())
    headers = [EXISTENCIAS_COLUMNAS_LABELS[c] for c in columnas]

    wb = Workbook()
    ws = wb.active
    ws.title = "Existencias"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row_num, lote in enumerate(lotes, 2):
        fila = _fila_lote_existencias_a_dict(lote)
        row_data = [fila.get(c, '') for c in columnas]

        for col_num, value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{row_num}"]
            cell.value = value
            cell.border = border
            if columnas[col_num - 1] in ('existencia', 'precio', 'importe'):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_num in range(1, len(columnas) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_existencias_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_inventario_detallado_excel(request):
    """
    Exporta el reporte de inventario detallado a Excel.
    GET: exporta con todas las columnas (incluye DESCRIPCIÓN).
    POST: exporta solo las columnas seleccionadas (columnas, orden_columnas y filtros en POST).
    """
    from_post = request.method == 'POST'
    lotes = _obtener_lotes_filtrados(request, from_post=from_post)

    # Columnas a exportar: desde POST (selector) o por defecto todas
    if from_post:
        columnas = request.POST.getlist('columnas')
        orden_columnas = (request.POST.get('orden_columnas') or '').strip()
        if orden_columnas:
            orden = [c for c in orden_columnas.split(',') if c in columnas]
            if orden:
                columnas = orden
        if not columnas:
            columnas = list(COLUMNAS_EXCEL_LABELS.keys())
    else:
        columnas = list(COLUMNAS_EXCEL_LABELS.keys())

    headers = [COLUMNAS_EXCEL_LABELS.get(c, c.upper()) for c in columnas]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Detallado"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row_num, lote in enumerate(lotes, 2):
        fila = _fila_lote_a_dict(lote)
        row_data = [fila.get(c, '') for c in columnas]

        for col_num, value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{row_num}"]
            cell.value = value
            cell.border = border
            if columnas[col_num - 1] == 'rfc':
                cell.number_format = '@'
            if columnas[col_num - 1] == 'inventario_disponible':
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_num in range(1, len(columnas) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18 if columnas[col_num - 1] != 'descripcion' else 40

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_inventario_detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


# --- Carga masiva desde Excel (mismo layout que el reporte) ---

ESTADO_INSUMO_MAP = {
    'disponible': 1,
    'suspendido': 4,
    'deteriorado': 5,
    'caducado': 6,
}

HEADERS_INVENTARIO_DETALLADO = [
    'ENTIDAD', 'CLUES', 'ORDEN DE SUMINISTRO', 'RFC', 'CLAVE', 'ESTADO DEL INSUMO',
    'INVENTARIO DISPONIBLE', 'LOTE', 'F_CAD', 'F_FAB', 'F_REC',
]


def _parse_fecha_celda(val):
    """Convierte celda Excel a date. Acepta str 'dd/mm/yyyy' o 'dd/mm/yyyy HH:MM' o objeto date/datetime."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _estado_to_int(estado_texto):
    if not estado_texto:
        return None
    key = str(estado_texto).strip().lower()
    return ESTADO_INSUMO_MAP.get(key)


@login_required
def carga_masiva_inventario_detallado(request):
    """
    Carga masiva: sube un Excel con el layout del reporte inventario detallado
    (ENTIDAD, CLUES, ORDEN DE SUMINISTRO, RFC, CLAVE, ESTADO DEL INSUMO,
     INVENTARIO DISPONIBLE, LOTE, F_CAD, F_FAB, F_REC) y actualiza los lotes
     existentes (identificados por CLUES + CLAVE + LOTE).
    """
    if request.method == 'GET':
        return render(request, 'inventario/reportes/carga_masiva_inventario_detallado.html', {})

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debe seleccionar un archivo Excel.')
        return redirect('reportes:carga_masiva_inventario_detallado')

    if not archivo.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'El archivo debe ser Excel (.xlsx).')
        return redirect('reportes:carga_masiva_inventario_detallado')

    dry_run = request.POST.get('dry_run') == '1'
    no_sobrescribir = request.POST.get('no_sobrescribir') == '1'

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f'No se pudo leer el Excel: {str(e)}')
        return redirect('reportes:carga_masiva_inventario_detallado')

    # Primera fila = encabezados
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        messages.error(request, 'El archivo no tiene encabezados.')
        wb.close()
        return redirect('reportes:carga_masiva_inventario_detallado')

    header_row = [str(c).strip() if c is not None else '' for c in header_row]
    col_index = {h: i for i, h in enumerate(header_row) if h}

    # Índices esperados (nombres exactos o muy parecidos)
    def col(name, alternatives=None):
        for n in [name] + (alternatives or []):
            if n in col_index:
                return col_index[n]
            for k in col_index:
                if n.lower() in k.lower() or k.lower() in n.lower():
                    return col_index[k]
        return None

    idx_entidad = col('ENTIDAD')
    idx_clues = col('CLUES')
    idx_orden = col('ORDEN DE SUMINISTRO')
    idx_rfc = col('RFC')
    idx_clave = col('CLAVE')
    idx_estado = col('ESTADO DEL INSUMO')
    idx_inventario = col('INVENTARIO DISPONIBLE')
    idx_lote = col('LOTE')
    idx_f_cad = col('F_CAD')
    idx_f_fab = col('F_FAB')
    idx_f_rec = col('F_REC')

    if idx_clues is None or idx_clave is None or idx_lote is None:
        messages.error(request, 'Faltan columnas obligatorias: CLUES, CLAVE, LOTE.')
        wb.close()
        return redirect('reportes:carga_masiva_inventario_detallado')

    actualizados = 0
    no_encontrados = []
    errores = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row) if row else []
        max_idx = max((i for i in [idx_clues, idx_clave, idx_lote] if i is not None), default=0)
        if len(row) <= max_idx:
            continue
        clues = str(row[idx_clues]).strip() if idx_clues is not None and len(row) > idx_clues and row[idx_clues] else ''
        clave = str(row[idx_clave]).strip() if idx_clave is not None and len(row) > idx_clave and row[idx_clave] else ''
        numero_lote = str(row[idx_lote]).strip() if idx_lote is not None and len(row) > idx_lote and row[idx_lote] else ''
        if not clues or not clave or not numero_lote:
            continue

        try:
            institucion = Institucion.objects.filter(clue=clues).first()
            producto = Producto.objects.filter(clave_cnis=clave).first()
            if not institucion:
                no_encontrados.append(f"Fila {row_num}: CLUES '{clues}' no encontrado")
                continue
            if not producto:
                no_encontrados.append(f"Fila {row_num}: CLAVE '{clave}' no encontrado")
                continue

            lote = Lote.objects.filter(
                institucion=institucion,
                producto=producto,
                numero_lote=numero_lote,
            ).first()
            if not lote:
                no_encontrados.append(f"Fila {row_num}: Lote '{numero_lote}' / CLAVE {clave} / CLUES {clues} no existe")
                continue

            # Helper: solo asignar si no_sobrescribir está desactivado, o si el campo actual está "vacío"
            def _valor_actual_vacio(model_obj, attr, consider_cero_vacio=True):
                val = getattr(model_obj, attr, None)
                if val is None:
                    return True
                if isinstance(val, str) and not val.strip():
                    return True
                if consider_cero_vacio and isinstance(val, (int, float)) and val == 0:
                    return True
                return False

            # Cantidad y estado: respetan no_sobrescribir (solo completar vacíos)
            if idx_inventario is not None and len(row) > idx_inventario and row[idx_inventario] is not None:
                if not no_sobrescribir or _valor_actual_vacio(lote, 'cantidad_disponible'):
                    try:
                        val = int(float(row[idx_inventario]))
                        if val >= 0:
                            lote.cantidad_disponible = val
                    except (ValueError, TypeError):
                        pass
            if idx_estado is not None and len(row) > idx_estado:
                estado_int = _estado_to_int(row[idx_estado])
                if estado_int is not None and (not no_sobrescribir or _valor_actual_vacio(lote, 'estado')):
                    lote.estado = estado_int

            # Orden, RFC y fechas: actualizar desde Excel cuando el archivo trae valor
            # RFC: no sobrescribir si el lote ya tiene dato
            if idx_rfc is not None and len(row) > idx_rfc and row[idx_rfc] is not None:
                rfc_val = str(row[idx_rfc]).strip()
                if rfc_val:
                    actual = (getattr(lote, 'rfc_proveedor', None) or '').strip()
                    if not actual:
                        lote.rfc_proveedor = rfc_val[:50]
            if idx_f_cad is not None and len(row) > idx_f_cad:
                f = _parse_fecha_celda(row[idx_f_cad])
                if f is not None:
                    lote.fecha_caducidad = f
            if idx_f_fab is not None and len(row) > idx_f_fab:
                f = _parse_fecha_celda(row[idx_f_fab])
                if f is not None:
                    lote.fecha_fabricacion = f
            if idx_f_rec is not None and len(row) > idx_f_rec:
                f = _parse_fecha_celda(row[idx_f_rec])
                if f is not None:
                    lote.fecha_recepcion = f

            # Orden de suministro: vincular existente o crear desde archivo (RFC + orden) si no existe
            if idx_orden is not None and len(row) > idx_orden and row[idx_orden]:
                orden_num = str(row[idx_orden]).strip()
                if orden_num:
                    orden = None
                    numero_orden_trunc = orden_num[:200]
                    # 1) Coincidencia exacta
                    orden = OrdenSuministro.objects.filter(numero_orden=numero_orden_trunc).first()
                    if not orden and len(orden_num) != len(numero_orden_trunc):
                        orden = OrdenSuministro.objects.filter(numero_orden=orden_num).first()
                    if not orden and lote.rfc_proveedor:
                        proveedor = Proveedor.objects.filter(rfc=lote.rfc_proveedor.strip()[:20]).first()
                        if proveedor:
                            orden = OrdenSuministro.objects.filter(
                                numero_orden=numero_orden_trunc,
                                proveedor=proveedor,
                            ).first()
                    # 2) Búsqueda flexible por primer token o inicio del texto
                    if not orden:
                        primer_token = (orden_num.split()[0] if orden_num.split() else orden_num)[:50]
                        if primer_token:
                            orden = OrdenSuministro.objects.filter(numero_orden__icontains=primer_token).first()
                    if not orden and len(orden_num) > 10:
                        for n in (30, 50, 80, 120):
                            if n <= len(orden_num):
                                orden = OrdenSuministro.objects.filter(numero_orden=orden_num[:n]).first()
                                if orden:
                                    break
                    # 3) No existe la orden: crearla con RFC y orden del archivo (partida 000, etc.)
                    if not orden and not dry_run:
                        proveedor = None
                        rfc_clean = (getattr(lote, 'rfc_proveedor', None) or '').strip()[:20]
                        if rfc_clean:
                            proveedor, _ = Proveedor.objects.get_or_create(
                                rfc=rfc_clean,
                                defaults={'razon_social': f'Proveedor carga masiva ({rfc_clean})'}
                            )
                        fecha_orden = lote.fecha_recepcion if getattr(lote, 'fecha_recepcion', None) else date.today()
                        orden, _ = OrdenSuministro.objects.get_or_create(
                            numero_orden=numero_orden_trunc,
                            defaults={
                                'proveedor': proveedor,
                                'partida_presupuestal': '000',
                                'fecha_orden': fecha_orden,
                            }
                        )
                    if orden:
                        lote.orden_suministro = orden

            if not dry_run:
                lote.save()
            actualizados += 1
        except Exception as e:
            errores.append(f"Fila {row_num}: {str(e)}")

    wb.close()

    if dry_run and actualizados:
        messages.info(request, f'[Dry-run] Se habrían actualizado {actualizados} lote(s). No se guardó nada.')
    elif actualizados:
        messages.success(request, f'Se actualizaron {actualizados} lote(s).')
    if no_encontrados:
        messages.warning(request, f'{len(no_encontrados)} registro(s) no encontrados (CLUES/CLAVE/Lote no existen). Ver detalle abajo.')
    if errores:
        messages.error(request, f'{len(errores)} error(es) al procesar filas.')

    context = {
        'actualizados': actualizados,
        'no_encontrados': no_encontrados[:100],
        'errores': errores[:100],
        'total_no_encontrados': len(no_encontrados),
        'total_errores': len(errores),
        'dry_run': dry_run,
        'no_sobrescribir': no_sobrescribir,
    }
    return render(request, 'inventario/reportes/carga_masiva_inventario_detallado.html', context)
