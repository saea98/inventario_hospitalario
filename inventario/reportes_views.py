from django.shortcuts import render
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from .llegada_models import ItemLlegada
from .reportes_forms import ReporteEntradasForm

class ReporteEntradasView(LoginRequiredMixin, View):
    def get(self, request):
        form = ReporteEntradasForm(request.GET)
        items = ItemLlegada.objects.select_related('llegada', 'producto', 'llegada__proveedor').all()

        if form.is_valid():
            if form.cleaned_data.get('fecha_inicio'):
                items = items.filter(llegada__fecha_llegada_real__gte=form.cleaned_data['fecha_inicio'])
            if form.cleaned_data.get('fecha_fin'):
                items = items.filter(llegada__fecha_llegada_real__lte=form.cleaned_data['fecha_fin'])
            if form.cleaned_data.get('proveedor'):
                items = items.filter(llegada__proveedor=form.cleaned_data['proveedor'])
            if form.cleaned_data.get('clave'):
                items = items.filter(producto__clave_cnis__icontains=form.cleaned_data['clave'])
            if form.cleaned_data.get('lote'):
                items = items.filter(numero_lote__icontains=form.cleaned_data['lote'])

        return render(request, 'inventario/reportes/reporte_entradas.html', {'form': form, 'items': items})


from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from weasyprint import HTML

class ReporteEntradasView(LoginRequiredMixin, View):
    def get(self, request):
        form = ReporteEntradasForm(request.GET)
        items = ItemLlegada.objects.select_related("llegada", "producto", "llegada__proveedor").all()

        if form.is_valid():
            if form.cleaned_data.get("fecha_inicio"):
                items = items.filter(llegada__fecha_llegada_real__gte=form.cleaned_data["fecha_inicio"])
            if form.cleaned_data.get("fecha_fin"):
                items = items.filter(llegada__fecha_llegada_real__lte=form.cleaned_data["fecha_fin"])
            if form.cleaned_data.get("proveedor"):
                items = items.filter(llegada__proveedor=form.cleaned_data["proveedor"])
            if form.cleaned_data.get("clave"):
                items = items.filter(producto__clave_cnis__icontains=form.cleaned_data["clave"])
            if form.cleaned_data.get("lote"):
                items = items.filter(numero_lote__icontains=form.cleaned_data["lote"])

        if "export" in request.GET:
            if request.GET["export"] == "excel":
                return self.export_to_excel(items)
            elif request.GET["export"] == "pdf":
                return self.export_to_pdf(items)

        return render(request, "inventario/reportes/reporte_entradas.html", {"form": form, "items": items})

    def export_to_excel(self, items):
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=reporte_entradas.xlsx"

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reporte de Entradas"

        headers = ["Fecha de Entrada", "Proveedor", "Clave", "Descripción", "Lote", "Fecha de Caducidad", "Cantidad Recibida"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}1"]
            cell.value = header

        for row_num, item in enumerate(items, 2):
            worksheet[f"A{row_num}"] = item.llegada.fecha_llegada_real.strftime("%d/%m/%Y")
            worksheet[f"B{row_num}"] = item.llegada.proveedor.nombre
            worksheet[f"C{row_num}"] = item.producto.clave_cnis
            worksheet[f"D{row_num}"] = item.producto.descripcion
            worksheet[f"E{row_num}"] = item.numero_lote
            worksheet[f"F{row_num}"] = item.fecha_caducidad.strftime("%d/%m/%Y")
            worksheet[f"G{row_num}"] = item.cantidad_recibida

        workbook.save(response)
        return response

    def export_to_pdf(self, items):
        html_string = render_to_string("inventario/reportes/reporte_entradas_pdf.html", {"items": items})
        html = HTML(string=html_string)
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = "attachment; filename=reporte_entradas.pdf"
        html.write_pdf(response)
        return response
