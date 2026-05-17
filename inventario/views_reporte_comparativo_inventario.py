"""
Reporte comparativo de inventario entre dos fechas.

Compara existencias por clave (u otra agrupación), muestra las mayores diferencias
y los movimientos del periodo para auditar variaciones (p. ej. ~700k entre semanas).
"""

from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .comparativo_inventario_utils import (
    agregar_diferencias_por_grupo,
    enriquecer_filas_con_movimientos,
    movimientos_en_periodo,
    resumen_movimientos_por_clave,
    totales_globales,
)
from .views_reporte_inventario_detallado import _obtener_lotes_filtrados


def _parse_fecha(s):
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


@login_required
def reporte_comparativo_inventario(request):
    fecha_a_str = request.GET.get('fecha_a', '').strip()
    fecha_b_str = request.GET.get('fecha_b', '').strip()
    agrupacion = request.GET.get('agrupacion', 'clave').strip() or 'clave'
    if agrupacion not in ('clave', 'clave_clues', 'clave_clues_almacen', 'lote'):
        agrupacion = 'clave'
    top_n = request.GET.get('top', '100').strip()
    try:
        top_n = max(10, min(500, int(top_n)))
    except ValueError:
        top_n = 100

    clave_detalle = request.GET.get('clave_detalle', '').strip()
    metrica = request.GET.get('metrica', 'fisica').strip() or 'fisica'
    if metrica not in ('fisica', 'disponible'):
        metrica = 'fisica'
    fecha_a = _parse_fecha(fecha_a_str)
    fecha_b = _parse_fecha(fecha_b_str)

    filas = []
    movimientos = []
    totales = {}
    usar_neto_b = False
    tiene_fechas = False

    if fecha_a and fecha_b:
        if fecha_a > fecha_b:
            fecha_a, fecha_b = fecha_b, fecha_a
            fecha_a_str, fecha_b_str = fecha_b_str, fecha_a_str
        tiene_fechas = True
        lotes_qs = _obtener_lotes_filtrados(request)

        filas, usar_neto_b = agregar_diferencias_por_grupo(
            lotes_qs,
            fecha_a,
            fecha_b,
            agrupacion=agrupacion,
            top_n=top_n,
            metrica=metrica,
        )
        claves_top = {f['clave_cnis'] for f in filas if f.get('clave_cnis')}
        mov_por_clave = resumen_movimientos_por_clave(
            lotes_qs, fecha_a, fecha_b, claves=claves_top
        )
        enriquecer_filas_con_movimientos(filas, mov_por_clave)
        totales = totales_globales(
            lotes_qs, fecha_a, fecha_b, usar_neto_b, metrica=metrica
        )
        totales['claves_en_top'] = len(filas)
        totales['claves_con_diff'] = sum(1 for f in filas if f['delta'] != 0)

        if clave_detalle:
            movimientos = movimientos_en_periodo(
                lotes_qs, fecha_a, fecha_b, clave_cnis=clave_detalle, limite=300
            )
    elif fecha_a_str or fecha_b_str:
        messages.warning(
            request,
            'Indique ambas fechas en formato válido (AAAA-MM-DD).',
        )

    paginator = Paginator(filas, 50)
    page_num = request.GET.get('page', 1)
    try:
        filas_pagina = paginator.page(page_num)
    except PageNotAnInteger:
        filas_pagina = paginator.page(1)
    except EmptyPage:
        filas_pagina = paginator.page(paginator.num_pages)

    context = {
        'fecha_a': fecha_a_str,
        'fecha_b': fecha_b_str,
        'agrupacion': agrupacion,
        'top_n': top_n,
        'clave_detalle': clave_detalle,
        'filas': filas_pagina,
        'filas_todas_count': len(filas),
        'movimientos': movimientos,
        'totales': totales,
        'usar_neto_b': usar_neto_b,
        'tiene_fechas': tiene_fechas,
        'filtro_entidad': request.GET.get('entidad', '').strip(),
        'filtro_clues': request.GET.get('clues', '').strip(),
        'filtro_almacen': request.GET.get('almacen', '').strip(),
        'filtro_clave': request.GET.get('clave', '').strip(),
        'filtro_lote': request.GET.get('lote', '').strip(),
        'filtro_orden': request.GET.get('orden', '').strip(),
        'filtro_rfc': request.GET.get('rfc', '').strip(),
        'filtro_proveedor': request.GET.get('proveedor', '').strip(),
        'excluir_sin_orden': request.GET.get('excluir_sin_orden', '') == 'si',
        'metrica': metrica,
        'query_export': request.GET.urlencode(),
    }
    return render(
        request,
        'inventario/reportes/reporte_comparativo_inventario.html',
        context,
    )


@login_required
def exportar_comparativo_inventario_excel(request):
    fecha_a = _parse_fecha(request.GET.get('fecha_a', ''))
    fecha_b = _parse_fecha(request.GET.get('fecha_b', ''))
    if not fecha_a or not fecha_b:
        messages.warning(request, 'Debe indicar fecha inicial y fecha final.')
        return redirect('reportes:reporte_comparativo_inventario')

    if fecha_a > fecha_b:
        fecha_a, fecha_b = fecha_b, fecha_a

    agrupacion = request.GET.get('agrupacion', 'clave').strip() or 'clave'
    metrica = request.GET.get('metrica', 'fisica').strip() or 'fisica'
    lotes_qs = _obtener_lotes_filtrados(request)
    filas, usar_neto_b = agregar_diferencias_por_grupo(
        lotes_qs,
        fecha_a,
        fecha_b,
        agrupacion=agrupacion,
        top_n=500,
        metrica=metrica,
    )
    mov_por_clave = resumen_movimientos_por_clave(
        lotes_qs, fecha_a, fecha_b, claves={f['clave_cnis'] for f in filas}
    )
    enriquecer_filas_con_movimientos(filas, mov_por_clave)
    totales = totales_globales(
        lotes_qs, fecha_a, fecha_b, usar_neto_b, metrica=metrica
    )

    wb = Workbook()
    ws_diff = wb.active
    ws_diff.title = 'Diferencias'
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    label_b = 'Inventario B (neto hoy)' if usar_neto_b else 'Inventario B (físico)'
    headers_diff = [
        'Grupo',
        'Clave CNIS',
        'CLUES',
        'Almacén',
        'Lote',
        f'Inventario A ({fecha_a})',
        label_b,
        'Delta B−A',
        'Entradas periodo',
        'Salidas periodo',
        'Neto movimientos',
        'Delta − neto mov.',
    ]
    for col, h in enumerate(headers_diff, 1):
        c = ws_diff.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_num, f in enumerate(filas, 2):
        ws_diff.cell(row=row_num, column=1, value=f['grupo'])
        ws_diff.cell(row=row_num, column=2, value=f['clave_cnis'])
        ws_diff.cell(row=row_num, column=3, value=f['clues'])
        ws_diff.cell(row=row_num, column=4, value=f['almacen'])
        ws_diff.cell(row=row_num, column=5, value=f['numero_lote'])
        ws_diff.cell(row=row_num, column=6, value=f['total_a'])
        ws_diff.cell(row=row_num, column=7, value=f['total_b'])
        ws_diff.cell(row=row_num, column=8, value=f['delta'])
        ws_diff.cell(row=row_num, column=9, value=f['mov_entradas'])
        ws_diff.cell(row=row_num, column=10, value=f['mov_salidas'])
        ws_diff.cell(row=row_num, column=11, value=f['mov_neto'])
        ws_diff.cell(row=row_num, column=12, value=f['diff_vs_movimientos'])

    tr = len(filas) + 2
    ws_diff.cell(row=tr, column=1, value='TOTALES (universo filtrado)')
    ws_diff.cell(row=tr, column=6, value=totales['total_a'])
    ws_diff.cell(row=tr, column=7, value=totales['total_b'])
    ws_diff.cell(row=tr, column=8, value=totales['delta'])

    ws_mov = wb.create_sheet('Movimientos')
    headers_mov = [
        'Fecha',
        'Clave CNIS',
        'Lote',
        'CLUES',
        'Almacén',
        'Tipo',
        'Cantidad',
        'Efecto',
        'Antes',
        'Después',
        'Motivo',
        'Documento',
        'Folio',
        'Usuario',
    ]
    for col, h in enumerate(headers_mov, 1):
        c = ws_mov.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    claves_con_diff = [f['clave_cnis'] for f in filas if f['delta'] != 0][:50]
    row_m = 2
    for clave in claves_con_diff:
        for m in movimientos_en_periodo(
            lotes_qs, fecha_a, fecha_b, clave_cnis=clave, limite=200
        ):
            ws_mov.cell(row=row_m, column=1, value=m['fecha'].strftime('%d/%m/%Y %H:%M'))
            ws_mov.cell(row=row_m, column=2, value=m['clave_cnis'])
            ws_mov.cell(row=row_m, column=3, value=m['numero_lote'])
            ws_mov.cell(row=row_m, column=4, value=m['clues'])
            ws_mov.cell(row=row_m, column=5, value=m['almacen'])
            ws_mov.cell(row=row_m, column=6, value=m['tipo'])
            ws_mov.cell(row=row_m, column=7, value=m['cantidad'])
            ws_mov.cell(row=row_m, column=8, value=m['efecto'])
            ws_mov.cell(row=row_m, column=9, value=m['cantidad_anterior'])
            ws_mov.cell(row=row_m, column=10, value=m['cantidad_nueva'])
            ws_mov.cell(row=row_m, column=11, value=m['motivo'])
            ws_mov.cell(row=row_m, column=12, value=m['documento'])
            ws_mov.cell(row=row_m, column=13, value=m['folio'])
            ws_mov.cell(row=row_m, column=14, value=m['usuario'])
            row_m += 1

    for ws in (ws_diff, ws_mov):
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="comparativo_inventario_{fecha_a}_{fecha_b}.xlsx"'
    )
    wb.save(response)
    return response
