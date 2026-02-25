"""
Fase 6 - Optimización de Picking para Suministro
Vistas para ordenar y visualizar propuestas de forma optimizada
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, F
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.template.loader import get_template
from django.conf import settings
from datetime import datetime, timedelta
import re
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os

from .pedidos_models import PropuestaPedido, ItemPropuesta, LoteAsignado
from .models import Lote, LoteUbicacion, UbicacionAlmacen, Almacen, MovimientoInventario
from django.urls import reverse
from .decorators_roles import requiere_rol
from .fase5_utils import generar_movimientos_suministro
from .excel_to_pdf_converter import convertir_excel_a_pdf
from .propuesta_utils import reservar_cantidad_lote, liberar_cantidad_lote
from django.db import transaction


def _natural_sort_key_codigo(codigo):
    """
    Genera una clave de ordenamiento natural para códigos de ubicación
    (ej. J6A.01.02, J6A.01.03, J6A.01.10) para que ordenen correctamente.
    """
    if not codigo:
        return (0, 0)
    parts = re.split(r'(\d+)', str(codigo))
    result = []
    for part in parts:
        if not part:
            continue
        if part.isdigit():
            result.append((1, int(part)))
        else:
            result.append((0, part))
    return tuple(result)


# ============================================================
# DASHBOARD DE PICKING
# ============================================================

@requiere_rol('Almacenista', 'Administrador', 'Gestor de Inventario')
def dashboard_picking(request):
    """
    Dashboard de picking - Muestra propuestas disponibles para picking
    Optimizado para tablet
    """
    
    # Obtener propuestas en estado REVISADA o EN_SURTIMIENTO
    propuestas = PropuestaPedido.objects.filter(
        estado__in=['REVISADA', 'EN_SURTIMIENTO']
    ).select_related('solicitud').order_by('-fecha_generacion')
    
    # Filtros
    almacen_filter = request.GET.get('almacen')
    if almacen_filter:
        propuestas = propuestas.filter(solicitud__almacen_destino_id=almacen_filter)
    
    estado_filter = request.GET.get('estado')
    if estado_filter:
        propuestas = propuestas.filter(estado=estado_filter)
    
    # Búsqueda
    busqueda = request.GET.get('q')
    if busqueda:
        propuestas = propuestas.filter(
            Q(solicitud__folio__icontains=busqueda) |
            Q(solicitud__institucion_solicitante__nombre__icontains=busqueda)
        )
    
    # Contar items por propuesta
    propuestas_con_items = []
    for prop in propuestas:
        total_items = prop.items.count()
        propuestas_con_items.append({
            'propuesta': prop,
            'total_items': total_items,
            'items_pendientes': prop.items.filter(estado__in=['DISPONIBLE', 'PARCIAL']).count()
        })
    
    # Obtener almacenes para filtro
    almacenes = Almacen.objects.all()
    
    context = {
        'propuestas': propuestas_con_items,
        'almacenes': almacenes,
        'almacen_filter': almacen_filter,
        'estado_filter': estado_filter,
        'busqueda': busqueda,
    }
    
    return render(request, 'inventario/picking/dashboard_picking.html', context)


# ============================================================
# MONITOR DE PICKING (electrónico vs manual)
# ============================================================

@login_required
@requiere_rol('Almacenista', 'Administrador', 'Gestor de Inventario', 'Supervisor')
def monitor_picking(request):
    """
    Monitor para ver qué usuarios están usando picking electrónico (tablet/web)
    y cuáles surten de forma manual (Excel/PDF + botón Surtir propuesta).
    - Electrónico: LoteAsignado con usuario_surtido no null (marcaron "Recoger" en la app).
    - Manual: Propuestas surtidas donde ningún ítem tiene usuario_surtido (surtida vía Surtir propuesta).
    """
    ahora = timezone.now()
    ultimas_24h = ahora - timedelta(hours=24)
    ultimas_2h = ahora - timedelta(hours=2)

    # --- Picking electrónico: detalle de cada ítem recogido (folio pedido = observaciones, insumos, ubicaciones)
    recogidas_electronicas = (
        LoteAsignado.objects
        .filter(surtido=True, usuario_surtido__isnull=False, fecha_surtimiento__gte=ultimas_24h)
        .select_related(
            'usuario_surtido',
            'item_propuesta__propuesta__solicitud',
            'lote_ubicacion__ubicacion__almacen',
            'lote_ubicacion__lote__producto',
        )
        .order_by('-fecha_surtimiento')
    )

    # Lista detallada: folio pedido (observaciones), insumos con ubicaciones
    detalle_electronicos = []
    usuarios_electronicos = {}
    for la in recogidas_electronicas:
        solicitud = la.item_propuesta.propuesta.solicitud if la.item_propuesta and la.item_propuesta.propuesta else None
        folio_pedido = (solicitud.observaciones_solicitud or '').strip() if solicitud else ''
        folio_solicitud = solicitud.folio if solicitud else '—'
        lu = la.lote_ubicacion
        producto = lu.lote.producto if lu and lu.lote else None
        insumo = producto.descripcion if producto else '—'
        clave_cnis = producto.clave_cnis if producto else '—'
        ubicacion = lu.ubicacion.codigo if lu and lu.ubicacion else '—'
        almacen = lu.ubicacion.almacen.nombre if lu and lu.ubicacion and lu.ubicacion.almacen else '—'

        detalle_electronicos.append({
            'usuario': la.usuario_surtido,
            'folio_pedido': folio_pedido or '—',
            'folio_solicitud': folio_solicitud,
            'insumo': insumo,
            'clave_cnis': clave_cnis,
            'ubicacion': ubicacion,
            'almacen': almacen,
            'cantidad': la.cantidad_asignada,
            'fecha_surtimiento': la.fecha_surtimiento,
        })

        u = la.usuario_surtido
        if u not in usuarios_electronicos:
            usuarios_electronicos[u] = {'total_items': 0, 'ultima_actividad': la.fecha_surtimiento}
        usuarios_electronicos[u]['total_items'] += 1
        usuarios_electronicos[u]['ultima_actividad'] = max(
            usuarios_electronicos[u]['ultima_actividad'] or la.fecha_surtimiento,
            la.fecha_surtimiento
        )

    lista_electronicos = [
        {
            'usuario': u,
            'total_items': datos['total_items'],
            'ultima_actividad': datos['ultima_actividad'],
            'activo_reciente': (datos['ultima_actividad'] or ahora) >= ultimas_2h,
        }
        for u, datos in usuarios_electronicos.items()
    ]
    lista_electronicos.sort(key=lambda x: (not x['activo_reciente'], -(x['ultima_actividad'] or ahora).timestamp()))

    # --- Picking manual: propuestas SURTIDA donde ningún ítem fue recogido por usuario (surtida vía "Surtir propuesta")
    # Propuestas que tienen al menos un LoteAsignado con usuario_surtido no null = tuvieron picking electrónico
    ids_con_electronico = (
        LoteAsignado.objects
        .filter(usuario_surtido__isnull=False, item_propuesta__propuesta__estado='SURTIDA')
        .values_list('item_propuesta__propuesta_id', flat=True)
        .distinct()
    )
    propuestas_manuales = (
        PropuestaPedido.objects
        .filter(estado='SURTIDA', fecha_surtimiento__gte=ultimas_24h, fecha_surtimiento__isnull=False)
        .exclude(id__in=ids_con_electronico)
        .select_related('solicitud', 'usuario_surtimiento')
        .order_by('-fecha_surtimiento')[:100]
    )

    context = {
        'lista_electronicos': lista_electronicos,
        'detalle_electronicos': detalle_electronicos,
        'propuestas_manuales': propuestas_manuales,
        'ultimas_24h': ultimas_24h,
        'ultimas_2h': ultimas_2h,
    }
    return render(request, 'inventario/picking/monitor_picking.html', context)


# ============================================================
# REDIRECCIÓN: picking/propuesta/<uuid>/ -> logistica/propuestas/<uuid>/picking/
# ============================================================

def redirect_picking_propuesta_a_logistica(request, propuesta_id):
    """Redirige la URL antigua picking/propuesta/<uuid>/ a logistica/propuestas/<uuid>/picking/."""
    url = reverse('logistica:picking_propuesta', kwargs={'propuesta_id': propuesta_id})
    return redirect(url)


# ============================================================
# VISTA DE PICKING OPTIMIZADA
# ============================================================

@requiere_rol('Almacenista', 'Administrador', 'Gestor de Inventario')
def picking_propuesta(request, propuesta_id):
    """
    Vista optimizada de picking para tablet/pantalla
    Muestra los items ordenados por ubicación
    """
    
    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id)
    
    # Validar que la propuesta esté en estado correcto
    if propuesta.estado not in ['REVISADA', 'EN_SURTIMIENTO']:
        messages.error(request, 'La propuesta no está lista para picking.')
        return redirect('logistica:detalle_propuesta', propuesta_id=propuesta_id)
    
    # Obtener orden de picking
    orden_picking = request.GET.get('orden', 'ubicacion')  # ubicacion, producto, cantidad
    
    # Obtener items de la propuesta con lotes asignados
    items_picking = []
    
    # Verificar si el usuario es administrador o supervisor
    es_administrador = request.user.is_staff or request.user.is_superuser
    es_supervisor = request.user.groups.filter(name='Supervisor').exists()
    puede_ver_todo = es_administrador or es_supervisor
    
    # Obtener almacén del usuario si no es administrador ni supervisor
    almacen_usuario_id = None
    if not puede_ver_todo and hasattr(request.user, 'almacen') and request.user.almacen:
        almacen_usuario_id = request.user.almacen.id
    
    for item in propuesta.items.all():
        # defer('usuario_surtido') por si la columna no existe aún en la BD
        for lote_asignado in item.lotes_asignados.filter(surtido=False).defer('usuario_surtido'):
            lote_ubicacion = lote_asignado.lote_ubicacion
            lote = lote_ubicacion.lote
            
            # Filtrar por almacén del usuario si no es administrador ni supervisor
            if not puede_ver_todo and almacen_usuario_id:
                if lote_ubicacion.ubicacion.almacen_id != almacen_usuario_id:
                    continue  # Saltar este item si no pertenece al almacén del usuario
            
            caducidad = lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A'
            items_picking.append({
                'item_id': item.id,
                'lote_asignado_id': lote_asignado.id,
                'producto': lote.producto.descripcion,
                'cantidad': lote_asignado.cantidad_asignada,
                'lote_numero': lote.numero_lote,
                'almacen': lote_ubicacion.ubicacion.almacen.nombre,
                'almacen_id': lote_ubicacion.ubicacion.almacen_id,
                'ubicacion': lote_ubicacion.ubicacion.codigo,
                'ubicacion_id': lote_ubicacion.ubicacion_id,
                'clave_cnis': lote.producto.clave_cnis,
                'caducidad': caducidad,
            })
    
    # Ordenar según parámetro
    if orden_picking == 'producto':
        items_picking.sort(key=lambda x: x['producto'])
    elif orden_picking == 'cantidad':
        items_picking.sort(key=lambda x: x['cantidad'], reverse=True)
    else:  # ubicacion (predeterminado): orden natural por código (J6A.01.02, .03, .10) y dentro por producto
        items_picking.sort(key=lambda x: (
            x['almacen_id'],
            _natural_sort_key_codigo(x['ubicacion']),
            (x['producto'] or '').lower(),
        ))
    
    # Agrupar por ubicación para vista (el dict conserva orden de inserción; los ítems dentro ya están ordenados)
    ubicaciones_agrupadas = {}
    for item in items_picking:
        key = f"{item['almacen']} - {item['ubicacion']}"
        if key not in ubicaciones_agrupadas:
            ubicaciones_agrupadas[key] = []
        ubicaciones_agrupadas[key].append(item)
    
    context = {
        'propuesta': propuesta,
        'items_picking': items_picking,
        'ubicaciones_agrupadas': ubicaciones_agrupadas,
        'orden_picking': orden_picking,
        'total_items': len(items_picking),
    }
    
    return render(request, 'inventario/picking/picking_propuesta.html', context)



# ============================================================
# MARCAR ITEM COMO RECOGIDO (AJAX)
# ============================================================

@login_required
@csrf_exempt
@require_http_methods(['POST'])
def marcar_item_recogido(request, lote_asignado_id):
    """
    Marca un item como recogido en la vista de picking
    Verifica si todos los items están recogidos y completa la propuesta
    """
    
    try:
        # defer por si la columna usuario_surtido_id no existe aún en la BD
        lote_asignado = LoteAsignado.objects.defer('usuario_surtido').get(id=lote_asignado_id)
        propuesta = lote_asignado.item_propuesta.propuesta
        
        # Evitar que otro usuario marque como recogido lo que ya recogió alguien
        if lote_asignado.surtido:
            return JsonResponse({
                'exito': False,
                'mensaje': 'Este ítem ya fue recogido por otro usuario. Actualice la página.',
                'ya_recogido': True,
            }, status=400)
        
        lote_asignado.surtido = True
        lote_asignado.fecha_surtimiento = timezone.now()
        try:
            lote_asignado.usuario_surtido = request.user
            lote_asignado.save()
        except Exception as e:
            from django.db.utils import ProgrammingError, OperationalError
            if (isinstance(e, (ProgrammingError, OperationalError)) and
                    'usuario_surtido' in str(e)):
                lote_asignado.save(update_fields=['surtido', 'fecha_surtimiento'])
            else:
                raise
        
        # Verificar si todos los items de la propuesta están recogidos
        total_lotes = LoteAsignado.objects.filter(
            item_propuesta__propuesta=propuesta
        ).count()
        
        lotes_recogidos = LoteAsignado.objects.filter(
            item_propuesta__propuesta=propuesta,
            surtido=True
        ).count()
        
        propuesta_completada = False
        
        if total_lotes == lotes_recogidos and total_lotes > 0:
            # Todos los items han sido recogidos
            # Cambiar a REVISADA (no a SURTIDA) para que el supervisor confirme
            propuesta.estado = 'REVISADA'
            propuesta.fecha_revision = timezone.now()
            propuesta.usuario_revision = request.user
            propuesta.save()
            
            propuesta_completada = True
            resultado = {'exito': True, 'mensaje': 'Propuesta lista para surtir'}
            
            return JsonResponse({
                'exito': resultado['exito'],
                'mensaje': f"Item marcado. {resultado['mensaje']}",
                'propuesta_completada': propuesta_completada,
                'lotes_recogidos': lotes_recogidos,
                'total_lotes': total_lotes,
                'movimientos_creados': resultado.get('movimientos_creados', 0)
            })
        
        return JsonResponse({
            'exito': True,
            'mensaje': 'Item marcado como recogido',
            'propuesta_completada': propuesta_completada,
            'lotes_recogidos': lotes_recogidos,
            'total_lotes': total_lotes
        })
    
    except LoteAsignado.DoesNotExist:
        return JsonResponse({
            'exito': False,
            'mensaje': 'Item no encontrado'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }, status=500)


# ============================================================
# CORRECCIÓN DE LOTE EN PROPUESTA DE SURTIMIENTO (sin perder avance)
# ============================================================

@login_required
@require_http_methods(['GET'])
def ubicaciones_para_corregir_lote(request):
    """
    Devuelve LoteUbicacion disponibles para el mismo producto y almacén que el ítem asignado.
    Se usa en el modal de "Corregir lote" en la vista de picking.
    """
    lote_asignado_id = request.GET.get('lote_asignado_id')
    if not lote_asignado_id:
        return JsonResponse({'error': 'lote_asignado_id es requerido'}, status=400)
    try:
        la = LoteAsignado.objects.select_related(
            'item_propuesta__producto',
            'lote_ubicacion__lote',
            'lote_ubicacion__ubicacion__almacen',
        ).get(id=lote_asignado_id)
    except LoteAsignado.DoesNotExist:
        return JsonResponse({'error': 'Asignación no encontrada'}, status=404)
    if la.surtido:
        return JsonResponse({'error': 'No se puede corregir un ítem ya recogido'}, status=400)
    producto = la.item_propuesta.producto
    almacen_id = None
    if la.lote_ubicacion and la.lote_ubicacion.ubicacion:
        almacen_id = la.lote_ubicacion.ubicacion.almacen_id
    if not almacen_id and la.lote_ubicacion and la.lote_ubicacion.lote and la.lote_ubicacion.lote.almacen_id:
        almacen_id = la.lote_ubicacion.lote.almacen_id
    if not almacen_id:
        return JsonResponse({'error': 'No se pudo determinar el almacén del ítem'}, status=400)
    cantidad_necesaria = la.cantidad_asignada
    lote_ubicacion_actual_id = la.lote_ubicacion_id
    # LoteUbicacion del mismo producto, mismo almacén, con stock suficiente (excluir el actual)
    from django.db.models import Q
    qs = LoteUbicacion.objects.filter(
        lote__producto=producto,
        lote__estado=1,
    ).filter(
        Q(ubicacion__almacen_id=almacen_id) | Q(lote__almacen_id=almacen_id)
    ).exclude(
        id=lote_ubicacion_actual_id
    ).select_related('lote', 'ubicacion__almacen').order_by('lote__fecha_caducidad')
    opciones = []
    for lu in qs:
        disp = lu.cantidad - lu.cantidad_reservada
        if disp < cantidad_necesaria:
            continue
        opciones.append({
            'id': str(lu.id),
            'lote': lu.lote.numero_lote,
            'codigo': lu.ubicacion.codigo if lu.ubicacion else '',
            'cantidad_disponible': disp,
            'fecha_caducidad': lu.lote.fecha_caducidad.strftime('%d/%m/%Y') if lu.lote.fecha_caducidad else '',
        })
    return JsonResponse({'opciones': opciones})


@login_required
@csrf_exempt
@require_http_methods(['POST'])
def corregir_lote_propuesta(request, lote_asignado_id):
    """
    Corrige el lote asignado en la propuesta de surtimiento: libera la reserva del lote erróneo,
    reserva en el lote correcto, actualiza LoteAsignado y crea movimientos de inventario
    tipo AJUSTE_DATOS_LOTE (igual que el resto de movimientos).
    Solo se permite si el ítem no ha sido marcado como recogido (surtido=False).
    """
    try:
        la = LoteAsignado.objects.select_related(
            'item_propuesta__propuesta__solicitud',
            'lote_ubicacion__lote',
            'lote_ubicacion__ubicacion__almacen',
        ).get(id=lote_asignado_id)
    except LoteAsignado.DoesNotExist:
        return JsonResponse({'exito': False, 'mensaje': 'Asignación no encontrada'}, status=404)
    if la.surtido:
        return JsonResponse({'exito': False, 'mensaje': 'No se puede corregir un ítem ya recogido'}, status=400)
    import json
    try:
        body = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        body = {}
    nuevo_lu_id = body.get('nuevo_lote_ubicacion_id') or request.POST.get('nuevo_lote_ubicacion_id')
    if not nuevo_lu_id:
        return JsonResponse({'exito': False, 'mensaje': 'Falta nuevo_lote_ubicacion_id'}, status=400)
    try:
        nueva_lu = LoteUbicacion.objects.select_related('lote', 'ubicacion__almacen').get(id=nuevo_lu_id)
    except LoteUbicacion.DoesNotExist:
        return JsonResponse({'exito': False, 'mensaje': 'Ubicación/lote destino no encontrada'}, status=404)
    if nueva_lu.id == la.lote_ubicacion_id:
        return JsonResponse({'exito': False, 'mensaje': 'Debe elegir un lote distinto al actual'}, status=400)
    if nueva_lu.lote.producto_id != la.item_propuesta.producto_id:
        return JsonResponse({'exito': False, 'mensaje': 'El lote debe ser del mismo producto'}, status=400)
    disp_nueva = nueva_lu.cantidad - nueva_lu.cantidad_reservada
    if disp_nueva < la.cantidad_asignada:
        return JsonResponse({
            'exito': False,
            'mensaje': f'No hay cantidad suficiente en el lote elegido (disponible: {disp_nueva}, necesario: {la.cantidad_asignada})'
        }, status=400)
    vieja_lu = la.lote_ubicacion
    viejo_lote = vieja_lu.lote
    nuevo_lote = nueva_lu.lote
    cantidad = la.cantidad_asignada
    solicitud = getattr(la.item_propuesta.propuesta, 'solicitud', None)
    folio_solicitud = ''
    if solicitud:
        folio_solicitud = (getattr(solicitud, 'observaciones_solicitud', None) or '').strip() or (getattr(solicitud, 'folio', None) or '')
    try:
        with transaction.atomic():
            resultado_liberacion = liberar_cantidad_lote(vieja_lu, cantidad)
            reserva_antigua_lote = resultado_liberacion.get('reserva_anterior_lote', 0)
            reserva_nueva_lote = resultado_liberacion.get('nueva_reservada_lote', 0)
            cantidad_real_liberada = resultado_liberacion.get('liberado_lote', cantidad)
            motivo_liberacion = (
                f"Cambio de lote en orden de picking - Liberación de reserva de {cantidad_real_liberada} unidades del lote {viejo_lote.numero_lote} "
                f"(Producto: {viejo_lote.producto.clave_cnis or 'N/A'}). "
                f"Folio pedido: {folio_solicitud}. Reserva anterior: {reserva_antigua_lote}, nueva: {reserva_nueva_lote}."
            )
            MovimientoInventario.objects.create(
                lote=viejo_lote,
                tipo_movimiento='AJUSTE_POSITIVO',
                cantidad=cantidad_real_liberada,
                cantidad_anterior=reserva_antigua_lote,
                cantidad_nueva=reserva_nueva_lote,
                motivo=motivo_liberacion,
                documento_referencia=f"PROP-SURT-{la.item_propuesta.propuesta_id}",
                folio=folio_solicitud or None,
                usuario=request.user,
            )
            if not reservar_cantidad_lote(nueva_lu, cantidad):
                raise ValueError("No se pudo reservar en el lote destino (posible condición de carrera).")
            nueva_lu.refresh_from_db()
            nuevo_lote.refresh_from_db()
            reserva_ant_nueva = nuevo_lote.cantidad_reservada - cantidad
            motivo_asig = (
                f"Ajuste a datos de lote en propuesta de surtimiento (corrección de captura). "
                f"Asignación al lote correcto {nuevo_lote.numero_lote} (producto: {nuevo_lote.producto.clave_cnis or 'N/A'}). "
                f"Folio pedido: {folio_solicitud}. Reserva anterior: {reserva_ant_nueva}, nueva: {nuevo_lote.cantidad_reservada}."
            )
            MovimientoInventario.objects.create(
                lote=nuevo_lote,
                tipo_movimiento='AJUSTE_DATOS_LOTE',
                cantidad=cantidad,
                cantidad_anterior=reserva_ant_nueva,
                cantidad_nueva=nuevo_lote.cantidad_reservada,
                motivo=motivo_asig,
                documento_referencia=f"PROP-SURT-{la.item_propuesta.propuesta_id}",
                folio=folio_solicitud or None,
                usuario=request.user,
            )
            la.lote_ubicacion = nueva_lu
            la.save(update_fields=['lote_ubicacion'])
        return JsonResponse({
            'exito': True,
            'mensaje': 'Lote corregido. Se generaron los registros de inventario (ajuste a datos de lote).',
            'nuevo_lote_numero': nuevo_lote.numero_lote,
            'nuevo_ubicacion_codigo': nueva_lu.ubicacion.codigo if nueva_lu.ubicacion else '',
        })
    except ValueError as e:
        return JsonResponse({'exito': False, 'mensaje': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'exito': False, 'mensaje': f'Error al corregir el lote: {str(e)}'}, status=500)


@login_required
@require_http_methods(['GET'])
def limites_cantidad_picking(request, lote_asignado_id):
    """
    Devuelve cantidad actual, máximo disponible y mínimo (por cantidad solicitada) para el modal de cambiar cantidad.
    """
    from django.db.models import Sum
    try:
        la = LoteAsignado.objects.select_related('item_propuesta', 'lote_ubicacion').get(id=lote_asignado_id)
    except LoteAsignado.DoesNotExist:
        return JsonResponse({'error': 'Asignación no encontrada'}, status=404)
    if la.surtido:
        return JsonResponse({'error': 'Ítem ya recogido'}, status=400)
    lu = la.lote_ubicacion
    cantidad_actual = la.cantidad_asignada
    disponible_efectivo = (lu.cantidad - lu.cantidad_reservada) + cantidad_actual
    cantidad_solicitada = la.item_propuesta.cantidad_solicitada
    total_otros = la.item_propuesta.lotes_asignados.exclude(id=la.id).aggregate(t=Sum('cantidad_asignada'))['t'] or 0
    minimo = 0
    return JsonResponse({
        'cantidad_actual': cantidad_actual,
        'maximo_disponible': disponible_efectivo,
        'minimo': minimo,
        'cantidad_solicitada': cantidad_solicitada,
        'total_otros': total_otros,
    })


@login_required
@csrf_exempt
@require_http_methods(['POST'])
def cambiar_cantidad_picking(request, lote_asignado_id):
    """
    Cambia la cantidad asignada de un ítem en la orden de picking.
    Permite cantidad menor a la solicitada o cero (ej. el hospital no se lleva el insumo o lleva menos).
    Actualiza la cantidad reservada (libera o reserva según el caso), genera movimiento de inventario
    como evidencia, y si la nueva cantidad es 0 elimina la asignación.
    Validación: no exceder cantidad disponible (cantidad - cantidad_reservada).
    """
    import json
    try:
        la = LoteAsignado.objects.select_related(
            'item_propuesta__propuesta__solicitud',
            'lote_ubicacion__lote',
            'lote_ubicacion__ubicacion',
        ).get(id=lote_asignado_id)
    except LoteAsignado.DoesNotExist:
        return JsonResponse({'exito': False, 'mensaje': 'Asignación no encontrada'}, status=404)
    if la.surtido:
        return JsonResponse({'exito': False, 'mensaje': 'No se puede cambiar la cantidad de un ítem ya recogido'}, status=400)
    try:
        body = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        body = {}
    try:
        nueva_cantidad = int(body.get('nueva_cantidad', -1))
    except (TypeError, ValueError):
        return JsonResponse({'exito': False, 'mensaje': 'nueva_cantidad debe ser un número entero'}, status=400)
    if nueva_cantidad < 0:
        return JsonResponse({'exito': False, 'mensaje': 'La cantidad no puede ser negativa.'}, status=400)

    lu = la.lote_ubicacion
    lote = lu.lote
    cantidad_actual = la.cantidad_asignada
    item = la.item_propuesta
    folio_solicitud = (item.propuesta.solicitud.observaciones_solicitud or '').strip() or item.propuesta.solicitud.folio or ''

    if nueva_cantidad == cantidad_actual:
        return JsonResponse({'exito': True, 'mensaje': 'La cantidad no ha cambiado.', 'nueva_cantidad': nueva_cantidad})

    disponible_efectivo = (lu.cantidad - lu.cantidad_reservada) + cantidad_actual
    if nueva_cantidad > disponible_efectivo:
        return JsonResponse({
            'exito': False,
            'mensaje': f'No hay cantidad suficiente. Disponible (cantidad - reservada) para este lote/ubicación: {disponible_efectivo}.'
        }, status=400)

    with transaction.atomic():
        if nueva_cantidad == 0:
            liberar = cantidad_actual
            resultado = liberar_cantidad_lote(lu, liberar)
            reserva_ant = resultado.get('reserva_anterior_lote', 0)
            reserva_nueva = resultado.get('nueva_reservada_lote', 0)
            motivo = (
                f"Cambio de cantidad en orden de picking - Liberación total de reserva ({liberar} unidades). "
                f"El hospital no se lleva este insumo o se corrige la línea. Lote {lote.numero_lote} "
                f"(Producto: {lote.producto.clave_cnis or 'N/A'}). Folio pedido: {folio_solicitud}. "
                f"Reserva anterior: {reserva_ant}, nueva: {reserva_nueva}."
            )
            MovimientoInventario.objects.create(
                lote=lote,
                tipo_movimiento='AJUSTE_POSITIVO',
                cantidad=liberar,
                cantidad_anterior=reserva_ant,
                cantidad_nueva=reserva_nueva,
                motivo=motivo,
                documento_referencia=f"PROP-SURT-{item.propuesta_id}",
                folio=folio_solicitud or None,
                usuario=request.user,
            )
            la.delete()
            return JsonResponse({
                'exito': True,
                'mensaje': 'Cantidad actualizada a cero. Se liberó la reserva, se generó el movimiento de inventario y se quitó la línea de la orden.',
                'nueva_cantidad': 0,
            })
        if nueva_cantidad < cantidad_actual:
            liberar = cantidad_actual - nueva_cantidad
            resultado = liberar_cantidad_lote(lu, liberar)
            reserva_ant = resultado.get('reserva_anterior_lote', 0)
            reserva_nueva = resultado.get('nueva_reservada_lote', 0)
            motivo = (
                f"Cambio de cantidad en orden de picking - Liberación de reserva de {liberar} unidades del lote {lote.numero_lote} "
                f"(Producto: {lote.producto.clave_cnis or 'N/A'}). El hospital lleva menos de lo solicitado. Folio pedido: {folio_solicitud}. "
                f"Reserva anterior: {reserva_ant}, nueva: {reserva_nueva}."
            )
            MovimientoInventario.objects.create(
                lote=lote,
                tipo_movimiento='AJUSTE_POSITIVO',
                cantidad=liberar,
                cantidad_anterior=reserva_ant,
                cantidad_nueva=reserva_nueva,
                motivo=motivo,
                documento_referencia=f"PROP-SURT-{item.propuesta_id}",
                folio=folio_solicitud or None,
                usuario=request.user,
            )
        else:
            reservar = nueva_cantidad - cantidad_actual
            if not reservar_cantidad_lote(lu, reservar):
                return JsonResponse({
                    'exito': False,
                    'mensaje': 'No se pudo reservar la cantidad adicional (condición de carrera o stock insuficiente).'
                }, status=400)
            lu.refresh_from_db()
            lote.refresh_from_db()
            reserva_ant = lote.cantidad_reservada - reservar
            motivo = (
                f"Cambio de cantidad en orden de picking - Incremento de reserva de {reservar} unidades en lote {lote.numero_lote} "
                f"(Producto: {lote.producto.clave_cnis or 'N/A'}). Folio pedido: {folio_solicitud}. "
                f"Reserva anterior: {reserva_ant}, nueva: {lote.cantidad_reservada}."
            )
            MovimientoInventario.objects.create(
                lote=lote,
                tipo_movimiento='AJUSTE_DATOS_LOTE',
                cantidad=reservar,
                cantidad_anterior=reserva_ant,
                cantidad_nueva=lote.cantidad_reservada,
                motivo=motivo,
                documento_referencia=f"PROP-SURT-{item.propuesta_id}",
                folio=folio_solicitud or None,
                usuario=request.user,
            )
        la.cantidad_asignada = nueva_cantidad
        la.save(update_fields=['cantidad_asignada'])
    return JsonResponse({
        'exito': True,
        'mensaje': 'Cantidad actualizada. Se ajustó la reserva y se generó el movimiento de inventario.',
        'nueva_cantidad': nueva_cantidad,
    })


# ============================================================
# IMPRIMIR HOJA DE SURTIDO
# ============================================================

@login_required
def imprimir_hoja_surtido(request, propuesta_id):
    """
    Genera un PDF con la hoja de picking ordenada por ubicación.
    Genera el Excel primero y luego lo convierte a PDF usando weasyprint.
    """
    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id)
    
    # Verificar si el usuario es administrador o supervisor
    es_administrador = request.user.is_staff or request.user.is_superuser
    es_supervisor = request.user.groups.filter(name='Supervisor').exists()
    puede_ver_todo = es_administrador or es_supervisor
    
    # Obtener almacén del usuario si no es administrador ni supervisor
    almacen_usuario_id = None
    if not puede_ver_todo and hasattr(request.user, 'almacen') and request.user.almacen:
        almacen_usuario_id = request.user.almacen.id
    
    # Obtener items de la propuesta con lotes asignados
    items_picking = []
    
    for item in propuesta.items.all():
        for lote_asignado in item.lotes_asignados.filter(surtido=False):
            lote_ubicacion = lote_asignado.lote_ubicacion
            lote = lote_ubicacion.lote
            
            # Filtrar por almacén del usuario si no es administrador ni supervisor
            if not puede_ver_todo and almacen_usuario_id:
                if lote_ubicacion.ubicacion.almacen_id != almacen_usuario_id:
                    continue  # Saltar este item si no pertenece al almacén del usuario
            
            caducidad = lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A'
            items_picking.append({
                'item_id': item.id,
                'lote_asignado_id': lote_asignado.id,
                'producto': lote.producto.descripcion,
                'cantidad': lote_asignado.cantidad_asignada,
                'lote_numero': lote.numero_lote,
                'almacen': lote_ubicacion.ubicacion.almacen.nombre,
                'almacen_id': lote_ubicacion.ubicacion.almacen_id,
                'ubicacion': lote_ubicacion.ubicacion.codigo,
                'ubicacion_id': lote_ubicacion.ubicacion_id,
                'clave_cnis': lote.producto.clave_cnis,
                'caducidad': caducidad,
            })
    
    # Ordenar por ubicación
    items_picking.sort(key=lambda x: (x['almacen_id'], x['ubicacion_id']))
    
    try:
        # Generar Excel
        excel_buffer = exportar_picking_excel_interno(propuesta, items_picking)
        
        # Convertir Excel a PDF usando weasyprint
        pdf_buffer = convertir_excel_a_pdf(excel_buffer)
        
        # Retornar PDF
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="picking_{propuesta.solicitud.folio}.pdf"'
        return response
        
    except Exception as e:
        return HttpResponse(f'Error al generar PDF: {str(e)}', status=500)



# ============================================================
# FUNCIÓN INTERNA PARA GENERAR EXCEL
# ============================================================

def exportar_picking_excel_interno(propuesta, items_picking):
    """
    Función interna que genera el Excel sin retornar HttpResponse.
    Usada tanto por exportar_picking_excel como por imprimir_hoja_surtido.
    
    Returns:
        BytesIO: Buffer con el contenido del Excel
    """
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Picking"
    
    # Definir estilos
    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar información de la propuesta
    ws['A1'] = "HOJA DE PICKING"
    ws['A1'].font = Font(bold=True, size=14, color="8B1538")
    ws.merge_cells('A1:G1')
    
    ws['A3'] = "Propuesta:"
    ws['B3'] = str(propuesta.solicitud.folio)
    ws['C3'] = "Institución Solicitante:"
    ws['D3'] = propuesta.solicitud.institucion_solicitante.denominacion if propuesta.solicitud.institucion_solicitante else "N/A"
    
    ws['A4'] = "Fecha:"
    ws['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws['C4'] = "Folio de Pedido:"
    ws['D4'] = propuesta.solicitud.observaciones_solicitud or "N/A"
    
    ws['A5'] = "Área:"
    ws['B5'] = propuesta.solicitud.almacen_destino.nombre if propuesta.solicitud.almacen_destino else "N/A"
    ws['C5'] = "Total Items:"
    ws['D5'] = len(items_picking)
    
    # Definir anchos de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12 * 6
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    
    # Agregar encabezados
    headers = ['UBICACIÓN', 'CLAVE CNIS', 'PRODUCTO', 'CADUCIDAD', 'LOTE', 'CANTIDAD', 'CANTIDAD SURTIDA']
    header_row = 8
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[header_row].height = 25
    
    # Agregar datos
    for row_num, item in enumerate(items_picking, header_row + 1):
        # Ubicación
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.value = item['ubicacion']
        cell_a.alignment = center_alignment
        cell_a.font = Font(bold=True)
        cell_a.border = border
        
        # Clave CNIS
        cell_b = ws.cell(row=row_num, column=2)
        cell_b.value = item['clave_cnis']
        cell_b.alignment = center_alignment
        cell_b.border = border
        
        # Producto
        cell_c = ws.cell(row=row_num, column=3)
        cell_c.value = item['producto']
        cell_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_c.border = border
        num_lines = len(str(item['producto']).split('\n')) + (len(str(item['producto'])) // 100)
        ws.row_dimensions[row_num].height = max(25, num_lines * 15)
        
        # Caducidad
        cell_d = ws.cell(row=row_num, column=4)
        cell_d.value = item['caducidad']
        cell_d.alignment = center_alignment
        cell_d.border = border
        
        # Lote
        cell_e = ws.cell(row=row_num, column=5)
        cell_e.value = item['lote_numero']
        cell_e.alignment = center_alignment
        cell_e.border = border
        
        # Cantidad
        cell_f = ws.cell(row=row_num, column=6)
        cell_f.value = item['cantidad']
        cell_f.alignment = center_alignment
        cell_f.font = Font(bold=True)
        cell_f.fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
        cell_f.border = border
        
        # Cantidad Surtida (vacío)
        cell_g = ws.cell(row=row_num, column=7)
        cell_g.value = ""
        cell_g.alignment = center_alignment
        cell_g.border = border
    
    # Crear respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


# ============================================================
# EXPORTAR HOJA DE PICKING A EXCEL
# ============================================================

@login_required
def exportar_picking_excel(request, propuesta_id):
    """
    Genera un archivo Excel con la hoja de picking ordenada por ubicación.
    Versión 2.0: Sin template, con encabezados en fila 8.
    """
    propuesta = get_object_or_404(PropuestaPedido, id=propuesta_id)
    
    # Verificar si el usuario es administrador o supervisor
    es_administrador = request.user.is_staff or request.user.is_superuser
    es_supervisor = request.user.groups.filter(name='Supervisor').exists()
    puede_ver_todo = es_administrador or es_supervisor
    
    # Obtener almacén del usuario si no es administrador ni supervisor
    almacen_usuario_id = None
    if not puede_ver_todo and hasattr(request.user, 'almacen') and request.user.almacen:
        almacen_usuario_id = request.user.almacen.id
    
    # Obtener items de la propuesta con lotes asignados
    items_picking = []
    
    for item in propuesta.items.all():
        for lote_asignado in item.lotes_asignados.filter(surtido=False):
            lote_ubicacion = lote_asignado.lote_ubicacion
            lote = lote_ubicacion.lote
            
            # Filtrar por almacén del usuario si no es administrador ni supervisor
            if not puede_ver_todo and almacen_usuario_id:
                if lote_ubicacion.ubicacion.almacen_id != almacen_usuario_id:
                    continue  # Saltar este item si no pertenece al almacén del usuario
            
            caducidad = lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A'
            items_picking.append({
                'item_id': item.id,
                'lote_asignado_id': lote_asignado.id,
                'producto': lote.producto.descripcion,
                'cantidad': lote_asignado.cantidad_asignada,
                'lote_numero': lote.numero_lote,
                'almacen': lote_ubicacion.ubicacion.almacen.nombre,
                'almacen_id': lote_ubicacion.ubicacion.almacen_id,
                'ubicacion': lote_ubicacion.ubicacion.codigo,
                'ubicacion_id': lote_ubicacion.ubicacion_id,
                'clave_cnis': lote.producto.clave_cnis,
                'caducidad': caducidad,
            })
    
    # Ordenar por ubicación
    items_picking.sort(key=lambda x: (x['almacen_id'], x['ubicacion_id']))
    
    # Crear workbook - Versión sin template
    wb = Workbook()
    ws = wb.active
    ws.title = "Picking"
    
    # Definir estilos
    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    # Agregar información de la propuesta
    ws['A1'] = "HOJA DE PICKING"
    ws['A1'].font = Font(bold=True, size=14, color="8B1538")
    ws.merge_cells('A1:G1')
    
    ws['A3'] = "Propuesta:"
    ws['B3'] = str(propuesta.solicitud.folio)
    ws['C3'] = "Institución Solicitante:"
    ws['D3'] = propuesta.solicitud.institucion_solicitante.denominacion if propuesta.solicitud.institucion_solicitante else "N/A"
    
    ws['A4'] = "Fecha:"
    ws['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws['C4'] = "Folio de Pedido:"
    ws['D4'] = propuesta.solicitud.observaciones_solicitud or "N/A"
    
    ws['A5'] = "Área:"
    ws['B5'] = propuesta.solicitud.almacen_destino.nombre if propuesta.solicitud.almacen_destino else "N/A"
    ws['C5'] = "Total Items:"
    ws['D5'] = len(items_picking)
    
    # Definir anchos de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12 * 6  # Aumentar 6 veces el ancho para PRODUCTO
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    
    # Agregar encabezados
    headers = ['UBICACIÓN', 'CLAVE CNIS', 'PRODUCTO', 'CADUCIDAD', 'LOTE', 'CANTIDAD', 'CANTIDAD SURTIDA']
    header_row = 8
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Establecer altura del encabezado
    ws.row_dimensions[header_row].height = 25
    
    # Agregar datos
    for row_num, item in enumerate(items_picking, header_row + 1):
        # Ubicación
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.value = item['ubicacion']
        cell_a.alignment = center_alignment
        cell_a.font = Font(bold=True)
        cell_a.border = border
        
        # Clave CNIS
        cell_b = ws.cell(row=row_num, column=2)
        cell_b.value = item['clave_cnis']
        cell_b.alignment = center_alignment
        cell_b.border = border
        
        # Producto
        cell_c = ws.cell(row=row_num, column=3)
        cell_c.value = item['producto']
        cell_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_c.border = border
        # Calcular altura basada en el número de líneas de texto
        num_lines = len(str(item['producto']).split('\n')) + (len(str(item['producto'])) // 100)
        ws.row_dimensions[row_num].height = max(25, num_lines * 15)
        
        # Caducidad
        cell_d = ws.cell(row=row_num, column=4)
        cell_d.value = item['caducidad']
        cell_d.alignment = center_alignment
        cell_d.border = border
        
        # Lote
        cell_e = ws.cell(row=row_num, column=5)
        cell_e.value = item['lote_numero']
        cell_e.alignment = center_alignment
        cell_e.border = border
        
        # Cantidad
        cell_f = ws.cell(row=row_num, column=6)
        cell_f.value = item['cantidad']
        cell_f.alignment = center_alignment
        cell_f.font = Font(bold=True)
        cell_f.fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
        cell_f.border = border
        
        # Cantidad Surtida (vacío)
        cell_g = ws.cell(row=row_num, column=7)
        cell_g.value = ""
        cell_g.alignment = center_alignment
        cell_g.border = border
        

    # Crear respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="picking_{propuesta.solicitud.folio}.xlsx"'
    
    return response
