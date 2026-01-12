"""
Reporte de Productos sin Ubicación Asignada

Muestra todos los lotes que no tienen ubicación asignada en el almacén.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Lote, Institucion, Almacen


@login_required
def reporte_sin_ubicacion(request):
    """
    Reporte de lotes sin ubicación asignada.
    Muestra lotes que no tienen ninguna ubicación registrada.
    """
    
    # Obtener todos los lotes
    todos_lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'almacen'
    ).all()
    
    # Lotes sin ubicación
    lotes_sin_ubicacion = []
    
    for lote in todos_lotes:
        # Verificar si tiene ubicaciones asignadas
        tiene_ubicacion = lote.ubicaciones_detalle.exists()
        
        if not tiene_ubicacion:
            lotes_sin_ubicacion.append({
                'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                'lote': lote.numero_lote,
                'institucion': lote.institucion.denominacion if lote.institucion else '-',
                'almacen': lote.almacen.nombre if lote.almacen else '-',
                'cantidad_disponible': lote.cantidad_disponible,
                'cantidad_inicial': lote.cantidad_inicial,
                'precio_unitario': lote.precio_unitario or 0,
                'valor_total': (lote.cantidad_disponible * lote.precio_unitario) if lote.precio_unitario else 0,
                'estado': lote.get_estado_display() if hasattr(lote, 'get_estado_display') else '-',
                'fecha_recepcion': lote.fecha_recepcion.strftime('%d/%m/%Y') if lote.fecha_recepcion else '-',
                'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else '-',
                'fecha_fabricacion': lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else '-',
                'lote_id': lote.id,
            })
    
    # Filtros
    filtro_clave = request.GET.get('clave', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_institucion = request.GET.get('institucion', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_estado = request.GET.get('estado', '')
    
    # Aplicar filtros
    if filtro_clave:
        lotes_sin_ubicacion = [l for l in lotes_sin_ubicacion if filtro_clave.lower() in l['clave_cnis'].lower()]
    
    if filtro_lote:
        lotes_sin_ubicacion = [l for l in lotes_sin_ubicacion if filtro_lote.lower() in l['lote'].lower()]
    
    if filtro_institucion:
        lotes_sin_ubicacion = [l for l in lotes_sin_ubicacion if l['institucion'] == filtro_institucion]
    
    if filtro_almacen:
        lotes_sin_ubicacion = [l for l in lotes_sin_ubicacion if l['almacen'] == filtro_almacen]
    
    if filtro_estado:
        lotes_sin_ubicacion = [l for l in lotes_sin_ubicacion if l['estado'] == filtro_estado]
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(lotes_sin_ubicacion, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener opciones de filtro
    instituciones = Institucion.objects.all().order_by('denominacion')
    almacenes = Almacen.objects.all().order_by('nombre')
    
    # Estadísticas
    total_cantidad = sum(l['cantidad_disponible'] for l in lotes_sin_ubicacion)
    total_valor = sum(l['valor_total'] for l in lotes_sin_ubicacion)
    
    context = {
        'page_obj': page_obj,
        'total_registros': len(lotes_sin_ubicacion),
        'total_cantidad': total_cantidad,
        'total_valor': total_valor,
        'instituciones': instituciones,
        'almacenes': almacenes,
        'filtro_clave': filtro_clave,
        'filtro_lote': filtro_lote,
        'filtro_institucion': filtro_institucion,
        'filtro_almacen': filtro_almacen,
        'filtro_estado': filtro_estado,
    }
    
    return render(request, 'inventario/reporte_sin_ubicacion.html', context)


@login_required
def exportar_sin_ubicacion_excel(request):
    """
    Exporta el reporte de lotes sin ubicación a Excel.
    """
    
    # Obtener todos los lotes
    todos_lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'almacen'
    ).all()
    
    # Lotes sin ubicación
    lotes_sin_ubicacion = []
    
    for lote in todos_lotes:
        # Verificar si tiene ubicaciones asignadas
        tiene_ubicacion = lote.ubicaciones_detalle.exists()
        
        if not tiene_ubicacion:
            lotes_sin_ubicacion.append({
                'clave_cnis': lote.producto.clave_cnis if lote.producto else '-',
                'descripcion_producto': lote.producto.descripcion if lote.producto else '-',
                'lote': lote.numero_lote,
                'institucion': lote.institucion.denominacion if lote.institucion else '-',
                'almacen': lote.almacen.nombre if lote.almacen else '-',
                'cantidad_disponible': lote.cantidad_disponible,
                'cantidad_inicial': lote.cantidad_inicial,
                'precio_unitario': lote.precio_unitario or 0,
                'valor_total': (lote.cantidad_disponible * lote.precio_unitario) if lote.precio_unitario else 0,
                'estado': lote.get_estado_display() if hasattr(lote, 'get_estado_display') else '-',
                'fecha_recepcion': lote.fecha_recepcion.strftime('%d/%m/%Y') if lote.fecha_recepcion else '-',
                'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else '-',
                'fecha_fabricacion': lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else '-',
            })
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sin Ubicación"
    
    # Estilos
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'CLAVE CNIS',
        'DESCRIPCIÓN PRODUCTO',
        'LOTE',
        'INSTITUCIÓN',
        'ALMACÉN',
        'CANTIDAD DISPONIBLE',
        'CANTIDAD INICIAL',
        'PRECIO UNITARIO',
        'VALOR TOTAL',
        'ESTADO',
        'FECHA RECEPCIÓN',
        'FECHA CADUCIDAD',
        'FECHA FABRICACIÓN'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    total_cantidad = 0
    total_valor = 0
    
    for row, lote in enumerate(lotes_sin_ubicacion, 2):
        total_cantidad += lote['cantidad_disponible']
        total_valor += lote['valor_total']
        
        ws.cell(row=row, column=1).value = lote['clave_cnis']
        ws.cell(row=row, column=2).value = lote['descripcion_producto']
        ws.cell(row=row, column=3).value = lote['lote']
        ws.cell(row=row, column=4).value = lote['institucion']
        ws.cell(row=row, column=5).value = lote['almacen']
        ws.cell(row=row, column=6).value = lote['cantidad_disponible']
        ws.cell(row=row, column=7).value = lote['cantidad_inicial']
        ws.cell(row=row, column=8).value = lote['precio_unitario']
        ws.cell(row=row, column=9).value = lote['valor_total']
        ws.cell(row=row, column=10).value = lote['estado']
        ws.cell(row=row, column=11).value = lote['fecha_recepcion']
        ws.cell(row=row, column=12).value = lote['fecha_caducidad']
        ws.cell(row=row, column=13).value = lote['fecha_fabricacion']
        
        # Aplicar bordes
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in [6, 7, 8, 9]:  # Columnas numéricas
                cell.alignment = Alignment(horizontal='right')
    
    # Fila de totales
    total_row = len(lotes_sin_ubicacion) + 2
    ws.cell(row=total_row, column=1).value = "TOTALES"
    ws.cell(row=total_row, column=1).font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    
    ws.cell(row=total_row, column=6).value = total_cantidad
    ws.cell(row=total_row, column=6).font = total_font
    ws.cell(row=total_row, column=6).fill = total_fill
    
    ws.cell(row=total_row, column=9).value = total_valor
    ws.cell(row=total_row, column=9).font = total_font
    ws.cell(row=total_row, column=9).fill = total_fill
    
    # Aplicar bordes a la fila de totales
    for col in range(1, 14):
        ws.cell(row=total_row, column=col).border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 18
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_sin_ubicacion.xlsx"'
    wb.save(response)
    
    return response
