"""
Módulo para generar Acuse de Entrega en Excel usando template
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment, numbers
from copy import copy
from datetime import datetime
import os
from django.conf import settings
from io import BytesIO


def copiar_estilo_celda(celda_origen, celda_destino):
    """
    Copia el estilo de una celda a otra de forma segura
    """
    if celda_origen.font:
        celda_destino.font = copy(celda_origen.font)
    if celda_origen.border:
        celda_destino.border = copy(celda_origen.border)
    if celda_origen.fill:
        celda_destino.fill = copy(celda_origen.fill)
    if celda_origen.number_format:
        celda_destino.number_format = copy(celda_origen.number_format)
    if celda_origen.alignment:
        celda_destino.alignment = copy(celda_origen.alignment)


def _acuse_rgb_relleno_solido(celda):
    """Devuelve el RGB del fg en relleno sólido (ej. 'FF8B1538') o None."""
    fill = celda.fill
    if fill is None or getattr(fill, 'patternType', None) != 'solid':
        return None
    fg = getattr(fill, 'fgColor', None)
    if fg is None or not getattr(fg, 'rgb', None):
        return None
    return str(fg.rgb).upper()


def _acuse_es_relleno_guinda(rgb_upper):
    """Colores guinda del template (xl/styles.xml: 8B1538 y variante 8A1538)."""
    if not rgb_upper:
        return False
    core = rgb_upper[-6:] if len(rgb_upper) >= 6 else rgb_upper
    return core in ('8B1538', '8A1538')


def _acuse_texto_blanco_en_rellenos_guinda(ws, max_row=40, max_col=11):
    """
    Fuerza texto blanco en cualquier celda con fondo guinda del acuse.
    En la plantilla, la banda con letra negra suele ser la fila 9 (no la 8:
    la 8 solo tiene A8 con fondo blanco); así cubrimos todas las variantes.
    """
    last = min(max_row, ws.max_row or max_row)
    for fila in range(1, last + 1):
        for col in range(1, max_col + 1):
            celda = ws.cell(row=fila, column=col)
            if type(celda).__name__ == 'MergedCell':
                continue
            rgb = _acuse_rgb_relleno_solido(celda)
            if not rgb or not _acuse_es_relleno_guinda(rgb):
                continue
            f = celda.font
            celda.font = Font(
                name=(f.name if f and f.name else 'Calibri'),
                size=(f.size if f and f.size else 9),
                bold=bool(f and f.bold),
                italic=bool(f and f.italic),
                color='FFFFFF',
            )


def agregar_bordes_celda(celda):
    """
    Agrega bordes a una celda
    """
    from openpyxl.styles import Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    celda.border = thin_border


def generar_acuse_excel(propuesta, almacen_id=None, for_pdf=False):
    """
    Genera un archivo Excel con el acuse de entrega usando un template como base.

    Cada fila del acuse proviene de un LoteAsignado con cantidad_asignada > 0. Si almacen_id
    está definido (usuario sin perfil de administrador pero con almacén asignado en su cuenta),
    solo se incluyen asignaciones cuya ubicación pertenece a ese almacén. Eso puede hacer que
    el PDF/Excel muestre menos líneas que el detalle web de la propuesta (donde se listan todos
    los ítems y lotes). Sin almacen_id, se incluyen todas las asignaciones de la propuesta.

    Args:
        propuesta: Objeto PropuestaPedido
        almacen_id: Opcional. Filtra por lote_ubicacion__ubicacion__almacen_id.
        for_pdf: Si True, no altera la configuración de impresión del libro (el PDF
            se genera a partir del mismo Excel y ya luce bien en horizontal allí).

    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    
    # Cargar el template
    template_path = os.path.join(settings.BASE_DIR, 'inventario', 'templates', 'acuse_entrega_template.xlsx')
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template no encontrado en: {template_path}")
    
    # Cargar el workbook desde el template
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Obtener datos de la propuesta
    folio = propuesta.solicitud.folio
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    folio_pedido = propuesta.solicitud.observaciones_solicitud or 'N/A'
    institucion = propuesta.solicitud.institucion_solicitante.denominacion if propuesta.solicitud.institucion_solicitante else 'N/A'
    
    # ============ ACTUALIZAR ENCABEZADO ============
    
    # Fila 1: #FOLIO
    ws['I1'].value = f'#FOLIO: {folio}'
    
    # Fila 3: FOLIO DE PEDIDO
    ws['I3'].value = f'FOLIO DE PEDIDO: {folio_pedido}'
    
    # Fila 4: FECHA
    ws['I4'].value = f'FECHA: {fecha_actual}'
    
    # ============ ACTUALIZAR TABLA DE FIRMAS ============
    
    # Fila 10: Institución
    ws['A10'].value = f'INSTITUCIÓN: {institucion}'
    
    # Limpiar nombre y puesto en "AUTORIZA (ALMACEN)" (columna C)
    # Buscar en las filas típicas de la tabla de firmas (10-14)
    for row in range(10, 15):
        cell_c = ws.cell(row=row, column=3)  # Columna C (AUTORIZA ALMACEN)
        if cell_c.value:
            cell_value = str(cell_c.value)
            # Eliminar cualquier referencia a "Gerardo Anaya" y "MESA DE CONTROL"
            cell_value = cell_value.replace('Gerardo Anaya', '')
            cell_value = cell_value.replace('gerardo anaya', '')
            cell_value = cell_value.replace('GERARDO ANAYA', '')
            cell_value = cell_value.replace('MESA DE CONTROL', '')
            cell_value = cell_value.replace('mesa de control', '')
            cell_value = cell_value.replace('Mesa de Control', '')
            
            # Si la celda contiene "NOMBRE:" o "PUESTO:", limpiar el contenido después
            if 'NOMBRE:' in cell_value or 'PUESTO:' in cell_value:
                lines = cell_value.split('\n')
                new_lines = []
                for line in lines:
                    line_stripped = line.strip()
                    # Eliminar líneas que contengan "Gerardo Anaya" o "MESA DE CONTROL"
                    if 'Gerardo' in line_stripped or 'Anaya' in line_stripped or 'MESA DE CONTROL' in line_stripped:
                        continue
                    if line_stripped.startswith('NOMBRE:'):
                        new_lines.append('NOMBRE:')
                        new_lines.append('')
                    elif line_stripped.startswith('PUESTO:'):
                        new_lines.append('PUESTO:')
                        new_lines.append('')
                    elif line_stripped and not line_stripped.startswith('NOMBRE:') and not line_stripped.startswith('PUESTO:'):
                        # Mantener otras líneas como "FIRMA:"
                        new_lines.append(line)
                if new_lines:
                    cell_c.value = '\n'.join(new_lines)
                else:
                    # Si no quedó nada, establecer valores por defecto limpios
                    cell_c.value = 'NOMBRE:\n\nPUESTO:\n\nFIRMA: __________________'
    
    # ============ ACTUALIZAR TABLA DE ITEMS ============
    
    # Refrescar la propuesta desde la base de datos para asegurar datos actualizados
    propuesta.refresh_from_db()
    
    # Recolectar todos los items primero
    # Obtener items directamente desde la base de datos para evitar problemas de caché
    # NO usar prefetch_related aquí porque puede usar datos en caché
    from .pedidos_models import ItemPropuesta, LoteAsignado
    
    items_data = []
    idx = 1
    
    # Obtener items directamente desde la base de datos con consulta fresca
    items = ItemPropuesta.objects.filter(
        propuesta=propuesta
    ).select_related('producto').all()
    
    for item in items:
        # Refrescar el item desde la base de datos
        item.refresh_from_db()
        
        # Obtener lotes asignados directamente desde la base de datos (consulta fresca)
        # Si almacen_id está definido, solo insumos del almacén del usuario (no admin)
        filtro_lotes = {
            'item_propuesta': item,
            'cantidad_asignada__gt': 0,
        }
        if almacen_id is not None:
            filtro_lotes['lote_ubicacion__ubicacion__almacen_id'] = almacen_id
        lotes_asignados = LoteAsignado.objects.filter(
            **filtro_lotes
        ).select_related(
            'lote_ubicacion__lote',
            'lote_ubicacion__ubicacion'
        ).all()
        
        for lote_asignado in lotes_asignados:
            # Refrescar el lote_asignado para asegurar datos actualizados
            lote_asignado.refresh_from_db()
            
            # Obtener las relaciones directamente desde la base de datos para asegurar datos actualizados
            # En lugar de usar lote_asignado.lote_ubicacion que puede estar en caché
            from .models import LoteUbicacion, Lote
            lote_ubicacion = LoteUbicacion.objects.select_related('lote', 'ubicacion').get(
                id=lote_asignado.lote_ubicacion_id
            )
            lote = Lote.objects.get(id=lote_ubicacion.lote_id)
            
            ubicacion = lote_ubicacion.ubicacion.codigo if lote_ubicacion.ubicacion else 'N/A'
            caducidad = lote.fecha_caducidad.strftime("%d/%m/%Y") if lote.fecha_caducidad else 'N/A'
            lote_info = lote.numero_lote if lote.numero_lote else 'N/A'
            descripcion = item.producto.descripcion
            # Usar siempre cantidad_asignada (ya filtramos los que tienen cantidad > 0)
            cantidad = lote_asignado.cantidad_asignada
            
            items_data.append({
                'idx': idx,
                'clave': item.producto.clave_cnis,
                'descripcion': descripcion,
                'um': item.producto.unidad_medida,
                'tipo': 'ORDINARIO',
                'lote': lote_info,
                'caducidad': caducidad,
                # 'clasificacion': 'MEDICAMENTO',  # Removido - ya no se incluye
                'ubicacion': ubicacion,
                'cantidad': cantidad,
                'folio_pedido': propuesta.solicitud.observaciones_solicitud or ''  # Usar observaciones_solicitud
            })
            idx += 1
    
    # Eliminar filas de datos del template (mantener encabezados)
    # El template tiene datos de ejemplo en filas 18 en adelante
    # Vamos a eliminar todas las filas después de la 17 (encabezados)
    while ws.max_row > 17:
        ws.delete_rows(18, 1)
    
    # Agregar las filas de datos
    row_num = 18
    for item_data in items_data:
        # Agregar fila con datos
        ws.cell(row=row_num, column=1).value = item_data['idx']
        ws.cell(row=row_num, column=2).value = item_data['clave']
        ws.cell(row=row_num, column=3).value = item_data['descripcion']
        ws.cell(row=row_num, column=4).value = item_data['um']
        ws.cell(row=row_num, column=5).value = item_data['tipo']
        ws.cell(row=row_num, column=6).value = item_data['lote']
        ws.cell(row=row_num, column=7).value = item_data['caducidad']
        # Columna 8: CLASIFICACIÓN (dejar vacío en el detalle)
        ws.cell(row=row_num, column=8).value = ''  # CLASIFICACIÓN vacía
        # Columna 9: UBICACIÓN (movido de columna 8)
        ws.cell(row=row_num, column=9).value = item_data['ubicacion']
        # Columna 10: CANTIDAD (movido de columna 9)
        ws.cell(row=row_num, column=10).value = item_data['cantidad']
        # Columna 11: FOLIO PEDIDO (movido de columna 10)
        ws.cell(row=row_num, column=11).value = item_data['folio_pedido']
        
        # Copiar estilos de la fila 18 del template (primera fila de datos)
        fila_referencia = 18 if row_num == 18 else row_num - 1
        
        # Ajustar rango de columnas (ahora son 11 columnas)
        for col in range(1, 12):
            celda_referencia = ws.cell(row=fila_referencia, column=col)
            celda_destino = ws.cell(row=row_num, column=col)
            copiar_estilo_celda(celda_referencia, celda_destino)
            # Agregar bordes
            agregar_bordes_celda(celda_destino)
        
        row_num += 1
    
    # ============ ACTUALIZAR ENCABEZADOS DE TABLA DE DETALLE ============
    
    # Actualizar encabezados:
    # Columna 8: CLASIFICACIÓN (dejar el título, pero el detalle queda vacío)
    ws.cell(row=17, column=8).value = 'CLASIFICACIÓN'
    # Columna 9: UBICACIÓN (movido de columna 8)
    ws.cell(row=17, column=9).value = 'UBICACIÓN'
    # Columna 10: CANTIDAD (movido de columna 9)
    ws.cell(row=17, column=10).value = 'CANTIDAD'
    # Columna 11: FOLIO PEDIDO (movido de columna 10)
    ws.cell(row=17, column=11).value = 'FOLIO PEDIDO'
    
    # Celdas con fondo guinda (p. ej. fila de encabezados de firmas / detalle): texto blanco
    _acuse_texto_blanco_en_rellenos_guinda(ws)

    # Solo descarga Excel: orientación horizontal al imprimir / vista previa
    if not for_pdf:
        ws.page_setup.orientation = 'landscape'
        # 1 = carta (US Letter), estándar openpyxl / OOXML
        ws.page_setup.paperSize = 1

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
