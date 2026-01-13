"""
Reporte de Ubicaciones Vacías en Almacenes

Muestra todas las ubicaciones que no tienen lotes asignados.
Permite filtrar por almacén e institución.
Exporta a Excel y PDF.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Count
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from io import BytesIO

from .models import UbicacionAlmacen, LoteUbicacion, Almacen, Institucion


@login_required
def reporte_ubicaciones_vacias(request):
    """
    Reporte de ubicaciones vacías (sin lotes asignados).
    Muestra ubicaciones que no tienen ningún lote registrado.
    """
    
    # Obtener todas las ubicaciones
    todas_ubicaciones = UbicacionAlmacen.objects.select_related(
        'almacen',
        'almacen__institucion'
    ).all()
    
    # Ubicaciones vacías (sin lotes asignados)
    ubicaciones_vacias = []
    
    for ubicacion in todas_ubicaciones:
        # Verificar si tiene lotes asignados
        tiene_lotes = LoteUbicacion.objects.filter(ubicacion=ubicacion).exists()
        
        if not tiene_lotes:
            ubicaciones_vacias.append({
                'id': ubicacion.id,
                'codigo': ubicacion.codigo,
                'descripcion': ubicacion.descripcion or '-',
                'nivel': ubicacion.nivel or '-',
                'pasillo': ubicacion.pasillo or '-',
                'rack': ubicacion.rack or '-',
                'seccion': ubicacion.seccion or '-',
                'almacen': ubicacion.almacen.nombre if ubicacion.almacen else '-',
                'almacen_id': ubicacion.almacen.id if ubicacion.almacen else None,
                'institucion': ubicacion.almacen.institucion.denominacion if ubicacion.almacen and ubicacion.almacen.institucion else '-',
                'institucion_id': ubicacion.almacen.institucion.id if ubicacion.almacen and ubicacion.almacen.institucion else None,
                'estado': ubicacion.get_estado_display() if hasattr(ubicacion, 'get_estado_display') else ubicacion.estado,
                'activo': 'Sí' if ubicacion.activo else 'No',
            })
    
    # Filtros
    filtro_codigo = request.GET.get('codigo', '').strip()
    filtro_almacen = request.GET.get('almacen', '')
    filtro_institucion = request.GET.get('institucion', '')
    filtro_estado = request.GET.get('estado', '')
    
    # Aplicar filtros
    if filtro_codigo:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if filtro_codigo.lower() in u['codigo'].lower()]
    
    if filtro_almacen:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['almacen_id'] == int(filtro_almacen)]
    
    if filtro_institucion:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['institucion_id'] == int(filtro_institucion)]
    
    if filtro_estado:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['estado'] == filtro_estado]
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(ubicaciones_vacias, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener opciones de filtro
    instituciones = Institucion.objects.all().order_by('denominacion')
    almacenes = Almacen.objects.all().order_by('nombre')
    
    # Estados disponibles
    estados = [
        ('disponible', 'Disponible'),
        ('ocupada', 'Ocupada'),
        ('bloqueada', 'Bloqueada'),
        ('cuarentena', 'Cuarentena'),
        ('caducados', 'Caducados'),
        ('devoluciones', 'Devoluciones'),
    ]
    
    context = {
        'page_obj': page_obj,
        'total_registros': len(ubicaciones_vacias),
        'instituciones': instituciones,
        'almacenes': almacenes,
        'estados': estados,
        'filtro_codigo': filtro_codigo,
        'filtro_almacen': filtro_almacen,
        'filtro_institucion': filtro_institucion,
        'filtro_estado': filtro_estado,
    }
    
    return render(request, 'inventario/reporte_ubicaciones_vacias.html', context)


@login_required
def exportar_ubicaciones_vacias_excel(request):
    """
    Exporta el reporte de ubicaciones vacías a Excel.
    """
    
    # Obtener todas las ubicaciones
    todas_ubicaciones = UbicacionAlmacen.objects.select_related(
        'almacen',
        'almacen__institucion'
    ).all()
    
    # Ubicaciones vacías
    ubicaciones_vacias = []
    
    for ubicacion in todas_ubicaciones:
        tiene_lotes = LoteUbicacion.objects.filter(ubicacion=ubicacion).exists()
        
        if not tiene_lotes:
            ubicaciones_vacias.append({
                'codigo': ubicacion.codigo,
                'descripcion': ubicacion.descripcion or '-',
                'nivel': ubicacion.nivel or '-',
                'pasillo': ubicacion.pasillo or '-',
                'rack': ubicacion.rack or '-',
                'seccion': ubicacion.seccion or '-',
                'almacen': ubicacion.almacen.nombre if ubicacion.almacen else '-',
                'institucion': ubicacion.almacen.institucion.denominacion if ubicacion.almacen and ubicacion.almacen.institucion else '-',
                'estado': ubicacion.get_estado_display() if hasattr(ubicacion, 'get_estado_display') else ubicacion.estado,
                'activo': 'Sí' if ubicacion.activo else 'No',
            })
    
    # Aplicar filtros
    filtro_codigo = request.GET.get('codigo', '').strip()
    filtro_almacen = request.GET.get('almacen', '')
    filtro_institucion = request.GET.get('institucion', '')
    filtro_estado = request.GET.get('estado', '')
    
    if filtro_codigo:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if filtro_codigo.lower() in u['codigo'].lower()]
    
    if filtro_almacen:
        almacen_obj = Almacen.objects.get(id=int(filtro_almacen))
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['almacen'] == almacen_obj.nombre]
    
    if filtro_institucion:
        institucion_obj = Institucion.objects.get(id=int(filtro_institucion))
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['institucion'] == institucion_obj.denominacion]
    
    if filtro_estado:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['estado'] == filtro_estado]
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ubicaciones Vacías"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'CÓDIGO UBICACIÓN',
        'DESCRIPCIÓN',
        'NIVEL',
        'PASILLO',
        'RACK',
        'SECCIÓN',
        'ALMACÉN',
        'INSTITUCIÓN',
        'ESTADO',
        'ACTIVO'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    for row, ubicacion in enumerate(ubicaciones_vacias, 2):
        ws.cell(row=row, column=1).value = ubicacion['codigo']
        ws.cell(row=row, column=2).value = ubicacion['descripcion']
        ws.cell(row=row, column=3).value = ubicacion['nivel']
        ws.cell(row=row, column=4).value = ubicacion['pasillo']
        ws.cell(row=row, column=5).value = ubicacion['rack']
        ws.cell(row=row, column=6).value = ubicacion['seccion']
        ws.cell(row=row, column=7).value = ubicacion['almacen']
        ws.cell(row=row, column=8).value = ubicacion['institucion']
        ws.cell(row=row, column=9).value = ubicacion['estado']
        ws.cell(row=row, column=10).value = ubicacion['activo']
        
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
    
    # Fila de totales
    total_row = len(ubicaciones_vacias) + 2
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=2).value = len(ubicaciones_vacias)
    ws.cell(row=total_row, column=2).font = total_font
    ws.cell(row=total_row, column=2).fill = total_fill
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ubicaciones_vacias_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def exportar_ubicaciones_vacias_pdf(request):
    """
    Exporta el reporte de ubicaciones vacías a PDF.
    """
    
    # Obtener todas las ubicaciones
    todas_ubicaciones = UbicacionAlmacen.objects.select_related(
        'almacen',
        'almacen__institucion'
    ).all()
    
    # Ubicaciones vacías
    ubicaciones_vacias = []
    
    for ubicacion in todas_ubicaciones:
        tiene_lotes = LoteUbicacion.objects.filter(ubicacion=ubicacion).exists()
        
        if not tiene_lotes:
            ubicaciones_vacias.append({
                'codigo': ubicacion.codigo,
                'descripcion': ubicacion.descripcion or '-',
                'nivel': ubicacion.nivel or '-',
                'pasillo': ubicacion.pasillo or '-',
                'rack': ubicacion.rack or '-',
                'seccion': ubicacion.seccion or '-',
                'almacen': ubicacion.almacen.nombre if ubicacion.almacen else '-',
                'institucion': ubicacion.almacen.institucion.denominacion if ubicacion.almacen and ubicacion.almacen.institucion else '-',
                'estado': ubicacion.get_estado_display() if hasattr(ubicacion, 'get_estado_display') else ubicacion.estado,
                'activo': 'Sí' if ubicacion.activo else 'No',
            })
    
    # Aplicar filtros
    filtro_codigo = request.GET.get('codigo', '').strip()
    filtro_almacen = request.GET.get('almacen', '')
    filtro_institucion = request.GET.get('institucion', '')
    filtro_estado = request.GET.get('estado', '')
    
    if filtro_codigo:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if filtro_codigo.lower() in u['codigo'].lower()]
    
    if filtro_almacen:
        almacen_obj = Almacen.objects.get(id=int(filtro_almacen))
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['almacen'] == almacen_obj.nombre]
    
    if filtro_institucion:
        institucion_obj = Institucion.objects.get(id=int(filtro_institucion))
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['institucion'] == institucion_obj.denominacion]
    
    if filtro_estado:
        ubicaciones_vacias = [u for u in ubicaciones_vacias if u['estado'] == filtro_estado]
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=6,
        alignment=1
    )
    
    title = Paragraph('REPORTE DE UBICACIONES VACÍAS EN ALMACENES', title_style)
    elements.append(title)
    
    # Información del reporte
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12
    )
    
    info_text = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de ubicaciones vacías: {len(ubicaciones_vacias)}'
    info = Paragraph(info_text, info_style)
    elements.append(info)
    
    # Tabla de datos
    headers = [
        'CÓDIGO UBICACIÓN',
        'DESCRIPCIÓN',
        'NIVEL',
        'PASILLO',
        'RACK',
        'SECCIÓN',
        'ALMACÉN',
        'INSTITUCIÓN',
        'ESTADO',
        'ACTIVO'
    ]
    data = [headers]
    
    for ubicacion in ubicaciones_vacias:
        data.append([
            ubicacion['codigo'],
            ubicacion['descripcion'],
            ubicacion['nivel'],
            ubicacion['pasillo'],
            ubicacion['rack'],
            ubicacion['seccion'],
            ubicacion['almacen'],
            ubicacion['institucion'],
            ubicacion['estado'],
            ubicacion['activo'],
        ])
    
    table = Table(data, colWidths=[1.2*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1.5*inch, 1*inch, 0.7*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ubicaciones_vacias_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(buffer.getvalue())
    
    return response
