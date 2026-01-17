"""
Módulo para convertir archivos Excel a PDF
Usa openpyxl para leer el Excel y weasyprint para generar el PDF
Sin dependencias externas como LibreOffice
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
from weasyprint import HTML, CSS
from io import BytesIO
import base64


def convertir_excel_a_pdf(excel_buffer):
    """
    Convierte un buffer de Excel a PDF usando openpyxl y weasyprint.
    
    Args:
        excel_buffer: BytesIO con contenido del archivo Excel
        
    Returns:
        BytesIO: Buffer con contenido del PDF
        
    Raises:
        Exception: Si falla la conversión
    """
    
    try:
        # Leer el Excel
        excel_buffer.seek(0)
        workbook = load_workbook(excel_buffer)
        worksheet = workbook.active
        
        # Convertir a HTML
        html_content = _excel_a_html(worksheet)
        
        # Generar PDF desde HTML
        pdf_buffer = BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        return pdf_buffer
        
    except Exception as e:
        raise Exception(f"Error al convertir Excel a PDF: {str(e)}")


def _excel_a_html(worksheet):
    """
    Convierte una hoja de Excel a HTML.
    
    Args:
        worksheet: Hoja de openpyxl
        
    Returns:
        str: HTML con el contenido formateado
    """
    
    html_parts = [
        '<!DOCTYPE html>',
        '<html>',
        '<head>',
        '<meta charset="UTF-8">',
        '<title>Picking</title>',
        '<style>',
        _generar_css(),
        '</style>',
        '</head>',
        '<body>',
        '<div class="container">',
    ]
    
    # Procesar cada fila
    for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
        # Determinar clase de fila
        row_class = _obtener_clase_fila(row_idx, worksheet)
        
        html_parts.append(f'<div class="row {row_class}">')
        
        for cell in row:
            cell_class = _obtener_clase_celda(cell)
            cell_value = _obtener_valor_celda(cell)
            cell_style = _obtener_estilo_celda(cell)
            
            html_parts.append(f'<div class="cell {cell_class}" style="{cell_style}">')
            html_parts.append(str(cell_value) if cell_value else '')
            html_parts.append('</div>')
        
        html_parts.append('</div>')
    
    html_parts.extend([
        '</div>',
        '</body>',
        '</html>',
    ])
    
    return '\n'.join(html_parts)


def _generar_css():
    """Genera el CSS para el HTML."""
    
    css = """
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }
    
    body {
        font-family: Arial, sans-serif;
        font-size: 10px;
        line-height: 1.3;
    }
    
    .container {
        width: 100%;
        padding: 8mm;
        display: flex;
        flex-direction: column;
    }
    
    .row {
        display: flex;
        flex-direction: row;
        width: 100%;
        min-height: 20px;
        border-bottom: 1px solid #ddd;
    }
    
    .row.header {
        background-color: #1f77b4;
        color: white;
        font-weight: bold;
        min-height: 25px;
    }
    
    .row.info {
        background-color: #f9f9f9;
        border-left: 3px solid #8B1538;
        padding: 5mm;
        min-height: auto;
    }
    
    .cell {
        display: flex;
        align-items: center;
        justify-content: center;
        padding: 2mm;
        border-right: 1px solid #ddd;
        word-wrap: break-word;
        overflow-wrap: break-word;
    }
    
    .cell.header {
        text-align: center;
        font-weight: bold;
        font-size: 11px;
    }
    
    .cell.ubicacion {
        font-weight: bold;
        color: #8B1538;
        text-align: center;
    }
    
    .cell.producto {
        text-align: left;
        font-size: 9px;
        padding: 3mm;
    }
    
    .cell.cantidad {
        text-align: center;
        font-weight: bold;
        background-color: #e8f4f8;
    }
    
    .cell.cantidad-surtida {
        text-align: center;
        min-height: 15mm;
        background-color: #ffffff;
    }
    
    .cell.lote {
        text-align: center;
        font-size: 9px;
    }
    
    @page {
        size: A4 landscape;
        margin: 8mm;
    }
    
    @media print {
        body {
            margin: 0;
            padding: 0;
        }
        .container {
            padding: 8mm;
        }
    }
    """
    
    return css


def _obtener_clase_fila(row_idx, worksheet):
    """Determina la clase CSS para una fila."""
    
    # Las primeras filas son de información
    if row_idx <= 7:
        return "info"
    
    # Fila 8 es el encabezado
    if row_idx == 8:
        return "header"
    
    # Filas de datos
    return "data"


def _obtener_clase_celda(cell):
    """Determina la clase CSS para una celda."""
    
    # Verificar si es encabezado
    if cell.fill and cell.fill.start_color:
        if cell.fill.start_color.rgb in ['FF1f77b4', '1f77b4']:
            return "header"
    
    # Verificar por contenido o posición
    return ""


def _obtener_valor_celda(cell):
    """Obtiene el valor de una celda."""
    
    value = cell.value
    
    if value is None:
        return ""
    
    # Formatear fechas
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y')
    
    return str(value)


def _obtener_estilo_celda(cell):
    """Genera estilos CSS inline para una celda."""
    
    styles = []
    
    # Ancho de columna
    if cell.column_letter:
        col_width = cell.parent.column_dimensions.get(cell.column_letter)
        if col_width and col_width.width:
            # Convertir ancho de Excel a píxeles aproximadamente
            width_px = int(col_width.width * 8)
            styles.append(f"width: {width_px}px;")
    
    # Color de fondo
    if cell.fill and cell.fill.start_color:
        color = cell.fill.start_color.rgb
        if color and color != '00000000':
            # Convertir color de Excel a hex
            if len(str(color)) == 8:
                hex_color = f"#{str(color)[2:]}"
            else:
                hex_color = f"#{str(color)}"
            styles.append(f"background-color: {hex_color};")
    
    # Color de texto
    if cell.font and cell.font.color:
        color = cell.font.color.rgb
        if color and color != '00000000':
            if len(str(color)) == 8:
                hex_color = f"#{str(color)[2:]}"
            else:
                hex_color = f"#{str(color)}"
            styles.append(f"color: {hex_color};")
    
    # Alineación
    if cell.alignment:
        if cell.alignment.horizontal:
            text_align = _mapear_alineacion(cell.alignment.horizontal)
            if text_align:
                styles.append(f"text-align: {text_align};")
        
        if cell.alignment.vertical:
            vertical_align = _mapear_alineacion_vertical(cell.alignment.vertical)
            if vertical_align:
                styles.append(f"vertical-align: {vertical_align};")
    
    # Negrita
    if cell.font and cell.font.bold:
        styles.append("font-weight: bold;")
    
    # Tamaño de fuente
    if cell.font and cell.font.size:
        styles.append(f"font-size: {cell.font.size}px;")
    
    # Bordes
    if cell.border:
        border_style = _generar_bordes(cell.border)
        if border_style:
            styles.append(border_style)
    
    return " ".join(styles)


def _mapear_alineacion(alignment):
    """Mapea alineación de Excel a CSS."""
    
    mapping = {
        'left': 'left',
        'center': 'center',
        'right': 'right',
        'justify': 'justify',
        'distributed': 'justify',
    }
    
    return mapping.get(alignment, 'left')


def _mapear_alineacion_vertical(alignment):
    """Mapea alineación vertical de Excel a CSS."""
    
    mapping = {
        'top': 'top',
        'center': 'middle',
        'bottom': 'bottom',
        'justify': 'middle',
        'distributed': 'middle',
    }
    
    return mapping.get(alignment, 'middle')


def _generar_bordes(border):
    """Genera estilos de borde CSS."""
    
    if not border:
        return ""
    
    border_parts = []
    
    if border.left and border.left.style:
        border_parts.append("border-left: 1px solid #999;")
    
    if border.right and border.right.style:
        border_parts.append("border-right: 1px solid #999;")
    
    if border.top and border.top.style:
        border_parts.append("border-top: 1px solid #999;")
    
    if border.bottom and border.bottom.style:
        border_parts.append("border-bottom: 1px solid #999;")
    
    return " ".join(border_parts)
