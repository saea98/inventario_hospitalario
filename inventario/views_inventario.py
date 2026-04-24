"""
Vistas para Fase 2.3: Gestión de Inventario
Dashboard, consulta de lotes, movimientos y reportes
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.db import transaction
from django.db.models import Q, Sum, Count, F, IntegerField
from django.db.models.functions import Coalesce
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import datetime, timedelta, date
from decimal import Decimal
import pandas as pd

from .models import Lote, MovimientoInventario, Producto, LoteUbicacion, Almacen, Institucion, UbicacionAlmacen
from .propuesta_utils import (
    totales_reserva_activa_por_lote_ids,
    totales_reserva_activa_por_lote_ubicacion_ids,
)
from .llegada_models import ItemLlegada
from .forms_entrada_salida import ItemEntradaForm, ItemSalidaForm


# ============================================================
# DASHBOARD DE INVENTARIO
# ============================================================

@login_required
def dashboard_inventario(request):
    """Dashboard principal de inventario con resumen y alertas"""
    
    # Obtener institución del usuario (si aplica)
    institucion = request.user.institucion if hasattr(request.user, 'institucion') else None
    
    # Filtrar por institución si el usuario tiene una asignada
    if institucion:
        lotes = Lote.objects.filter(institucion=institucion)
    else:
        lotes = Lote.objects.all()
    
    # Resumen de stock
    total_lotes = lotes.count()
    total_cantidad = lotes.aggregate(Sum('cantidad_disponible'))['cantidad_disponible__sum'] or 0
    
    # Lotes por estado
    lotes_disponibles = lotes.filter(estado=1).count()
    lotes_suspendidos = lotes.filter(estado=4).count()
    lotes_deteriorados = lotes.filter(estado=5).count()
    lotes_caducados = lotes.filter(estado=6).count()
    
    # Alertas de caducidad
    hoy = timezone.now().date()
    fecha_alerta = hoy + timedelta(days=90)
    
    lotes_proximos_caducar = lotes.filter(
        fecha_caducidad__lte=fecha_alerta,
        fecha_caducidad__gt=hoy,
        estado=1
    ).order_by('fecha_caducidad')[:10]
    
    lotes_caducados_alert = lotes.filter(
        fecha_caducidad__lt=hoy,
        estado__in=[1, 4, 5]
    ).order_by('fecha_caducidad')[:10]
    
    # Movimientos recientes
    movimientos_recientes = MovimientoInventario.objects.select_related('lote').order_by('-fecha_movimiento')[:10]
    
    # Productos con bajo stock
    productos_bajo_stock = []
    for producto in Producto.objects.filter(activo=True):
        stock_total = lotes.filter(producto=producto, estado=1).aggregate(
            Sum('cantidad_disponible')
        )['cantidad_disponible__sum'] or 0
        
        if stock_total < 10:  # Umbral de bajo stock
            productos_bajo_stock.append({
                'producto': producto,
                'stock': stock_total
            })
    
    # Stock por almacén
    almacenes = Almacen.objects.filter(activo=True)
    stock_almacenes = []
    for almacen in almacenes:
        cantidad = lotes.filter(almacen=almacen, estado=1).aggregate(
            Sum('cantidad_disponible')
        )['cantidad_disponible__sum'] or 0
        stock_almacenes.append({
            'almacen': almacen,
            'cantidad': cantidad
        })
    
    context = {
        'total_lotes': total_lotes,
        'total_cantidad': total_cantidad,
        'lotes_disponibles': lotes_disponibles,
        'lotes_suspendidos': lotes_suspendidos,
        'lotes_deteriorados': lotes_deteriorados,
        'lotes_caducados': lotes_caducados,
        'lotes_proximos_caducar': lotes_proximos_caducar,
        'lotes_caducados_alert': lotes_caducados_alert,
        'movimientos_recientes': movimientos_recientes,
        'productos_bajo_stock': productos_bajo_stock,
        'stock_almacenes': stock_almacenes,
    }
    
    return render(request, 'inventario/dashboard_inventario.html', context)


# ============================================================
# CONSULTA DE LOTES
# ============================================================

# Fecha mínima considerada coherente para fecha_recepcion (evitar fechas antes del arranque del sistema)
FECHA_RECEPCION_MIN = date(2010, 1, 1)


def _fecha_recepcion_es_coherente(fecha_recepcion, hoy=None):
    """True si la fecha de recepción es coherente: no futura y no anterior a FECHA_RECEPCION_MIN."""
    if fecha_recepcion is None:
        return True
    hoy = hoy or timezone.now().date()
    return FECHA_RECEPCION_MIN <= fecha_recepcion <= hoy


@login_required
def lista_lotes(request):
    """Lista de lotes con filtros y búsqueda"""
    
    # Obtener institución del usuario
    institucion = request.user.institucion if hasattr(request.user, 'institucion') else None
    
    # Filtro base
    if institucion:
        lotes = Lote.objects.filter(institucion=institucion).select_related(
            'producto', 'almacen', 'ubicacion'
        )
    else:
        lotes = Lote.objects.all().select_related(
            'producto', 'almacen', 'ubicacion'
        )
    
    # Filtros
    filtro_institucion = request.GET.get('institucion', '')
    filtro_estado = request.GET.get('estado', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_ubicacion = request.GET.get('ubicacion', '')
    filtro_producto = request.GET.get('producto', '')
    filtro_caducidad = request.GET.get('caducidad', '')
    filtro_fecha_recepcion = request.GET.get('fecha_recepcion', '')  # '', 'coherentes', 'incoherentes'
    filtro_fecha_recepcion_desde = request.GET.get('fecha_recepcion_desde', '').strip()
    filtro_fecha_recepcion_hasta = request.GET.get('fecha_recepcion_hasta', '').strip()
    filtro_con_remision = request.GET.get('con_remision', '')  # '' o '1'
    filtro_con_orden_suministro = request.GET.get('con_orden_suministro', '')  # '' o '1'
    filtro_fecha_cita_desde = request.GET.get('fecha_cita_desde', '').strip()
    filtro_fecha_cita_hasta = request.GET.get('fecha_cita_hasta', '').strip()
    busqueda_lote = request.GET.get('busqueda_lote', '')
    busqueda_cnis = request.GET.get('busqueda_cnis', '')
    busqueda_producto = request.GET.get("busqueda_producto", "")
    filtro_partida = request.GET.get("partida", "")
    
    # Aplicar filtros
    if filtro_institucion:
        lotes = lotes.filter(institucion_id=int(filtro_institucion))
    
    if filtro_estado:
        lotes = lotes.filter(estado=int(filtro_estado))
    
    if filtro_almacen:
        lotes = lotes.filter(almacen_id=int(filtro_almacen))
    
    # Filtro por ubicación: considerar asignaciones en LoteUbicacion (ubicaciones_detalle)
    if filtro_ubicacion:
        lotes = lotes.filter(ubicaciones_detalle__ubicacion_id=int(filtro_ubicacion)).distinct()
    
    if filtro_producto:
        lotes = lotes.filter(producto_id=int(filtro_producto))
    
    # Filtro de caducidad
    hoy = timezone.now().date()
    if filtro_caducidad:
        if filtro_caducidad == 'caducados':
            lotes = lotes.filter(fecha_caducidad__lt=hoy)
        elif filtro_caducidad == 'menos_30':
            lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=30))
        elif filtro_caducidad == 'menos_60':
            lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=60))
        elif filtro_caducidad == 'menos_90':
            lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=90))
    
    # Búsquedas separadas
    if busqueda_lote:
        lotes = lotes.filter(numero_lote__icontains=busqueda_lote)
    
    if busqueda_cnis:
        lotes = lotes.filter(producto__clave_cnis__icontains=busqueda_cnis)
    
    if busqueda_producto:
        lotes = lotes.filter(producto__descripcion__icontains=busqueda_producto)

    if filtro_partida:
        lotes = lotes.filter(partida__icontains=filtro_partida)

    # Filtro por coherencia de fecha de recepción (no futura, no anterior al arranque del sistema)
    hoy = timezone.now().date()
    if filtro_fecha_recepcion == 'coherentes':
        lotes = lotes.filter(
            fecha_recepcion__gte=FECHA_RECEPCION_MIN,
            fecha_recepcion__lte=hoy,
        )
    elif filtro_fecha_recepcion == 'incoherentes':
        lotes = lotes.filter(
            Q(fecha_recepcion__gt=hoy) | Q(fecha_recepcion__lt=FECHA_RECEPCION_MIN)
        )

    # Rango de fecha de recepción (desde / hasta)
    if filtro_fecha_recepcion_desde:
        try:
            fd = datetime.strptime(filtro_fecha_recepcion_desde, '%Y-%m-%d').date()
            lotes = lotes.filter(fecha_recepcion__gte=fd)
        except ValueError:
            pass
    if filtro_fecha_recepcion_hasta:
        try:
            fh = datetime.strptime(filtro_fecha_recepcion_hasta, '%Y-%m-%d').date()
            lotes = lotes.filter(fecha_recepcion__lte=fh)
        except ValueError:
            pass

    # Solo con remisión (campo remision no vacío)
    if filtro_con_remision == '1':
        lotes = lotes.filter(remision__isnull=False).exclude(remision='')

    # Solo con orden de suministro
    if filtro_con_orden_suministro == '1':
        lotes = lotes.filter(orden_suministro__isnull=False)

    # Rango de fecha de cita (lotes que vienen de llegada con cita en ese rango)
    if filtro_fecha_cita_desde:
        try:
            fd = datetime.strptime(filtro_fecha_cita_desde, '%Y-%m-%d').date()
            lotes = lotes.filter(item_llegada__llegada__cita__fecha_cita__date__gte=fd)
        except ValueError:
            pass
    if filtro_fecha_cita_hasta:
        try:
            fh = datetime.strptime(filtro_fecha_cita_hasta, '%Y-%m-%d').date()
            lotes = lotes.filter(item_llegada__llegada__cita__fecha_cita__date__lte=fh)
        except ValueError:
            pass
    
    # Ordenar
    lotes = lotes.order_by('-fecha_recepcion')

    # Detalle por ubicación (misma fila = un lote; badges = LoteUbicacion)
    lotes = lotes.prefetch_related('ubicaciones_detalle__ubicacion__almacen')
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(lotes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Reserva real = suma LoteAsignado (surtido=False), no el campo Lote.cantidad_reservada
    lote_ids_pagina = [l.id for l in page_obj]
    reservas_reales_map = totales_reserva_activa_por_lote_ids(lote_ids_pagina)
    for lote in page_obj:
        lote.reserva_activa_calculada = reservas_reales_map.get(lote.id, 0)

    # Reserva por LoteUbicacion (asignación en propuesta = producto + lote + ubicación)
    lu_ids_pagina = []
    for lote in page_obj:
        for ubi in lote.ubicaciones_detalle.all():
            lu_ids_pagina.append(ubi.id)
    reserva_por_lu = totales_reserva_activa_por_lote_ubicacion_ids(lu_ids_pagina)
    for lote in page_obj:
        for ubi in lote.ubicaciones_detalle.all():
            ubi.reserva_activa_calculada = reserva_por_lu.get(ubi.id, 0)

    # IDs de lotes en esta página con fecha de recepción incoherente (para marcar en la tabla)
    lotes_fecha_incoherente = set()
    for lote in page_obj:
        if not _fecha_recepcion_es_coherente(lote.fecha_recepcion, hoy):
            lotes_fecha_incoherente.add(lote.id)
    
    # Opciones para filtros
    instituciones = Institucion.objects.filter(activo=True).order_by('clue')
    almacenes = Almacen.objects.filter(activo=True).select_related('institucion')
    ubicaciones = UbicacionAlmacen.objects.filter(activo=True).select_related('almacen').order_by('almacen__nombre', 'codigo')
    productos = Producto.objects.filter(activo=True)
    estados = Lote.ESTADOS_CHOICES
    
    # Columnas disponibles para exportación
    columnas_disponibles = [
        {'value': 'numero_lote', 'label': 'Número de Lote'},
        {'value': 'producto__descripcion', 'label': 'Producto'},
        {'value': 'institucion__denominacion', 'label': 'Institución'},
        {'value': 'cantidad_inicial', 'label': 'Cantidad Inicial'},
        {'value': 'cantidad_disponible', 'label': 'Cantidad Disponible'},
        {'value': 'cantidad_reservada', 'label': 'Cantidad reservada (suma LoteAsignado activo)'},
        {'value': 'precio_unitario', 'label': 'Precio Unitario'},
        {'value': 'valor_total', 'label': 'Valor Total'},
        {'value': 'fecha_fabricacion', 'label': 'Fecha de Fabricación'},
        {'value': 'fecha_caducidad', 'label': 'Fecha de Caducidad'},
        {'value': 'fecha_recepcion', 'label': 'Fecha de Recepción'},
        {'value': 'estado', 'label': 'Estado'},
        {'value': 'observaciones', 'label': 'Observaciones'},
        {'value': 'rfc_proveedor', 'label': 'RFC Proveedor'},
        {'value': 'proveedor', 'label': 'Proveedor'},
        {'value': 'orden_suministro__proveedor__rfc', 'label': 'RFC Proveedor (Orden suministro)'},
        {'value': 'orden_suministro__proveedor__razon_social', 'label': 'Proveedor (Orden suministro)'},
        {'value': 'partida', 'label': 'Partida'},
        {'value': 'clave_saica', 'label': 'Clave SAICA'},
        {'value': 'descripcion_saica', 'label': 'Descripción SAICA'},
        {'value': 'unidad_saica', 'label': 'Unidad SAICA'},
        {'value': 'fuente_datos', 'label': 'Fuente de Datos'},
        {'value': 'contrato', 'label': 'Contrato'},
        {'value': 'folio', 'label': 'Folio'},
        {'value': 'subtotal', 'label': 'Subtotal'},
        {'value': 'iva', 'label': 'IVA'},
        {'value': 'importe_total', 'label': 'Importe Total'},
        {'value': 'licitacion', 'label': 'Licitación / Procedimiento'},
        {'value': 'pedido', 'label': 'Pedido'},
        {'value': 'remision', 'label': 'Remisión'},
        {'value': 'responsable', 'label': 'Responsable'},
        {'value': 'reviso', 'label': 'Revisó'},
        {'value': 'tipo_entrega', 'label': 'Tipo de Entrega'},
        {'value': 'tipo_red', 'label': 'Tipo de Red'},
        {'value': 'epa', 'label': 'EPA'},
        {'value': 'producto__clave_cnis', 'label': 'Clave CNIS'},
        {'value': 'almacen__nombre', 'label': 'Almacén'},
        {'value': 'ubicacion__codigo', 'label': 'Ubicación'},
    ]
    
    context = {
        'page_obj': page_obj,
        'instituciones': instituciones,
        'almacenes': almacenes,
        'ubicaciones': ubicaciones,
        'productos': productos,
        'estados': estados,
        'filtro_institucion': filtro_institucion,
        'filtro_estado': filtro_estado,
        'filtro_almacen': filtro_almacen,
        'filtro_ubicacion': filtro_ubicacion,
        'filtro_producto': filtro_producto,
        'filtro_caducidad': filtro_caducidad,
        'busqueda_lote': busqueda_lote,
        'busqueda_cnis': busqueda_cnis,
        "busqueda_producto": busqueda_producto,
        "filtro_partida": filtro_partida,
        'filtro_fecha_recepcion': filtro_fecha_recepcion,
        'filtro_fecha_recepcion_desde': filtro_fecha_recepcion_desde,
        'filtro_fecha_recepcion_hasta': filtro_fecha_recepcion_hasta,
        'filtro_con_remision': filtro_con_remision,
        'filtro_con_orden_suministro': filtro_con_orden_suministro,
        'filtro_fecha_cita_desde': filtro_fecha_cita_desde,
        'filtro_fecha_cita_hasta': filtro_fecha_cita_hasta,
        'lotes_fecha_incoherente': lotes_fecha_incoherente,
        'columnas_disponibles': columnas_disponibles,
    }
    
    return render(request, 'inventario/lista_lotes.html', context)


@login_required
def reporte_lotes_personalizado(request):
    """
    Reporte independiente para exportar lotes a Excel con campos personalizados.
    Esta vista solo muestra filtros y el configurador de columnas, sin la tabla de existencias,
    para poder asignarla a otros roles sin dar acceso completo a 'lista_lotes'.
    """
    # Obtener institución del usuario (mismo criterio que lista_lotes)
    institucion = request.user.institucion if hasattr(request.user, 'institucion') else None

    # Filtros (mismos nombres que usa exportar_lotes_personalizado)
    filtro_institucion = request.GET.get('institucion', '')
    filtro_estado = request.GET.get('estado', '')
    filtro_almacen = request.GET.get('almacen', '')
    filtro_producto = request.GET.get('producto', '')
    filtro_caducidad = request.GET.get('caducidad', '')
    filtro_fecha_recepcion = request.GET.get('fecha_recepcion', '')
    filtro_fecha_recepcion_desde = request.GET.get('fecha_recepcion_desde', '').strip()
    filtro_fecha_recepcion_hasta = request.GET.get('fecha_recepcion_hasta', '').strip()
    filtro_con_remision = request.GET.get('con_remision', '')
    filtro_con_orden_suministro = request.GET.get('con_orden_suministro', '')
    filtro_fecha_cita_desde = request.GET.get('fecha_cita_desde', '').strip()
    filtro_fecha_cita_hasta = request.GET.get('fecha_cita_hasta', '').strip()
    busqueda_lote = request.GET.get('busqueda_lote', '')
    busqueda_cnis = request.GET.get('busqueda_cnis', '')
    busqueda_producto = request.GET.get("busqueda_producto", "")
    filtro_partida = request.GET.get("partida", "")
    filtro_excluir_sin_rfc = request.GET.get("excluir_sin_rfc", "")

    # Opciones para filtros (idénticas a lista_lotes)
    instituciones = Institucion.objects.filter(activo=True).order_by('clue')
    almacenes = Almacen.objects.filter(activo=True).select_related('institucion')
    productos = Producto.objects.filter(activo=True)
    estados = Lote.ESTADOS_CHOICES

    # Columnas disponibles para exportación (reutilizar la misma definición)
    columnas_disponibles = [
        {'value': 'numero_lote', 'label': 'Número de Lote'},
        {'value': 'producto__descripcion', 'label': 'Producto'},
        {'value': 'institucion__denominacion', 'label': 'Institución'},
        {'value': 'cantidad_inicial', 'label': 'Cantidad Inicial'},
        {'value': 'cantidad_disponible', 'label': 'Cantidad Disponible'},
        {'value': 'cantidad_reservada', 'label': 'Cantidad reservada (suma LoteAsignado activo)'},
        {'value': 'precio_unitario', 'label': 'Precio Unitario'},
        {'value': 'valor_total', 'label': 'Valor Total'},
        {'value': 'fecha_fabricacion', 'label': 'Fecha de Fabricación'},
        {'value': 'fecha_caducidad', 'label': 'Fecha de Caducidad'},
        {'value': 'fecha_recepcion', 'label': 'Fecha de Recepción'},
        {'value': 'estado', 'label': 'Estado'},
        {'value': 'observaciones', 'label': 'Observaciones'},
        {'value': 'rfc_proveedor', 'label': 'RFC Proveedor'},
        {'value': 'proveedor', 'label': 'Proveedor'},
        {'value': 'orden_suministro__proveedor__rfc', 'label': 'RFC Proveedor (Orden suministro)'},
        {'value': 'orden_suministro__proveedor__razon_social', 'label': 'Proveedor (Orden suministro)'},
        {'value': 'partida', 'label': 'Partida'},
        {'value': 'clave_saica', 'label': 'Clave SAICA'},
        {'value': 'descripcion_saica', 'label': 'Descripción SAICA'},
        {'value': 'unidad_saica', 'label': 'Unidad SAICA'},
        {'value': 'fuente_datos', 'label': 'Fuente de Datos'},
        {'value': 'contrato', 'label': 'Contrato'},
        {'value': 'folio', 'label': 'Folio'},
        {'value': 'subtotal', 'label': 'Subtotal'},
        {'value': 'iva', 'label': 'IVA'},
        {'value': 'importe_total', 'label': 'Importe Total'},
        {'value': 'licitacion', 'label': 'Licitación / Procedimiento'},
        {'value': 'pedido', 'label': 'Pedido'},
        {'value': 'remision', 'label': 'Remisión'},
        {'value': 'responsable', 'label': 'Responsable'},
        {'value': 'reviso', 'label': 'Revisó'},
        {'value': 'tipo_entrega', 'label': 'Tipo de Entrega'},
        {'value': 'tipo_red', 'label': 'Tipo de Red'},
        {'value': 'epa', 'label': 'EPA'},
        {'value': 'producto__clave_cnis', 'label': 'Clave CNIS'},
        {'value': 'almacen__nombre', 'label': 'Almacén'},
        {'value': 'ubicacion__codigo', 'label': 'Ubicación'},
    ]

    context = {
        'instituciones': instituciones,
        'almacenes': almacenes,
        'productos': productos,
        'estados': estados,
        'filtro_institucion': filtro_institucion,
        'filtro_estado': filtro_estado,
        'filtro_almacen': filtro_almacen,
        'filtro_producto': filtro_producto,
        'filtro_caducidad': filtro_caducidad,
        'busqueda_lote': busqueda_lote,
        'busqueda_cnis': busqueda_cnis,
        "busqueda_producto": busqueda_producto,
        "filtro_partida": filtro_partida,
        "filtro_fecha_recepcion": filtro_fecha_recepcion,
        "filtro_fecha_recepcion_desde": filtro_fecha_recepcion_desde,
        "filtro_fecha_recepcion_hasta": filtro_fecha_recepcion_hasta,
        "filtro_con_remision": filtro_con_remision,
        "filtro_con_orden_suministro": filtro_con_orden_suministro,
        "filtro_fecha_cita_desde": filtro_fecha_cita_desde,
        "filtro_fecha_cita_hasta": filtro_fecha_cita_hasta,
        "filtro_excluir_sin_rfc": filtro_excluir_sin_rfc,
        'columnas_disponibles': columnas_disponibles,
        'institucion_usuario': institucion,
    }

    return render(request, 'inventario/reportes/reporte_lotes_personalizado.html', context)


@login_required
@require_http_methods(["POST"])
def corregir_fecha_recepcion_lote(request, lote_id):
    """
    Corrige la fecha de recepción de un lote (desde la lista, para lotes con etiqueta REVISAR).
    Espera POST con 'nueva_fecha' (YYYY-MM-DD). Devuelve JSON.
    """
    lote = get_object_or_404(Lote, id=lote_id)
    nueva_fecha_str = (request.POST.get('nueva_fecha') or request.GET.get('nueva_fecha') or '').strip()
    if not nueva_fecha_str:
        return JsonResponse({'success': False, 'error': 'Falta la nueva fecha (nueva_fecha).'}, status=400)
    try:
        nueva_fecha = datetime.strptime(nueva_fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Formato de fecha inválido. Use YYYY-MM-DD.'}, status=400)
    # Opcional: acotar a rango coherente para no permitir de nuevo fechas futuras
    hoy = timezone.now().date()
    if nueva_fecha > hoy:
        return JsonResponse({
            'success': False,
            'error': f'La fecha no puede ser futura. Use hoy ({hoy.strftime("%d/%m/%Y")}) o anterior.'
        }, status=400)
    if nueva_fecha < FECHA_RECEPCION_MIN:
        return JsonResponse({
            'success': False,
            'error': f'La fecha no puede ser anterior a {FECHA_RECEPCION_MIN.strftime("%d/%m/%Y")}.'
        }, status=400)
    lote.fecha_recepcion = nueva_fecha
    lote.save(update_fields=['fecha_recepcion'])
    return JsonResponse({
        'success': True,
        'fecha_mostrar': nueva_fecha.strftime('%d/%m/%Y'),
        'es_coherente': _fecha_recepcion_es_coherente(nueva_fecha, hoy),
    })


def _remisiones_distintas_lote(lote):
    """
    Todas las remisiones distintas asociadas al lote: campo del lote, llegada que lo creó
    y cada movimiento de ENTRADA (p. ej. recepciones adicionales con otra remisión).
    """
    out = []
    seen = set()

    def agregar(txt):
        s = (txt or "").strip()
        if s and s not in seen:
            seen.add(s)
            out.append(s)

    agregar(lote.remision)
    item = ItemLlegada.objects.filter(lote_creado=lote).select_related("llegada").first()
    if item and item.llegada_id:
        agregar(item.llegada.remision)
    for mov in (
        lote.movimientos.filter(anulado=False, tipo_movimiento="ENTRADA")
        .exclude(Q(remision__isnull=True) | Q(remision__exact=""))
        .order_by("fecha_movimiento", "id")
    ):
        agregar(mov.remision)
    return out


@login_required
def detalle_lote(request, lote_id):
    """Detalle de un lote específico"""
    
    lote = get_object_or_404(
        Lote.objects.select_related('producto', 'almacen', 'ubicacion', 'institucion'),
        id=lote_id
    )
    
    # Ubicaciones del lote (con reserva real por ubicación = LoteAsignado)
    ubicaciones = list(
        LoteUbicacion.objects.filter(lote=lote).select_related('ubicacion', 'ubicacion__almacen')
    )
    lu_ids = [u.id for u in ubicaciones]
    reserva_lu = totales_reserva_activa_por_lote_ubicacion_ids(lu_ids)
    for u in ubicaciones:
        u.reserva_activa_calculada = reserva_lu.get(u.id, 0)

    # Movimientos del lote
    movimientos = lote.movimientos.all().order_by('-fecha_movimiento')
    remisiones_lote = _remisiones_distintas_lote(lote)
    
    # Información adicional
    dias_para_caducidad = lote.dias_para_caducidad
    esta_proximo_caducar = lote.esta_proximo_a_caducar
    esta_caducado = lote.esta_caducado
    
    # Usuario con perfil gestión de almacén puede editar cantidad (ajuste por almacén)
    user_groups = set(request.user.groups.values_list('name', flat=True))
    puede_editar_cantidad_almacen = (
        request.user.is_superuser
        or request.user.has_perm('inventario.change_lote')
        or any('almacen' in g.lower() or 'almacén' in g.lower() for g in user_groups)
    )
    
    context = {
        'lote': lote,
        'ubicaciones': ubicaciones,
        'movimientos': movimientos,
        'dias_para_caducidad': dias_para_caducidad,
        'esta_proximo_caducar': esta_proximo_caducar,
        'esta_caducado': esta_caducado,
        'puede_editar_cantidad_almacen': puede_editar_cantidad_almacen,
        'remisiones_lote': remisiones_lote,
    }
    
    return render(request, 'inventario/detalle_lote.html', context)


@login_required
def ajustar_cantidad_lote(request, lote_id):
    """
    Ajuste de cantidad disponible desde el detalle del lote.
    Solo usuarios con perfil de gestión de almacén (o permiso change_lote).
    Registra el movimiento como ajuste de inventario por almacén.
    """
    lote = get_object_or_404(
        Lote.objects.select_related('producto', 'institucion'),
        id=lote_id
    )
    user_groups = set(request.user.groups.values_list('name', flat=True))
    puede_editar = (
        request.user.is_superuser
        or request.user.has_perm('inventario.change_lote')
        or any('almacen' in g.lower() or 'almacén' in g.lower() for g in user_groups)
    )
    if not puede_editar:
        messages.error(request, 'No tienes permiso para ajustar la cantidad de este lote.')
        return redirect('detalle_lote', lote_id=lote_id)
    
    if request.method != 'POST':
        return redirect('detalle_lote', lote_id=lote_id)
    
    try:
        nueva_cantidad = int(request.POST.get('nueva_cantidad', ''))
    except (ValueError, TypeError):
        messages.error(request, 'La cantidad debe ser un número entero.')
        return redirect('detalle_lote', lote_id=lote_id)
    
    if nueva_cantidad < 0:
        messages.error(request, 'La cantidad no puede ser negativa.')
        return redirect('detalle_lote', lote_id=lote_id)
    
    motivo_usuario = (request.POST.get('motivo', '') or '').strip()
    motivo_movimiento = 'Ajuste de inventario por almacén.'
    if motivo_usuario:
        motivo_movimiento += f' {motivo_usuario}'
    
    cantidad_anterior = lote.cantidad_disponible
    if cantidad_anterior == nueva_cantidad:
        messages.info(request, 'La cantidad no ha cambiado.')
        return redirect('detalle_lote', lote_id=lote_id)
    
    if nueva_cantidad > cantidad_anterior:
        tipo_movimiento = 'AJUSTE_POSITIVO'
        cantidad_delta = nueva_cantidad - cantidad_anterior
    else:
        tipo_movimiento = 'AJUSTE_NEGATIVO'
        cantidad_delta = cantidad_anterior - nueva_cantidad
    
    MovimientoInventario.objects.create(
        lote=lote,
        tipo_movimiento=tipo_movimiento,
        cantidad=cantidad_delta,
        cantidad_anterior=cantidad_anterior,
        cantidad_nueva=nueva_cantidad,
        motivo=motivo_movimiento,
        usuario=request.user,
        institucion_destino=lote.institucion,
    )
    lote.cantidad_disponible = nueva_cantidad
    lote.save()
    
    messages.success(request, f'Cantidad actualizada de {cantidad_anterior} a {nueva_cantidad}. Movimiento registrado como ajuste de inventario por almacén.')
    return redirect('detalle_lote', lote_id=lote_id)


# ============================================================
# MOVIMIENTOS DE INVENTARIO
# ============================================================

def _cantidad_fisica_lote_ubicaciones(lote):
    """Suma LoteUbicacion.cantidad; si no hay filas, Lote.cantidad_disponible (misma regla que reportes)."""
    ubs = lote.ubicaciones_detalle.all()
    if ubs:
        return sum(lu.cantidad for lu in ubs)
    return int(lote.cantidad_disponible or 0)


def _registrar_salida_lote_con_movimiento(lote, cantidad_salida, usuario, motivo, observaciones_extra):
    """
    Descuenta existencias: por ubicaciones si existen (FIFO por id), alineado con surtimiento;
    si no hay ubicaciones, descuenta solo el lote. Actualiza cantidad_disponible/reservada del lote
    y crea MovimientoInventario tipo SALIDA.
    """
    obs = (observaciones_extra or '').strip()
    motivo_txt = motivo
    if obs:
        motivo_txt = f'{motivo}. Obs.: {obs}'

    with transaction.atomic():
        lote = Lote.objects.select_for_update().select_related('producto').get(pk=lote.pk)
        ubs = list(
            LoteUbicacion.objects.filter(lote=lote).select_for_update().order_by('pk')
        )

        if ubs:
            cantidad_anterior = sum(u.cantidad for u in ubs)
            if cantidad_salida > cantidad_anterior:
                raise ValueError(
                    f'Cantidad solicitada ({cantidad_salida}) mayor que la existencia física ({cantidad_anterior}).'
                )
            remaining = cantidad_salida
            for lu in ubs:
                if remaining <= 0:
                    break
                take = min(lu.cantidad, remaining)
                if take <= 0:
                    continue
                lu.cantidad = lu.cantidad - take
                lu.cantidad_reservada = max(0, lu.cantidad_reservada - take)
                lu.save(update_fields=['cantidad', 'cantidad_reservada'])
                remaining -= take

            tot_cant = (
                LoteUbicacion.objects.filter(lote=lote).aggregate(t=Sum('cantidad'))['t'] or 0
            )
            tot_res = (
                LoteUbicacion.objects.filter(lote=lote).aggregate(t=Sum('cantidad_reservada'))['t']
                or 0
            )
            Lote.objects.filter(pk=lote.pk).update(
                cantidad_disponible=tot_cant, cantidad_reservada=tot_res
            )
            cantidad_nueva = tot_cant
        else:
            cantidad_anterior = int(lote.cantidad_disponible or 0)
            if cantidad_salida > cantidad_anterior:
                raise ValueError(
                    f'Cantidad solicitada ({cantidad_salida}) mayor que disponible ({cantidad_anterior}).'
                )
            cantidad_nueva = cantidad_anterior - cantidad_salida
            Lote.objects.filter(pk=lote.pk).update(cantidad_disponible=cantidad_nueva)

        movimiento = MovimientoInventario.objects.create(
            lote=lote,
            tipo_movimiento='SALIDA',
            cantidad=cantidad_salida,
            cantidad_anterior=cantidad_anterior,
            cantidad_nueva=cantidad_nueva,
            motivo=motivo_txt,
            usuario=usuario,
            institucion_destino=lote.institucion,
        )
        return movimiento


@login_required
def registrar_salida(request):
    """Registrar salida de inventario (lista desde /lotes/ puede pasar ?lote=<id> para preseleccionar)."""

    if request.method == 'POST':
        form = ItemSalidaForm(request.POST)
        if form.is_valid():
            try:
                lote = Lote.objects.get(id=form.cleaned_data['lote_id'])
                if lote.estado != 1:
                    messages.error(request, 'El lote no está disponible para salida.')
                    return redirect(f"{reverse('registrar_salida')}?lote={lote.pk}")

                cantidad_salida = form.cleaned_data['cantidad_salida']
                fisico = _cantidad_fisica_lote_ubicaciones(lote)
                if cantidad_salida > fisico:
                    messages.error(
                        request,
                        f'Cantidad solicitada ({cantidad_salida}) mayor que la existencia disponible ({fisico}).'
                    )
                    return redirect(f"{reverse('registrar_salida')}?lote={lote.pk}")

                movimiento = _registrar_salida_lote_con_movimiento(
                    lote,
                    cantidad_salida,
                    request.user,
                    form.cleaned_data['motivo_salida'],
                    form.cleaned_data.get('observaciones', ''),
                )

                messages.success(
                    request,
                    f'Salida registrada correctamente. Movimiento ID: {movimiento.id}. '
                    f'Quedó reflejada en movimientos de inventario.',
                )
                return redirect('lista_lotes')

            except Lote.DoesNotExist:
                messages.error(request, 'Lote no encontrado')
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error al registrar salida: {str(e)}')
    else:
        form = ItemSalidaForm()

    lote_preseleccionado_id = None
    raw_lote = (request.GET.get('lote') or '').strip()
    if raw_lote.isdigit():
        lote_preseleccionado_id = int(raw_lote)
    if request.method == 'POST':
        raw_post = (request.POST.get('lote_id') or '').strip()
        if raw_post.isdigit():
            lote_preseleccionado_id = int(raw_post)

    # Lotes con existencia física > 0; siempre incluir el preseleccionado por ?lote= (p. ej. bordes de sincronización)
    lotes_base = (
        Lote.objects.filter(estado=1)
        .select_related('producto')
        .annotate(sum_ubic=Sum('ubicaciones_detalle__cantidad'))
        .annotate(
            fisico=Coalesce(
                F('sum_ubic'),
                F('cantidad_disponible'),
                output_field=IntegerField(),
            )
        )
    )
    q_disponible = Q(fisico__gt=0)
    if lote_preseleccionado_id:
        q_disponible |= Q(pk=lote_preseleccionado_id)
    lotes = lotes_base.filter(q_disponible).order_by('producto__clave_cnis', 'numero_lote')

    context = {
        'form': form,
        'lotes': lotes,
        'titulo': 'Registrar Salida de Inventario',
        'lote_preseleccionado_id': lote_preseleccionado_id,
    }

    return render(request, 'inventario/registrar_salida.html', context)


@login_required
def registrar_ajuste(request):
    """Registrar ajuste de inventario"""
    
    if request.method == 'POST':
        lote_id = request.POST.get('lote_id')
        tipo_ajuste = request.POST.get('tipo_ajuste')  # AJUSTE_POSITIVO o AJUSTE_NEGATIVO
        cantidad = int(request.POST.get('cantidad', 0))
        motivo = request.POST.get('motivo', '')
        observaciones = request.POST.get('observaciones', '')
        
        try:
            lote = Lote.objects.get(id=lote_id)
            
            # Validar cantidad para ajuste negativo
            if tipo_ajuste == 'AJUSTE_NEGATIVO' and cantidad > lote.cantidad_disponible:
                messages.error(
                    request,
                    f'Cantidad a restar ({cantidad}) mayor que disponible ({lote.cantidad_disponible})'
                )
                return redirect('registrar_ajuste')
            
            # Calcular nueva cantidad
            if tipo_ajuste == 'AJUSTE_POSITIVO':
                cantidad_nueva = lote.cantidad_disponible + cantidad
            else:
                cantidad_nueva = lote.cantidad_disponible - cantidad
            
            # Crear movimiento
            movimiento = MovimientoInventario.objects.create(
                lote=lote,
                tipo_movimiento=tipo_ajuste,
                cantidad=cantidad,
                cantidad_anterior=lote.cantidad_disponible,
                cantidad_nueva=cantidad_nueva,
                motivo=motivo,
                observaciones=observaciones,
                usuario=request.user,
                institucion=lote.institucion,
            )
            
            # Actualizar lote
            lote.cantidad_disponible = cantidad_nueva
            lote.save()
            
            messages.success(request, f'Ajuste registrado correctamente. Movimiento ID: {movimiento.id}')
            return redirect('lista_lotes')
            
        except Lote.DoesNotExist:
            messages.error(request, 'Lote no encontrado')
        except Exception as e:
            messages.error(request, f'Error al registrar ajuste: {str(e)}')
    
    # Obtener lotes
    lotes = Lote.objects.filter(estado=1).select_related('producto')
    
    context = {
        'lotes': lotes,
        'titulo': 'Registrar Ajuste de Inventario'
    }
    
    return render(request, 'inventario/registrar_ajuste.html', context)


def _movimientos_filtrados_desde_request(request):
    """
    Queryset de movimientos con los mismos criterios GET que la lista (sin paginar).
    Parámetros: tipo, fecha_desde, fecha_hasta, busqueda_lote, busqueda_clave,
    busqueda_producto, busqueda_motivo. Ignora 'page'.
    """
    institucion = request.user.institucion if hasattr(request.user, 'institucion') else None

    if institucion:
        movimientos = MovimientoInventario.objects.filter(
            Q(institucion_destino=institucion) | Q(lote__institucion=institucion)
        ).select_related(
            'lote', 'lote__producto', 'lote__institucion', 'institucion_destino', 'usuario'
        )
    else:
        movimientos = MovimientoInventario.objects.all().select_related(
            'lote', 'lote__producto', 'lote__institucion', 'institucion_destino', 'usuario'
        )

    filtro_tipo = request.GET.get('tipo', '')
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    busqueda_lote = request.GET.get('busqueda_lote', '').strip()
    busqueda_clave = request.GET.get('busqueda_clave', '').strip()
    busqueda_producto = request.GET.get('busqueda_producto', '').strip()
    busqueda_motivo = request.GET.get('busqueda_motivo', '').strip()

    if filtro_tipo:
        movimientos = movimientos.filter(tipo_movimiento=filtro_tipo)

    if filtro_fecha_desde:
        try:
            fecha = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha)
        except ValueError:
            pass

    if filtro_fecha_hasta:
        try:
            fecha = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha)
        except ValueError:
            pass

    if busqueda_lote:
        movimientos = movimientos.filter(lote__numero_lote__icontains=busqueda_lote)

    if busqueda_clave:
        movimientos = movimientos.filter(lote__producto__clave_cnis__icontains=busqueda_clave)

    if busqueda_producto:
        movimientos = movimientos.filter(lote__producto__descripcion__icontains=busqueda_producto)

    if busqueda_motivo:
        movimientos = movimientos.filter(motivo__icontains=busqueda_motivo)

    return movimientos.order_by('-fecha_movimiento')


@login_required
def lista_movimientos(request):
    """Lista de movimientos de inventario"""
    movimientos = _movimientos_filtrados_desde_request(request)

    filtro_tipo = request.GET.get('tipo', '')
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    busqueda_lote = request.GET.get('busqueda_lote', '').strip()
    busqueda_clave = request.GET.get('busqueda_clave', '').strip()
    busqueda_producto = request.GET.get('busqueda_producto', '').strip()
    busqueda_motivo = request.GET.get('busqueda_motivo', '').strip()

    from django.core.paginator import Paginator

    paginator = Paginator(movimientos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    tipos_movimiento = MovimientoInventario.TIPOS_MOVIMIENTO

    context = {
        'page_obj': page_obj,
        'tipos_movimiento': tipos_movimiento,
        'filtro_tipo': filtro_tipo,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
        'busqueda_lote': busqueda_lote,
        'busqueda_clave': busqueda_clave,
        'busqueda_producto': busqueda_producto,
        'busqueda_motivo': busqueda_motivo,
    }

    return render(request, 'inventario/lista_movimientos.html', context)


@login_required
def exportar_movimientos_excel(request):
    """Exporta a Excel los movimientos con los mismos filtros GET que la lista (sin límite de página)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    movimientos = _movimientos_filtrados_desde_request(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos'

    headers = [
        'Fecha y hora',
        'Tipo',
        'Clave CNIS',
        'Número de lote',
        'Producto',
        'CLUES',
        'Institución',
        'Cant. anterior',
        'Cant. nueva',
        'Cantidad movimiento',
        'Remisión',
        'Folio del pedido',
        'CLUES destino (pedido)',
        'Institución destino (pedido)',
        'Motivo',
        'Usuario',
        'Anulado',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for m in movimientos.iterator(chunk_size=500):
        lote = m.lote
        prod = lote.producto if lote else None
        inst = lote.institucion if lote else None
        inst_clue = getattr(inst, 'clue', '') or ''
        inst_nom = getattr(inst, 'denominacion', '') or ''
        remision = m.remision or (getattr(lote, 'remision', None) or '') or ''
        inst_dest = m.institucion_destino
        if m.mostrar_bloque_destino_pedido_lista and inst_dest:
            dest_clue = (getattr(inst_dest, 'clue', None) or '')
            dest_nom = (getattr(inst_dest, 'denominacion', None) or '')
        else:
            dest_clue = '-'
            dest_nom = '-'
        folio_pedido = m.folio_pedido_lista_movimientos
        ws.append(
            [
                m.fecha_movimiento.strftime('%d/%m/%Y %H:%M') if m.fecha_movimiento else '',
                m.get_tipo_movimiento_display(),
                (prod.clave_cnis or '') if prod else '',
                lote.numero_lote if lote else '',
                (prod.descripcion or '') if prod else '',
                inst_clue,
                inst_nom,
                m.cantidad_anterior,
                m.cantidad_nueva,
                m.cantidad,
                str(remision) if remision else '',
                folio_pedido,
                dest_clue,
                dest_nom,
                (m.motivo or '')[:5000],
                m.usuario.username if m.usuario_id else '',
                'Sí' if m.anulado else 'No',
            ]
        )

    widths = [18, 16, 14, 14, 36, 12, 28, 12, 12, 14, 14, 24, 14, 32, 40, 16, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f"movimientos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


# ============================================================
# CAMBIO DE ESTADO DE LOTE
# ============================================================

@login_required
@require_http_methods(["POST"])
def cambiar_estado_lote(request, lote_id):
    """Cambiar estado de un lote"""
    
    lote = get_object_or_404(Lote, id=lote_id)
    nuevo_estado = request.POST.get('nuevo_estado')
    motivo = request.POST.get('motivo', '')
    
    try:
        nuevo_estado = int(nuevo_estado)
        
        # Validar que el estado sea válido
        estados_validos = [choice[0] for choice in Lote.ESTADOS_CHOICES]
        if nuevo_estado not in estados_validos:
            messages.error(request, 'Estado no válido')
            return redirect('detalle_lote', lote_id=lote_id)
        
        # Actualizar lote
        lote.estado = nuevo_estado
        lote.motivo_cambio_estado = motivo
        lote.fecha_cambio_estado = timezone.now()
        lote.usuario_cambio_estado = request.user
        lote.save()
        
        # Crear movimiento de auditoría
        tipo_movimiento_map = {
            4: 'SUSPENSIÓN',
            5: 'DETERIORO',
            6: 'CADUCIDAD'
        }
        
        if nuevo_estado in tipo_movimiento_map:
            MovimientoInventario.objects.create(
                lote=lote,
                tipo_movimiento=tipo_movimiento_map[nuevo_estado],
                cantidad=0,
                cantidad_anterior=lote.cantidad_disponible,
                cantidad_nueva=lote.cantidad_disponible,
                motivo=motivo,
                usuario=request.user,
                institucion_destino=lote.institucion,
            )
        
        estado_nombre = dict(Lote.ESTADOS_CHOICES)[nuevo_estado]
        messages.success(request, f'Estado del lote actualizado a: {estado_nombre}')
        
    except Exception as e:
        messages.error(request, f'Error al cambiar estado: {str(e)}')
    
    return redirect('detalle_lote', lote_id=lote_id)



# ============================================================
# EXPORTACIÓN PERSONALIZADA DE LOTES
# ============================================================

@login_required
def exportar_lotes_personalizado(request):
    """
    Vista para exportar lotes a Excel con campos personalizados.
    Respeta los filtros aplicados en la lista.
    """
    if request.method == "POST":
        try:
            # 1️⃣ Recuperar campos seleccionados y ordenados
            campos = request.POST.getlist("columnas")  # checkboxes seleccionados
            orden_columnas = request.POST.get("orden_columnas", "")
            
            # Si hay un orden personalizado, lo respetamos
            if orden_columnas:
                orden = [c for c in orden_columnas.split(",") if c in campos]
                if orden:
                    campos = orden

            if not campos:
                return JsonResponse({"error": "No se seleccionaron columnas válidas"}, status=400)

            # 2️⃣ Obtener institución del usuario
            institucion = request.user.institucion if hasattr(request.user, 'institucion') else None
            
            # 3️⃣ Filtro base
            if institucion:
                lotes = Lote.objects.filter(institucion=institucion).select_related(
                    'producto', 'almacen', 'ubicacion', 'institucion', 'orden_suministro', 'orden_suministro__proveedor'
                ).prefetch_related('ubicaciones_detalle__ubicacion')
            else:
                lotes = Lote.objects.all().select_related(
                    'producto', 'almacen', 'ubicacion', 'institucion', 'orden_suministro', 'orden_suministro__proveedor'
                ).prefetch_related('ubicaciones_detalle__ubicacion')
            
            # 4✍⃣ Aplicar filtros (si vienen en la petición)
            filtro_institucion = request.POST.get('filtro_institucion', '')
            filtro_estado = request.POST.get('filtro_estado', '')
            filtro_almacen = request.POST.get('filtro_almacen', '')
            filtro_ubicacion = request.POST.get('filtro_ubicacion', '')
            filtro_producto = request.POST.get('filtro_producto', '')
            filtro_caducidad = request.POST.get('filtro_caducidad', '')
            busqueda_lote = request.POST.get('busqueda_lote', '')
            busqueda_cnis = request.POST.get('busqueda_cnis', '')
            busqueda_producto = request.POST.get('busqueda_producto', '')
            filtro_partida = request.POST.get('filtro_partida', '')
            filtro_fecha_recepcion = request.POST.get('filtro_fecha_recepcion', '')
            filtro_fecha_recepcion_desde = request.POST.get('filtro_fecha_recepcion_desde', '').strip()
            filtro_fecha_recepcion_hasta = request.POST.get('filtro_fecha_recepcion_hasta', '').strip()
            filtro_con_remision = request.POST.get('filtro_con_remision', '')
            filtro_con_orden_suministro = request.POST.get('filtro_con_orden_suministro', '')
            filtro_fecha_cita_desde = request.POST.get('filtro_fecha_cita_desde', '').strip()
            filtro_fecha_cita_hasta = request.POST.get('filtro_fecha_cita_hasta', '').strip()
            
            if filtro_institucion:
                lotes = lotes.filter(institucion_id=int(filtro_institucion))
            
            if filtro_estado:
                lotes = lotes.filter(estado=int(filtro_estado))
            
            if filtro_almacen:
                lotes = lotes.filter(almacen_id=int(filtro_almacen))
            
            # Filtro por ubicación también respetando asignaciones en LoteUbicacion
            if filtro_ubicacion:
                lotes = lotes.filter(ubicaciones_detalle__ubicacion_id=int(filtro_ubicacion)).distinct()
            
            if filtro_producto:
                lotes = lotes.filter(producto_id=int(filtro_producto))
            
            # Filtro de caducidad
            hoy = timezone.now().date()
            if filtro_caducidad:
                if filtro_caducidad == 'caducados':
                    lotes = lotes.filter(fecha_caducidad__lt=hoy)
                elif filtro_caducidad == 'menos_30':
                    lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=30))
                elif filtro_caducidad == 'menos_60':
                    lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=60))
                elif filtro_caducidad == 'menos_90':
                    lotes = lotes.filter(fecha_caducidad__gte=hoy, fecha_caducidad__lte=hoy + timedelta(days=90))
            
            if busqueda_lote:
                lotes = lotes.filter(numero_lote__icontains=busqueda_lote)
            
            if busqueda_cnis:
                lotes = lotes.filter(producto__clave_cnis__icontains=busqueda_cnis)
            
            if busqueda_producto:
                lotes = lotes.filter(producto__descripcion__icontains=busqueda_producto)

            if filtro_partida:
                lotes = lotes.filter(partida__icontains=filtro_partida)

            # Filtro por coherencia de fecha de recepción (igual que lista_lotes)
            if filtro_fecha_recepcion == 'coherentes':
                lotes = lotes.filter(
                    fecha_recepcion__gte=FECHA_RECEPCION_MIN,
                    fecha_recepcion__lte=hoy,
                )
            elif filtro_fecha_recepcion == 'incoherentes':
                lotes = lotes.filter(
                    Q(fecha_recepcion__gt=hoy) | Q(fecha_recepcion__lt=FECHA_RECEPCION_MIN)
                )

            if filtro_fecha_recepcion_desde:
                try:
                    fd = datetime.strptime(filtro_fecha_recepcion_desde, '%Y-%m-%d').date()
                    lotes = lotes.filter(fecha_recepcion__gte=fd)
                except ValueError:
                    pass
            if filtro_fecha_recepcion_hasta:
                try:
                    fh = datetime.strptime(filtro_fecha_recepcion_hasta, '%Y-%m-%d').date()
                    lotes = lotes.filter(fecha_recepcion__lte=fh)
                except ValueError:
                    pass

            if filtro_con_remision == '1':
                lotes = lotes.filter(remision__isnull=False).exclude(remision='')
            if filtro_con_orden_suministro == '1':
                lotes = lotes.filter(orden_suministro__isnull=False)
            if filtro_fecha_cita_desde:
                try:
                    fd = datetime.strptime(filtro_fecha_cita_desde, '%Y-%m-%d').date()
                    lotes = lotes.filter(item_llegada__llegada__cita__fecha_cita__date__gte=fd)
                except ValueError:
                    pass
            if filtro_fecha_cita_hasta:
                try:
                    fh = datetime.strptime(filtro_fecha_cita_hasta, '%Y-%m-%d').date()
                    lotes = lotes.filter(item_llegada__llegada__cita__fecha_cita__date__lte=fh)
                except ValueError:
                    pass

            # Excluir lotes sin RFC del proveedor (orden de suministro)
            excluir_sin_rfc = request.POST.get('excluir_sin_rfc_proveedor', '')
            if excluir_sin_rfc:
                lotes = lotes.filter(
                    orden_suministro__isnull=False,
                    orden_suministro__proveedor__isnull=False,
                    orden_suministro__proveedor__rfc__isnull=False,
                ).exclude(orden_suministro__proveedor__rfc='')
            
            # 5️⃣ Procesar datos manualmente para manejar ubicaciones correctamente
            datos_lista = []
            estado_map = dict(Lote.ESTADOS_CHOICES)
            # Mapeo de columnas a datos complementarios (citas, llegadas, pedidos)
            from .lote_utils import get_datos_complementarios_lote
            complementarios_map = {
                'rfc_proveedor': 'rfc_proveedor',
                'orden_suministro__proveedor__rfc': 'rfc_proveedor',
                'proveedor': 'proveedor',
                'orden_suministro__proveedor__razon_social': 'proveedor',
                'partida': 'partida',
                'contrato': 'contrato',
                'folio': 'folio',
                'subtotal': 'subtotal',
                'iva': 'iva',
                'importe_total': 'importe_total',
                'licitacion': 'licitacion',
                'pedido': 'pedido',
                'remision': 'remision',
                'responsable': 'responsable',
                'reviso': 'reviso',
                'tipo_entrega': 'tipo_entrega',
                'tipo_red': 'tipo_red',
                'orden_suministro': 'orden_suministro_numero',
            }

            reserva_por_lote = totales_reserva_activa_por_lote_ids(
                list(lotes.values_list('id', flat=True))
            )

            def _valor_para_campo(lote, campo, comp):
                if campo == 'cantidad_reservada':
                    return reserva_por_lote.get(lote.id, 0)
                if campo == 'ubicacion__codigo':
                    return None  # se asigna en el bloque por ubi
                if campo == 'almacen__nombre':
                    return lote.almacen.nombre if lote.almacen else ""
                if campo == 'producto__clave_cnis':
                    return lote.producto.clave_cnis if lote.producto else ""
                if campo == 'institucion__denominacion':
                    return lote.institucion.denominacion if lote.institucion else ""
                if campo == 'producto__descripcion':
                    return lote.producto.descripcion if lote.producto else ""
                if campo == 'estado':
                    return estado_map.get(lote.estado, str(lote.estado))
                valor = getattr(lote, campo.replace('__', '_'), None)
                if valor is None and '__' in campo:
                    partes = campo.split('__')
                    obj = lote
                    for parte in partes:
                        obj = getattr(obj, parte, None)
                        if obj is None:
                            break
                    valor = obj
                if (valor is None or (isinstance(valor, str) and not str(valor).strip())) and campo in complementarios_map:
                    valor = comp.get(complementarios_map[campo], valor)
                return valor

            for lote in lotes.order_by('-fecha_recepcion'):
                comp = get_datos_complementarios_lote(lote)
                ubicaciones = lote.ubicaciones_detalle.all()

                if ubicaciones.exists():
                    for ubi in ubicaciones:
                        registro = {}
                        for campo in campos:
                            if campo == 'ubicacion__codigo':
                                registro[campo] = ubi.ubicacion.codigo if ubi.ubicacion else ""
                            elif campo == 'cantidad_disponible':
                                # En exporte por ubicación, reflejar la existencia de esa ubicación
                                # para evitar repetir el total global del lote en cada renglón.
                                registro[campo] = int(ubi.cantidad or 0)
                            else:
                                registro[campo] = _valor_para_campo(lote, campo, comp)
                        datos_lista.append(registro)
                else:
                    registro = {}
                    for campo in campos:
                        if campo == 'ubicacion__codigo':
                            registro[campo] = ""
                        else:
                            registro[campo] = _valor_para_campo(lote, campo, comp)
                    datos_lista.append(registro)
            
            if not datos_lista:
                return JsonResponse({"error": "No hay datos para exportar"}, status=404)

            # 6️⃣ Procesar campos legibles
            for registro in datos_lista:
                # Fechas legibles
                for k, v in registro.items():
                    if isinstance(v, Decimal):
                        registro[k] = float(v)
                    elif isinstance(v, (date, datetime)):
                        registro[k] = v.strftime("%Y-%m-%d %H:%M") if isinstance(v, datetime) else v.strftime("%Y-%m-%d")
                    elif v is None:
                        registro[k] = ""

            # 7️⃣ Exportar a Excel respetando el orden de columnas
            df = pd.DataFrame(datos_lista)

            # Si hay orden definido, reordenamos las columnas
            columnas_finales = [col for col in campos if col in df.columns]
            df = df[columnas_finales]

            # Generar Excel
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="lotes.xlsx"'
            df.to_excel(response, index=False, sheet_name='Lotes')

            return response

        except Exception as e:
            return JsonResponse({"error": f"Error generando el reporte: {str(e)}"}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)
