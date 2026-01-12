"""
Reporte de Conteo de Almacén

Muestra un resumen de productos con sus cantidades, importes y cifras de conteos.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Q
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

from .models import Lote, LoteUbicacion, RegistroConteoFisico, Producto


@login_required
def reporte_conteo_almacen(request):
    """
    Reporte de conteo de almacén con layout específico.
    Muestra: Consecutivo, Fuente de Financiamiento, Clave, Descripción, 
    Unidad de Medida, Cantidad, Importe, Cifra Primer Conteo, 
    Cifra Segundo Conteo, Tercer Conteo
    """
    
    # Obtener filtros
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    filtro_almacen = request.GET.get('almacen', '')
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'almacen'
    ).all()
    
    # Aplicar filtro de almacén
    if filtro_almacen:
        lotes = lotes.filter(almacen__nombre=filtro_almacen)
    
    # Agrupar por producto
    productos_dict = defaultdict(lambda: {
        'clave_cnis': '',
        'descripcion': '',
        'unidad_medida': '',
        'cantidad_total': 0,
        'importe_total': 0,
        'cifra_primer_conteo': 0,
        'cifra_segundo_conteo': 0,
        'cifra_tercer_conteo': 0,
        'lotes': []
    })
    
    for lote in lotes:
        if not lote.producto:
            continue
        
        clave = lote.producto.clave_cnis
        
        # Llenar información del producto
        productos_dict[clave]['clave_cnis'] = clave
        productos_dict[clave]['descripcion'] = lote.producto.descripcion
        productos_dict[clave]['unidad_medida'] = lote.producto.unidad_medida or 'PIEZA'
        
        # Sumar cantidad e importe
        cantidad_lote = lote.cantidad_disponible
        importe_lote = cantidad_lote * (lote.precio_unitario or 0)
        
        productos_dict[clave]['cantidad_total'] += cantidad_lote
        productos_dict[clave]['importe_total'] += importe_lote
        productos_dict[clave]['lotes'].append(lote)
    
    # Obtener conteos
    conteos = RegistroConteoFisico.objects.select_related(
        'lote',
        'lote__producto'
    ).all()
    
    # Aplicar filtro de fechas a conteos
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            conteos = conteos.filter(fecha_conteo__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            fecha_hasta = fecha_hasta + timedelta(days=1)
            conteos = conteos.filter(fecha_conteo__date__lt=fecha_hasta)
        except ValueError:
            pass
    
    # Agrupar conteos por producto
    conteos_dict = defaultdict(lambda: {
        'primer_conteo': 0,
        'segundo_conteo': 0,
        'tercer_conteo': 0
    })
    
    for conteo in conteos:
        if not conteo.lote or not conteo.lote.producto:
            continue
        
        clave = conteo.lote.producto.clave_cnis
        
        if conteo.numero_conteo == 1:
            conteos_dict[clave]['primer_conteo'] += conteo.cantidad_conteo
        elif conteo.numero_conteo == 2:
            conteos_dict[clave]['segundo_conteo'] += conteo.cantidad_conteo
        elif conteo.numero_conteo == 3:
            conteos_dict[clave]['tercer_conteo'] += conteo.cantidad_conteo
    
    # Combinar datos
    reporte_data = []
    consecutivo = 1
    
    for clave in sorted(productos_dict.keys()):
        producto_info = productos_dict[clave]
        conteo_info = conteos_dict.get(clave, {
            'primer_conteo': 0,
            'segundo_conteo': 0,
            'tercer_conteo': 0
        })
        
        reporte_data.append({
            'consecutivo': consecutivo,
            'fuente_financiamiento': 'U013',  # Por el momento fijo
            'clave_cnis': producto_info['clave_cnis'],
            'descripcion': producto_info['descripcion'],
            'unidad_medida': producto_info['unidad_medida'],
            'cantidad': producto_info['cantidad_total'],
            'importe': producto_info['importe_total'],
            'cifra_primer_conteo': conteo_info.get('primer_conteo', 0),
            'cifra_segundo_conteo': conteo_info.get('segundo_conteo', 0),
            'cifra_tercer_conteo': conteo_info.get('tercer_conteo', 0),
        })
        
        consecutivo += 1
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(reporte_data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener almacenes
    from .models import Almacen
    almacenes = Almacen.objects.all().order_by('nombre')
    
    # Calcular totales
    total_cantidad = sum(r['cantidad'] for r in reporte_data)
    total_importe = sum(r['importe'] for r in reporte_data)
    total_primer_conteo = sum(r['cifra_primer_conteo'] for r in reporte_data)
    total_segundo_conteo = sum(r['cifra_segundo_conteo'] for r in reporte_data)
    total_tercer_conteo = sum(r['cifra_tercer_conteo'] for r in reporte_data)
    
    context = {
        'page_obj': page_obj,
        'total_registros': len(reporte_data),
        'total_cantidad': total_cantidad,
        'total_importe': total_importe,
        'total_primer_conteo': total_primer_conteo,
        'total_segundo_conteo': total_segundo_conteo,
        'total_tercer_conteo': total_tercer_conteo,
        'almacenes': almacenes,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
        'filtro_almacen': filtro_almacen,
        'reporte_data': reporte_data,  # Para exportación
    }
    
    return render(request, 'inventario/reporte_conteo_almacen.html', context)


@login_required
def exportar_conteo_almacen_excel(request):
    """
    Exporta el reporte de conteo de almacén a Excel.
    """
    
    # Obtener filtros
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    filtro_almacen = request.GET.get('almacen', '')
    
    # Obtener todos los lotes
    lotes = Lote.objects.select_related(
        'producto',
        'institucion',
        'almacen'
    ).all()
    
    # Aplicar filtro de almacén
    if filtro_almacen:
        lotes = lotes.filter(almacen__nombre=filtro_almacen)
    
    # Agrupar por producto
    productos_dict = defaultdict(lambda: {
        'clave_cnis': '',
        'descripcion': '',
        'unidad_medida': '',
        'cantidad_total': 0,
        'importe_total': 0,
        'cifra_primer_conteo': 0,
        'cifra_segundo_conteo': 0,
        'cifra_tercer_conteo': 0,
    })
    
    for lote in lotes:
        if not lote.producto:
            continue
        
        clave = lote.producto.clave_cnis
        
        # Llenar información del producto
        productos_dict[clave]['clave_cnis'] = clave
        productos_dict[clave]['descripcion'] = lote.producto.descripcion
        productos_dict[clave]['unidad_medida'] = lote.producto.unidad_medida or 'PIEZA'
        
        # Sumar cantidad e importe
        cantidad_lote = lote.cantidad_disponible
        importe_lote = cantidad_lote * (lote.precio_unitario or 0)
        
        productos_dict[clave]['cantidad_total'] += cantidad_lote
        productos_dict[clave]['importe_total'] += importe_lote
    
    # Obtener conteos
    conteos = RegistroConteoFisico.objects.select_related(
        'lote',
        'lote__producto'
    ).all()
    
    # Aplicar filtro de fechas a conteos
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            conteos = conteos.filter(fecha_conteo__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            fecha_hasta = fecha_hasta + timedelta(days=1)
            conteos = conteos.filter(fecha_conteo__date__lt=fecha_hasta)
        except ValueError:
            pass
    
    # Agrupar conteos por producto
    conteos_dict = defaultdict(lambda: {
        'primer_conteo': 0,
        'segundo_conteo': 0,
        'tercer_conteo': 0
    })
    
    for conteo in conteos:
        if not conteo.lote or not conteo.lote.producto:
            continue
        
        clave = conteo.lote.producto.clave_cnis
        
        if conteo.numero_conteo == 1:
            conteos_dict[clave]['primer_conteo'] += conteo.cantidad_conteo
        elif conteo.numero_conteo == 2:
            conteos_dict[clave]['segundo_conteo'] += conteo.cantidad_conteo
        elif conteo.numero_conteo == 3:
            conteos_dict[clave]['tercer_conteo'] += conteo.cantidad_conteo
    
    # Combinar datos
    reporte_data = []
    consecutivo = 1
    
    for clave in sorted(productos_dict.keys()):
        producto_info = productos_dict[clave]
        conteo_info = conteos_dict.get(clave, {
            'primer_conteo': 0,
            'segundo_conteo': 0,
            'tercer_conteo': 0
        })
        
        reporte_data.append({
            'consecutivo': consecutivo,
            'fuente_financiamiento': 'U013',
            'clave_cnis': producto_info['clave_cnis'],
            'descripcion': producto_info['descripcion'],
            'unidad_medida': producto_info['unidad_medida'],
            'cantidad': producto_info['cantidad_total'],
            'importe': producto_info['importe_total'],
            'cifra_primer_conteo': conteo_info.get('primer_conteo', 0),
            'cifra_segundo_conteo': conteo_info.get('segundo_conteo', 0),
            'cifra_tercer_conteo': conteo_info.get('tercer_conteo', 0),
        })
        
        consecutivo += 1
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo Almacén"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'CONSECUTIVO',
        'FUENTE DE FINANCIAMIENTO',
        'CLAVE',
        'DESCRIPCIÓN',
        'UNIDAD DE MEDIDA',
        'CANTIDAD',
        'IMPORTE',
        'CIFRA PRIMER CONTEO',
        'CIFRA SEGUNDO CONTEO',
        'TERCER CONTEO'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    total_cantidad = 0
    total_importe = 0
    total_primer_conteo = 0
    total_segundo_conteo = 0
    total_tercer_conteo = 0
    
    for row, item in enumerate(reporte_data, 2):
        total_cantidad += item['cantidad']
        total_importe += item['importe']
        total_primer_conteo += item['cifra_primer_conteo']
        total_segundo_conteo += item['cifra_segundo_conteo']
        total_tercer_conteo += item['cifra_tercer_conteo']
        
        ws.cell(row=row, column=1).value = item['consecutivo']
        ws.cell(row=row, column=2).value = item['fuente_financiamiento']
        ws.cell(row=row, column=3).value = item['clave_cnis']
        ws.cell(row=row, column=4).value = item['descripcion']
        ws.cell(row=row, column=5).value = item['unidad_medida']
        ws.cell(row=row, column=6).value = item['cantidad']
        ws.cell(row=row, column=7).value = item['importe']
        ws.cell(row=row, column=8).value = item['cifra_primer_conteo']
        ws.cell(row=row, column=9).value = item['cifra_segundo_conteo']
        ws.cell(row=row, column=10).value = item['cifra_tercer_conteo']
        
        # Aplicar bordes
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in [6, 7, 8, 9, 10]:  # Columnas numéricas
                cell.alignment = Alignment(horizontal='right')
    
    # Fila de totales
    total_row = len(reporte_data) + 2
    ws.cell(row=total_row, column=1).value = "TOTALES"
    ws.cell(row=total_row, column=1).font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    
    ws.cell(row=total_row, column=6).value = total_cantidad
    ws.cell(row=total_row, column=6).font = total_font
    ws.cell(row=total_row, column=6).fill = total_fill
    
    ws.cell(row=total_row, column=7).value = total_importe
    ws.cell(row=total_row, column=7).font = total_font
    ws.cell(row=total_row, column=7).fill = total_fill
    
    ws.cell(row=total_row, column=8).value = total_primer_conteo
    ws.cell(row=total_row, column=8).font = total_font
    ws.cell(row=total_row, column=8).fill = total_fill
    
    ws.cell(row=total_row, column=9).value = total_segundo_conteo
    ws.cell(row=total_row, column=9).font = total_font
    ws.cell(row=total_row, column=9).fill = total_fill
    
    ws.cell(row=total_row, column=10).value = total_tercer_conteo
    ws.cell(row=total_row, column=10).font = total_font
    ws.cell(row=total_row, column=10).fill = total_fill
    
    # Aplicar bordes a la fila de totales
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_conteo_almacen.xlsx"'
    wb.save(response)
    
    return response
