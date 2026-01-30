"""
Vistas para Reportes de Salidas y Análisis
Basados en MovimientoInventario generados por Fase 5
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q, F, DecimalField
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import timedelta, datetime
import json
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    MovimientoInventario, Lote, Institucion, Almacen, Producto, LoteUbicacion
)
from .pedidos_models import PropuestaPedido, ItemPropuesta, LoteAsignado, SolicitudPedido
from .decorators_roles import requiere_rol
from .propuesta_utils import liberar_cantidad_lote
from django.db import transaction

logger = logging.getLogger(__name__)


# ============================================================
# REPORTE GENERAL DE SALIDAS
# ============================================================

@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista')
def reporte_general_salidas(request):
    """Reporte general de salidas basado en MovimientoInventario"""
    
    # Obtener institución del usuario
    institucion = request.user.almacen.institucion if request.user.almacen else None
    if not institucion:
        messages.error(request, 'No tienes una institución asignada.')
        return redirect('dashboard')
    
    # Filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Consulta base: Movimientos de tipo SALIDA de la institución
    movimientos = MovimientoInventario.objects.filter(
        tipo_movimiento='SALIDA',
        lote__institucion=institucion
    )
    
    # Aplicar filtros de fecha
    if fecha_inicio:
        movimientos = movimientos.filter(fecha_movimiento__gte=fecha_inicio)
    if fecha_fin:
        movimientos = movimientos.filter(fecha_movimiento__lte=fecha_fin)
    
    # Estadísticas
    total_salidas = movimientos.count()
    total_cantidad = movimientos.aggregate(Sum('cantidad'))['cantidad__sum'] or 0
    total_monto = movimientos.aggregate(
        total=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField())
    )['total'] or 0
    
    # Salidas por almacén
    salidas_por_almacen = movimientos.values('lote__almacen__nombre').annotate(
        cantidad_movimientos=Count('id'),
        total_cantidad=Sum('cantidad'),
        total_monto=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField())
    ).order_by('-total_cantidad')
    
    # Top 10 productos más salidos
    top_productos = movimientos.values('lote__producto__descripcion').annotate(
        total_cantidad=Sum('cantidad'),
        total_monto=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField()),
        cantidad_movimientos=Count('id')
    ).order_by('-total_cantidad')[:10]
    
    context = {
        'total_salidas': total_salidas,
        'total_cantidad': total_cantidad,
        'total_monto': total_monto,
        'promedio_cantidad': total_cantidad / total_salidas if total_salidas > 0 else 0,
        'salidas_por_almacen': salidas_por_almacen,
        'top_productos': top_productos,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    
    return render(request, 'inventario/reportes_salidas/reporte_general.html', context)


# ============================================================
# ANÁLISIS DE DISTRIBUCIONES
# ============================================================

@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista')
def analisis_distribuciones(request):
    """Análisis de distribuciones basado en MovimientoInventario"""
    
    # Obtener institución del usuario
    institucion = request.user.almacen.institucion if request.user.almacen else None
    if not institucion:
        messages.error(request, 'No tienes una institución asignada.')
        return redirect('dashboard')
    
    # Filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Consulta base
    movimientos = MovimientoInventario.objects.filter(
        tipo_movimiento='SALIDA',
        lote__institucion=institucion
    )
    
    # Aplicar filtros de fecha
    if fecha_inicio:
        movimientos = movimientos.filter(fecha_movimiento__gte=fecha_inicio)
    if fecha_fin:
        movimientos = movimientos.filter(fecha_movimiento__lte=fecha_fin)
    
    # Estadísticas
    total_movimientos = movimientos.count()
    total_cantidad = movimientos.aggregate(Sum('cantidad'))['cantidad__sum'] or 0
    total_monto = movimientos.aggregate(
        total=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField())
    )['total'] or 0
    
    # Distribuciones por almacén origen
    distribuciones_almacen = movimientos.values('lote__almacen__nombre').annotate(
        cantidad_movimientos=Count('id'),
        total_cantidad=Sum('cantidad'),
        total_monto=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField())
    ).order_by('-total_cantidad')
    
    # Análisis por motivo (propuesta)
    distribuciones_propuesta = movimientos.values('motivo').annotate(
        cantidad_movimientos=Count('id'),
        total_cantidad=Sum('cantidad'),
        total_monto=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField())
    ).order_by('-total_cantidad')[:10]
    
    context = {
        'total_movimientos': total_movimientos,
        'total_cantidad': total_cantidad,
        'total_monto': total_monto,
        'distribuciones_almacen': distribuciones_almacen,
        'distribuciones_propuesta': distribuciones_propuesta,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    
    return render(request, 'inventario/reportes_salidas/analisis_distribuciones.html', context)


# ============================================================
# ANÁLISIS TEMPORAL
# ============================================================

@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista')
def analisis_temporal(request):
    """Análisis temporal de salidas (últimos 30 días)"""
    
    # Obtener institución del usuario
    institucion = request.user.almacen.institucion if request.user.almacen else None
    if not institucion:
        messages.error(request, 'No tienes una institución asignada.')
        return redirect('dashboard')
    
    # Últimos 30 días
    fecha_inicio = timezone.now() - timedelta(days=30)
    
    # Consulta base
    movimientos = MovimientoInventario.objects.filter(
        tipo_movimiento='SALIDA',
        lote__institucion=institucion,
        fecha_movimiento__gte=fecha_inicio
    )
    
    # Datos temporales por día
    datos_temporales = []
    for i in range(30, -1, -1):
        fecha = timezone.now() - timedelta(days=i)
        fecha_inicio_dia = fecha.replace(hour=0, minute=0, second=0, microsecond=0)
        fecha_fin_dia = fecha.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        movimientos_dia = movimientos.filter(
            fecha_movimiento__gte=fecha_inicio_dia,
            fecha_movimiento__lte=fecha_fin_dia
        )
        
        cantidad_movimientos = movimientos_dia.count()
        total_cantidad = movimientos_dia.aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        total_monto = movimientos_dia.aggregate(
            total=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField())
        )['total'] or 0
        
        if cantidad_movimientos > 0:
            datos_temporales.append({
                'fecha': fecha.strftime('%d/%m/%Y'),
                'cantidad': cantidad_movimientos,
                'items': total_cantidad,
                'monto': float(total_monto) if total_monto else 0
            })
    
    # Estadísticas generales
    total_movimientos = movimientos.count()
    total_cantidad = movimientos.aggregate(Sum('cantidad'))['cantidad__sum'] or 0
    total_monto = movimientos.aggregate(
        total=Sum(F('cantidad') * F('lote__precio_unitario'), output_field=DecimalField())
    )['total'] or 0
    
    context = {
        'datos_temporales': datos_temporales,
        'total_movimientos': total_movimientos,
        'total_cantidad': total_cantidad,
        'total_monto': total_monto,
        'promedio_diario': total_movimientos / 30 if total_movimientos > 0 else 0,
    }
    
    return render(request, 'inventario/reportes_salidas/analisis_temporal.html', context)


# ============================================================
# REPORTE DE SALIDAS - ÓRDENES DE SURTIMIENTO SURTIDAS
# ============================================================

@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista', 'Supervisor')
def reporte_salidas_surtidas(request):
    """
    Reporte detallado de órdenes de surtimiento ya surtidas.
    Muestra información completa de cada item surtido.
    """
    # Filtros (filtro_lote para no sobrescribir con la variable "lote" del bucle)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    folio = request.GET.get('folio', '').strip()
    clave_cnis = request.GET.get('clave_cnis', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    institucion_id = request.GET.get('institucion')
    estatus_movimiento = request.GET.get('estatus_movimiento', '').strip()
    
    # Obtener propuestas surtidas con sus relaciones (filtro_lote evita sobrescribir variable en el bucle)
    propuestas = PropuestaPedido.objects.filter(
        estado='SURTIDA',
        fecha_surtimiento__isnull=False
    ).select_related(
        'solicitud',
        'solicitud__institucion_solicitante',
        'solicitud__almacen_destino',
        'usuario_surtimiento'
    ).prefetch_related(
        'items__producto',
        'items__item_solicitud',
        'items__lotes_asignados__lote_ubicacion__lote',
        'items__lotes_asignados__lote_ubicacion__ubicacion'
    )
    
    # Aplicar filtros
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            propuestas = propuestas.filter(fecha_surtimiento__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            # Incluir todo el día
            from datetime import time as dt_time
            fecha_fin_obj = datetime.combine(fecha_fin_obj.date(), dt_time.max)
            propuestas = propuestas.filter(fecha_surtimiento__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    if folio:
        folio = folio.strip()
        if folio:
            propuestas = propuestas.filter(
                Q(solicitud__folio__icontains=folio) |
                Q(solicitud__observaciones_solicitud__icontains=folio)
            )
    
    if clave_cnis:
        clave_cnis = clave_cnis.strip()
        if clave_cnis:
            # Filtrar propuestas que tengan items con productos que coincidan con la clave CNIS
            propuestas = propuestas.filter(
                items__producto__clave_cnis__icontains=clave_cnis
            ).distinct()
    
    if filtro_lote:
        propuestas = propuestas.filter(
            items__lotes_asignados__surtido=True,
            items__lotes_asignados__lote_ubicacion__lote__numero_lote__icontains=filtro_lote
        ).distinct()
    
    if institucion_id:
        try:
            institucion_id = int(institucion_id)
            propuestas = propuestas.filter(
                solicitud__institucion_solicitante_id=institucion_id
            )
        except (ValueError, TypeError):
            pass
    
    # Construir datos del reporte
    datos_reporte = []
    partida_counter = 1
    
    # Normalizar filtros para comparación (guardar valores originales antes de modificar)
    clave_cnis_filter_value = clave_cnis.strip().upper() if clave_cnis else None
    lote_filter_value = filtro_lote.strip().upper() if filtro_lote else None
    
    for propuesta in propuestas.order_by('-fecha_surtimiento', 'solicitud__folio'):
        solicitud = propuesta.solicitud
        
        for item_propuesta in propuesta.items.all():
            producto = item_propuesta.producto
            
            # Aplicar filtro de clave_cnis a nivel de item
            if clave_cnis_filter_value:
                if not producto.clave_cnis or clave_cnis_filter_value not in producto.clave_cnis.upper():
                    continue
            
            # Obtener lotes asignados surtidos
            lotes_surtidos = item_propuesta.lotes_asignados.filter(surtido=True)
            
            if not lotes_surtidos.exists():
                continue
            
            for lote_asignado in lotes_surtidos:
                lote_ubicacion = lote_asignado.lote_ubicacion
                lote = lote_ubicacion.lote
                ubicacion = lote_ubicacion.ubicacion
                
                # Aplicar filtro de lote a nivel de lote individual
                if lote_filter_value:
                    if not lote.numero_lote or lote_filter_value not in lote.numero_lote.upper():
                        continue
                
                # Obtener movimiento de inventario relacionado (si existe, para remision_ingreso)
                movimiento = MovimientoInventario.objects.filter(
                    lote=lote,
                    tipo_movimiento='SALIDA',
                    folio=str(propuesta.id) if propuesta.id else ''
                ).filter(cantidad=lote_asignado.cantidad_asignada).first()
                if not movimiento:
                    movimiento = MovimientoInventario.objects.filter(
                        lote=lote,
                        tipo_movimiento='SALIDA',
                        folio=str(propuesta.id) if propuesta.id else ''
                    ).first()
                
                # Cantidad previa = cantidad_disponible (inventario_lote) + cantidad surtida
                cantidad_surtida = lote_asignado.cantidad_asignada
                cantidad_disponible_lote = getattr(lote, 'cantidad_disponible', None) or 0
                cantidad_previa = cantidad_disponible_lote + cantidad_surtida
                
                # Obtener orden de suministro del lote (si existe)
                orden_reposicion = lote.orden_suministro.numero_orden if lote.orden_suministro else ''
                
                # Calcular días para caducidad
                if lote.fecha_caducidad:
                    dias_caducidad = (lote.fecha_caducidad - timezone.now().date()).days
                else:
                    dias_caducidad = None
                
                tiene_movimiento = movimiento is not None
                datos_reporte.append({
                    'partida': partida_counter,
                    'clave_cnis': producto.clave_cnis,
                    'descripcion': producto.descripcion,
                    'unidad_medida': producto.unidad_medida if producto.unidad_medida else '',
                    'lote': lote.numero_lote,
                    'lote_id': lote.pk,
                    'lote_ubicacion_id': lote_ubicacion.pk,
                    'propuesta_id': propuesta.id,
                    'caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else '',
                    'dias_caducidad': dias_caducidad,
                    'cantidad_solicitada': item_propuesta.cantidad_solicitada,
                    'cantidad_disponible': cantidad_disponible_lote,
                    'cantidad_previa': cantidad_previa,
                    'cantidad_surtida': lote_asignado.cantidad_asignada,
                    'observaciones': solicitud.observaciones_solicitud or '',
                    'recurso': solicitud.institucion_solicitante.nombre if solicitud.institucion_solicitante else '',
                    'destino': (solicitud.almacen_destino.institucion.nombre or solicitud.almacen_destino.institucion.denominacion) if solicitud.almacen_destino and solicitud.almacen_destino.institucion else '',
                    'ubicacion': ubicacion.codigo if ubicacion else '',
                    'fecha_captura': solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_solicitud else '',
                    'folio': solicitud.observaciones_solicitud or solicitud.folio,
                    'fecha_entrega_programada': solicitud.fecha_entrega_programada.strftime('%d/%m/%Y') if solicitud.fecha_entrega_programada else '',
                    'status': propuesta.get_estado_display(),
                    'tiene_movimiento': tiene_movimiento,
                    'remision_ingreso': movimiento.remision if movimiento else '',
                    'orden_reposicion': orden_reposicion,
                    'usuario': propuesta.usuario_surtimiento.get_full_name() if propuesta.usuario_surtimiento else propuesta.usuario_surtimiento.username if propuesta.usuario_surtimiento else '',
                })
                
                partida_counter += 1
    
    total_sin_movimiento = sum(1 for d in datos_reporte if not d.get('tiene_movimiento', True))
    
    # Filtro por estatus movimiento (EST. MOV.)
    if estatus_movimiento == 'sin_movimiento':
        datos_reporte = [d for d in datos_reporte if not d.get('tiene_movimiento')]
    elif estatus_movimiento == 'con_movimiento':
        datos_reporte = [d for d in datos_reporte if d.get('tiene_movimiento')]
    # Renumerar partidas después del filtro
    for i, d in enumerate(datos_reporte, 1):
        d['partida'] = i
    
    # Paginación
    paginator = Paginator(datos_reporte, 50)
    page = request.GET.get('page', 1)
    try:
        datos_paginados = paginator.page(page)
    except PageNotAnInteger:
        datos_paginados = paginator.page(1)
    except EmptyPage:
        datos_paginados = paginator.page(paginator.num_pages)
    
    # Obtener instituciones para el filtro
    instituciones = Institucion.objects.all().order_by('denominacion')
    
    # Query string sin 'page' para que los enlaces de paginación conserven los filtros
    get_copy = request.GET.copy()
    if 'page' in get_copy:
        get_copy.pop('page')
    query_string_sin_page = get_copy.urlencode()
    
    context = {
        'datos': datos_paginados,
        'total_registros': len(datos_reporte),
        'total_sin_movimiento': total_sin_movimiento,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'folio': folio,
        'clave_cnis': clave_cnis,
        'lote': filtro_lote,
        'institucion_id': institucion_id,
        'instituciones': instituciones,
        'estatus_movimiento': estatus_movimiento,
        'query_string_sin_page': query_string_sin_page,
    }
    
    return render(request, 'inventario/reportes_salidas/reporte_salidas_surtidas.html', context)


@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista', 'Supervisor')
@require_http_methods(["GET", "POST"])
def aplicar_salida_surtida(request):
    """
    Aplica la salida para un registro surtido que no tiene movimiento de inventario:
    descontar de cantidad disponible del lote, rebajar cantidad reservada y crear
    MovimientoInventario con motivo "Ajuste por sistema".
    """
    if request.method != 'POST':
        logger.info("aplicar_salida_surtida: método no POST, redirigiendo")
        return redirect(reverse('reportes_salidas:reporte_salidas_surtidas'))

    propuesta_id = request.POST.get('propuesta_id')
    lote_id = request.POST.get('lote_id')
    lote_ubicacion_id = request.POST.get('lote_ubicacion_id')
    cantidad = request.POST.get('cantidad')
    return_url = request.POST.get('return_url', '').strip()

    logger.info(
        "aplicar_salida_surtida: POST recibido | propuesta_id=%s lote_id=%s lote_ubicacion_id=%s cantidad=%s return_url=%s",
        propuesta_id, lote_id, lote_ubicacion_id, cantidad, return_url[:80] if return_url else ""
    )

    if not all([propuesta_id, lote_id, lote_ubicacion_id, cantidad]):
        logger.warning("aplicar_salida_surtida: faltan datos en POST")
        messages.error(request, 'Faltan datos para aplicar la salida (propuesta, lote, ubicación o cantidad).')
        return redirect(return_url or reverse('reportes_salidas:reporte_salidas_surtidas'))

    try:
        cantidad = int(cantidad)
        if cantidad <= 0:
            raise ValueError('La cantidad debe ser mayor que cero.')
    except (ValueError, TypeError) as e:
        logger.warning("aplicar_salida_surtida: cantidad inválida cantidad=%s error=%s", cantidad, e)
        messages.error(request, 'Cantidad inválida.')
        return redirect(return_url or reverse('reportes_salidas:reporte_salidas_surtidas'))

    # Cargar LoteUbicacion con lote, producto (clave) y ubicación para afectar y auditar por (clave, lote, ubicación)
    lote_ubicacion = get_object_or_404(
        LoteUbicacion.objects.select_related('lote__producto', 'ubicacion'),
        pk=lote_ubicacion_id
    )
    lote = lote_ubicacion.lote
    ubicacion = lote_ubicacion.ubicacion
    codigo_ubicacion = ubicacion.codigo if ubicacion else ''
    clave_cnis = (lote.producto.clave_cnis or '') if lote.producto else ''

    logger.info(
        "aplicar_salida_surtida: LoteUbicacion cargado | pk=%s lote_id=%s lote_numero=%s ubicacion=%s clave_cnis=%s "
        "cantidad_actual=%s cantidad_reservada_actual=%s",
        lote_ubicacion.pk, lote.pk, lote.numero_lote, codigo_ubicacion, clave_cnis,
        lote_ubicacion.cantidad, getattr(lote_ubicacion, 'cantidad_reservada', None)
    )
    logger.info(
        "aplicar_salida_surtida: Lote cargado | lote_id=%s cantidad_disponible=%s cantidad_reservada=%s",
        lote.pk, lote.cantidad_disponible, getattr(lote, 'cantidad_reservada', None)
    )

    if str(lote.pk) != str(lote_id):
        logger.warning("aplicar_salida_surtida: lote_id no coincide lote.pk=%s lote_id_POST=%s", lote.pk, lote_id)
        messages.error(request, 'El lote no coincide con la ubicación.')
        return redirect(return_url or reverse('reportes_salidas:reporte_salidas_surtidas'))

    propuesta = get_object_or_404(
        PropuestaPedido.objects.select_related('solicitud'),
        id=propuesta_id
    )
    solicitud = propuesta.solicitud

    with transaction.atomic():
        # Afectación exclusiva de esta ubicación (lote_ubicacion = par lote+ubicación)
        cantidad_anterior_ubicacion = lote_ubicacion.cantidad
        cantidad_nueva_ubicacion = cantidad_anterior_ubicacion - cantidad
        if cantidad_nueva_ubicacion < 0:
            logger.warning(
                "aplicar_salida_surtida: cantidad insuficiente en ubicación | lote=%s ubicacion=%s "
                "cantidad_anterior_ubicacion=%s cantidad_a_descontar=%s",
                lote.numero_lote, codigo_ubicacion, cantidad_anterior_ubicacion, cantidad
            )
            messages.error(
                request,
                f'Cantidad insuficiente en lote {lote.numero_lote}, ubicación {codigo_ubicacion}. '
                f'Disponible en esa ubicación: {cantidad_anterior_ubicacion}, a descontar: {cantidad}.'
            )
            return redirect(return_url or reverse('reportes_salidas:reporte_salidas_surtidas'))

        cantidad_anterior_lote = lote.cantidad_disponible or 0
        cantidad_nueva_lote = cantidad_anterior_lote - cantidad
        if cantidad_nueva_lote < 0:
            logger.warning(
                "aplicar_salida_surtida: cantidad insuficiente a nivel lote | lote_id=%s cantidad_disponible=%s cantidad=%s",
                lote.pk, cantidad_anterior_lote, cantidad
            )
            messages.error(request, f'La cantidad disponible del lote ({cantidad_anterior_lote}) es menor que la cantidad a descontar ({cantidad}).')
            return redirect(return_url or reverse('reportes_salidas:reporte_salidas_surtidas'))

        reserva_antes_ubicacion = lote_ubicacion.cantidad_reservada or 0
        reserva_nueva_ubicacion = max(0, reserva_antes_ubicacion - cantidad)

        # Descontar solo en esta ubicación y rebajar reserva de esta ubicación
        lote_ubicacion.cantidad = cantidad_nueva_ubicacion
        lote_ubicacion.cantidad_reservada = reserva_nueva_ubicacion
        logger.info(
            "aplicar_salida_surtida: ANTES save LoteUbicacion | pk=%s cantidad %s -> %s cantidad_reservada %s -> %s",
            lote_ubicacion.pk, cantidad_anterior_ubicacion, cantidad_nueva_ubicacion, reserva_antes_ubicacion, reserva_nueva_ubicacion
        )
        lote_ubicacion.save(update_fields=['cantidad', 'cantidad_reservada'])
        logger.info("aplicar_salida_surtida: LoteUbicacion guardado pk=%s", lote_ubicacion.pk)

        # Recalcular cantidad disponible del lote y rebajar cantidad reservada a nivel lote
        cantidad_total_ubicaciones = sum(lu.cantidad for lu in lote.ubicaciones_detalle.all())
        reserva_antes_lote = lote.cantidad_reservada or 0
        reserva_nueva_lote = max(0, reserva_antes_lote - cantidad)
        lote.cantidad_disponible = cantidad_total_ubicaciones
        lote.cantidad_reservada = reserva_nueva_lote
        logger.info(
            "aplicar_salida_surtida: ANTES save Lote | lote_id=%s cantidad_disponible %s -> %s (suma_ubicaciones=%s) "
            "cantidad_reservada %s -> %s",
            lote.pk, cantidad_anterior_lote, cantidad_total_ubicaciones, cantidad_total_ubicaciones, reserva_antes_lote, reserva_nueva_lote
        )
        lote.save(update_fields=['cantidad_disponible', 'cantidad_reservada'])
        logger.info("aplicar_salida_surtida: Lote guardado lote_id=%s", lote.pk)

        # Registrar movimiento con leyenda "Ajuste por sistema" y tripleta (clave, lote, ubicación)
        motivo_mov = f"Ajuste por sistema. Clave: {clave_cnis}, Lote: {lote.numero_lote}, Ubicación: {codigo_ubicacion}"
        mov = MovimientoInventario.objects.create(
            lote=lote,
            tipo_movimiento='SALIDA',
            cantidad=cantidad,
            cantidad_anterior=cantidad_anterior_lote,
            cantidad_nueva=cantidad_nueva_lote,
            motivo=motivo_mov,
            documento_referencia=(solicitud.folio or '')[:100] if solicitud.folio else '',
            pedido=(solicitud.folio or '')[:255] if solicitud.folio else '',
            folio=str(propuesta.id),
            institucion_destino=solicitud.institucion_solicitante,
            usuario=request.user
        )
        logger.info(
            "aplicar_salida_surtida: MovimientoInventario creado id=%s lote_id=%s cantidad=%s cantidad_anterior=%s cantidad_nueva=%s motivo=%s",
            mov.id, lote.pk, cantidad, cantidad_anterior_lote, cantidad_nueva_lote, motivo_mov[:80]
        )

    logger.info(
        "aplicar_salida_surtida: OK | afectación aplicada clave=%s lote=%s ubicacion=%s cantidad=%s",
        clave_cnis, lote.numero_lote, codigo_ubicacion, cantidad
    )
    messages.success(
        request,
        f'Salida aplicada correctamente: se descontaron {cantidad} unidades (Clave: {clave_cnis}, Lote: {lote.numero_lote}, Ubicación: {codigo_ubicacion}) y se registró el movimiento con motivo "Ajuste por sistema".'
    )
    if return_url and return_url.startswith('/') and not return_url.startswith('//'):
        return redirect(return_url)
    return redirect(reverse('reportes_salidas:reporte_salidas_surtidas'))


@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista', 'Supervisor')
def movimientos_surtimiento(request):
    """
    Muestra los movimientos de inventario generados por un registro de surtimiento
    (propuesta + lote). Recibe propuesta (uuid) y lote (pk) por GET.
    """
    propuesta_id = request.GET.get('propuesta')
    lote_id = request.GET.get('lote')
    if not propuesta_id or not lote_id:
        messages.warning(request, 'Faltan parámetros (propuesta y lote) para ver los movimientos.')
        return redirect('reportes_salidas:reporte_salidas_surtidas')
    movimientos = MovimientoInventario.objects.filter(
        lote_id=lote_id,
        tipo_movimiento='SALIDA',
        folio=str(propuesta_id)
    ).select_related('lote', 'lote__producto', 'usuario').order_by('-fecha_movimiento')
    # Contexto para el breadcrumb / volver
    try:
        lote_obj = Lote.objects.select_related('producto').get(pk=lote_id)
        lote_numero = lote_obj.numero_lote
        producto_desc = (lote_obj.producto.descripcion or '')[:60] if lote_obj.producto else ''
    except Lote.DoesNotExist:
        lote_numero = ''
        producto_desc = ''
    context = {
        'movimientos': movimientos,
        'propuesta_id': propuesta_id,
        'lote_id': lote_id,
        'lote_numero': lote_numero,
        'producto_desc': producto_desc,
    }
    return render(request, 'inventario/reportes_salidas/movimientos_surtimiento.html', context)


@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista', 'Supervisor')
def exportar_salidas_surtidas_excel(request):
    """
    Exporta el reporte de salidas surtidas a Excel.
    """
    # Aplicar los mismos filtros que la vista principal
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    folio = request.GET.get('folio', '').strip()
    clave_cnis = request.GET.get('clave_cnis', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    institucion_id = request.GET.get('institucion')
    
    # Obtener propuestas surtidas (misma lógica que la vista)
    # Solo propuestas que tienen fecha_surtimiento (realmente surtidas)
    propuestas = PropuestaPedido.objects.filter(
        estado='SURTIDA',
        fecha_surtimiento__isnull=False
    ).select_related(
        'solicitud',
        'solicitud__institucion_solicitante',
        'solicitud__almacen_destino',
        'usuario_surtimiento'
    ).prefetch_related(
        'items__producto',
        'items__item_solicitud',
        'items__lotes_asignados__lote_ubicacion__lote',
        'items__lotes_asignados__lote_ubicacion__ubicacion'
    )
    
    # Aplicar filtros (misma lógica que la vista principal)
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            propuestas = propuestas.filter(fecha_surtimiento__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            from datetime import time as dt_time
            fecha_fin_obj = datetime.combine(fecha_fin_obj.date(), dt_time.max)
            propuestas = propuestas.filter(fecha_surtimiento__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    if folio:
        folio = folio.strip()
        if folio:
            propuestas = propuestas.filter(
                Q(solicitud__folio__icontains=folio) |
                Q(solicitud__observaciones_solicitud__icontains=folio)
            )
    
    if clave_cnis:
        clave_cnis = clave_cnis.strip()
        if clave_cnis:
            # Filtrar propuestas que tengan items con productos que coincidan con la clave CNIS
            propuestas = propuestas.filter(
                items__producto__clave_cnis__icontains=clave_cnis
            ).distinct()
    
    if filtro_lote:
        propuestas = propuestas.filter(
            items__lotes_asignados__surtido=True,
            items__lotes_asignados__lote_ubicacion__lote__numero_lote__icontains=filtro_lote
        ).distinct()
    
    if institucion_id:
        try:
            institucion_id = int(institucion_id)
            propuestas = propuestas.filter(
                solicitud__institucion_solicitante_id=institucion_id
            )
        except (ValueError, TypeError):
            pass
    
    # Construir datos del reporte
    datos_reporte = []
    partida_counter = 1
    
    # Normalizar filtros para comparación (guardar valores originales antes de modificar)
    clave_cnis_filter_value = clave_cnis.strip().upper() if clave_cnis else None
    lote_filter_value = filtro_lote.strip().upper() if filtro_lote else None
    
    for propuesta in propuestas.order_by('-fecha_surtimiento', 'solicitud__folio'):
        solicitud = propuesta.solicitud
        
        for item_propuesta in propuesta.items.all():
            producto = item_propuesta.producto
            
            # Aplicar filtro de clave_cnis a nivel de item
            if clave_cnis_filter_value:
                if not producto.clave_cnis or clave_cnis_filter_value not in producto.clave_cnis.upper():
                    continue
            
            # Obtener lotes asignados surtidos
            lotes_surtidos = item_propuesta.lotes_asignados.filter(surtido=True)
            
            if not lotes_surtidos.exists():
                continue
            
            for lote_asignado in lotes_surtidos:
                lote_ubicacion = lote_asignado.lote_ubicacion
                lote = lote_ubicacion.lote
                ubicacion = lote_ubicacion.ubicacion
                
                # Aplicar filtro de lote a nivel de lote individual
                if lote_filter_value:
                    if not lote.numero_lote or lote_filter_value not in lote.numero_lote.upper():
                        continue
                
                # Movimiento (para remision_ingreso)
                movimiento = MovimientoInventario.objects.filter(
                    lote=lote,
                    tipo_movimiento='SALIDA',
                    folio=str(propuesta.id) if propuesta.id else ''
                ).filter(cantidad=lote_asignado.cantidad_asignada).first()
                if not movimiento:
                    movimiento = MovimientoInventario.objects.filter(
                        lote=lote,
                        tipo_movimiento='SALIDA',
                        folio=str(propuesta.id) if propuesta.id else ''
                    ).first()
                # Cantidad previa = cantidad disponible actual + cantidad surtida
                cantidad_previa = (lote.cantidad_disponible or 0) + lote_asignado.cantidad_asignada
                tiene_movimiento = movimiento is not None
                orden_reposicion = lote.orden_suministro.numero_orden if lote.orden_suministro else ''
                
                datos_reporte.append({
                    'partida': partida_counter,
                    'clave_cnis': producto.clave_cnis,
                    'descripcion': producto.descripcion,
                    'unidad_medida': producto.unidad_medida if producto.unidad_medida else '',
                    'lote': lote.numero_lote,
                    'caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else '',
                    'cantidad_solicitada': item_propuesta.cantidad_solicitada,
                    'cantidad_previa': cantidad_previa,
                    'cantidad_surtida': lote_asignado.cantidad_asignada,
                    'observaciones': solicitud.observaciones_solicitud or '',
                    'recurso': solicitud.institucion_solicitante.nombre if solicitud.institucion_solicitante else '',
                    'destino': (solicitud.almacen_destino.institucion.nombre or solicitud.almacen_destino.institucion.denominacion) if solicitud.almacen_destino and solicitud.almacen_destino.institucion else '',
                    'ubicacion': ubicacion.codigo if ubicacion else '',
                    'fecha_captura': solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_solicitud else '',
                    'folio': solicitud.observaciones_solicitud or solicitud.folio,
                    'fecha_entrega_programada': solicitud.fecha_entrega_programada.strftime('%d/%m/%Y') if solicitud.fecha_entrega_programada else '',
                    'status': propuesta.get_estado_display(),
                    'estatus_movimiento': 'Con movimiento' if tiene_movimiento else 'Sin movimiento - Revisar',
                    'remision_ingreso': movimiento.remision if movimiento else '',
                    'orden_reposicion': orden_reposicion,
                    'usuario': propuesta.usuario_surtimiento.get_full_name() if propuesta.usuario_surtimiento else propuesta.usuario_surtimiento.username if propuesta.usuario_surtimiento else '',
                })
                
                partida_counter += 1
    
    # Aplicar filtro por estatus movimiento si viene en la URL (mismo criterio que la vista)
    estatus_movimiento_exp = request.GET.get('estatus_movimiento', '').strip()
    if estatus_movimiento_exp == 'sin_movimiento':
        datos_reporte = [d for d in datos_reporte if d.get('estatus_movimiento') == 'Sin movimiento - Revisar']
    elif estatus_movimiento_exp == 'con_movimiento':
        datos_reporte = [d for d in datos_reporte if d.get('estatus_movimiento') == 'Con movimiento']
    for i, d in enumerate(datos_reporte, 1):
        d['partida'] = i
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Salidas Surtidas"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Encabezados
    headers = [
        'PARTIDA', 'CLAVE (CNIS)', 'DESCRIPCION', 'UNIDAD DE MEDIDA', 'LOTE',
        'CADUCIDAD', 'CANTIDAD SOLICITADA', 'CANT. PREVIA AL SURTIMIENTO', 'CANTIDAD SURTIDA', 'OBSERVACIONES',
        'RECURSO', 'DESTINO', 'UBICACIÓN', 'FECHA CAPTURA', 'FOLIO',
        'FECHA ENTREGA PROGRAMADA', 'STATUS', 'REMISION DE INGRESO',
        'ORDEN DE REPOSICION', 'USUARIO'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos
    for row_num, dato in enumerate(datos_reporte, 2):
        ws.cell(row=row_num, column=1, value=dato['partida'])
        ws.cell(row=row_num, column=2, value=dato['clave_cnis'])
        ws.cell(row=row_num, column=3, value=dato['descripcion'])
        ws.cell(row=row_num, column=4, value=dato['unidad_medida'])
        ws.cell(row=row_num, column=5, value=dato['lote'])
        ws.cell(row=row_num, column=6, value=dato['caducidad'])
        ws.cell(row=row_num, column=7, value=dato['cantidad_solicitada'])
        ws.cell(row=row_num, column=8, value=dato.get('cantidad_previa'))
        ws.cell(row=row_num, column=9, value=dato['cantidad_surtida'])
        ws.cell(row=row_num, column=10, value=dato['observaciones'])
        ws.cell(row=row_num, column=11, value=dato['recurso'])
        ws.cell(row=row_num, column=12, value=dato['destino'])
        ws.cell(row=row_num, column=13, value=dato['ubicacion'])
        ws.cell(row=row_num, column=14, value=dato['fecha_captura'])
        ws.cell(row=row_num, column=15, value=dato['folio'])
        ws.cell(row=row_num, column=16, value=dato['fecha_entrega_programada'])
        ws.cell(row=row_num, column=17, value=dato['status'])
        ws.cell(row=row_num, column=18, value=dato['remision_ingreso'])
        ws.cell(row=row_num, column=19, value=dato['orden_reposicion'])
        ws.cell(row=row_num, column=20, value=dato['usuario'])
        
        # Aplicar bordes
        for col_num in range(1, 21):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Ajustar ancho de columnas
    column_widths = [10, 15, 50, 15, 15, 12, 18, 22, 18, 30, 30, 25, 15, 18, 20, 22, 15, 20, 20, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_salidas_surtidas_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# ============================================================
# REPORTE DE RESERVAS
# ============================================================

@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista', 'Supervisor')
def reporte_reservas(request):
    """
    Reporte detallado de todas las reservas activas (LoteAsignado con surtido=False).
    Permite al usuario validar y liberar reservas si es necesario.
    """
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    folio = request.GET.get('folio', '').strip()
    clave_cnis = request.GET.get('clave_cnis', '').strip()
    institucion_id = request.GET.get('institucion')
    estado_propuesta = request.GET.get('estado_propuesta', '')
    
    # Obtener todas las reservas activas (LoteAsignado con surtido=False)
    reservas = LoteAsignado.objects.filter(
        surtido=False
    ).select_related(
        'item_propuesta__propuesta__solicitud',
        'item_propuesta__propuesta__solicitud__institucion_solicitante',
        'item_propuesta__propuesta__solicitud__almacen_destino',
        'item_propuesta__producto',
        'lote_ubicacion__lote',
        'lote_ubicacion__lote__producto',
        'lote_ubicacion__ubicacion'
    )
    
    # Aplicar filtros
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            reservas = reservas.filter(fecha_asignacion__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            from datetime import time as dt_time
            fecha_fin_obj = datetime.combine(fecha_fin_obj.date(), dt_time.max)
            reservas = reservas.filter(fecha_asignacion__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    if folio:
        reservas = reservas.filter(
            Q(item_propuesta__propuesta__solicitud__folio__icontains=folio) |
            Q(item_propuesta__propuesta__solicitud__observaciones_solicitud__icontains=folio)
        )
    
    if clave_cnis:
        reservas = reservas.filter(
            item_propuesta__producto__clave_cnis__icontains=clave_cnis
        )
    
    if institucion_id:
        reservas = reservas.filter(
            item_propuesta__propuesta__solicitud__institucion_solicitante_id=institucion_id
        )
    
    if estado_propuesta:
        reservas = reservas.filter(
            item_propuesta__propuesta__estado=estado_propuesta
        )
    
    # Construir datos del reporte
    datos_reporte = []
    partida_counter = 1
    
    for reserva in reservas.order_by('-fecha_asignacion', 'item_propuesta__propuesta__solicitud__folio'):
        propuesta = reserva.item_propuesta.propuesta
        solicitud = propuesta.solicitud
        producto = reserva.item_propuesta.producto
        lote_ubicacion = reserva.lote_ubicacion
        lote = lote_ubicacion.lote
        ubicacion = lote_ubicacion.ubicacion
        
        # Calcular días para caducidad
        if lote.fecha_caducidad:
            dias_caducidad = (lote.fecha_caducidad - timezone.now().date()).days
        else:
            dias_caducidad = None
        
        datos_reporte.append({
            'id': reserva.id,
            'partida': partida_counter,
            'clave_cnis': producto.clave_cnis,
            'descripcion': producto.descripcion,
            'unidad_medida': producto.unidad_medida if producto.unidad_medida else '',
            'lote': lote.numero_lote,
            'caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else '',
            'dias_caducidad': dias_caducidad,
            'cantidad_reservada': reserva.cantidad_asignada,
            'cantidad_solicitada': reserva.item_propuesta.cantidad_solicitada,
            'observaciones': solicitud.observaciones_solicitud or '',
            'recurso': solicitud.institucion_solicitante.nombre if solicitud.institucion_solicitante else '',
            'destino': (solicitud.almacen_destino.institucion.nombre or solicitud.almacen_destino.institucion.denominacion) if solicitud.almacen_destino and solicitud.almacen_destino.institucion else '',
            'ubicacion': ubicacion.codigo if ubicacion else '',
            'fecha_reserva': reserva.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if reserva.fecha_asignacion else '',
            'folio': solicitud.observaciones_solicitud or solicitud.folio,
            'fecha_entrega_programada': solicitud.fecha_entrega_programada.strftime('%d/%m/%Y') if solicitud.fecha_entrega_programada else '',
            'estado_propuesta': propuesta.get_estado_display(),
            'estado_propuesta_codigo': propuesta.estado,
            'propuesta_id': propuesta.id,
        })
        
        partida_counter += 1
    
    # Paginación
    paginator = Paginator(datos_reporte, 50)
    page = request.GET.get('page', 1)
    try:
        datos_paginados = paginator.page(page)
    except PageNotAnInteger:
        datos_paginados = paginator.page(1)
    except EmptyPage:
        datos_paginados = paginator.page(paginator.num_pages)
    
    # Obtener instituciones para el filtro
    instituciones = Institucion.objects.all().order_by('nombre')
    
    # Estados de propuesta para el filtro
    estados_propuesta = [
        ('GENERADA', 'Generada'),
        ('REVISADA', 'Revisada'),
        ('EN_SURTIMIENTO', 'En Surtimiento'),
        ('SURTIDA', 'Surtida'),
        ('CANCELADA', 'Cancelada'),
    ]
    
    context = {
        'datos': datos_paginados,
        'total_registros': len(datos_reporte),
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'folio': folio,
        'clave_cnis': clave_cnis,
        'institucion_id': institucion_id,
        'estado_propuesta': estado_propuesta,
        'instituciones': instituciones,
        'estados_propuesta': estados_propuesta,
    }
    
    return render(request, 'inventario/reportes_salidas/reporte_reservas.html', context)


@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Supervisor')
@require_http_methods(["POST"])
@transaction.atomic
def liberar_reserva(request, reserva_id):
    """
    Libera una reserva específica (LoteAsignado).
    Elimina el LoteAsignado y libera la cantidad reservada.
    """
    try:
        reserva = get_object_or_404(
            LoteAsignado.objects.select_related(
                'lote_ubicacion__lote',
                'item_propuesta__propuesta'
            ),
            id=reserva_id,
            surtido=False  # Solo se pueden liberar reservas no surtidas
        )
        
        propuesta = reserva.item_propuesta.propuesta
        
        # Verificar que la propuesta esté en un estado que permita liberar
        estados_liberables = ['GENERADA', 'REVISADA', 'EN_SURTIMIENTO']
        if propuesta.estado not in estados_liberables:
            return JsonResponse({
                'exito': False,
                'mensaje': f'No se puede liberar la reserva. La propuesta está en estado "{propuesta.get_estado_display()}"'
            }, status=400)
        
        lote_ubicacion = reserva.lote_ubicacion
        lote = lote_ubicacion.lote
        cantidad_liberar = reserva.cantidad_asignada
        item_propuesta = reserva.item_propuesta  # Guardar referencia antes de eliminar
        propuesta = reserva.item_propuesta.propuesta
        solicitud = propuesta.solicitud if propuesta else None
        
        # Obtener cantidades antes de liberar para el movimiento
        cantidad_disponible_anterior = lote.cantidad_disponible
        cantidad_reservada_anterior = lote_ubicacion.cantidad_reservada
        
        # Liberar la cantidad reservada
        resultado_liberacion = liberar_cantidad_lote(lote_ubicacion, cantidad_liberar)
        
        # Refrescar para obtener valores actualizados
        lote_ubicacion.refresh_from_db()
        lote.refresh_from_db()
        
        # Crear movimiento de inventario para registrar la liberación
        # Nota: La cantidad disponible no cambia al liberar una reserva,
        # solo se libera la cantidad reservada, pero registramos el movimiento para auditoría
        MovimientoInventario.objects.create(
            lote=lote,
            tipo_movimiento='AJUSTE_POSITIVO',
            cantidad=cantidad_liberar,
            cantidad_anterior=cantidad_disponible_anterior,
            cantidad_nueva=cantidad_disponible_anterior,  # La cantidad disponible no cambia al liberar reserva
            motivo=f"Liberación de reserva - Se liberaron {cantidad_liberar} unidades reservadas. Propuesta: {propuesta.id.hex[:8] if propuesta else 'N/A'}, Folio: {solicitud.folio if solicitud and solicitud.folio else 'N/A'}",
            documento_referencia=solicitud.folio if solicitud and solicitud.folio else '',
            pedido=solicitud.folio if solicitud and solicitud.folio else '',
            folio=str(propuesta.id) if propuesta else '',
            institucion_destino=solicitud.institucion_solicitante if solicitud else None,
            usuario=request.user
        )
        
        # Eliminar el LoteAsignado
        reserva.delete()
        
        # Actualizar cantidad_propuesta del item si es necesario
        item_propuesta.refresh_from_db()
        
        # Recalcular cantidad_propuesta sumando las cantidades asignadas restantes (excluyendo la que se eliminó)
        cantidad_restante = item_propuesta.lotes_asignados.filter(surtido=False).aggregate(
            total=Sum('cantidad_asignada')
        )['total'] or 0
        
        item_propuesta.cantidad_propuesta = cantidad_restante
        item_propuesta.save(update_fields=['cantidad_propuesta'])
        
        messages.success(
            request,
            f'Reserva liberada exitosamente. Se liberaron {cantidad_liberar} unidades del lote {lote_ubicacion.lote.numero_lote}'
        )
        
        return JsonResponse({
            'exito': True,
            'mensaje': f'Reserva liberada exitosamente. Se liberaron {cantidad_liberar} unidades.',
            'cantidad_liberada': cantidad_liberar,
            'lote': lote_ubicacion.lote.numero_lote
        })
        
    except LoteAsignado.DoesNotExist:
        return JsonResponse({
            'exito': False,
            'mensaje': 'La reserva no existe o ya fue surtida'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'exito': False,
            'mensaje': f'Error al liberar la reserva: {str(e)}'
        }, status=500)


@login_required
@requiere_rol('Administrador', 'Gestor de Inventario', 'Analista', 'Supervisor')
def exportar_reservas_excel(request):
    """
    Exporta el reporte de reservas a Excel.
    """
    # Aplicar los mismos filtros que la vista principal
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    folio = request.GET.get('folio', '').strip()
    clave_cnis = request.GET.get('clave_cnis', '').strip()
    institucion_id = request.GET.get('institucion')
    estado_propuesta = request.GET.get('estado_propuesta', '')
    
    # Obtener reservas (misma lógica que la vista)
    reservas = LoteAsignado.objects.filter(
        surtido=False
    ).select_related(
        'item_propuesta__propuesta__solicitud',
        'item_propuesta__propuesta__solicitud__institucion_solicitante',
        'item_propuesta__propuesta__solicitud__almacen_destino',
        'item_propuesta__producto',
        'lote_ubicacion__lote',
        'lote_ubicacion__lote__producto',
        'lote_ubicacion__ubicacion'
    )
    
    # Aplicar filtros (misma lógica que la vista)
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            reservas = reservas.filter(fecha_asignacion__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            from datetime import time as dt_time
            fecha_fin_obj = datetime.combine(fecha_fin_obj.date(), dt_time.max)
            reservas = reservas.filter(fecha_asignacion__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    if folio:
        reservas = reservas.filter(
            Q(item_propuesta__propuesta__solicitud__folio__icontains=folio) |
            Q(item_propuesta__propuesta__solicitud__observaciones_solicitud__icontains=folio)
        )
    
    if clave_cnis:
        reservas = reservas.filter(
            item_propuesta__producto__clave_cnis__icontains=clave_cnis
        )
    
    if institucion_id:
        reservas = reservas.filter(
            item_propuesta__propuesta__solicitud__institucion_solicitante_id=institucion_id
        )
    
    if estado_propuesta:
        reservas = reservas.filter(
            item_propuesta__propuesta__estado=estado_propuesta
        )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'PARTIDA', 'CLAVE CNIS', 'DESCRIPCIÓN', 'UNIDAD DE MEDIDA', 'LOTE', 'CADUCIDAD',
        'CANTIDAD RESERVADA', 'CANTIDAD SOLICITADA', 'OBSERVACIONES', 'RECURSO', 'DESTINO',
        'UBICACIÓN', 'FECHA RESERVA', 'FOLIO', 'FECHA ENTREGA PROGRAMADA', 'ESTADO PROPUESTA'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    partida_counter = 1
    for reserva in reservas.order_by('-fecha_asignacion', 'item_propuesta__propuesta__solicitud__folio'):
        propuesta = reserva.item_propuesta.propuesta
        solicitud = propuesta.solicitud
        producto = reserva.item_propuesta.producto
        lote = reserva.lote_ubicacion.lote
        ubicacion = reserva.lote_ubicacion.ubicacion
        
        row_num = partida_counter + 1
        data = [
            partida_counter,
            producto.clave_cnis,
            producto.descripcion,
            producto.unidad_medida or '',
            lote.numero_lote,
            lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else '',
            reserva.cantidad_asignada,
            reserva.item_propuesta.cantidad_solicitada,
            solicitud.observaciones_solicitud or '',
            solicitud.institucion_solicitante.nombre if solicitud.institucion_solicitante else '',
            (solicitud.almacen_destino.institucion.nombre or solicitud.almacen_destino.institucion.denominacion) if solicitud.almacen_destino and solicitud.almacen_destino.institucion else '',
            ubicacion.codigo if ubicacion else '',
            reserva.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if reserva.fecha_asignacion else '',
            solicitud.observaciones_solicitud or solicitud.folio,
            solicitud.fecha_entrega_programada.strftime('%d/%m/%Y') if solicitud.fecha_entrega_programada else '',
            propuesta.get_estado_display(),
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        partida_counter += 1
    
    # Ajustar anchos de columna
    column_widths = [10, 18, 50, 15, 15, 12, 18, 18, 30, 30, 30, 15, 18, 20, 20, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_reservas_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
